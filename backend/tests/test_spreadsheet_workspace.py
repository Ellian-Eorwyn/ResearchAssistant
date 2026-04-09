from __future__ import annotations

import csv
import io
import json
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from backend.models.repository import RepositoryColumnOutputConstraint
from backend.models.settings import EffectiveSettings, LLMBackendConfig
from backend.models.spreadsheets import (
    SpreadsheetColumnCreateRequest,
    SpreadsheetColumnUpdateRequest,
    SpreadsheetExportRequest,
    SpreadsheetRowPatchRequest,
    SpreadsheetSessionTargetSelectRequest,
)
from backend.storage.attached_repository import AttachedRepositoryService
from backend.storage.file_store import FileStore
from backend.storage.spreadsheet_workspace import SpreadsheetWorkspaceService


def _build_csv_bytes(rows: list[dict[str, str]]) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8")


class SpreadsheetWorkspaceServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory(prefix="sheet-workspace-test-")
        root = Path(self.temp_dir.name)
        self.store = FileStore(base_dir=root / "data")
        self.repository = AttachedRepositoryService(store=self.store)
        self.repository.create(str(root / "repo"))
        self.service = SpreadsheetWorkspaceService(
            store=self.store,
            repository_service=self.repository,
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_csv_roundtrip_with_custom_column(self) -> None:
        upload = self.service.upload_session(
            "people.csv",
            _build_csv_bytes(
                [
                    {"Name": "Ada", "Age": "37"},
                    {"Name": "Grace", "Age": "45"},
                ]
            ),
        )
        session_id = upload.session.session.session_id
        created = self.service.create_column(
            session_id,
            SpreadsheetColumnCreateRequest(label="Summary"),
        )
        self.service.patch_row(
            session_id,
            "row_000001",
            SpreadsheetRowPatchRequest(values={created.id: "Pioneer", "source_002": "38"}),
        )

        content, _headers, _media_type = self.service.export_session(
            SpreadsheetExportRequest(session_id=session_id)
        )
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        rows = list(reader)

        self.assertEqual(rows[0]["Name"], "Ada")
        self.assertEqual(rows[0]["Age"], "38")
        self.assertEqual(rows[0]["Summary"], "Pioneer")
        self.assertEqual(rows[1]["Summary"], "")

    def test_xlsx_export_preserves_other_sheets(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["Name", "Score"])
        worksheet.append(["Ada", 1])
        summary = workbook.create_sheet("Summary")
        summary["A1"] = "Keep Me"
        stream = io.BytesIO()
        workbook.save(stream)
        workbook.close()

        upload = self.service.upload_session("book.xlsx", stream.getvalue())
        session_id = upload.session.session.session_id
        self.service.patch_row(
            session_id,
            "row_000001",
            SpreadsheetRowPatchRequest(values={"source_002": 9}),
        )
        content, _headers, _media_type = self.service.export_session(
            SpreadsheetExportRequest(session_id=session_id)
        )

        exported = load_workbook(io.BytesIO(content))
        try:
            self.assertEqual(exported["Data"]["B2"].value, 9)
            self.assertEqual(exported["Summary"]["A1"].value, "Keep Me")
        finally:
            exported.close()

    def test_json_export_preserves_outer_structure(self) -> None:
        payload = {
            "meta": {"project": "alpha"},
            "items": [{"name": "Ada", "score": 1}],
            "flags": [1, 2, 3],
        }
        upload = self.service.upload_session(
            "items.json",
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        )
        session_id = upload.session.session.session_id
        self.service.patch_row(
            session_id,
            "row_000001",
            SpreadsheetRowPatchRequest(values={"source_002": 7}),
        )

        content, _headers, _media_type = self.service.export_session(
            SpreadsheetExportRequest(session_id=session_id)
        )
        exported = json.loads(content.decode("utf-8"))

        self.assertEqual(exported["meta"]["project"], "alpha")
        self.assertEqual(exported["flags"], [1, 2, 3])
        self.assertEqual(exported["items"][0]["score"], 7)

    def test_jsonl_roundtrip(self) -> None:
        payload = b'{"name":"Ada","score":1}\n{"name":"Grace","score":2}\n'
        upload = self.service.upload_session("scores.jsonl", payload)
        session_id = upload.session.session.session_id
        self.service.patch_row(
            session_id,
            "row_000002",
            SpreadsheetRowPatchRequest(values={"source_002": 3}),
        )

        content, _headers, _media_type = self.service.export_session(
            SpreadsheetExportRequest(session_id=session_id)
        )
        lines = [json.loads(line) for line in content.decode("utf-8").splitlines() if line.strip()]

        self.assertEqual(lines[1]["score"], 3)
        self.assertEqual(lines[0]["name"], "Ada")

    def test_sqlite_export_preserves_other_tables(self) -> None:
        db_path = Path(self.temp_dir.name) / "sample.sqlite"
        with sqlite3.connect(db_path) as conn:
            conn.execute("CREATE TABLE items (name TEXT, score INTEGER)")
            conn.execute("INSERT INTO items(name, score) VALUES ('Ada', 1)")
            conn.execute("CREATE TABLE notes (body TEXT)")
            conn.execute("INSERT INTO notes(body) VALUES ('unchanged')")
            conn.commit()

        upload = self.service.upload_session("sample.sqlite", db_path.read_bytes())
        session_id = upload.session.session.session_id
        items_target = next(
          target for target in upload.session.targets if target.label == "items"
        )
        self.service.select_target(
            session_id,
            SpreadsheetSessionTargetSelectRequest(target_id=items_target.id),
        )
        created = self.service.create_column(
            session_id,
            SpreadsheetColumnCreateRequest(label="Summary"),
        )
        self.service.patch_row(
            session_id,
            "row_000001",
            SpreadsheetRowPatchRequest(values={"source_002": 5, created.id: "done"}),
        )

        content, _headers, _media_type = self.service.export_session(
            SpreadsheetExportRequest(session_id=session_id)
        )
        exported_path = Path(self.temp_dir.name) / "exported.sqlite"
        exported_path.write_bytes(content)

        with sqlite3.connect(exported_path) as conn:
            score = conn.execute("SELECT score FROM items").fetchone()[0]
            summary = conn.execute('SELECT "Summary" FROM items').fetchone()[0]
            note = conn.execute("SELECT body FROM notes").fetchone()[0]

        self.assertEqual(score, 5)
        self.assertEqual(summary, "done")
        self.assertEqual(note, "unchanged")

    def test_invalid_json_shape_is_rejected(self) -> None:
        with self.assertRaisesRegex(ValueError, "array of objects"):
            self.service.upload_session(
                "bad.json",
                json.dumps({"name": "Ada", "score": 1}).encode("utf-8"),
            )

    def test_llm_run_uses_only_selected_input_columns(self) -> None:
        upload = self.service.upload_session(
            "people.csv",
            _build_csv_bytes(
                [
                    {"Name": "Ada", "Role": "Engineer", "Note": "Ignore me"},
                ]
            ),
        )
        session_id = upload.session.session.session_id
        created = self.service.create_column(
            session_id,
            SpreadsheetColumnCreateRequest(label="Summary"),
        )
        updated = self.service.update_column(
            session_id,
            created.id,
            SpreadsheetColumnUpdateRequest(
                instruction_prompt="Summarize the selected fields in one short phrase.",
                output_constraint=RepositoryColumnOutputConstraint(kind="text", max_words=4),
                input_column_ids=["source_001", "source_002"],
            ),
        )
        columns = self.service.get_session(session_id).columns
        row = self.service.list_manifest(session_id).rows[0]
        captured_prompts: list[str] = []

        def fake_chat_completion(*, user_prompt: str, **_kwargs: object) -> str:
            captured_prompts.append(user_prompt)
            return '{"value":"Ada engineer","status":"ok"}'

        settings = EffectiveSettings(
            use_llm=True,
            llm_backend=LLMBackendConfig(model="demo"),
            research_purpose="Test spreadsheet prompt selection",
        )

        with patch("backend.storage.spreadsheet_workspace.UnifiedLLMClient.sync_chat_completion", side_effect=fake_chat_completion):
            with patch("backend.storage.spreadsheet_workspace.UnifiedLLMClient.sync_close", return_value=None):
                value = self.service._generate_column_value_for_row(
                    settings=settings,
                    column=updated,
                    columns=columns,
                    row=row,
                )

        self.assertEqual(value, "Ada engineer")
        self.assertEqual(len(captured_prompts), 1)
        prompt = captured_prompts[0]
        self.assertIn('"Name": "Ada"', prompt)
        self.assertIn('"Role": "Engineer"', prompt)
        self.assertNotIn('"Note": "Ignore me"', prompt)

    def test_parquet_roundtrip_when_pyarrow_is_available(self) -> None:
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except Exception:
            self.skipTest("pyarrow is not installed")
            return

        sink = io.BytesIO()
        table = pa.Table.from_pylist([{"name": "Ada", "score": 1}])
        pq.write_table(table, sink)

        upload = self.service.upload_session("scores.parquet", sink.getvalue())
        session_id = upload.session.session.session_id
        self.service.patch_row(
            session_id,
            "row_000001",
            SpreadsheetRowPatchRequest(values={"source_002": 2}),
        )
        content, _headers, _media_type = self.service.export_session(
            SpreadsheetExportRequest(session_id=session_id)
        )
        exported = pq.read_table(io.BytesIO(content)).to_pylist()

        self.assertEqual(exported[0]["score"], 2)


if __name__ == "__main__":
    unittest.main()
