"""Attached repository service for persistent source/citation expansion."""

from __future__ import annotations

import csv
import base64
import hashlib
import io
import json
import mimetypes
import re
import shutil
import tempfile
import threading
import unicodedata
import uuid
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.models.bibliography import BibliographyArtifact, BibliographyEntry, ReferencesSection
from backend.models.agent import (
    AgentResourceContent,
    AgentResourceRecord,
    AgentRunCounts,
    AgentRunCurrentItem,
    AgentRunRecord,
    AgentRunSnapshot,
    AgentSourceContentChunk,
    AgentSourceArtifactUris,
    AgentSourceFreshness,
    AgentSourceProvenance,
    AgentSourceRecord,
)
from backend.models.citation_metadata import CitationFieldEvidence, CitationMetadata
from backend.models.export import ExportArtifact, ExportRow
from backend.models.common import ProcessingConfig
from backend.models.ingestion import IngestedDocument
from backend.models.ingestion_profiles import (
    DocumentNormalizationResult,
    IngestionProfile,
    IngestionProfileActionResponse,
    IngestionProfileListResponse,
    IngestionProfileSuggestion,
    IngestionProfileSuggestionActionResponse,
    IngestionProfileSuggestionListResponse,
)
from backend.models.repository import (
    RepositoryActionResponse,
    RepositoryBundleExportRequest,
    RepositoryCitationRisExportRequest,
    RepositoryColumnConfig,
    RepositoryColumnCreateRequest,
    RepositoryColumnOutputConstraint,
    RepositoryColumnPromptFixResponse,
    RepositoryColumnRunRequest,
    RepositoryColumnRunRowError,
    RepositoryColumnRunStartResponse,
    RepositoryColumnRunStatus,
    RepositoryColumnUpdateRequest,
    RepositoryDocumentImportDocument,
    RepositoryDocumentImportListResponse,
    RepositoryDocumentImportRecord,
    RepositoryDuplicateCandidateGroup,
    RepositoryDuplicateCandidateResponse,
    RepositoryDuplicateCandidateRow,
    RepositoryExportJobResponse,
    RepositoryHealth,
    RepositoryImportResponse,
    RepositoryMergeResponse,
    RepositoryManifestFilterRequest,
    RepositoryManifestExportRequest,
    RepositoryProcessDocumentsResponse,
    RepositoryReprocessDocumentsResponse,
    RepositorySourceBulkRisReadyResponse,
    RepositoryScanSummary,
    RepositorySourceDeleteResponse,
    RepositorySourceExportResponse,
    RepositorySourceTaskRequest,
    RepositorySourceTaskResponse,
    RepositoryStatusResponse,
)
from backend.models.project_profiles import (
    ProjectProfileGenerateResponse,
    ProjectProfileSaveResponse,
)
from backend.models.settings import AppSettings, EffectiveSettings, RepoSettings
from backend.models.sources import (
    SOURCE_MANIFEST_COLUMNS,
    SourceDownloadRequest,
    SourceDownloadStatus,
    SourcePhaseMetadata,
    SourceManifestArtifact,
    SourceManifestRow,
    SourceOutputOptions,
)
from backend.llm.client import UnifiedLLMClient
from backend.llm.prompts import (
    COLUMN_PROMPT_FIX_SYSTEM,
    COLUMN_PROMPT_FIX_USER,
    COLUMN_RUN_SYSTEM,
    COLUMN_RUN_USER,
    PROJECT_PROFILE_GENERATION_SYSTEM,
    PROJECT_PROFILE_GENERATION_USER,
)
from backend.pipeline.orchestrator import PipelineOrchestrator
from backend.pipeline.standardized_markdown import (
    NormalizedDocumentOutput,
    builtin_ingestion_profiles,
    normalize_document_to_standardized_markdown,
    standardized_metadata_filename,
    standardized_markdown_filename,
)
from backend.pipeline.source_downloader import (
    MANIFEST_DERIVED_COLUMNS,
    PHASE_CATALOG,
    PHASE_CITATION_VERIFY,
    PHASE_CLEANUP,
    PHASE_CONVERT,
    PHASE_FETCH,
    PHASE_RATING,
    PHASE_SUMMARY,
    PHASE_TITLE,
    _apply_citation_manual_overrides,
    _coerce_citation_metadata,
    _effective_manual_override_fields,
    _finalize_citation_metadata,
    _merge_citation_metadata,
    _normalize_phase_name,
    _citation_reference_url_for_row,
    _normalize_row_phase_metadata,
    _row_citation_verification_status,
    SourceDownloadOrchestrator,
    build_ris_records,
    build_manifest_record,
    build_manifest_csv,
    build_manifest_xlsx,
    clean_url_candidate,
    dedupe_url_key,
    llm_backend_ready_for_chat,
    normalize_url,
    normalize_citation_authors,
    summarize_output_rows,
)
from backend.pipeline.source_list_parser import parse_source_list_upload
from backend.pipeline.stage_bibliography import (
    build_entries_from_inline_urls,
    merge_inline_urls_into_entries,
    parse_bibliography,
)
from backend.pipeline.stage_export import write_csv
from backend.pipeline.stage_export_sqlite import build_wikiclaude_sqlite_db
from backend.pipeline.stage_ingest import run_ingestion
from backend.pipeline.stage_references import detect_references_section
from backend.parsers.text_utils import normalize_unicode
from backend.storage.file_store import FileStore
from backend.storage.project_profiles import (
    DEFAULT_PROJECT_PROFILE_FILENAME,
    resolve_project_profile_yaml,
)
from backend.storage.repository_bundle_viewer import build_repository_bundle_viewer_html

try:  # pragma: no cover - POSIX only
    import fcntl
except Exception:  # pragma: no cover - Windows fallback
    fcntl = None


SCHEMA_VERSION = 4
INTERNAL_DIR_NAME = ".ra_repo"
META_FILE_NAME = "repository.json"
STATE_FILE_NAME = "repository_state.json"
LOCK_FILE_NAME = "repository.lock"
REPO_SETTINGS_FILE_NAME = "settings.json"
INGESTION_PROFILES_FILE_NAME = "ingestion_profiles.json"
INGESTION_PROFILE_SUGGESTIONS_FILE_NAME = "ingestion_profile_suggestions.json"
BUNDLED_INGESTION_PROFILES_FILE_NAME = "bundled_ingestion_profiles.json"
PROJECT_PROFILES_DIR_NAME = "project_profiles"
REPO_JOBS_DIR_NAME = "jobs"
AGENT_RESOURCES_FILE_NAME = "agent_resources.json"
AGENT_TOKENS_FILE_NAME = "agent_tokens.json"
AGENT_IDEMPOTENCY_FILE_NAME = "agent_idempotency.json"
AGENT_AUDIT_FILE_NAME = "agent_audit.jsonl"
DOCUMENTS_DIR_NAME = "documents"
SOURCES_DIR_NAME = "sources"
MANIFEST_CSV_NAME = "manifest.csv"
MANIFEST_XLSX_NAME = "manifest.xlsx"
CITATIONS_CSV_NAME = "citations.csv"
CITATIONS_XLSX_NAME = "citations.xlsx"
REPO_JOB_PREFIX = "repo"

JOB_SEED_FILE_FIELDS = [
    "raw_file",
    "rendered_file",
    "rendered_pdf_file",
    "markdown_file",
    "llm_cleanup_file",
    "catalog_file",
    "summary_file",
    "rating_file",
    "metadata_file",
]

TRACKING_PARAM_EXACT = {"gclid", "fbclid", "msclkid"}
TRACKING_PARAM_PREFIXES = ("utm_",)
DUPLICATE_SCAN_MAX_GROUPS = 100
DUPLICATE_TITLE_STOP_WORDS = {
    "a",
    "an",
    "and",
    "at",
    "by",
    "for",
    "from",
    "in",
    "into",
    "of",
    "on",
    "or",
    "the",
    "to",
    "with",
}

FILE_FIELDS = [
    "raw_file",
    "rendered_file",
    "rendered_pdf_file",
    "markdown_file",
    "llm_cleanup_file",
    "catalog_file",
    "summary_file",
    "rating_file",
    "metadata_file",
]

REPOSITORY_BUNDLE_FILE_KINDS = ("pdf", "rendered", "html", "md")
REPOSITORY_BUNDLE_RESEARCH_COLUMNS: list[tuple[str, str]] = [
    ("title", "Title"),
    ("author_names", "Authors"),
    ("publication_date", "Publication Date"),
    ("organization_name", "Organization"),
    ("organization_type", "Organization Type"),
    ("export_url", "URL"),
    ("markdown_char_count", "Markdown Char Count"),
    ("citation_report_number", "Report Number"),
    ("document_type", "Document Type"),
    ("citation_type", "Citation Type"),
    ("export_overall_rating", "Overall Rating"),
]

SUPPORTED_DOCUMENT_IMPORT_EXTENSIONS = {".pdf", ".docx", ".md"}
SUPPORTED_SEED_IMPORT_EXTENSIONS = {".csv", ".xlsx", ".pdf", ".docx", ".md"}
SUPPORTED_MANUAL_SOURCE_EXTENSIONS = {
    ".pdf",
    ".doc",
    ".docx",
    ".html",
    ".htm",
    ".md",
    ".rtf",
    ".txt",
}


@dataclass
class _MergedScan:
    rows: list[SourceManifestRow]
    citations: list[ExportRow]
    next_source_id: int
    duplicate_urls_removed: int


class AttachedRepositoryService:
    def __init__(self, store: FileStore):
        self.store = store
        self._path: Path | None = None
        self._job_store: FileStore | None = None
        self._mutex = threading.RLock()
        self._download_thread: threading.Thread | None = None
        self._download_state = "idle"
        self._download_message = ""
        self._last_scan: RepositoryScanSummary | None = None

    @property
    def is_attached(self) -> bool:
        return self._path is not None

    @property
    def path(self) -> Path:
        if self._path is None:
            raise ValueError("No repository attached")
        return self._path

    @property
    def project_profiles_dir(self) -> Path:
        return self.path / PROJECT_PROFILES_DIR_NAME

    @property
    def documents_dir(self) -> Path:
        return self.path / DOCUMENTS_DIR_NAME

    @property
    def sources_dir(self) -> Path:
        return self.path / SOURCES_DIR_NAME

    def repo_job_store(self) -> FileStore:
        if not self.is_attached:
            raise ValueError("No repository attached")
        jobs_dir = self._repo_jobs_dir()
        if self._job_store is None or self._job_store.base_dir != jobs_dir:
            self._job_store = FileStore(
                base_dir=jobs_dir,
                sync_project_profiles=False,
            )
        return self._job_store

    def job_store_for(self, job_id: str = "") -> FileStore:
        normalized = str(job_id or "").strip()
        if self.is_attached:
            repo_store = self.repo_job_store()
            if (
                (normalized and normalized.startswith(f"{REPO_JOB_PREFIX}_"))
                or (normalized and repo_store.job_exists(normalized))
            ):
                return repo_store
        return self.store

    def job_exists(self, job_id: str) -> bool:
        return self.job_store_for(job_id).job_exists(job_id)

    # ---- Per-repo settings ----

    def load_repo_settings(self) -> RepoSettings:
        """Load settings from {repo}/.ra_repo/settings.json, returning defaults if missing."""
        if not self.is_attached:
            return RepoSettings()
        settings_path = self._internal_dir() / REPO_SETTINGS_FILE_NAME
        if not settings_path.exists():
            return RepoSettings()
        try:
            raw = json.loads(settings_path.read_text(encoding="utf-8"))
            return RepoSettings(**raw)
        except (json.JSONDecodeError, OSError, Exception):
            return RepoSettings()

    def save_repo_settings(self, settings: RepoSettings) -> None:
        """Write settings to {repo}/.ra_repo/settings.json."""
        if not self.is_attached:
            raise ValueError("No repository attached")
        settings_path = self._internal_dir() / REPO_SETTINGS_FILE_NAME
        settings_path.write_text(
            json.dumps(settings.model_dump(mode="json"), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def load_effective_settings(
        self, settings: EffectiveSettings | None = None,
    ) -> EffectiveSettings:
        """Merge app-level and repo-level settings into one object."""
        if settings is not None:
            return settings
        app_settings = self.store.load_app_settings()
        repo_settings = self.load_repo_settings()
        return EffectiveSettings.from_app_and_repo(app_settings, repo_settings)

    def generate_project_profile(
        self,
        *,
        research_purpose: str = "",
        profile_name: str = "",
        filename: str = "",
        settings: EffectiveSettings | None = None,
    ) -> ProjectProfileGenerateResponse:
        if not self.is_attached:
            raise ValueError("No repository attached")

        effective = self.load_effective_settings(settings)
        if not effective.use_llm:
            raise ValueError("Project profile generation requires LLM features to be enabled.")
        if not llm_backend_ready_for_chat(effective.llm_backend):
            raise ValueError("Project profile generation requires a configured chat-capable LLM backend.")

        template_filename = (
            effective.default_project_profile_name.strip()
            or DEFAULT_PROJECT_PROFILE_FILENAME
        )
        template_yaml = self._load_project_profile_template(template_filename)
        effective_research_purpose = (research_purpose or effective.research_purpose or "").strip()

        client = UnifiedLLMClient(effective.llm_backend)
        try:
            raw_response = client.sync_chat_completion(
                system_prompt=PROJECT_PROFILE_GENERATION_SYSTEM,
                user_prompt=PROJECT_PROFILE_GENERATION_USER.format(
                    research_purpose=effective_research_purpose or "No explicit research purpose was provided.",
                    profile_name=(profile_name or "").strip(),
                    template_yaml=template_yaml,
                ),
                response_format="json",
            ).strip()
        except Exception as exc:
            raise ValueError(f"Project profile generation failed: {type(exc).__name__}: {exc}") from exc
        finally:
            client.sync_close()

        try:
            payload = json.loads(raw_response or "{}")
        except json.JSONDecodeError as exc:
            raise ValueError("Project profile generation returned invalid JSON.") from exc
        if not isinstance(payload, dict):
            raise ValueError("Project profile generation returned an invalid payload.")

        content = str(payload.get("yaml") or "").strip()
        if not content:
            raise ValueError("Project profile generation returned an empty YAML draft.")

        resolved_profile_name = (
            (profile_name or "").strip()
            or str(payload.get("profile_name") or "").strip()
            or "Generated Project Profile"
        )
        resolved_filename = self._normalize_project_profile_filename(
            filename=filename,
            filename_stem=str(payload.get("filename_stem") or "").strip(),
            profile_name=resolved_profile_name,
        )
        return ProjectProfileGenerateResponse(
            status="generated",
            filename=resolved_filename,
            profile_name=resolved_profile_name,
            content=content.rstrip() + "\n",
        )

    def save_project_profile(
        self,
        *,
        filename: str,
        content: str,
    ) -> ProjectProfileSaveResponse:
        if not self.is_attached:
            raise ValueError("No repository attached")
        safe_name = self._normalize_project_profile_filename(
            filename=filename,
            filename_stem="",
            profile_name=Path(filename or "profile").stem,
        )
        normalized_content = str(content or "").strip()
        if not normalized_content:
            raise ValueError("Project profile content cannot be empty.")

        destination = self.project_profiles_dir / safe_name
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_text(normalized_content + "\n", encoding="utf-8")
        return ProjectProfileSaveResponse(
            status="saved",
            filename=safe_name,
            name=Path(safe_name).stem,
            content=normalized_content + "\n",
        )

    def _normalize_column_label(self, value: str, fallback: str = "New Column") -> str:
        normalized = " ".join(str(value or "").split()).strip()
        return normalized or fallback

    def _resolve_column_config_locked(
        self,
        column_id: str,
        column_configs: list[RepositoryColumnConfig],
        *,
        create_builtin: bool = False,
    ) -> tuple[RepositoryColumnConfig, int | None]:
        normalized_id = str(column_id or "").strip()
        if not normalized_id:
            raise ValueError("column_id is required")

        for index, config in enumerate(column_configs):
            builtin_key = str(config.builtin_key or config.id or "").strip()
            if config.id == normalized_id or (
                config.kind == "builtin" and builtin_key == normalized_id
            ):
                return config.model_copy(deep=True), index

        if create_builtin and normalized_id in PROCESSABLE_BUILTIN_COLUMNS:
            return (
                RepositoryColumnConfig(
                    id=normalized_id,
                    label=_manifest_column_label(normalized_id),
                    kind="builtin",
                    builtin_key=normalized_id,
                ),
                None,
            )

        raise ValueError(f"Unknown or unsupported column_id: {normalized_id}")

    def create_column(self, label: str = "") -> RepositoryColumnConfig:
        if not self.is_attached:
            raise ValueError("No repository attached")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            column_configs = _load_column_configs(state.get("column_configs", []))
            existing_ids = {config.id for config in column_configs}

            column_id = ""
            while not column_id or column_id in existing_ids:
                column_id = f"custom_{uuid.uuid4().hex[:8]}"

            config = RepositoryColumnConfig(
                id=column_id,
                label=self._normalize_column_label(label, fallback="New Column"),
                kind="custom",
            )
            column_configs.append(config)
            self._save_state_locked(
                sources=rows,
                citations=citations,
                imports=state.get("imports", []),
                column_configs=column_configs,
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(rows, citations)
            return config

    def update_column(
        self,
        column_id: str,
        *,
        patch: dict[str, Any],
    ) -> RepositoryColumnConfig:
        if not self.is_attached:
            raise ValueError("No repository attached")
        if not patch:
            raise ValueError("At least one column field is required.")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            column_configs = _load_column_configs(state.get("column_configs", []))
            config, index = self._resolve_column_config_locked(
                column_id,
                column_configs,
                create_builtin=True,
            )

            if "label" in patch and patch.get("label") is not None:
                if config.kind != "custom":
                    raise ValueError("Only custom columns can be renamed.")
                config.label = self._normalize_column_label(str(patch.get("label") or ""))

            if "instruction_prompt" in patch and patch.get("instruction_prompt") is not None:
                config.instruction_prompt = self._normalize_source_patch_multiline_text(
                    patch.get("instruction_prompt")
                )

            if "output_constraint" in patch and patch.get("output_constraint") is not None:
                candidate = patch.get("output_constraint")
                config.output_constraint = (
                    candidate
                    if isinstance(candidate, RepositoryColumnOutputConstraint)
                    else RepositoryColumnOutputConstraint.model_validate(candidate)
                )
            elif "instruction_prompt" in patch and patch.get("instruction_prompt") is not None:
                if config.instruction_prompt:
                    config.output_constraint = _infer_column_output_constraint(
                        config.instruction_prompt,
                        existing=config.output_constraint,
                    )
                else:
                    config.output_constraint = None

            if "include_row_context" in patch and patch.get("include_row_context") is not None:
                config.include_row_context = bool(patch.get("include_row_context"))

            if "include_source_text" in patch and patch.get("include_source_text") is not None:
                config.include_source_text = bool(patch.get("include_source_text"))

            if index is None:
                column_configs.append(config)
            else:
                column_configs[index] = config

            self._save_state_locked(
                sources=rows,
                citations=citations,
                imports=state.get("imports", []),
                column_configs=column_configs,
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "updated_at": _utc_now_iso(),
                }
            )
            return config

    def fix_column_prompt(
        self,
        column_id: str,
        *,
        draft_prompt: str,
    ) -> RepositoryColumnPromptFixResponse:
        if not self.is_attached:
            raise ValueError("No repository attached")

        normalized_prompt = self._normalize_source_patch_multiline_text(draft_prompt)
        if not normalized_prompt:
            raise ValueError("draft_prompt is required")

        repo_settings = self.load_effective_settings()
        if not repo_settings.use_llm or not llm_backend_ready_for_chat(repo_settings.llm_backend):
            raise ValueError("Prompt fix requires an enabled chat-capable LLM backend.")

        with self._writer_lock():
            column_configs = _load_column_configs(
                self._load_state_locked().get("column_configs", [])
            )
            config, _ = self._resolve_column_config_locked(
                column_id,
                column_configs,
                create_builtin=True,
            )

        client = UnifiedLLMClient(repo_settings.llm_backend)
        try:
            current_prompt = _effective_column_instruction_prompt(config)
            current_constraint = _effective_column_output_constraint(config)
            raw_response = client.sync_chat_completion(
                system_prompt=COLUMN_PROMPT_FIX_SYSTEM,
                user_prompt=COLUMN_PROMPT_FIX_USER.format(
                    column_label=config.label,
                    current_prompt=current_prompt,
                    current_constraint_json=json.dumps(
                        _serialize_column_output_constraint(current_constraint),
                        ensure_ascii=False,
                    ),
                    draft_prompt=normalized_prompt,
                ),
                response_format="json",
            )
        finally:
            client.sync_close()

        payload = _parse_json_object(raw_response)
        rewritten_prompt = self._normalize_source_patch_multiline_text(
            payload.get("prompt") or normalized_prompt
        )
        output_constraint = _coerce_column_output_constraint_payload(
            payload.get("output_constraint")
        ) or _infer_column_output_constraint(
            rewritten_prompt,
            existing=current_constraint,
        )
        notes = [
            str(item).strip()
            for item in payload.get("notes", [])
            if str(item or "").strip()
        ] if isinstance(payload.get("notes"), list) else []
        return RepositoryColumnPromptFixResponse(
            status="completed",
            column_id=str(config.builtin_key or config.id or column_id).strip(),
            prompt=rewritten_prompt,
            output_constraint=output_constraint,
            notes=notes,
        )

    def start_column_run(
        self,
        column_id: str,
        *,
        payload: RepositoryColumnRunRequest,
    ) -> RepositoryColumnRunStartResponse:
        if not self.is_attached:
            raise ValueError("No repository attached")

        repo_settings = self.load_effective_settings()

        with self._writer_lock():
            if self._download_thread and self._download_thread.is_alive():
                raise ValueError("A repository operation is already running")

            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            column_configs = _load_column_configs(state.get("column_configs", []))
            config, index = self._resolve_column_config_locked(
                column_id,
                column_configs,
                create_builtin=True,
            )
            requires_llm = _column_requires_llm(config)
            if requires_llm and (
                not repo_settings.use_llm
                or not llm_backend_ready_for_chat(repo_settings.llm_backend)
            ):
                raise ValueError("Column runs require an enabled chat-capable LLM backend.")
            config.instruction_prompt = self._normalize_source_patch_multiline_text(config.instruction_prompt)
            effective_prompt = _effective_column_instruction_prompt(config)
            if requires_llm and not effective_prompt:
                raise ValueError("Save instructions for this column before running it.")
            if requires_llm and config.output_constraint is None:
                config.output_constraint = _effective_column_output_constraint(config)

            records = [
                build_manifest_record(row, base_dir=self.path, column_configs=column_configs)
                for row in rows
            ]
            columns = _build_manifest_column_metadata(records, column_configs)
            row_by_id = {row.id: row.model_copy(deep=True) for row in rows}
            normalized_scope = str(payload.scope or "filtered").strip().lower()
            if normalized_scope == "filtered":
                base_records, _, _ = self._filter_manifest_records(
                    records,
                    columns,
                    **payload.filters.model_dump(mode="json"),
                    sort_by="",
                    sort_dir="",
                )
            elif normalized_scope == "selected":
                selected_ids = [
                    str(source_id or "").strip()
                    for source_id in payload.source_ids
                    if str(source_id or "").strip() in row_by_id
                ]
                selected_id_set = set(selected_ids)
                base_records = [
                    record for record in records if str(record.get("id") or "") in selected_id_set
                ]
            elif normalized_scope in {"all", "empty_only"}:
                base_records = list(records)
            else:
                raise ValueError(f"Unsupported column run scope: {normalized_scope}")

            if normalized_scope == "empty_only":
                target_records = [
                    record
                    for record in base_records
                    if not _manifest_record_has_column_value(
                        record,
                        str(config.builtin_key or config.id),
                    )
                ]
            else:
                target_records = base_records

            if not target_records:
                if normalized_scope == "selected":
                    raise ValueError("No selected repository rows are available for this column run.")
                if normalized_scope == "empty_only":
                    raise ValueError(f"No blank {config.label} cells are available for this column run.")
                raise ValueError("No repository rows match the current filters.")

            target_row_ids = [
                str(record.get("id") or "")
                for record in target_records
                if str(record.get("id") or "") in row_by_id
            ]
            populated_rows = sum(
                1
                for record in target_records
                if _manifest_record_has_column_value(
                    record,
                    str(config.builtin_key or config.id),
                )
            )
            normalized_column_id = str(config.builtin_key or config.id or column_id).strip()

            if populated_rows > 0 and not payload.confirm_overwrite:
                return RepositoryColumnRunStartResponse(
                    job_id="",
                    status="confirmation_required",
                    column_id=normalized_column_id,
                    total_rows=len(target_row_ids),
                    populated_rows=populated_rows,
                    message=(
                        f"{populated_rows} matching row(s) already have values in "
                        f"{config.label}. Confirm overwrite to continue."
                    ),
                )

            if index is None:
                column_configs.append(config)
            else:
                column_configs[index] = config
            column_configs[
                next(
                    idx
                    for idx, item in enumerate(column_configs)
                    if item.id == config.id
                )
            ].last_run_status = "running"

            self._save_state_locked(
                sources=rows,
                citations=citations,
                imports=state.get("imports", []),
                column_configs=column_configs,
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "updated_at": _utc_now_iso(),
                }
            )

            job_store = self.repo_job_store()
            job_id = job_store.create_job(prefix=REPO_JOB_PREFIX)
            initial_status = RepositoryColumnRunStatus(
                job_id=job_id,
                column_id=normalized_column_id,
                column_label=config.label,
                state="pending",
                total_rows=len(target_row_ids),
                message=f"Queued {len(target_row_ids)} row(s) for {config.label}.",
            )
            job_store.save_artifact(
                job_id,
                "repo_column_run_context",
                {
                    "column_id": normalized_column_id,
                    "column_label": config.label,
                    "target_row_ids": target_row_ids,
                    "scope": normalized_scope,
                    "filters": payload.filters.model_dump(mode="json"),
                    "repository_path": str(self.path),
                },
            )
            job_store.save_source_status(job_id, initial_status.model_dump(mode="json"))

            self._download_state = "running"
            self._download_message = initial_status.message
            self._download_thread = threading.Thread(
                target=self._repository_column_run_worker,
                args=(job_id, normalized_column_id, target_row_ids, repo_settings),
                daemon=True,
            )
            self._download_thread.start()

            return RepositoryColumnRunStartResponse(
                job_id=job_id,
                status="started",
                column_id=normalized_column_id,
                total_rows=len(target_row_ids),
                populated_rows=populated_rows,
                message=initial_status.message,
            )

    def get_column_run_status(self, job_id: str) -> RepositoryColumnRunStatus:
        if not self.is_attached:
            raise ValueError("No repository attached")
        normalized_job_id = str(job_id or "").strip()
        if not normalized_job_id:
            raise ValueError("job_id is required")
        raw_status = self.job_store_for(normalized_job_id).get_source_status(normalized_job_id)
        if raw_status is None:
            raise ValueError("Column run status not found.")
        return RepositoryColumnRunStatus.model_validate(raw_status)

    def _repository_column_run_worker(
        self,
        job_id: str,
        column_id: str,
        source_ids: list[str],
        repo_settings: EffectiveSettings,
    ) -> None:
        job_store = self.job_store_for(job_id)
        raw_status = job_store.get_source_status(job_id) or {}
        try:
            status = RepositoryColumnRunStatus.model_validate(raw_status)
        except Exception:
            status = RepositoryColumnRunStatus(
                job_id=job_id,
                column_id=column_id,
                total_rows=len(source_ids),
            )

        status.state = "running"
        status.started_at = status.started_at or _utc_now_iso()
        status.message = f"Running {status.column_label or column_id} for {len(source_ids)} row(s)."
        job_store.save_source_status(job_id, status.model_dump(mode="json"))

        try:
            for source_id in source_ids:
                with self._writer_lock():
                    state = self._load_state_locked()
                    rows = _load_source_rows(state.get("sources", []))
                    column_configs = _load_column_configs(state.get("column_configs", []))
                    row = next((item for item in rows if item.id == source_id), None)
                    if row is None:
                        raise ValueError(f"Unknown source_id: {source_id}")
                    config, _ = self._resolve_column_config_locked(
                        column_id,
                        column_configs,
                        create_builtin=True,
                    )
                    manifest_record = build_manifest_record(
                        row,
                        base_dir=self.path,
                        column_configs=column_configs,
                    )

                status.current_source_id = source_id
                status.current_source_title = str(manifest_record.get("title") or "")
                status.message = f"Running {status.processed_rows + 1} of {status.total_rows}."
                job_store.save_source_status(job_id, status.model_dump(mode="json"))

                try:
                    if _column_requires_llm(config):
                        next_value = self._generate_column_value_for_row(
                            config=config,
                            row=row,
                            manifest_record=manifest_record,
                            repo_settings=repo_settings,
                        )
                        self.update_source(
                            source_id,
                            patch=_column_value_patch_for_field(config, next_value),
                        )
                    else:
                        self._run_non_llm_column_for_row(
                            config=config,
                            source_id=source_id,
                        )
                    status.succeeded_rows += 1
                except Exception as exc:
                    status.failed_rows += 1
                    if len(status.row_errors) < 10:
                        status.row_errors.append(
                            RepositoryColumnRunRowError(
                                source_id=source_id,
                                message=str(exc),
                            )
                        )
                status.processed_rows += 1
                job_store.save_source_status(job_id, status.model_dump(mode="json"))

            status.state = "failed" if status.failed_rows == status.total_rows else "completed"
            status.current_source_id = ""
            status.current_source_title = ""
            status.completed_at = _utc_now_iso()
            status.message = (
                f"Processed {status.total_rows} row(s): "
                f"{status.succeeded_rows} succeeded, {status.failed_rows} failed."
            )
            job_store.save_source_status(job_id, status.model_dump(mode="json"))

            with self._writer_lock():
                state = self._load_state_locked()
                rows = _load_source_rows(state.get("sources", []))
                citations = _load_citation_rows(state.get("citations", []))
                column_configs = _load_column_configs(state.get("column_configs", []))
                config, index = self._resolve_column_config_locked(
                    column_id,
                    column_configs,
                    create_builtin=True,
                )
                config.last_run_at = status.completed_at
                config.last_run_status = "failed" if status.state == "failed" else "completed"
                if index is None:
                    column_configs.append(config)
                else:
                    column_configs[index] = config
                self._save_state_locked(
                    sources=rows,
                    citations=citations,
                    imports=state.get("imports", []),
                    column_configs=column_configs,
                )
                self._save_meta_locked(
                    {
                        **self._load_meta_locked(),
                        "updated_at": _utc_now_iso(),
                    }
                )
        finally:
            raw_status = job_store.get_source_status(job_id) or {}
            final_state = str(raw_status.get("state") or "completed").strip().lower()
            final_message = str(raw_status.get("message") or "Repository column run completed")
            with self._mutex:
                self._download_state = "failed" if final_state == "failed" else "completed"
                self._download_message = final_message

    def _generate_column_value_for_row(
        self,
        *,
        config: RepositoryColumnConfig,
        row: SourceManifestRow,
        manifest_record: dict[str, str | int | float | bool],
        repo_settings: EffectiveSettings,
    ) -> str:
        instruction_prompt = _effective_column_instruction_prompt(config)
        output_constraint = _effective_column_output_constraint(config) or RepositoryColumnOutputConstraint(
            kind="text",
            fallback_value="",
        )
        current_value = _cell_string_value(
            manifest_record.get(str(config.builtin_key or config.id))
        )
        include_source_text = _effective_column_include_source_text(config)
        include_row_context = _effective_column_include_row_context(config)
        source_text = ""
        if include_source_text:
            source_text = self._load_repository_text_artifact(row, "llm_cleanup_file")
            if not source_text:
                source_text = self._load_repository_text_artifact(row, "markdown_file")
            source_text = _truncate_column_run_text(
                source_text,
                int(repo_settings.llm_backend.max_source_chars or 0),
            )

        row_metadata: dict[str, str | int | float | bool] = {}
        if include_row_context:
            row_metadata = {
                key: value
                for key, value in manifest_record.items()
                if key
                not in {
                    "raw_file",
                    "rendered_file",
                    "rendered_pdf_file",
                    "markdown_file",
                    "llm_cleanup_file",
                    "catalog_file",
                    "summary_file",
                    "rating_file",
                    "metadata_file",
                    "rating_raw_json",
                    "citation_field_evidence_json",
                }
            }
        hard_rules = "\n".join(
            [
                "- Output only the target cell value.",
                "- No markdown.",
                "- No explanations.",
                "- Keep the value single-cell-safe.",
                f"- Use `{output_constraint.fallback_value}` when evidence is insufficient."
                if output_constraint.fallback_value
                else "- Use a blank value when evidence is insufficient.",
            ]
        )

        client = UnifiedLLMClient(repo_settings.llm_backend)
        try:
            raw_response = client.sync_chat_completion(
                system_prompt=COLUMN_RUN_SYSTEM,
                user_prompt=COLUMN_RUN_USER.format(
                    research_purpose=repo_settings.research_purpose or "",
                    column_label=config.label,
                    column_prompt=instruction_prompt,
                    output_constraint_json=json.dumps(
                        _serialize_column_output_constraint(output_constraint),
                        ensure_ascii=False,
                    ),
                    hard_rules=hard_rules,
                    current_value=current_value,
                    row_metadata_json=json.dumps(row_metadata, ensure_ascii=False, indent=2),
                    document_text=source_text or "",
                ),
                response_format="json",
            )
        finally:
            client.sync_close()

        payload = _parse_json_object(raw_response)
        normalized_status = str(payload.get("status") or "").strip().lower()
        normalized_value = _coerce_column_output_value(
            payload.get("value"),
            output_constraint,
        )
        normalized_value = _column_run_author_fallback(
            config=config,
            value=normalized_value,
            row=row,
            manifest_record=manifest_record,
        )
        if normalized_status == "insufficient_evidence" and not normalized_value:
            return str(output_constraint.fallback_value or "")
        if not normalized_value and normalized_status == "insufficient_evidence":
            return normalized_value
        if normalized_status not in {"ok", "insufficient_evidence"} and not normalized_value:
            return str(output_constraint.fallback_value or "")
        return normalized_value

    def _run_non_llm_column_for_row(
        self,
        *,
        config: RepositoryColumnConfig,
        source_id: str,
    ) -> None:
        target_key = str(config.builtin_key or config.id).strip()
        if target_key == "citation_ready":
            self.refresh_source_citation_readiness(source_id)
            return
        raise ValueError(f"Unsupported non-LLM column: {target_key}")

    def update_source(
        self,
        source_id: str,
        *,
        patch: dict[str, Any],
    ) -> dict[str, Any]:
        if not self.is_attached:
            raise ValueError("No repository attached")

        normalized_id = str(source_id or "").strip()
        if not normalized_id:
            raise ValueError("source_id is required")

        requested_fields = {key for key, value in (patch or {}).items() if key and value is not ...}
        if not requested_fields:
            raise ValueError("At least one editable field is required.")

        metadata_fields = {
            "title",
            "author_names",
            "publication_date",
            "publication_year",
            "document_type",
            "organization_name",
            "organization_type",
            "tags_text",
            "notes",
        }
        rating_fields = {
            "overall_relevance",
            "depth_score",
            "relevant_detail_score",
            "rating_rationale",
            "relevant_sections",
        }
        citation_fields = {
            "citation_title",
            "citation_authors",
            "citation_issued",
            "citation_type",
            "citation_url",
            "citation_publisher",
            "citation_container_title",
            "citation_volume",
            "citation_issue",
            "citation_pages",
            "citation_doi",
            "citation_report_number",
            "citation_standard_number",
            "citation_language",
            "citation_accessed",
        }

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            column_configs = _load_column_configs(state.get("column_configs", []))
            row_index = next((index for index, item in enumerate(rows) if item.id == normalized_id), -1)
            if row_index < 0:
                raise ValueError(f"Unknown source_id: {normalized_id}")

            row = rows[row_index].model_copy(deep=True)

            if requested_fields.intersection(metadata_fields):
                for field_name in metadata_fields:
                    if field_name not in requested_fields:
                        continue
                    setattr(row, field_name, self._normalize_source_patch_text(patch.get(field_name)))
                if "publication_date" in requested_fields and "publication_year" not in requested_fields:
                    row.publication_year = _publication_year_from_date_text(row.publication_date)
                self._write_catalog_patch(row)

            if requested_fields.intersection(citation_fields) or "citation_override_fields" in requested_fields:
                override_fields = patch.get("citation_override_fields")
                self._write_citation_patch(
                    row,
                    patch={
                        key: patch.get(key)
                        for key in citation_fields
                        if key in requested_fields
                    },
                    override_fields=override_fields if isinstance(override_fields, list) else [],
                )

            if "summary_text" in requested_fields:
                self._write_summary_patch(
                    row,
                    self._normalize_source_patch_multiline_text(patch.get("summary_text")),
                )

            if requested_fields.intersection(rating_fields):
                self._write_rating_patch(
                    row,
                    patch={
                        key: patch.get(key)
                        for key in rating_fields
                        if key in requested_fields
                    },
                )

            if "custom_fields" in requested_fields:
                custom_updates = patch.get("custom_fields") or {}
                if not isinstance(custom_updates, dict):
                    raise ValueError("custom_fields must be an object.")
                allowed_custom_ids = {
                    config.id for config in column_configs if config.kind == "custom"
                }
                next_custom_fields = dict(row.custom_fields or {})
                for field_name, raw_value in custom_updates.items():
                    normalized_field_name = str(field_name or "").strip()
                    if not normalized_field_name:
                        continue
                    if normalized_field_name not in allowed_custom_ids:
                        raise ValueError(f"Unknown custom column: {normalized_field_name}")
                    normalized_value = self._normalize_source_patch_text(raw_value)
                    if normalized_value:
                        next_custom_fields[normalized_field_name] = normalized_value
                    else:
                        next_custom_fields.pop(normalized_field_name, None)
                row.custom_fields = next_custom_fields

            self._write_repository_source_metadata(row)
            rows[row_index] = row
            rows = self._sort_rows(rows)
            self._save_state_locked(
                sources=rows,
                citations=citations,
                imports=state.get("imports", []),
                column_configs=column_configs,
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "next_source_id": _next_source_id_from_rows(rows),
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(rows, citations)
            return build_manifest_record(
                row,
                base_dir=self.path,
                column_configs=column_configs,
            )

    def refresh_source_citation_readiness(self, source_id: str) -> dict[str, Any]:
        if not self.is_attached:
            raise ValueError("No repository attached")

        normalized_id = str(source_id or "").strip()
        if not normalized_id:
            raise ValueError("source_id is required")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            column_configs = _load_column_configs(state.get("column_configs", []))
            row_index = next((index for index, item in enumerate(rows) if item.id == normalized_id), -1)
            if row_index < 0:
                raise ValueError(f"Unknown source_id: {normalized_id}")

            row = rows[row_index].model_copy(deep=True)
            catalog_payload = self._load_repository_json_artifact(row, "catalog_file")
            citation = _coerce_citation_metadata(catalog_payload.get("citation"))
            citation = _merge_citation_metadata(
                citation,
                self._row_citation_defaults_payload(row),
                overwrite_existing=False,
            )
            markdown_path = self._resolve_source_file_path_for_kind(row, "md")
            citation.verification_content_digest = _file_sha256(markdown_path)
            citation = _finalize_citation_metadata(citation)
            self._write_catalog_citation_payload(
                row,
                catalog_payload=catalog_payload,
                citation=citation,
            )

            self._write_repository_source_metadata(row)
            rows[row_index] = row
            rows = self._sort_rows(rows)
            self._save_state_locked(
                sources=rows,
                citations=citations,
                imports=state.get("imports", []),
                column_configs=column_configs,
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "next_source_id": _next_source_id_from_rows(rows),
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(rows, citations)
            return build_manifest_record(
                row,
                base_dir=self.path,
                column_configs=column_configs,
            )

    def bulk_mark_sources_ris_ready(
        self,
        source_ids: list[str],
    ) -> RepositorySourceBulkRisReadyResponse:
        if not self.is_attached:
            raise ValueError("No repository attached")

        normalized_ids = _dedupe_strings(
            [str(source_id or "").strip() for source_id in source_ids if str(source_id or "").strip()]
        )
        if not normalized_ids:
            raise ValueError("Select at least one repository source.")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            column_configs = _load_column_configs(state.get("column_configs", []))
            row_by_id = {row.id: row for row in rows}
            missing_ids = [source_id for source_id in normalized_ids if source_id not in row_by_id]
            if missing_ids:
                raise ValueError(f"Unknown source_id: {missing_ids[0]}")
            manifest_by_id = {
                source_id: build_manifest_record(
                    row_by_id[source_id],
                    base_dir=self.path,
                    column_configs=column_configs,
                )
                for source_id in normalized_ids
            }

        ready_sources = 0
        blocked_sources = 0
        for source_id in normalized_ids:
            manifest_record = manifest_by_id[source_id]
            patch: dict[str, Any] = {}
            override_fields: list[str] = []
            title_value = str(manifest_record.get("citation_title") or manifest_record.get("title") or "").strip()
            authors_value = str(
                manifest_record.get("citation_authors") or manifest_record.get("author_names") or ""
            ).strip()
            issued_value = str(
                manifest_record.get("citation_issued")
                or manifest_record.get("publication_date")
                or manifest_record.get("publication_year")
                or ""
            ).strip()
            if title_value:
                patch["citation_title"] = title_value
                override_fields.append("title")
            if authors_value:
                patch["citation_authors"] = authors_value
                override_fields.append("authors")
            if issued_value:
                patch["citation_issued"] = issued_value
                override_fields.append("issued")
            if override_fields:
                patch["citation_override_fields"] = override_fields
                updated = self.update_source(source_id, patch=patch)
            else:
                updated = self.refresh_source_citation_readiness(source_id)
            if bool(updated.get("citation_ready")):
                ready_sources += 1
            else:
                blocked_sources += 1

        message = (
            f"Marked {ready_sources} source(s) RIS ready."
            if blocked_sources == 0
            else (
                f"Marked {ready_sources} source(s) RIS ready. "
                f"{blocked_sources} source(s) still need title, authors, or publication year."
            )
        )
        return RepositorySourceBulkRisReadyResponse(
            requested_sources=len(normalized_ids),
            ready_sources=ready_sources,
            blocked_sources=blocked_sources,
            message=message,
        )

    def _load_project_profile_template(self, filename: str) -> str:
        safe_name = Path(filename or DEFAULT_PROJECT_PROFILE_FILENAME).name
        if safe_name != (filename or DEFAULT_PROJECT_PROFILE_FILENAME):
            raise ValueError(f"Invalid project profile filename: {filename}")
        template_path = self.project_profiles_dir / safe_name
        if not template_path.is_file():
            template_path = self.project_profiles_dir / DEFAULT_PROJECT_PROFILE_FILENAME
        if not template_path.is_file():
            raise ValueError("No compatible project profile template is available.")
        return template_path.read_text(encoding="utf-8")

    def _normalize_project_profile_filename(
        self,
        *,
        filename: str,
        filename_stem: str,
        profile_name: str,
    ) -> str:
        requested = Path(filename or "").name
        if requested:
            if requested != (filename or ""):
                raise ValueError(f"Invalid project profile filename: {filename}")
            if not requested.endswith((".yaml", ".yml")):
                raise ValueError("Project profile filename must end in .yaml or .yml")
            return requested

        stem_source = filename_stem or profile_name or "generated_project_profile"
        normalized_stem = re.sub(r"[^a-zA-Z0-9]+", "_", stem_source).strip("_").lower()
        normalized_stem = normalized_stem or "generated_project_profile"
        return f"{normalized_stem}.yaml"

    def _normalize_source_patch_text(self, value: Any) -> str:
        return " ".join(str(value or "").split()).strip()

    def _normalize_source_patch_multiline_text(self, value: Any) -> str:
        text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
        lines = [line.rstrip() for line in text.split("\n")]
        return "\n".join(lines).strip()

    def _source_catalog_rel(self, row: SourceManifestRow) -> Path:
        return Path(SOURCES_DIR_NAME) / row.id / f"{row.id}_catalog.json"

    def _source_summary_rel(self, row: SourceManifestRow) -> Path:
        return Path(SOURCES_DIR_NAME) / row.id / f"{row.id}_summary.md"

    def _source_rating_rel(self, row: SourceManifestRow) -> Path:
        return Path(SOURCES_DIR_NAME) / row.id / f"{row.id}_rating.json"

    def _load_repository_json_artifact(
        self,
        row: SourceManifestRow,
        field_name: str,
    ) -> dict[str, Any]:
        rel_value = str(getattr(row, field_name) or "").strip()
        source_path = self._resolve_repository_artifact_path(row, field_name, rel_value)
        if source_path is None or not source_path.is_file():
            return {}
        try:
            payload = json.loads(source_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        return payload if isinstance(payload, dict) else {}

    def _load_repository_text_artifact(
        self,
        row: SourceManifestRow,
        field_name: str,
    ) -> str:
        rel_value = str(getattr(row, field_name) or "").strip()
        source_path = self._resolve_repository_artifact_path(row, field_name, rel_value)
        if source_path is None or not source_path.is_file():
            return ""
        try:
            return source_path.read_text(encoding="utf-8", errors="replace").strip()
        except OSError:
            return ""

    def _write_repository_text_file(self, rel_path: Path, content: str) -> None:
        destination = self.path / rel_path
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_text(content, encoding="utf-8")

    def _write_repository_json_file(self, rel_path: Path, payload: dict[str, Any]) -> None:
        destination = self.path / rel_path
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    def _delete_repository_artifact(self, row: SourceManifestRow, field_name: str) -> None:
        rel_value = str(getattr(row, field_name) or "").strip()
        source_path = self._resolve_repository_artifact_path(row, field_name, rel_value)
        if source_path is not None and source_path.is_file():
            try:
                source_path.unlink()
            except OSError:
                pass
        setattr(row, field_name, "")

    def _write_catalog_patch(self, row: SourceManifestRow) -> None:
        catalog_payload = self._load_repository_json_artifact(row, "catalog_file")
        citation = _coerce_citation_metadata(catalog_payload.get("citation"))
        self._write_catalog_citation_payload(
            row,
            catalog_payload=catalog_payload,
            citation=citation,
        )

    def _row_citation_defaults_payload(self, row: SourceManifestRow) -> dict[str, Any]:
        return {
            "title": row.title,
            "authors": [
                author.model_dump(mode="json")
                for author in normalize_citation_authors(row.author_names)
            ],
            "issued": row.publication_date or row.publication_year,
            "publisher": row.organization_name,
            "item_type": row.document_type,
            "url": _citation_reference_url_for_row(row),
        }

    def _write_catalog_citation_payload(
        self,
        row: SourceManifestRow,
        *,
        catalog_payload: dict[str, Any],
        citation: CitationMetadata,
    ) -> None:
        catalog_payload.update(
            {
                "title": row.title,
                "title_status": "existing" if row.title else row.title_status,
                "author_names": row.author_names,
                "publication_date": row.publication_date,
                "publication_year": row.publication_year,
                "document_type": row.document_type,
                "organization_name": row.organization_name,
                "organization_type": row.organization_type,
                "source_kind": row.source_kind,
                "citation": citation.model_dump(mode="json"),
            }
        )
        catalog_rel = self._source_catalog_rel(row)
        self._write_repository_json_file(catalog_rel, catalog_payload)
        row.catalog_file = catalog_rel.as_posix()
        row.catalog_status = "existing"
        row.title_status = "existing" if row.title else row.title_status

    def _write_citation_patch(
        self,
        row: SourceManifestRow,
        *,
        patch: dict[str, Any],
        override_fields: list[str],
    ) -> None:
        catalog_payload = self._load_repository_json_artifact(row, "catalog_file")
        citation = _coerce_citation_metadata(catalog_payload.get("citation"))
        citation = _merge_citation_metadata(
            citation,
            self._row_citation_defaults_payload(row),
            overwrite_existing=False,
        )

        field_map = {
            "citation_title": "title",
            "citation_authors": "authors",
            "citation_issued": "issued",
            "citation_type": "item_type",
            "citation_url": "url",
            "citation_publisher": "publisher",
            "citation_container_title": "container_title",
            "citation_volume": "volume",
            "citation_issue": "issue",
            "citation_pages": "pages",
            "citation_doi": "doi",
            "citation_report_number": "report_number",
            "citation_standard_number": "standard_number",
            "citation_language": "language",
            "citation_accessed": "accessed",
        }
        effective_override_fields = set(_effective_manual_override_fields(citation))
        effective_override_fields.update(
            str(item).strip() for item in override_fields if str(item).strip()
        )

        for patch_key, citation_field in field_map.items():
            if patch_key not in patch:
                continue
            normalized_value = self._normalize_source_patch_text(patch.get(patch_key))
            if citation_field == "authors":
                citation.authors = normalize_citation_authors(normalized_value)
            else:
                setattr(citation, citation_field, normalized_value)
            effective_override_fields.add(citation_field)

        for field_name in effective_override_fields:
            field_data = citation.field_evidence.get(field_name)
            if field_data is None:
                field_data = CitationFieldEvidence()
            if field_name == "authors":
                field_data.value = "; ".join(
                    author.literal or ", ".join(part for part in [author.family, author.given] if part)
                    for author in citation.authors
                    if author.literal or author.family or author.given
                )
            else:
                field_data.value = str(getattr(citation, field_name, "") or "")
            field_data.manual_override = True
            field_data.source_type = "manual_override"
            field_data.source_label = "Manual override"
            field_data.confidence = 1.0
            citation.field_evidence[field_name] = field_data

        citation.manual_override_fields = sorted(effective_override_fields)
        citation.verification_status = "verified"
        citation.verification_confidence = max(citation.verification_confidence, 1.0)
        citation.verified_at = _utc_now_iso()
        markdown_path = self._resolve_source_file_path_for_kind(row, "md")
        citation.verification_content_digest = _file_sha256(markdown_path)
        citation = _finalize_citation_metadata(citation)
        citation = _apply_citation_manual_overrides(citation, citation)

        self._write_catalog_citation_payload(
            row,
            catalog_payload=catalog_payload,
            citation=citation,
        )

    def _write_summary_patch(self, row: SourceManifestRow, summary_text: str) -> None:
        if not summary_text:
            self._delete_repository_artifact(row, "summary_file")
            row.summary_status = ""
            return
        summary_rel = self._source_summary_rel(row)
        self._write_repository_text_file(summary_rel, summary_text + "\n")
        row.summary_file = summary_rel.as_posix()
        row.summary_status = "existing"

    def _write_rating_patch(self, row: SourceManifestRow, *, patch: dict[str, Any]) -> None:
        rating_payload = self._load_repository_json_artifact(row, "rating_file")
        ratings = rating_payload.get("ratings")
        if not isinstance(ratings, dict):
            ratings = {}
        ratings = dict(ratings)

        if "overall_relevance" in patch:
            value = _normalize_rating_score(patch.get("overall_relevance"))
            if value is None:
                ratings.pop("overall_relevance", None)
                rating_payload.pop("overall_score", None)
            else:
                ratings["overall_relevance"] = value
                rating_payload["overall_score"] = value
        if "depth_score" in patch:
            value = _normalize_rating_score(patch.get("depth_score"))
            if value is None:
                ratings.pop("depth_score", None)
            else:
                ratings["depth_score"] = value
        if "relevant_detail_score" in patch:
            value = _normalize_rating_score(patch.get("relevant_detail_score"))
            if value is None:
                ratings.pop("relevant_detail_score", None)
            else:
                ratings["relevant_detail_score"] = value
        if ratings:
            rating_payload["ratings"] = ratings
        else:
            rating_payload.pop("ratings", None)

        if "rating_rationale" in patch:
            rationale = self._normalize_source_patch_multiline_text(patch.get("rating_rationale"))
            if rationale:
                rating_payload["rationale"] = rationale
            else:
                rating_payload.pop("rationale", None)

        if "relevant_sections" in patch:
            relevant_sections = _split_relevant_sections_text(patch.get("relevant_sections"))
            if relevant_sections:
                rating_payload["relevant_sections"] = relevant_sections
            else:
                rating_payload.pop("relevant_sections", None)

        if not rating_payload:
            self._delete_repository_artifact(row, "rating_file")
            row.rating_status = ""
            return

        rating_rel = self._source_rating_rel(row)
        self._write_repository_json_file(rating_rel, rating_payload)
        row.rating_file = rating_rel.as_posix()
        row.rating_status = "existing"

    def list_ingestion_profiles(self) -> IngestionProfileListResponse:
        profiles = self._load_bundled_ingestion_profiles()
        profiles.extend(self._load_custom_ingestion_profiles())
        return IngestionProfileListResponse(
            default_profile_id="generic_numeric_academic",
            profiles=profiles,
        )

    def _validate_profile_override(self, profile_override: str) -> str:
        selected_profile_id = (profile_override or "").strip()
        if not selected_profile_id:
            return ""

        available_profile_ids = {
            profile.profile_id for profile in self.list_ingestion_profiles().profiles
        }
        if selected_profile_id not in available_profile_ids:
            raise ValueError(f"Unknown ingestion profile: {selected_profile_id}")
        return selected_profile_id

    def _document_record_display_name(self, document_record: dict[str, Any]) -> str:
        source_name = str(document_record.get("source_document_name") or "").strip()
        if source_name:
            return Path(source_name).name or source_name

        repository_path = str(document_record.get("repository_path") or "").strip()
        if repository_path:
            return Path(repository_path).name

        filename = str(document_record.get("filename") or "").strip()
        return Path(filename).name or filename or "document"

    def _pending_normalization_result(
        self,
        *,
        filename: str,
        source_document_path: str,
        selected_profile_id: str,
    ) -> DocumentNormalizationResult:
        return DocumentNormalizationResult(
            filename=filename,
            source_document_path=source_document_path,
            selected_profile_id=selected_profile_id or "auto_detect",
            selected_profile_label=selected_profile_id or "Auto-detect",
            status="pending",
        )

    def _load_import_document_records(
        self,
        import_record: dict[str, Any],
    ) -> list[dict[str, str]]:
        import_id = str(import_record.get("import_id") or "").strip()
        resolved_records: list[dict[str, str]] = []

        raw_documents = import_record.get("documents", [])
        if isinstance(raw_documents, list):
            for raw_document in raw_documents:
                if not isinstance(raw_document, dict):
                    continue
                repository_path = str(raw_document.get("repository_path") or "").strip()
                filename = str(raw_document.get("filename") or "").strip()
                if not repository_path and import_id and filename:
                    repository_path = (
                        Path(DOCUMENTS_DIR_NAME) / import_id / Path(filename).name
                    ).as_posix()
                if not filename and repository_path:
                    filename = Path(repository_path).name
                if not repository_path or not filename:
                    continue
                source_abs = self.path / Path(repository_path)
                if not source_abs.is_file():
                    continue
                sha256 = str(raw_document.get("sha256") or "").strip()
                if not sha256:
                    sha256 = hashlib.sha256(source_abs.read_bytes()).hexdigest()
                resolved_records.append(
                    {
                        "filename": filename,
                        "source_document_name": self._document_record_display_name(
                            raw_document
                        ),
                        "repository_path": repository_path,
                        "sha256": sha256,
                        "document_import_id": str(
                            raw_document.get("document_import_id") or import_id
                        ).strip()
                        or import_id,
                    }
                )

        if resolved_records:
            return resolved_records

        if not import_id:
            return []

        documents_dir = self.documents_dir / import_id
        if not documents_dir.is_dir():
            return []

        for path in sorted(documents_dir.iterdir()):
            if not path.is_file() or not _is_original_document_import_path(path):
                continue
            resolved_records.append(
                {
                    "filename": path.name,
                    "source_document_name": path.name,
                    "repository_path": (Path(DOCUMENTS_DIR_NAME) / import_id / path.name).as_posix(),
                    "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
                    "document_import_id": import_id,
                }
            )
        return resolved_records

    def list_document_imports(self) -> RepositoryDocumentImportListResponse:
        if not self.is_attached:
            raise ValueError("No repository attached")

        with self._writer_lock():
            state = self._load_state_locked()
            imports = list(state.get("imports", []))

        response_records: list[RepositoryDocumentImportRecord] = []
        for raw_import in sorted(
            imports,
            key=lambda item: str(item.get("imported_at") or ""),
            reverse=True,
        ):
            if str(raw_import.get("import_type") or "").strip() != "document_process":
                continue

            document_records = self._load_import_document_records(raw_import)
            if not document_records:
                continue

            response_records.append(
                RepositoryDocumentImportRecord(
                    import_id=str(raw_import.get("import_id") or "").strip(),
                    import_type="document_process",
                    imported_at=str(raw_import.get("imported_at") or "").strip(),
                    provenance=str(raw_import.get("provenance") or "").strip(),
                    selected_profile_id=str(
                        raw_import.get("selected_profile_id") or ""
                    ).strip(),
                    processing_job_id=str(
                        raw_import.get("processing_job_id") or ""
                    ).strip(),
                    document_count=len(document_records),
                    rerunnable=True,
                    documents=[
                        RepositoryDocumentImportDocument(
                            filename=self._document_record_display_name(document_record),
                            repository_path=str(
                                document_record.get("repository_path") or ""
                            ).strip(),
                            sha256=str(document_record.get("sha256") or "").strip(),
                        )
                        for document_record in document_records
                    ],
                )
            )

        return RepositoryDocumentImportListResponse(imports=response_records)

    def save_ingestion_profile(self, profile: IngestionProfile) -> IngestionProfileActionResponse:
        if not self.is_attached:
            raise ValueError("No repository attached")
        if profile.built_in:
            raise ValueError("Built-in profiles cannot be overwritten")
        if profile.profile_id in {
            bundled.profile_id for bundled in self._load_bundled_ingestion_profiles()
        }:
            raise ValueError("Built-in profiles cannot be overwritten")

        profiles = self._load_custom_ingestion_profiles()
        replaced = False
        for index, current in enumerate(profiles):
            if current.profile_id == profile.profile_id:
                profiles[index] = profile
                replaced = True
                break
        if not replaced:
            profiles.append(profile)
        self._save_custom_ingestion_profiles(profiles)
        action = "Updated" if replaced else "Created"
        return IngestionProfileActionResponse(
            status="completed",
            message=f"{action} ingestion profile `{profile.profile_id}`.",
            profile=profile,
        )

    def delete_ingestion_profile(self, profile_id: str) -> IngestionProfileActionResponse:
        if not self.is_attached:
            raise ValueError("No repository attached")
        profiles = self._load_custom_ingestion_profiles()
        remaining = [profile for profile in profiles if profile.profile_id != profile_id]
        if len(remaining) == len(profiles):
            raise ValueError(f"Ingestion profile not found: {profile_id}")
        self._save_custom_ingestion_profiles(remaining)
        return IngestionProfileActionResponse(
            status="completed",
            message=f"Deleted ingestion profile `{profile_id}`.",
        )

    def list_ingestion_profile_suggestions(self) -> IngestionProfileSuggestionListResponse:
        return IngestionProfileSuggestionListResponse(
            suggestions=self._load_ingestion_profile_suggestions()
        )

    def accept_ingestion_profile_suggestion(
        self,
        suggestion_id: str,
    ) -> IngestionProfileSuggestionActionResponse:
        suggestions = self._load_ingestion_profile_suggestions()
        accepted: IngestionProfileSuggestion | None = None
        for suggestion in suggestions:
            if suggestion.suggestion_id == suggestion_id:
                accepted = suggestion
                suggestion.status = "accepted"
                break
        if accepted is None:
            raise ValueError(f"Ingestion profile suggestion not found: {suggestion_id}")

        self.save_ingestion_profile(accepted.proposed_profile)
        self._save_ingestion_profile_suggestions(suggestions)
        return IngestionProfileSuggestionActionResponse(
            status="completed",
            message=f"Accepted ingestion profile suggestion `{suggestion_id}`.",
            suggestion=accepted,
            accepted_profile=accepted.proposed_profile,
        )

    def reject_ingestion_profile_suggestion(
        self,
        suggestion_id: str,
    ) -> IngestionProfileSuggestionActionResponse:
        suggestions = self._load_ingestion_profile_suggestions()
        rejected: IngestionProfileSuggestion | None = None
        for suggestion in suggestions:
            if suggestion.suggestion_id == suggestion_id:
                suggestion.status = "rejected"
                rejected = suggestion
                break
        if rejected is None:
            raise ValueError(f"Ingestion profile suggestion not found: {suggestion_id}")
        self._save_ingestion_profile_suggestions(suggestions)
        return IngestionProfileSuggestionActionResponse(
            status="completed",
            message=f"Rejected ingestion profile suggestion `{suggestion_id}`.",
            suggestion=rejected,
        )

    def _load_custom_ingestion_profiles(self) -> list[IngestionProfile]:
        if not self.is_attached:
            return []
        path = self._internal_dir() / INGESTION_PROFILES_FILE_NAME
        if not path.exists():
            return []
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return []
        if not isinstance(raw, list):
            return []
        profiles: list[IngestionProfile] = []
        for item in raw:
            try:
                profile = IngestionProfile.model_validate(item)
            except Exception:
                continue
            if profile.built_in:
                profile.built_in = False
            profiles.append(profile)
        return profiles

    def _default_ingestion_profiles(self) -> list[IngestionProfile]:
        return [
            profile.model_copy(update={"built_in": True}) for profile in builtin_ingestion_profiles()
        ]

    def _load_bundled_ingestion_profiles(self) -> list[IngestionProfile]:
        if not self.is_attached:
            return self._default_ingestion_profiles()
        path = self._internal_dir() / BUNDLED_INGESTION_PROFILES_FILE_NAME
        if not path.exists():
            return self._default_ingestion_profiles()
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return self._default_ingestion_profiles()
        if not isinstance(raw, list):
            return self._default_ingestion_profiles()

        profiles: list[IngestionProfile] = []
        seen_profile_ids: set[str] = set()
        for item in raw:
            try:
                profile = IngestionProfile.model_validate(item)
            except Exception:
                continue
            profile = profile.model_copy(update={"built_in": True})
            if not profile.profile_id or profile.profile_id in seen_profile_ids:
                continue
            seen_profile_ids.add(profile.profile_id)
            profiles.append(profile)
        return profiles or self._default_ingestion_profiles()

    def _sync_bundled_ingestion_profiles(self) -> None:
        if not self.is_attached:
            raise ValueError("No repository attached")
        path = self._internal_dir() / BUNDLED_INGESTION_PROFILES_FILE_NAME
        if path.exists():
            return
        payload = [
            profile.model_dump(mode="json") for profile in self._default_ingestion_profiles()
        ]
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _save_custom_ingestion_profiles(
        self,
        profiles: list[IngestionProfile],
    ) -> None:
        if not self.is_attached:
            raise ValueError("No repository attached")
        path = self._internal_dir() / INGESTION_PROFILES_FILE_NAME
        payload = [profile.model_dump(mode="json") for profile in profiles if not profile.built_in]
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _load_ingestion_profile_suggestions(self) -> list[IngestionProfileSuggestion]:
        if not self.is_attached:
            return []
        path = self._internal_dir() / INGESTION_PROFILE_SUGGESTIONS_FILE_NAME
        if not path.exists():
            return []
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return []
        if not isinstance(raw, list):
            return []
        suggestions: list[IngestionProfileSuggestion] = []
        for item in raw:
            try:
                suggestions.append(IngestionProfileSuggestion.model_validate(item))
            except Exception:
                continue
        return suggestions

    def _save_ingestion_profile_suggestions(
        self,
        suggestions: list[IngestionProfileSuggestion],
    ) -> None:
        if not self.is_attached:
            raise ValueError("No repository attached")
        path = self._internal_dir() / INGESTION_PROFILE_SUGGESTIONS_FILE_NAME
        payload = [suggestion.model_dump(mode="json") for suggestion in suggestions]
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # ---- Create new repository ----

    def create(self, path_str: str) -> RepositoryStatusResponse:
        """Scaffold a new repository at the given path and attach it."""
        resolved = Path(path_str).expanduser().resolve()
        resolved.mkdir(parents=True, exist_ok=True)

        # Don't create over a non-empty directory that already has .ra_repo
        internal = resolved / INTERNAL_DIR_NAME
        if internal.exists() and (internal / STATE_FILE_NAME).exists():
            raise ValueError(
                f"A repository already exists at {resolved}. Use 'Open' instead."
            )

        # Create full scaffold
        internal.mkdir(parents=True, exist_ok=True)

        # Write initial metadata and state
        now = _utc_now_iso()
        meta = {
            "schema_version": SCHEMA_VERSION,
            "created_at": now,
            "updated_at": now,
            "last_scan_at": now,
            "next_source_id": 1,
        }
        (internal / META_FILE_NAME).write_text(
            json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        state = {"sources": [], "citations": [], "imports": [], "column_configs": []}
        (internal / STATE_FILE_NAME).write_text(
            json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        # Write default settings
        default_settings = RepoSettings()
        (internal / REPO_SETTINGS_FILE_NAME).write_text(
            json.dumps(default_settings.model_dump(mode="json"), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        self._path = resolved
        self._job_store = None
        with self._writer_lock():
            self._ensure_scaffold_locked()
            self._rebuild_outputs_locked([], [])

        # Attach the newly created repo
        return self.attach(str(resolved))

    def attach(self, path_value: str) -> RepositoryStatusResponse:
        resolved = self._resolve_path(path_value)
        self._path = resolved
        self._job_store = None
        self._ensure_internal_dirs()

        with self._writer_lock():
            self._ensure_scaffold_locked()
            meta = self._load_meta_locked()
            if not self._meta_path().exists():
                self._create_backup_snapshot_locked("first_attach")
                meta = self._default_meta()
            elif int(meta.get("schema_version") or 1) < SCHEMA_VERSION:
                self._create_backup_snapshot_locked(
                    f"schema_upgrade_v{int(meta.get('schema_version') or 1)}"
                )
                meta["schema_version"] = SCHEMA_VERSION

            merged, scan = self._scan_and_merge_locked()
            self._save_state_locked(
                sources=merged.rows,
                citations=merged.citations,
                imports=self._load_state_locked().get("imports", []),
            )
            self._save_meta_locked(
                {
                    **meta,
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": merged.next_source_id,
                    "last_scan_at": scan.scanned_at,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(merged.rows, merged.citations)
            self._last_scan = scan
            self._download_state = "idle"
            self._download_message = "Repository attached"

        # One-time migration: if app-level settings have default LLM config
        # but this repo has real config from the old per-repo format, adopt it.
        self._maybe_migrate_repo_settings_to_app()

        return self.get_status()

    def _maybe_migrate_repo_settings_to_app(self) -> None:
        """Copy infrastructure settings from repo to app level if app has defaults."""
        settings_path = self._internal_dir() / REPO_SETTINGS_FILE_NAME
        if not settings_path.exists():
            return
        try:
            raw = json.loads(settings_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return
        # Only migrate if the repo file has old-style infrastructure fields
        if "llm_backend" not in raw:
            return
        app_settings = self.store.load_app_settings()
        # Only adopt if the app-level LLM config is still at defaults
        if app_settings.llm_backend.model:
            return
        try:
            from backend.models.settings import LLMBackendConfig
            llm_config = LLMBackendConfig(**raw.get("llm_backend", {}))
        except Exception:
            return
        app_settings.llm_backend = llm_config
        app_settings.use_llm = raw.get("use_llm", app_settings.use_llm)
        app_settings.fetch_delay = raw.get("fetch_delay", app_settings.fetch_delay)
        app_settings.searxng_base_url = raw.get("searxng_base_url", app_settings.searxng_base_url)
        self.store.save_app_settings(app_settings)

    def get_status(self) -> RepositoryStatusResponse:
        if not self.is_attached:
            return RepositoryStatusResponse(attached=False, message="No repository attached")

        with self._writer_lock():
            state = self._load_state_locked()
            meta = self._load_meta_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            health = self._compute_health(rows, citations)
            queued_count = sum(1 for row in rows if (row.fetch_status or "") in {"", "queued"})

            return RepositoryStatusResponse(
                attached=True,
                path=str(self.path),
                schema_version=int(meta.get("schema_version") or SCHEMA_VERSION),
                next_source_id=int(meta.get("next_source_id") or 1),
                total_sources=len(rows),
                total_citations=len(citations),
                queued_count=queued_count,
                download_state=self._download_state,
                message=self._download_message,
                last_scan_at=str(meta.get("last_scan_at") or ""),
                last_updated_at=str(meta.get("updated_at") or ""),
                health=health,
                output_summary=summarize_output_rows(rows),
                scan=self._last_scan,
            )

    def get_dashboard(
        self,
        recent_imports_limit: int = 10,
        recent_jobs_limit: int = 20,
    ) -> dict[str, Any]:
        if not self.is_attached:
            raise ValueError("No repository attached")

        status = self.get_status()
        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            imports = list(state.get("imports", []))

        output_formats = {
            "raw": sum(1 for row in rows if (row.raw_file or "").strip()),
            "rendered_html": sum(1 for row in rows if (row.rendered_file or "").strip()),
            "rendered_pdf": sum(1 for row in rows if (row.rendered_pdf_file or "").strip()),
            "markdown": sum(1 for row in rows if (row.markdown_file or "").strip()),
            "catalogs": sum(1 for row in rows if (row.catalog_file or "").strip()),
            "summaries": sum(1 for row in rows if (row.summary_file or "").strip()),
            "ratings": sum(1 for row in rows if (row.rating_file or "").strip()),
        }

        warning_aggregates = {
            "missing_files": int(status.health.missing_files),
            "orphaned_citation_rows": int(status.health.orphaned_citation_rows),
            "incomplete_summaries": sum(
                1
                for row in rows
                if (row.summary_status or "").strip().lower() in {"", "missing", "failed"}
            ),
            "failed_ratings": sum(
                1 for row in rows if (row.rating_status or "").strip().lower() == "failed"
            ),
            "failed_catalogs": sum(
                1 for row in rows if (row.catalog_status or "").strip().lower() == "failed"
            ),
            "failed_fetches": sum(
                1 for row in rows if (row.fetch_status or "").strip().lower() == "failed"
            ),
        }

        recent_imports = sorted(
            imports,
            key=lambda item: str(item.get("imported_at") or ""),
            reverse=True,
        )[: max(1, recent_imports_limit)]

        return {
            "status": status.model_dump(mode="json"),
            "metrics": {
                "total_sources": int(status.total_sources),
                "total_citations": int(status.total_citations),
                "queued_count": int(status.queued_count),
                "next_source_id": int(status.next_source_id),
            },
            "output_formats": output_formats,
            "warning_aggregates": warning_aggregates,
            "recent_imports": recent_imports,
            "recent_jobs": self._collect_recent_jobs(recent_jobs_limit),
        }

    def list_manifest(
        self,
        *,
        q: str = "",
        fetch_status: str = "",
        detected_type: str = "",
        source_kind: str = "",
        document_type: str = "",
        organization_type: str = "",
        organization_name: str = "",
        author_names: str = "",
        publication_date: str = "",
        tags_text: str = "",
        has_summary: bool | None = None,
        has_rating: bool | None = None,
        rating_overall_min: float | None = None,
        rating_overall_max: float | None = None,
        rating_overall_relevance_min: float | None = None,
        rating_overall_relevance_max: float | None = None,
        rating_depth_score_min: float | None = None,
        rating_depth_score_max: float | None = None,
        rating_relevant_detail_score_min: float | None = None,
        rating_relevant_detail_score_max: float | None = None,
        citation_type: str = "",
        citation_doi: str = "",
        citation_report_number: str = "",
        citation_standard_number: str = "",
        citation_missing_fields: str = "",
        citation_ready: bool | None = None,
        citation_confidence_min: float | None = None,
        citation_confidence_max: float | None = None,
        sort_by: str = "",
        sort_dir: str = "",
        limit: int = 50,
        offset: int = 0,
    ) -> dict[str, Any]:
        if not self.is_attached:
            raise ValueError("No repository attached")

        safe_limit = max(1, min(int(limit), 500))
        safe_offset = max(0, int(offset))
        records, columns = self._manifest_records()
        filtered, normalized_sort_by, normalized_sort_dir = self._filter_manifest_records(
            records,
            columns,
            q=q,
            fetch_status=fetch_status,
            detected_type=detected_type,
            source_kind=source_kind,
            document_type=document_type,
            organization_type=organization_type,
            organization_name=organization_name,
            author_names=author_names,
            publication_date=publication_date,
            tags_text=tags_text,
            has_summary=has_summary,
            has_rating=has_rating,
            rating_overall_min=rating_overall_min,
            rating_overall_max=rating_overall_max,
            rating_overall_relevance_min=rating_overall_relevance_min,
            rating_overall_relevance_max=rating_overall_relevance_max,
            rating_depth_score_min=rating_depth_score_min,
            rating_depth_score_max=rating_depth_score_max,
            rating_relevant_detail_score_min=rating_relevant_detail_score_min,
            rating_relevant_detail_score_max=rating_relevant_detail_score_max,
            citation_type=citation_type,
            citation_doi=citation_doi,
            citation_report_number=citation_report_number,
            citation_standard_number=citation_standard_number,
            citation_missing_fields=citation_missing_fields,
            citation_ready=citation_ready,
            citation_confidence_min=citation_confidence_min,
            citation_confidence_max=citation_confidence_max,
            sort_by=sort_by,
            sort_dir=sort_dir,
        )
        paged_rows = filtered[safe_offset : safe_offset + safe_limit]

        return {
            "rows": paged_rows,
            "total": len(filtered),
            "limit": safe_limit,
            "offset": safe_offset,
            "sort_by": normalized_sort_by,
            "sort_dir": normalized_sort_dir,
            "columns": columns,
            "filters": {
                "q": q,
                "fetch_status": fetch_status,
                "detected_type": detected_type,
                "source_kind": source_kind,
                "document_type": document_type,
                "organization_type": organization_type,
                "organization_name": organization_name,
                "author_names": author_names,
                "publication_date": publication_date,
                "tags_text": tags_text,
                "has_summary": has_summary,
                "has_rating": has_rating,
                "rating_overall_min": rating_overall_min,
                "rating_overall_max": rating_overall_max,
                "rating_overall_relevance_min": rating_overall_relevance_min,
                "rating_overall_relevance_max": rating_overall_relevance_max,
                "rating_depth_score_min": rating_depth_score_min,
                "rating_depth_score_max": rating_depth_score_max,
                "rating_relevant_detail_score_min": rating_relevant_detail_score_min,
                "rating_relevant_detail_score_max": rating_relevant_detail_score_max,
                "citation_type": citation_type,
                "citation_doi": citation_doi,
                "citation_report_number": citation_report_number,
                "citation_standard_number": citation_standard_number,
                "citation_missing_fields": citation_missing_fields,
                "citation_ready": citation_ready,
                "citation_confidence_min": citation_confidence_min,
                "citation_confidence_max": citation_confidence_max,
            },
        }

    def _manifest_records(
        self,
    ) -> tuple[list[dict[str, str | int | float | bool]], list[dict[str, Any]]]:
        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            column_configs = _load_column_configs(state.get("column_configs", []))
        records = [
            build_manifest_record(row, base_dir=self.path, column_configs=column_configs)
            for row in rows
        ]
        return records, _build_manifest_column_metadata(records, column_configs)

    def _filter_manifest_records(
        self,
        records: list[dict[str, str | int | float | bool]],
        columns: list[dict[str, Any]],
        *,
        q: str = "",
        fetch_status: str = "",
        detected_type: str = "",
        source_kind: str = "",
        document_type: str = "",
        organization_type: str = "",
        organization_name: str = "",
        author_names: str = "",
        publication_date: str = "",
        tags_text: str = "",
        has_summary: bool | None = None,
        has_rating: bool | None = None,
        rating_overall_min: float | None = None,
        rating_overall_max: float | None = None,
        rating_overall_relevance_min: float | None = None,
        rating_overall_relevance_max: float | None = None,
        rating_depth_score_min: float | None = None,
        rating_depth_score_max: float | None = None,
        rating_relevant_detail_score_min: float | None = None,
        rating_relevant_detail_score_max: float | None = None,
        citation_type: str = "",
        citation_doi: str = "",
        citation_report_number: str = "",
        citation_standard_number: str = "",
        citation_missing_fields: str = "",
        citation_ready: bool | None = None,
        citation_confidence_min: float | None = None,
        citation_confidence_max: float | None = None,
        sort_by: str = "",
        sort_dir: str = "",
    ) -> tuple[list[dict[str, str | int | float | bool]], str, str]:
        normalized_sort_by = (sort_by or "").strip()
        normalized_sort_dir = (sort_dir or "").strip().lower()
        if normalized_sort_dir and normalized_sort_dir not in {"asc", "desc"}:
            raise ValueError("Invalid sort_dir. Use `asc` or `desc`.")
        allowed_sort_fields = {str(item.get("key") or "") for item in columns}
        if normalized_sort_by and normalized_sort_by not in allowed_sort_fields:
            raise ValueError(
                f"Invalid sort_by. Allowed: {sorted(item for item in allowed_sort_fields if item)}"
            )
        if not normalized_sort_by:
            normalized_sort_dir = ""

        q_norm = (q or "").strip().lower()
        fetch_status_norm = (fetch_status or "").strip().lower()
        detected_type_norm = (detected_type or "").strip().lower()
        source_kind_norm = (source_kind or "").strip().lower()
        document_type_norm = (document_type or "").strip().lower()
        organization_type_norm = (organization_type or "").strip().lower()
        organization_name_norm = (organization_name or "").strip().lower()
        author_names_norm = (author_names or "").strip().lower()
        publication_date_norm = (publication_date or "").strip().lower()
        tags_text_norm = (tags_text or "").strip().lower()
        citation_type_norm = (citation_type or "").strip().lower()
        citation_doi_norm = (citation_doi or "").strip().lower()
        citation_report_number_norm = (citation_report_number or "").strip().lower()
        citation_standard_number_norm = (citation_standard_number or "").strip().lower()
        citation_missing_fields_norm = (citation_missing_fields or "").strip().lower()
        rating_filters = {
            "rating_overall": (rating_overall_min, rating_overall_max),
            "rating_overall_relevance": (
                rating_overall_relevance_min,
                rating_overall_relevance_max,
            ),
            "rating_depth_score": (rating_depth_score_min, rating_depth_score_max),
            "rating_relevant_detail_score": (
                rating_relevant_detail_score_min,
                rating_relevant_detail_score_max,
            ),
            "citation_confidence": (citation_confidence_min, citation_confidence_max),
        }

        filtered: list[dict[str, str | int | float | bool]] = []
        for record in records:
            if fetch_status_norm and str(record.get("fetch_status") or "").strip().lower() != fetch_status_norm:
                continue
            if detected_type_norm and str(record.get("detected_type") or "").strip().lower() != detected_type_norm:
                continue
            if source_kind_norm and str(record.get("source_kind") or "").strip().lower() != source_kind_norm:
                continue
            if document_type_norm and document_type_norm not in str(record.get("document_type") or "").strip().lower():
                continue
            if organization_type_norm and organization_type_norm not in str(record.get("organization_type") or "").strip().lower():
                continue
            if organization_name_norm and organization_name_norm not in str(record.get("organization_name") or "").strip().lower():
                continue
            if author_names_norm and author_names_norm not in str(record.get("author_names") or "").strip().lower():
                continue
            if publication_date_norm and publication_date_norm not in str(record.get("publication_date") or "").strip().lower():
                continue
            if tags_text_norm and tags_text_norm not in str(record.get("tags_text") or "").strip().lower():
                continue
            if citation_type_norm and citation_type_norm not in str(record.get("citation_type") or "").strip().lower():
                continue
            if citation_doi_norm and citation_doi_norm not in str(record.get("citation_doi") or "").strip().lower():
                continue
            if citation_report_number_norm and citation_report_number_norm not in str(record.get("citation_report_number") or "").strip().lower():
                continue
            if citation_standard_number_norm and citation_standard_number_norm not in str(record.get("citation_standard_number") or "").strip().lower():
                continue
            if citation_missing_fields_norm and citation_missing_fields_norm not in str(record.get("citation_missing_fields") or "").strip().lower():
                continue

            summary_present = bool(str(record.get("summary_file") or "").strip()) or (
                str(record.get("summary_status") or "").strip().lower()
                in {"generated", "existing"}
            )
            rating_present = bool(str(record.get("rating_file") or "").strip()) or (
                str(record.get("rating_status") or "").strip().lower()
                in {"generated", "existing"}
            )

            if has_summary is True and not summary_present:
                continue
            if has_summary is False and summary_present:
                continue
            if has_rating is True and not rating_present:
                continue
            if has_rating is False and rating_present:
                continue
            if citation_ready is True and not bool(record.get("citation_ready")):
                continue
            if citation_ready is False and bool(record.get("citation_ready")):
                continue

            if not _manifest_record_matches_thresholds(record, rating_filters):
                continue

            if q_norm:
                haystack = " ".join(
                    [
                        str(record.get("id") or ""),
                        str(record.get("source_kind") or ""),
                        str(record.get("title") or ""),
                        str(record.get("author_names") or ""),
                        str(record.get("publication_date") or ""),
                        str(record.get("document_type") or ""),
                        str(record.get("organization_name") or ""),
                        str(record.get("organization_type") or ""),
                        str(record.get("tags_text") or ""),
                        str(record.get("original_url") or ""),
                        str(record.get("final_url") or ""),
                        str(record.get("source_document_name") or ""),
                        str(record.get("provenance_ref") or ""),
                        str(record.get("summary_text") or ""),
                        str(record.get("rating_rationale") or ""),
                        str(record.get("relevant_sections") or ""),
                        str(record.get("citation_type") or ""),
                        str(record.get("citation_doi") or ""),
                        str(record.get("citation_report_number") or ""),
                        str(record.get("citation_standard_number") or ""),
                        str(record.get("citation_missing_fields") or ""),
                        str(record.get("notes") or ""),
                        str(record.get("error_message") or ""),
                    ]
                ).lower()
                if q_norm not in haystack:
                    continue

            filtered.append(record)

        if normalized_sort_by:
            reverse = normalized_sort_dir == "desc"
            sort_type = next(
                (
                    str(item.get("sort_type") or "text")
                    for item in columns
                    if str(item.get("key") or "") == normalized_sort_by
                ),
                "text",
            )
            filtered = _sort_manifest_records(
                filtered,
                sort_by=normalized_sort_by,
                reverse=reverse,
                sort_type=sort_type,
            )
        return filtered, normalized_sort_by, normalized_sort_dir

    def list_agent_sources(
        self,
        *,
        q: str = "",
        status: str = "",
        fetch_status: str = "",
        convert_status: str = "",
        tag_status: str = "",
        summarize_status: str = "",
        import_id: str = "",
        has_summary: bool | None = None,
        has_rating: bool | None = None,
        min_relevance: float | None = None,
        sort_by: str = "rating_overall",
        sort_dir: str = "desc",
        limit: int = 50,
        cursor: str = "",
    ) -> dict[str, Any]:
        if not self.is_attached:
            raise ValueError("No repository attached")

        normalized_sort_by = (sort_by or "rating_overall").strip().lower() or "rating_overall"
        normalized_sort_dir = (sort_dir or "desc").strip().lower()
        if normalized_sort_dir not in {"asc", "desc"}:
            raise ValueError("Invalid sort_dir. Use `asc` or `desc`.")
        allowed_sort_fields = {"rating_overall", "updated_at", "source_id", "title"}
        if normalized_sort_by not in allowed_sort_fields:
            raise ValueError(
                f"Invalid sort_by. Allowed: {sorted(allowed_sort_fields)}"
            )

        safe_limit = max(1, min(int(limit), 500))
        offset = _decode_agent_offset_cursor(cursor)
        q_norm = (q or "").strip().lower()
        status_norm = (status or "").strip().lower()
        fetch_status_norm = (fetch_status or "").strip().lower()
        convert_status_norm = (convert_status or "").strip().lower()
        tag_status_norm = (tag_status or "").strip().lower()
        summarize_status_norm = (summarize_status or "").strip().lower()
        import_id_norm = (import_id or "").strip()

        with self._writer_lock():
            state = self._load_state_locked()
            rows = self._sort_rows(_load_source_rows(state.get("sources", [])))
            imports = list(state.get("imports", []))

        paired_records: list[tuple[AgentSourceRecord, dict[str, Any]]] = []
        for row in rows:
            manifest_record = build_manifest_record(row, base_dir=self.path)
            source_record = self._build_agent_source_record(
                row=row,
                imports=imports,
                manifest_record=manifest_record,
            )
            paired_records.append((source_record, manifest_record))

        filtered: list[tuple[AgentSourceRecord, dict[str, Any]]] = []
        for source_record, manifest_record in paired_records:
            if status_norm and source_record.fetch_status.lower() != status_norm:
                continue
            if fetch_status_norm and source_record.fetch_status.lower() != fetch_status_norm:
                continue
            if convert_status_norm and source_record.convert_status.lower() != convert_status_norm:
                continue
            if tag_status_norm and source_record.tag_status.lower() != tag_status_norm:
                continue
            if summarize_status_norm and source_record.summarize_status.lower() != summarize_status_norm:
                continue
            if import_id_norm and source_record.provenance.import_id != import_id_norm:
                continue
            if has_summary is True and not source_record.summary_present:
                continue
            if has_summary is False and source_record.summary_present:
                continue
            if has_rating is True and not source_record.rating_present:
                continue
            if has_rating is False and source_record.rating_present:
                continue

            relevance_value = _coerce_optional_float(
                manifest_record.get("rating_overall_relevance")
            )
            if relevance_value is None:
                relevance_value = source_record.rating_overall
            if min_relevance is not None and (
                relevance_value is None or relevance_value < float(min_relevance)
            ):
                continue

            if q_norm:
                haystack = " ".join(
                    [
                        source_record.source_id,
                        source_record.title,
                        source_record.original_url,
                        source_record.final_url,
                        source_record.provenance.source_document_name,
                        source_record.provenance.provenance_ref,
                        str(manifest_record.get("summary_text") or ""),
                        str(manifest_record.get("rating_rationale") or ""),
                        str(manifest_record.get("relevant_sections") or ""),
                    ]
                ).lower()
                if q_norm not in haystack:
                    continue

            filtered.append((source_record, manifest_record))

        sorted_records = _sort_agent_source_records(
            filtered,
            sort_by=normalized_sort_by,
            sort_dir=normalized_sort_dir,
        )
        page = sorted_records[offset : offset + safe_limit]
        next_offset = offset + safe_limit
        next_cursor = (
            _encode_agent_offset_cursor(next_offset) if next_offset < len(sorted_records) else ""
        )
        return {
            "items": [record.model_dump(mode="json") for record, _ in page],
            "total": len(sorted_records),
            "limit": safe_limit,
            "cursor": cursor,
            "next_cursor": next_cursor,
            "sort_by": normalized_sort_by,
            "sort_dir": normalized_sort_dir,
        }

    def get_agent_source(self, source_id: str) -> AgentSourceRecord:
        if not self.is_attached:
            raise ValueError("No repository attached")
        normalized_id = str(source_id or "").strip()
        if not normalized_id:
            raise ValueError("source_id is required")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            imports = list(state.get("imports", []))

        row = next((item for item in rows if item.id == normalized_id), None)
        if row is None:
            raise ValueError(f"Unknown source_id: {normalized_id}")
        manifest_record = build_manifest_record(row, base_dir=self.path)
        return self._build_agent_source_record(
            row=row,
            imports=imports,
            manifest_record=manifest_record,
        )

    def get_agent_source_content(
        self,
        source_id: str,
        *,
        kind: str,
        cursor: str = "",
        chunk_size: int = 8000,
    ) -> AgentSourceContentChunk:
        if not self.is_attached:
            raise ValueError("No repository attached")
        normalized_id = str(source_id or "").strip()
        if not normalized_id:
            raise ValueError("source_id is required")
        normalized_kind = str(kind or "").strip().lower()
        if normalized_kind not in {"markdown", "clean_markdown", "summary", "rating", "metadata"}:
            raise ValueError(
                "Invalid content kind. Use `markdown`, `clean_markdown`, `summary`, `rating`, or `metadata`."
            )

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))

        row = next((item for item in rows if item.id == normalized_id), None)
        if row is None:
            raise ValueError(f"Unknown source_id: {normalized_id}")

        artifact_uri = _agent_source_artifact_uri(normalized_id, normalized_kind)
        content = ""
        mime_type = "text/plain"
        if normalized_kind == "markdown":
            content = self._load_source_text_artifact(row, "markdown_file")
            mime_type = "text/markdown"
        elif normalized_kind == "clean_markdown":
            content = self._load_source_text_artifact(
                row,
                "llm_cleanup_file",
                fallback_field="markdown_file",
            )
            mime_type = "text/markdown"
        elif normalized_kind == "summary":
            content = self._load_source_text_artifact(row, "summary_file")
            mime_type = "text/markdown"
        elif normalized_kind == "rating":
            content = self._load_source_text_artifact(row, "rating_file")
            mime_type = "application/json"
        else:
            metadata_path = self._resolve_repository_artifact_path(
                row,
                "metadata_file",
                row.metadata_file,
            )
            if metadata_path is not None and metadata_path.is_file():
                content = metadata_path.read_text(encoding="utf-8", errors="replace")
            else:
                content = json.dumps(row.model_dump(mode="json"), ensure_ascii=False, indent=2)
            mime_type = "application/json"

        if not content.strip():
            raise ValueError(
                f"No `{normalized_kind}` content available for source `{normalized_id}`"
            )

        safe_chunk_size = max(1, min(int(chunk_size), 50000))
        offset = _decode_agent_offset_cursor(cursor)
        if offset > len(content):
            raise ValueError("Cursor is out of range for the requested content.")
        end = min(len(content), offset + safe_chunk_size)
        next_cursor = _encode_agent_offset_cursor(end) if end < len(content) else ""
        return AgentSourceContentChunk(
            source_id=normalized_id,
            kind=normalized_kind,
            mime_type=mime_type,
            artifact_uri=artifact_uri,
            cursor=cursor,
            next_cursor=next_cursor,
            total_length=len(content),
            offset_start=offset,
            offset_end=end,
            content=content[offset:end],
        )

    def get_agent_run(
        self,
        run_id: str,
        *,
        live_jobs: dict[str, SourceDownloadOrchestrator] | None = None,
        live_jobs_lock: threading.Lock | None = None,
    ) -> AgentRunRecord:
        if not self.is_attached:
            raise ValueError("No repository attached")
        normalized_run_id = str(run_id or "").strip()
        if not normalized_run_id:
            raise ValueError("run_id is required")

        store = self.job_store_for(normalized_run_id)
        if not store.job_exists(normalized_run_id):
            raise ValueError(f"Unknown run_id: {normalized_run_id}")

        raw_status = store.get_source_status(normalized_run_id)
        if raw_status is None and live_jobs is not None and live_jobs_lock is not None:
            with live_jobs_lock:
                orchestrator = live_jobs.get(normalized_run_id)
            if orchestrator is not None:
                pending = _build_pending_source_status(
                    job_id=normalized_run_id,
                    store=store,
                    orchestrator=orchestrator,
                )
                raw_status = pending.model_dump(mode="json")

        status = (
            SourceDownloadStatus.model_validate(raw_status)
            if isinstance(raw_status, dict)
            else None
        )
        context = store.load_artifact(normalized_run_id, "repo_source_task_context") or {}
        artifact = store.load_artifact(normalized_run_id, "06_sources_manifest") or {}
        rows = _load_source_rows(artifact.get("rows", []))
        selected_phases = _normalize_agent_phase_names(
            context.get("selected_phases", []),
            run_download=bool(getattr(status, "run_download", False)),
            run_convert=bool(getattr(status, "run_convert", False)),
            run_cleanup=bool(getattr(status, "run_llm_cleanup", False)),
            run_title=bool(getattr(status, "run_llm_title", False)),
            run_catalog=bool(getattr(status, "run_catalog", False)),
            run_citation_verify=bool(getattr(status, "run_citation_verify", False)),
            run_rating=bool(getattr(status, "run_llm_rating", False)),
            run_summary=bool(getattr(status, "run_llm_summary", False)),
        )
        counts = _build_agent_run_counts(
            rows=rows,
            selected_phases=selected_phases,
            fallback_status=status,
        )

        manifest_csv_path = store.get_sources_manifest_csv_path(normalized_run_id)
        manifest_xlsx_path = store.get_sources_manifest_xlsx_path(normalized_run_id)
        bundle_path = store.get_sources_bundle_path(normalized_run_id)
        output_summary = {}
        if status is not None:
            output_summary = status.output_summary.model_dump(mode="json")
        elif rows:
            output_summary = summarize_output_rows(rows).model_dump(mode="json")

        phase_states: dict[str, SourcePhaseMetadata] = {}
        if status is not None:
            phase_states = {
                key: value.model_copy(deep=True)
                for key, value in status.phase_states.items()
            }
        elif selected_phases:
            phase_states = {
                phase: SourcePhaseMetadata(phase=phase, status="pending")
                for phase in selected_phases
            }

        selected_source_ids = [
            str(item).strip()
            for item in context.get("selected_ids", [])
            if str(item).strip()
        ]
        snapshot = AgentRunSnapshot(
            manifest_csv=str(manifest_csv_path) if manifest_csv_path.exists() else "",
            manifest_xlsx=str(manifest_xlsx_path) if manifest_xlsx_path.exists() else "",
            bundle_file=str(bundle_path) if bundle_path.exists() else "",
            repository_path=str(context.get("repository_path") or self.path),
            output_summary=output_summary,
        )
        current_item = AgentRunCurrentItem(
            source_id=str(getattr(status, "current_item_id", "") or ""),
            url=str(getattr(status, "current_url", "") or ""),
        )
        return AgentRunRecord(
            run_id=normalized_run_id,
            scope=str(context.get("scope") or getattr(status, "selected_scope", "") or ""),
            import_id=str(
                context.get("import_id") or getattr(status, "selected_import_id", "") or ""
            ),
            phase_states=phase_states,
            counts=counts,
            current_item=current_item,
            selected_source_ids=selected_source_ids,
            started_at=str(getattr(status, "started_at", "") or ""),
            completed_at=str(getattr(status, "completed_at", "") or ""),
            cancel_requested=bool(getattr(status, "cancel_requested", False)),
            result_snapshot=snapshot,
        )

    def cancel_agent_run(
        self,
        run_id: str,
        *,
        live_jobs: dict[str, SourceDownloadOrchestrator] | None = None,
        live_jobs_lock: threading.Lock | None = None,
    ) -> AgentRunRecord:
        normalized_run_id = str(run_id or "").strip()
        if not normalized_run_id:
            raise ValueError("run_id is required")

        store = self.job_store_for(normalized_run_id)
        if not store.job_exists(normalized_run_id):
            raise ValueError(f"Unknown run_id: {normalized_run_id}")

        orchestrator: SourceDownloadOrchestrator | None = None
        if live_jobs is not None and live_jobs_lock is not None:
            with live_jobs_lock:
                orchestrator = live_jobs.get(normalized_run_id)
        if orchestrator is not None:
            orchestrator.request_cancel()
            message = (
                getattr(getattr(orchestrator, "status", None), "message", "")
                or "Stop requested. Finishing the current item before stopping."
            )
            if getattr(orchestrator, "writes_to_repository", False):
                self.mark_source_tasks_cancelling(message)
        else:
            status = store.get_source_status(normalized_run_id) or {}
            state = str(status.get("state") or "").strip().lower()
            if state not in {"cancelled", "completed", "failed", "cancelling"}:
                raise RuntimeError(
                    "Run is not currently cancellable because no live worker handle is available."
                )

        return self.get_agent_run(
            normalized_run_id,
            live_jobs=live_jobs,
            live_jobs_lock=live_jobs_lock,
        )

    def _load_source_text_artifact(
        self,
        row: SourceManifestRow,
        field_name: str,
        *,
        fallback_field: str = "",
    ) -> str:
        rel_value = str(getattr(row, field_name) or "").strip()
        source_path = self._resolve_repository_artifact_path(row, field_name, rel_value)
        if source_path is None and fallback_field:
            fallback_value = str(getattr(row, fallback_field) or "").strip()
            source_path = self._resolve_repository_artifact_path(
                row,
                fallback_field,
                fallback_value,
            )
        if source_path is None or not source_path.is_file():
            return ""
        return source_path.read_text(encoding="utf-8", errors="replace")

    def _build_agent_source_record(
        self,
        *,
        row: SourceManifestRow,
        imports: list[dict[str, Any]],
        manifest_record: dict[str, Any] | None = None,
    ) -> AgentSourceRecord:
        derived = manifest_record or build_manifest_record(row, base_dir=self.path)
        import_id = _row_import_id(row)
        rating_overall = _coerce_optional_float(derived.get("rating_overall"))
        rating_confidence = _coerce_optional_float(derived.get("rating_confidence"))
        markdown_digest = _file_sha256(
            self._resolve_repository_artifact_path(row, "markdown_file", row.markdown_file)
        )
        clean_markdown_path = self._resolve_repository_artifact_path(
            row,
            "llm_cleanup_file",
            row.llm_cleanup_file,
        )
        if clean_markdown_path is None:
            clean_markdown_path = self._resolve_repository_artifact_path(
                row,
                "markdown_file",
                row.markdown_file,
            )
        clean_markdown_digest = _file_sha256(clean_markdown_path)
        summary_digest = _file_sha256(
            self._resolve_repository_artifact_path(row, "summary_file", row.summary_file)
        )
        rating_digest = _file_sha256(
            self._resolve_repository_artifact_path(row, "rating_file", row.rating_file)
        )
        summary_metadata = row.phase_metadata.get(PHASE_SUMMARY)
        tag_metadata = row.phase_metadata.get(PHASE_RATING)
        summary_stale = _phase_is_stale(summary_metadata, clean_markdown_digest)
        rating_stale = _phase_is_stale(tag_metadata, clean_markdown_digest)

        return AgentSourceRecord(
            source_id=row.id,
            original_url=row.original_url,
            final_url=row.final_url or row.original_url,
            title=row.title,
            detected_type=row.detected_type,
            fetch_status=_agent_fetch_status(row),
            convert_status=_agent_phase_status(row, PHASE_CONVERT),
            tag_status=_agent_phase_status(row, PHASE_RATING),
            summarize_status=_agent_phase_status(row, PHASE_SUMMARY),
            rating_overall=rating_overall,
            rating_confidence=rating_confidence,
            summary_present=bool(summary_digest),
            rating_present=bool(rating_digest),
            content_digests={
                "fetch": row.sha256,
                "markdown": markdown_digest,
                "clean_markdown": clean_markdown_digest,
                "summary": summary_digest,
                "summary_source": str(getattr(summary_metadata, "content_digest", "") or ""),
                "rating": rating_digest,
                "rating_source": str(getattr(tag_metadata, "content_digest", "") or ""),
            },
            artifact_uris=AgentSourceArtifactUris(
                markdown=(
                    _agent_source_artifact_uri(row.id, "markdown") if markdown_digest else ""
                ),
                clean_markdown=(
                    _agent_source_artifact_uri(row.id, "clean_markdown")
                    if clean_markdown_digest
                    else ""
                ),
                summary=(
                    _agent_source_artifact_uri(row.id, "summary") if summary_digest else ""
                ),
                rating=(
                    _agent_source_artifact_uri(row.id, "rating") if rating_digest else ""
                ),
                metadata=_agent_source_artifact_uri(row.id, "metadata"),
            ),
            provenance=AgentSourceProvenance(
                import_id=import_id,
                import_type=row.import_type,
                imported_at=row.imported_at,
                provenance_ref=row.provenance_ref,
                repository_path=str(self.path),
                repository_source_id=row.repository_source_id or row.id,
                source_document_name=row.source_document_name,
                citation_number=row.citation_number,
            ),
            freshness=AgentSourceFreshness(
                summary_stale=summary_stale,
                rating_stale=rating_stale,
            ),
            phase_metadata={
                key: value.model_copy(deep=True) for key, value in row.phase_metadata.items()
            },
            updated_at=_row_updated_at(row, self.path, imports=imports),
        )

    def resolve_source_file(
        self,
        *,
        source_id: str,
        kind: str,
    ) -> tuple[Path, str, dict[str, str]]:
        if not self.is_attached:
            raise ValueError("No repository attached")

        normalized_kind = _normalize_source_file_kind(kind)
        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))

        row = next((item for item in rows if item.id == source_id), None)
        if row is None:
            raise ValueError(f"Source `{source_id}` not found")

        source_path = self._resolve_source_file_path_for_kind(row, normalized_kind)
        if source_path is None or not source_path.is_file():
            raise ValueError(
                f"No file available for `{normalized_kind}` on source `{source_id}`"
            )

        media_type = _media_type_for_repository_source_path(source_path, normalized_kind)
        headers = _repository_source_file_headers(source_path)
        return source_path, media_type, headers

    def find_duplicate_source_candidates(self) -> RepositoryDuplicateCandidateResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before scanning for duplicate sources")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))

        if len(rows) < 2:
            return RepositoryDuplicateCandidateResponse(
                status="completed",
                scanned_sources=len(rows),
                total_groups=0,
                total_candidate_rows=0,
                truncated=False,
                message="No likely duplicates found.",
                groups=[],
            )

        contexts = [
            _build_duplicate_candidate_context(row, base_dir=self.path)
            for row in self._sort_rows(rows)
        ]
        assigned_ids: set[str] = set()
        groups: list[RepositoryDuplicateCandidateGroup] = []
        truncated = False

        def append_group(
            candidates: list[_DuplicateCandidateContext],
            *,
            reason: str,
            confidence: str,
        ) -> None:
            nonlocal truncated
            available = [
                candidate
                for candidate in candidates
                if candidate.row.id not in assigned_ids
            ]
            unique_available = sorted(
                {candidate.row.id: candidate for candidate in available}.values(),
                key=lambda candidate: _sortable_source_id(candidate.row.id),
            )
            if len(unique_available) < 2:
                return
            if len(groups) >= DUPLICATE_SCAN_MAX_GROUPS:
                truncated = True
                return
            groups.append(
                _build_duplicate_candidate_group(
                    unique_available,
                    group_id=f"dup-{len(groups) + 1:04d}",
                    match_reason=reason,
                    confidence=confidence,
                )
            )
            assigned_ids.update(candidate.row.id for candidate in unique_available)

        doi_buckets: dict[str, list[_DuplicateCandidateContext]] = {}
        url_buckets: dict[str, list[_DuplicateCandidateContext]] = {}
        sha_buckets: dict[str, list[_DuplicateCandidateContext]] = {}
        for context in contexts:
            if context.doi_key:
                doi_buckets.setdefault(context.doi_key, []).append(context)
            for url_key in context.url_keys:
                url_buckets.setdefault(url_key, []).append(context)
            if context.sha_key:
                sha_buckets.setdefault(context.sha_key, []).append(context)

        for bucket in _duplicate_contexts_by_bucket_size(doi_buckets):
            append_group(bucket, reason="Matching DOI", confidence="high")
        for bucket in _duplicate_contexts_by_bucket_size(url_buckets):
            append_group(bucket, reason="Matching normalized URL", confidence="high")
        for bucket in _duplicate_contexts_by_bucket_size(sha_buckets):
            append_group(bucket, reason="Matching source content digest", confidence="high")

        remaining_contexts = [
            context for context in contexts if context.row.id not in assigned_ids
        ]
        author_year_buckets: dict[str, list[_DuplicateCandidateContext]] = {}
        organization_year_buckets: dict[str, list[_DuplicateCandidateContext]] = {}
        for context in remaining_contexts:
            if context.year and context.author_signature and len(context.title_token_set) >= 3:
                author_year_buckets.setdefault(
                    f"{context.year}|{context.author_signature}",
                    [],
                ).append(context)
            if (
                context.year
                and context.organization_signature
                and len(context.title_token_set) >= 3
            ):
                organization_year_buckets.setdefault(
                    f"{context.year}|{context.organization_signature}",
                    [],
                ).append(context)

        for bucket in _duplicate_contexts_by_bucket_size(author_year_buckets):
            for cluster in _cluster_duplicate_title_contexts(bucket):
                append_group(
                    cluster,
                    reason="Similar title, year, and authors",
                    confidence="medium",
                )
        for bucket in _duplicate_contexts_by_bucket_size(organization_year_buckets):
            for cluster in _cluster_duplicate_title_contexts(bucket):
                append_group(
                    cluster,
                    reason="Similar title, year, and organization",
                    confidence="medium",
                )

        total_candidate_rows = sum(len(group.rows) for group in groups)
        if groups:
            message = f"Found {len(groups)} potential duplicate group(s)."
            if truncated:
                message += " Showing the first 100 groups."
        else:
            message = "No likely duplicates found."

        return RepositoryDuplicateCandidateResponse(
            status="completed",
            scanned_sources=len(contexts),
            total_groups=len(groups),
            total_candidate_rows=total_candidate_rows,
            truncated=truncated,
            message=message,
            groups=groups,
        )

    def delete_sources(self, source_ids: list[str]) -> RepositorySourceDeleteResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before deleting sources")

        normalized_ids = _normalize_source_ids(source_ids)
        if not normalized_ids:
            raise ValueError("At least one source id is required")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            imports = list(state.get("imports", []))
            meta = self._load_meta_locked()

            selected_rows = [row for row in rows if row.id in normalized_ids]
            if not selected_rows:
                raise ValueError("No matching repository sources were found")

            remaining_rows = [row for row in rows if row.id not in normalized_ids]
            protected_paths = self._protected_repository_file_paths(remaining_rows)

            deleted_files = 0
            for row in selected_rows:
                for field_name in FILE_FIELDS:
                    rel_value = str(getattr(row, field_name) or "").strip()
                    if not rel_value:
                        continue
                    source_path = self._resolve_repository_artifact_path(
                        row,
                        field_name,
                        rel_value,
                    )
                    if source_path is None or not source_path.is_file():
                        continue
                    if not self._is_path_within_repo(source_path):
                        continue
                    resolved_key = str(source_path.resolve())
                    if resolved_key in protected_paths:
                        continue
                    try:
                        source_path.unlink()
                    except OSError:
                        continue
                    deleted_files += 1
                    self._cleanup_empty_repository_dirs(source_path.parent)

            remaining_citations = [
                citation
                for citation in citations
                if (citation.repository_source_id or "") not in normalized_ids
            ]
            deleted_citations = len(citations) - len(remaining_citations)
            sorted_rows = self._sort_rows(remaining_rows)
            sorted_citations = self._sort_citations(remaining_citations)
            next_source_id = max(
                int(meta.get("next_source_id") or 1),
                _next_source_id_from_rows(sorted_rows),
            )

            self._save_state_locked(
                sources=sorted_rows,
                citations=sorted_citations,
                imports=imports,
            )
            self._save_meta_locked(
                {
                    **meta,
                    "next_source_id": next_source_id,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(sorted_rows, sorted_citations)
            self._download_message = (
                f"Deleted {len(selected_rows)} source(s), "
                f"{deleted_citations} citation row(s), and {deleted_files} file(s)"
            )

            return RepositorySourceDeleteResponse(
                status="completed",
                deleted_sources=len(selected_rows),
                deleted_citations=deleted_citations,
                deleted_files=deleted_files,
                total_sources=len(sorted_rows),
                total_citations=len(sorted_citations),
                message=self._download_message,
            )

    def export_source_files(
        self,
        *,
        source_ids: list[str],
        file_kinds: list[str],
        destination_path: str,
    ) -> RepositorySourceExportResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before exporting files")

        normalized_ids = _normalize_source_ids(source_ids)
        if not normalized_ids:
            raise ValueError("At least one source id is required")

        normalized_kinds = _normalize_source_file_kinds(file_kinds)
        if not normalized_kinds:
            raise ValueError("At least one file kind is required")

        destination_dir = _resolve_export_destination_path(destination_path)

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))

        selected_rows = self._sort_rows([row for row in rows if row.id in normalized_ids])
        if not selected_rows:
            raise ValueError("No matching repository sources were found")

        used_names = {item.name.lower() for item in destination_dir.iterdir() if item.is_file()}
        exported_files = 0
        missing_files = 0

        for row in selected_rows:
            for kind_name in normalized_kinds:
                source_path = self._resolve_source_file_path_for_kind(row, kind_name)
                if source_path is None or not source_path.is_file():
                    missing_files += 1
                    continue
                target_name = _build_flat_export_filename(
                    source_id=row.id,
                    title=row.title,
                    extension=source_path.suffix or _default_extension_for_source_kind(kind_name),
                    used_names=used_names,
                )
                shutil.copy2(source_path, destination_dir / target_name)
                exported_files += 1

        message = (
            f"Exported {exported_files} file(s) from {len(selected_rows)} source(s) "
            f"to {destination_dir}"
        )
        if missing_files:
            message += f" ({missing_files} missing file(s) skipped)"

        return RepositorySourceExportResponse(
            status="completed",
            requested_sources=len(selected_rows),
            exported_files=exported_files,
            missing_files=missing_files,
            destination_path=str(destination_dir),
            message=message,
        )

    def export_repository_bundle(
        self,
        payload: RepositoryBundleExportRequest,
    ) -> tuple[bytes, dict[str, str]]:
        if not self.is_attached:
            raise ValueError("Attach a repository before exporting a repository bundle")

        normalized_scope = str(payload.scope or "all").strip().lower() or "all"
        if normalized_scope not in {"all", "selected"}:
            raise ValueError("Invalid scope. Use `all` or `selected`.")
        normalized_mode = str(payload.mode or "offline").strip().lower() or "offline"
        if normalized_mode not in {"offline", "cloud"}:
            raise ValueError("Invalid mode. Use `offline` or `cloud`.")
        normalized_base_url = (
            _normalize_repository_bundle_base_url(payload.base_url)
            if normalized_mode == "cloud"
            else ""
        )
        normalized_file_kinds = _normalize_repository_bundle_file_kinds(payload.file_kinds)

        _scope, target_rows = self._resolve_manifest_export_rows(
            scope=normalized_scope,
            source_ids=payload.source_ids,
            filters=RepositoryManifestFilterRequest(),
        )
        if not target_rows:
            raise ValueError("No repository rows are available for export.")

        with self._writer_lock():
            column_configs = _load_column_configs(self._load_state_locked().get("column_configs", []))

        manifest_records = {
            row.id: build_manifest_record(row, base_dir=self.path, column_configs=column_configs)
            for row in target_rows
        }
        custom_configs, base_headers, custom_headers, csv_headers = _repository_bundle_csv_layout(column_configs)
        csv_export_records = [
            _build_repository_bundle_csv_record(
                manifest_records[row.id],
                custom_configs=custom_configs,
                base_headers=base_headers,
                custom_headers=custom_headers,
            )
            for row in target_rows
        ]
        csv_text = _render_repository_bundle_csv(csv_headers, csv_export_records)
        csv_records_by_id = {
            row.id: export_record
            for row, export_record in zip(target_rows, csv_export_records)
        }
        ris_text, ris_exported_count, ris_skipped_count = build_ris_records(
            target_rows,
            base_dir=self.path,
        )
        zip_filename = (
            "selected-repository-cloud-export.zip"
            if normalized_mode == "cloud" and normalized_scope == "selected"
            else "repository-cloud-export.zip"
            if normalized_mode == "cloud"
            else "selected-repository-export.zip"
            if normalized_scope == "selected"
            else "repository-export.zip"
        )

        selected_bundle_kinds = set(normalized_file_kinds)
        use_type_directories = normalized_mode == "offline" and selected_bundle_kinds in (
            {"pdf", "html", "md"},
            set(REPOSITORY_BUNDLE_FILE_KINDS),
        )
        directory_names = {
            "pdf": "PDF",
            "rendered": "RENDERED",
            "html": "HTML",
            "md": "MD",
        }
        used_names_by_directory: dict[str, set[str]] = {}
        artifact_entries_by_row: dict[str, dict[str, dict[str, str]]] = {}
        file_entries: list[tuple[str, Path]] = []
        exported_files = 0
        missing_files = 0

        for row in target_rows:
            record = manifest_records[row.id]
            row_artifact_entries = artifact_entries_by_row.setdefault(row.id, {})
            for kind in normalized_file_kinds:
                source_path = _resolve_repository_bundle_file_path(self, row, kind)
                if source_path is None or not source_path.is_file():
                    missing_files += 1
                    continue
                display_name = source_path.name
                if normalized_mode == "cloud":
                    storage_name = _build_repository_bundle_storage_filename(
                        row=row,
                        kind=kind,
                        display_name=display_name,
                        extension=source_path.suffix or _default_extension_for_source_kind(kind),
                    )
                    archive_name = f"files/{storage_name}"
                    row_artifact_entries[kind] = {
                        "kind": kind,
                        "displayName": display_name,
                        "storageName": storage_name,
                    }
                else:
                    directory = directory_names[kind] if use_type_directories else ""
                    used_names = used_names_by_directory.setdefault(directory, set())
                    filename = _build_repository_bundle_filename(
                        row=row,
                        record=record,
                        extension=source_path.suffix or _default_extension_for_source_kind(kind),
                        used_names=used_names,
                    )
                    archive_name = f"{directory}/{filename}" if directory else filename
                    row_artifact_entries[kind] = {
                        "kind": kind,
                        "displayName": display_name,
                        "relativePath": archive_name,
                    }
                file_entries.append((archive_name, source_path))
                exported_files += 1

        viewer_payload = {
            "repositoryName": self.path.name if self.path is not None else "Repository Browser",
            "exportMode": normalized_mode,
            "exportScope": normalized_scope,
            "bundleFileKinds": normalized_file_kinds,
            "exportedAt": datetime.now(timezone.utc).isoformat(),
            "csvHeaders": csv_headers,
            "rows": [
                _build_repository_bundle_viewer_row(
                    self,
                    row,
                    record=manifest_records[row.id],
                    csv_record=csv_records_by_id[row.id],
                    custom_configs=custom_configs,
                    custom_headers=custom_headers,
                    artifact_entries=artifact_entries_by_row.get(row.id, {}),
                )
                for row in target_rows
            ],
        }
        viewer_html = build_repository_bundle_viewer_html(
            viewer_payload,
            base_url=(
                normalized_base_url
                if normalized_mode == "cloud"
                else ""
            ),
        )

        archive = io.BytesIO()
        with zipfile.ZipFile(archive, mode="w", compression=zipfile.ZIP_DEFLATED) as bundle:
            bundle.writestr("index.html", viewer_html.encode("utf-8"))
            bundle.writestr("research-export.csv", csv_text.encode("utf-8-sig"))
            bundle.writestr("citations.ris", ris_text.encode("utf-8"))
            if normalized_mode == "cloud":
                bundle.writestr(
                    "manifest.json",
                    json.dumps(viewer_payload, ensure_ascii=False, indent=2).encode("utf-8"),
                )
            for archive_name, source_path in file_entries:
                bundle.writestr(archive_name, source_path.read_bytes())

        archive.seek(0)
        auxiliary_file_count = 4 if normalized_mode == "cloud" else 3
        headers = {
            "Content-Disposition": f'attachment; filename="{zip_filename}"',
            "X-ResearchAssistant-Requested-Count": str(len(target_rows)),
            "X-ResearchAssistant-Exported-Count": str(exported_files + auxiliary_file_count),
            "X-ResearchAssistant-Skipped-Count": str(missing_files + ris_skipped_count),
        }
        return archive.getvalue(), headers

    def _resolve_manifest_export_rows(
        self,
        *,
        scope: str,
        source_ids: list[str],
        filters: RepositoryManifestFilterRequest,
    ) -> tuple[str, list[SourceManifestRow]]:
        normalized_scope = str(scope or "all").strip().lower() or "all"
        if normalized_scope not in {"all", "filtered", "selected"}:
            raise ValueError("Invalid scope. Use `all`, `filtered`, or `selected`.")

        records, columns = self._manifest_records()
        filtered_records, _, _ = self._filter_manifest_records(
            records,
            columns,
            q=filters.q,
            fetch_status=filters.fetch_status,
            detected_type=filters.detected_type,
            source_kind=filters.source_kind,
            document_type=filters.document_type,
            organization_type=filters.organization_type,
            organization_name=filters.organization_name,
            author_names=filters.author_names,
            publication_date=filters.publication_date,
            tags_text=filters.tags_text,
            has_summary=filters.has_summary,
            has_rating=filters.has_rating,
            rating_overall_min=filters.rating_overall_min,
            rating_overall_max=filters.rating_overall_max,
            rating_overall_relevance_min=filters.rating_overall_relevance_min,
            rating_overall_relevance_max=filters.rating_overall_relevance_max,
            rating_depth_score_min=filters.rating_depth_score_min,
            rating_depth_score_max=filters.rating_depth_score_max,
            rating_relevant_detail_score_min=filters.rating_relevant_detail_score_min,
            rating_relevant_detail_score_max=filters.rating_relevant_detail_score_max,
            citation_type=filters.citation_type,
            citation_doi=filters.citation_doi,
            citation_report_number=filters.citation_report_number,
            citation_standard_number=filters.citation_standard_number,
            citation_missing_fields=filters.citation_missing_fields,
            citation_ready=filters.citation_ready,
            citation_confidence_min=filters.citation_confidence_min,
            citation_confidence_max=filters.citation_confidence_max,
            sort_by="id",
            sort_dir="asc",
        )

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))

        row_by_id = {row.id: row for row in rows}
        if normalized_scope == "all":
            target_records = _sort_manifest_records(records, sort_by="id", reverse=False)
            ordered_ids = [str(record.get("id") or "") for record in target_records]
        elif normalized_scope == "filtered":
            ordered_ids = [str(record.get("id") or "") for record in filtered_records]
        else:
            normalized_ids = [str(source_id or "").strip() for source_id in source_ids if str(source_id or "").strip()]
            if not normalized_ids:
                raise ValueError("At least one source id is required for selected export")
            missing = sorted(source_id for source_id in normalized_ids if source_id not in row_by_id)
            if missing:
                raise ValueError(f"Unknown source_ids: {', '.join(missing[:20])}")
            ordered_ids = normalized_ids

        target_rows = [row_by_id[source_id] for source_id in ordered_ids if source_id in row_by_id]
        return normalized_scope, target_rows

    def export_citations_ris(
        self,
        payload: RepositoryCitationRisExportRequest,
    ) -> tuple[bytes, dict[str, str]]:
        if not self.is_attached:
            raise ValueError("Attach a repository before exporting RIS citations")

        normalized_scope, target_rows = self._resolve_manifest_export_rows(
            scope=payload.scope,
            source_ids=payload.source_ids,
            filters=payload.filters,
        )
        ris_text, exported_count, skipped_count = build_ris_records(target_rows, base_dir=self.path)
        requested_count = len(target_rows)
        filename = {
            "all": "repository-citations.ris",
            "filtered": "filtered-citations.ris",
            "selected": "selected-citations.ris",
        }[normalized_scope]
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-ResearchAssistant-Requested-Count": str(requested_count),
            "X-ResearchAssistant-Exported-Count": str(exported_count),
            "X-ResearchAssistant-Skipped-Count": str(skipped_count),
        }
        return ris_text.encode("utf-8"), headers

    def export_manifest(
        self,
        payload: RepositoryManifestExportRequest,
    ) -> tuple[bytes, dict[str, str], str]:
        if not self.is_attached:
            raise ValueError("Attach a repository before exporting spreadsheet data")

        normalized_scope, target_rows = self._resolve_manifest_export_rows(
            scope=payload.scope,
            source_ids=payload.source_ids,
            filters=payload.filters,
        )
        normalized_format = str(payload.format or "csv").strip().lower() or "csv"
        if normalized_format not in {"csv", "xlsx"}:
            raise ValueError("Invalid format. Use `csv` or `xlsx`.")

        with self._writer_lock():
            column_configs = _load_column_configs(self._load_state_locked().get("column_configs", []))
        fieldnames, records = _build_manifest_export_rows(
            target_rows,
            base_dir=self.path,
            column_configs=column_configs,
            column_scope=payload.column_scope,
            column_keys=payload.column_keys,
        )

        if normalized_format == "xlsx":
            content = _build_manifest_export_xlsx(fieldnames, records)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = {
                "all": "repository-manifest.xlsx",
                "filtered": "filtered-manifest.xlsx",
                "selected": "selected-manifest.xlsx",
            }[normalized_scope]
        else:
            content = _build_manifest_export_csv(fieldnames, records).encode("utf-8-sig")
            media_type = "text/csv; charset=utf-8"
            filename = {
                "all": "repository-manifest.csv",
                "filtered": "filtered-manifest.csv",
                "selected": "selected-manifest.csv",
            }[normalized_scope]

        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-ResearchAssistant-Requested-Count": str(len(target_rows)),
            "X-ResearchAssistant-Exported-Count": str(len(target_rows)),
            "X-ResearchAssistant-Skipped-Count": "0",
        }
        return content, headers, media_type

    def get_citation_data(self) -> dict[str, Any]:
        if not self.is_attached:
            raise ValueError("No repository attached")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_citation_rows(state.get("citations", []))

        bibliography_by_ref: dict[int, dict[str, Any]] = {}
        for row in rows:
            ref_numbers = _extract_ref_numbers(row.citation_ref_numbers or row.citation_raw)
            for ref_number in ref_numbers:
                if ref_number in bibliography_by_ref:
                    continue
                bibliography_by_ref[ref_number] = {
                    "ref_number": ref_number,
                    "authors": [
                        chunk.strip()
                        for chunk in str(row.cited_authors or "").split(";")
                        if chunk.strip()
                    ],
                    "title": row.cited_title or "",
                    "year": row.cited_year or "",
                    "url": row.cited_url or "",
                    "doi": row.cited_doi or "",
                    "raw_text": row.cited_raw_entry or row.cited_title or row.cited_url or "",
                    "parse_confidence": _safe_float(row.match_confidence),
                }

        bibliography_entries = [
            bibliography_by_ref[ref_number]
            for ref_number in sorted(bibliography_by_ref.keys())
        ]
        ref_to_index = {
            int(entry["ref_number"]): idx
            for idx, entry in enumerate(bibliography_entries)
            if int(entry.get("ref_number") or 0) > 0
        }

        citations_payload: list[dict[str, Any]] = []
        sentences_payload: list[dict[str, Any]] = []
        matches_payload: list[dict[str, Any]] = []

        for idx, row in enumerate(rows, start=1):
            ref_numbers = _extract_ref_numbers(row.citation_ref_numbers or row.citation_raw)
            raw_marker = str(row.citation_raw or "").strip()
            if not raw_marker and ref_numbers:
                raw_marker = "[" + ", ".join(str(item) for item in ref_numbers) + "]"
            page_number = _parse_int(str(row.page_in_source or ""))
            style = "bracket" if raw_marker.startswith("[") else "unknown"

            citations_payload.append(
                {
                    "raw_marker": raw_marker,
                    "ref_numbers": ref_numbers,
                    "page_number": page_number,
                    "style": style,
                }
            )

            sentence_text = str(row.citing_sentence or "").strip()
            paragraph_text = str(row.citing_paragraph or "").strip()
            if sentence_text or paragraph_text:
                citation_id = f"repo-citation-{idx}"
                sentences_payload.append(
                    {
                        "page_number": page_number,
                        "text": paragraph_text or sentence_text,
                        "paragraph": paragraph_text or sentence_text,
                        "citation_ids": [citation_id],
                    }
                )

            confidence = _safe_float(row.match_confidence)
            for ref_number in ref_numbers:
                matches_payload.append(
                    {
                        "ref_number": ref_number,
                        "matched_bib_entry_index": ref_to_index.get(ref_number),
                        "match_confidence": confidence,
                        "match_method": row.match_method or "repository_citation",
                    }
                )

        return {
            "bibliography": {"entries": bibliography_entries},
            "citations": {
                "citations": citations_payload,
                "sentences": sentences_payload,
                "matches": matches_payload,
            },
        }

    def _collect_recent_jobs(self, limit: int) -> list[dict[str, Any]]:
        cap = max(1, int(limit))
        artifacts_dir = self.repo_job_store().artifacts_dir if self.is_attached else self.store.artifacts_dir
        if not artifacts_dir.is_dir():
            return []

        job_store = self.repo_job_store() if self.is_attached else self.store

        recent_jobs: list[dict[str, Any]] = []
        for job_dir in artifacts_dir.iterdir():
            if not job_dir.is_dir():
                continue
            job_id = job_dir.name

            extraction = job_store.get_job_status(job_id) or {}
            if extraction:
                preprocess_state = str(
                    extraction.get("repository_preprocess_state") or ""
                ).strip()
                preprocess_message = str(
                    extraction.get("repository_preprocess_message") or ""
                ).strip()
                finalize_state = str(
                    extraction.get("repository_finalize_state") or ""
                ).strip()
                finalize_message = str(
                    extraction.get("repository_finalize_message") or ""
                ).strip()
                updated_at = (
                    str(extraction.get("repository_preprocess_updated_at") or "")
                    or str(extraction.get("repository_finalize_updated_at") or "")
                    or str(extraction.get("completed_at") or "")
                    or _latest_stage_timestamp(extraction.get("stages", []))
                    or str(extraction.get("created_at") or "")
                )
                progress_pct = float(extraction.get("progress_pct") or 0.0)
                current_stage = str(extraction.get("current_stage") or "pending")
                if preprocess_state in {"pending", "running"}:
                    recent_jobs.append(
                        {
                            "job_id": job_id,
                            "kind": "citation_extraction",
                            "state": preprocess_state,
                            "updated_at": updated_at,
                            "message": preprocess_message
                            or "Preparing standardized markdown.",
                            "progress_pct": progress_pct,
                        }
                    )
                    continue

                updated_at = (
                    str(extraction.get("repository_finalize_updated_at") or "")
                    or str(extraction.get("completed_at") or "")
                    or _latest_stage_timestamp(extraction.get("stages", []))
                    or str(extraction.get("created_at") or "")
                )
                if finalize_state in {"pending", "running"}:
                    recent_jobs.append(
                        {
                            "job_id": job_id,
                            "kind": "citation_extraction",
                            "state": finalize_state,
                            "updated_at": updated_at,
                            "message": finalize_message
                            or "Finalizing repository updates.",
                            "progress_pct": progress_pct,
                        }
                    )
                    continue

                updated_at = (
                    str(extraction.get("completed_at") or "")
                    or _latest_stage_timestamp(extraction.get("stages", []))
                    or str(extraction.get("created_at") or "")
                )
                recent_jobs.append(
                        {
                            "job_id": job_id,
                            "kind": "citation_extraction",
                            "state": current_stage,
                            "updated_at": updated_at,
                            "message": (
                                preprocess_message
                                if preprocess_state == "failed" and preprocess_message
                                else finalize_message
                                if finalize_state == "failed" and finalize_message
                                else f"stage={current_stage} progress={progress_pct:.0f}%"
                            ),
                            "progress_pct": progress_pct,
                        }
                    )

            source = job_store.get_source_status(job_id) or {}
            if source:
                updated_at = (
                    str(source.get("completed_at") or "")
                    or str(source.get("started_at") or "")
                )
                state = str(source.get("state") or "pending")
                processed = int(source.get("processed_urls") or 0)
                total = int(source.get("total_urls") or 0)
                recent_jobs.append(
                    {
                        "job_id": job_id,
                        "kind": "source_capture",
                        "state": state,
                        "updated_at": updated_at,
                        "message": f"state={state} processed={processed}/{total}",
                        "processed_urls": processed,
                        "total_urls": total,
                    }
                )

        recent_jobs.sort(
            key=lambda item: str(item.get("updated_at") or ""),
            reverse=True,
        )
        return recent_jobs[:cap]

    def import_source_list(self, filename: str, content: bytes) -> RepositoryImportResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before importing")
        return self.import_seed_files([(filename, content)])

    def import_seed_files(
        self,
        files: list[tuple[str, bytes]],
    ) -> RepositoryImportResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before importing")

        all_entries: list[BibliographyEntry] = []
        provenance_parts: list[str] = []
        for filename, content in files:
            entries = self._extract_seed_entries_from_file(filename=filename, content=content)
            if entries:
                all_entries.extend(entries)
                provenance_parts.append(Path(filename or "seed").name)

        if not all_entries:
            raise ValueError(
                "No usable links were found. Seed uploads support .csv, .xlsx, .pdf, .docx, and .md."
            )

        return self._import_entries(
            entries=all_entries,
            import_type="source_seed",
            provenance_label=", ".join(provenance_parts) or "seed_upload",
            default_source_document="seed_upload",
            write_placeholder_citations=False,
            source_kind="url",
        )

    def import_manual_documents(
        self,
        files: list[tuple[str, bytes]],
    ) -> RepositoryImportResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before importing")

        import_id = uuid.uuid4().hex[:12]
        imported_at = _utc_now_iso()
        provenance_parts: list[str] = []

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            imports = list(state.get("imports", []))
            meta = self._load_meta_locked()
            next_source_id = int(meta.get("next_source_id") or _next_source_id_from_rows(rows))

            existing_by_sha = {
                str(row.sha256 or "").strip(): row
                for row in rows
                if row.source_kind == "uploaded_document" and str(row.sha256 or "").strip()
            }

            accepted_new = 0
            duplicates = 0
            total_candidates = 0
            document_records: list[dict[str, str]] = []

            for original_filename, content in files:
                ext = Path(original_filename or "").suffix.lower()
                if ext not in SUPPORTED_MANUAL_SOURCE_EXTENSIONS:
                    continue

                total_candidates += 1
                provenance_parts.append(Path(original_filename or "document").name)
                sha256 = hashlib.sha256(content).hexdigest()
                existing = existing_by_sha.get(sha256)
                if existing is not None:
                    duplicates += 1
                    document_records.append(
                        {
                            "filename": Path(original_filename or "document").name,
                            "repository_path": existing.raw_file,
                            "sha256": sha256,
                            "source_id": existing.id,
                        }
                    )
                    continue

                source_id = f"{next_source_id:06d}"
                next_source_id += 1
                source_name = Path(original_filename or f"{source_id}{ext}").name
                raw_rel = _repository_source_file_path(
                    source_id=source_id,
                    field="raw_file",
                    source_name=source_name,
                )
                raw_abs = self.path / raw_rel
                raw_abs.parent.mkdir(parents=True, exist_ok=True)
                raw_abs.write_bytes(content)

                detected_type = _local_document_detected_type(ext)
                title = ""
                if ext == ".md":
                    title = _extract_markdown_seed_title(content.decode("utf-8", errors="replace"))

                row = SourceManifestRow(
                    id=source_id,
                    repository_source_id=source_id,
                    source_kind="uploaded_document",
                    import_type="document_source",
                    imported_at=imported_at,
                    provenance_ref=f"{import_id}:{source_name}",
                    source_document_name=source_name,
                    original_url="",
                    final_url="",
                    fetch_status="not_applicable",
                    content_type=mimetypes.guess_type(source_name)[0] or "",
                    detected_type=detected_type,
                    fetch_method="local_upload",
                    title=title,
                    title_status="extracted" if title else "not_requested",
                    raw_file=raw_rel.as_posix(),
                    notes="local_document",
                    fetched_at=imported_at,
                    sha256=sha256,
                )
                self._write_repository_source_metadata(row)
                rows.append(row)
                existing_by_sha[sha256] = row
                accepted_new += 1
                document_records.append(
                    {
                        "filename": source_name,
                        "repository_path": raw_rel.as_posix(),
                        "sha256": sha256,
                        "source_id": source_id,
                    }
                )

            if total_candidates == 0:
                raise ValueError(
                    "No supported documents were provided. Use .pdf, .doc, .docx, .html, .md, .rtf, or .txt."
                )

            sorted_rows = self._sort_rows(rows)
            self._save_state_locked(
                sources=sorted_rows,
                citations=citations,
                imports=[
                    *imports,
                    {
                        "import_id": import_id,
                        "import_type": "document_source",
                        "provenance": ", ".join(provenance_parts),
                        "imported_at": imported_at,
                        "total_candidates": total_candidates,
                        "accepted_new": accepted_new,
                        "duplicates_skipped": duplicates,
                        "source_ids": _dedupe_strings(
                            [str(item.get("source_id") or "").strip() for item in document_records]
                        ),
                        "documents": document_records,
                    },
                ],
            )
            self._save_meta_locked(
                {
                    **meta,
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": next_source_id,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(sorted_rows, citations)

        queued_count = sum(
            1 for row in sorted_rows if (row.fetch_status or "") in {"", "queued"}
        )
        return RepositoryImportResponse(
            import_id=import_id,
            import_type="document_source",
            total_candidates=total_candidates,
            accepted_new=accepted_new,
            duplicates_skipped=duplicates,
            total_sources=len(sorted_rows),
            queued_count=queued_count,
            message=(
                f"Added {accepted_new} new repository document(s)"
                + (f" ({duplicates} duplicate documents skipped)" if duplicates else "")
            ),
        )

    def import_document(self, filename: str, content: bytes) -> RepositoryImportResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before importing")
        settings = self.load_effective_settings()

        ext = Path(filename or "").suffix.lower()
        if ext not in SUPPORTED_DOCUMENT_IMPORT_EXTENSIONS:
            raise ValueError("Unsupported document type. Use .pdf, .docx, or .md.")

        with tempfile.TemporaryDirectory(prefix="repo-import-doc-") as tmp:
            tmp_path = Path(tmp)
            file_path = tmp_path / (filename or f"document{ext}")
            file_path.write_bytes(content)

            ingestion = run_ingestion(tmp_path)
            if not ingestion.documents:
                raise ValueError("No usable document content found")

            doc = ingestion.documents[0]
            section = detect_references_section(doc)
            sections = [section] if section else []
            artifact = parse_bibliography(
                sections,
                use_llm=settings.use_llm,
                llm_backend=settings.llm_backend,
            )

            if doc.inline_citation_urls:
                if artifact.entries:
                    artifact.entries = merge_inline_urls_into_entries(
                        artifact.entries,
                        doc.inline_citation_urls,
                    )
                else:
                    artifact.entries = build_entries_from_inline_urls(doc.inline_citation_urls)

            entries = artifact.entries
            if not entries:
                raise ValueError(
                    "No reference URLs found in document references/inline citation links"
                )

            for entry in entries:
                if not entry.source_document_name:
                    entry.source_document_name = doc.filename

            return self._import_entries(
                entries=entries,
                import_type="document",
                provenance_label=filename,
                default_source_document=doc.filename,
            )

    def _extract_seed_entries_from_file(
        self,
        *,
        filename: str,
        content: bytes,
    ) -> list[BibliographyEntry]:
        ext = Path(filename or "").suffix.lower()
        if ext in {".csv", ".xlsx"}:
            parsed = parse_source_list_upload(filename=filename, content=content)
            for entry in parsed.entries:
                if not entry.source_document_name:
                    entry.source_document_name = Path(filename).name
            return parsed.entries
        if ext not in SUPPORTED_SEED_IMPORT_EXTENSIONS:
            return []
        if ext == ".md":
            return _extract_seed_entries_from_markdown(
                filename=filename,
                text=content.decode("utf-8", errors="replace"),
            )

        with tempfile.TemporaryDirectory(prefix="repo-seed-import-") as tmp:
            tmp_path = Path(tmp)
            file_path = tmp_path / (Path(filename).name or f"document{ext}")
            file_path.write_bytes(content)
            ingestion = run_ingestion(tmp_path)
        if not ingestion.documents:
            return []
        return _extract_seed_entries_from_document(
            filename=filename,
            document=ingestion.documents[0],
        )

    def process_documents(
        self,
        files: list[tuple[str, bytes]],
        settings: EffectiveSettings | None = None,
        profile_override: str = "",
    ) -> RepositoryProcessDocumentsResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before processing documents")

        repo_settings = self.load_effective_settings(settings)
        selected_profile_id = self._validate_profile_override(profile_override)
        import_id = uuid.uuid4().hex[:12]
        documents_dir = self.documents_dir / import_id
        documents_dir.mkdir(parents=True, exist_ok=True)

        job_store = self.repo_job_store()
        job_id = job_store.create_job(prefix=REPO_JOB_PREFIX)
        accepted_documents = 0
        used_filenames: set[str] = set()
        document_records: list[dict[str, str]] = []
        pending_normalization: list[DocumentNormalizationResult] = []

        for original_filename, content in files:
            ext = Path(original_filename or "").suffix.lower()
            if ext not in SUPPORTED_DOCUMENT_IMPORT_EXTENSIONS:
                continue

            safe_name = _next_unique_filename(original_filename, used_filenames)
            repo_path = documents_dir / safe_name
            repo_path.write_bytes(content)

            document_records.append(
                {
                    "filename": safe_name,
                    "source_document_name": safe_name,
                    "repository_path": (Path(DOCUMENTS_DIR_NAME) / import_id / safe_name).as_posix(),
                    "sha256": hashlib.sha256(content).hexdigest(),
                    "document_import_id": import_id,
                }
            )
            pending_normalization.append(
                self._pending_normalization_result(
                    filename=safe_name,
                    source_document_path=(
                        Path(DOCUMENTS_DIR_NAME) / import_id / safe_name
                    ).as_posix(),
                    selected_profile_id=selected_profile_id,
                )
            )
            accepted_documents += 1

        if accepted_documents == 0:
            raise ValueError("No supported documents were provided. Use .pdf, .docx, or .md.")

        job_store.save_artifact(
            job_id,
            "repo_processing_context",
            {
                "import_id": import_id,
                "documents": document_records,
                "repository_path": str(self.path),
                "profile_override": selected_profile_id,
                "processing_mode": "process_documents",
            },
        )
        status = job_store.get_job_status(job_id) or {}
        status["selected_profile_id"] = selected_profile_id or "auto_detect"
        status["document_normalization"] = [
            item.model_dump(mode="json") for item in pending_normalization
        ]
        status["processing_mode"] = "process_documents"
        status["repository_preprocess_state"] = "pending"
        status["repository_preprocess_message"] = "Waiting to prepare standardized markdown."
        status["repository_preprocess_updated_at"] = _utc_now_iso()
        status["repository_finalize_state"] = ""
        status["repository_finalize_message"] = ""
        status["repository_finalize_updated_at"] = ""
        job_store.save_job_status(job_id, status)

        thread = threading.Thread(
            target=self._process_documents_worker,
            args=(job_id, import_id, repo_settings, document_records, selected_profile_id),
            daemon=True,
        )
        thread.start()

        status = self.get_status()
        return RepositoryProcessDocumentsResponse(
            job_id=job_id,
            import_id=import_id,
            accepted_documents=accepted_documents,
            total_sources=status.total_sources,
            total_citations=status.total_citations,
            selected_profile_id=selected_profile_id or "auto_detect",
            document_normalization=pending_normalization,
            message=f"Queued {accepted_documents} document(s) for repository processing",
        )

    def reprocess_documents(
        self,
        target_import_ids: list[str],
        settings: EffectiveSettings | None = None,
        profile_override: str = "",
    ) -> RepositoryReprocessDocumentsResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before reprocessing documents")

        normalized_import_ids = _dedupe_strings(
            [str(item or "").strip() for item in target_import_ids]
        )
        if not normalized_import_ids:
            raise ValueError("Select at least one prior document import to reprocess.")

        repo_settings = self.load_effective_settings(settings)
        selected_profile_id = self._validate_profile_override(profile_override)
        reprocess_id = uuid.uuid4().hex[:12]
        job_store = self.repo_job_store()
        job_id = job_store.create_job(prefix=REPO_JOB_PREFIX)

        with self._writer_lock():
            state = self._load_state_locked()
            imports = list(state.get("imports", []))

        imports_by_id = {
            str(item.get("import_id") or "").strip(): item for item in imports
        }
        missing_import_ids = [
            import_id
            for import_id in normalized_import_ids
            if str(imports_by_id.get(import_id, {}).get("import_type") or "").strip()
            != "document_process"
        ]
        if missing_import_ids:
            raise ValueError(
                "Unknown rerunnable document import(s): "
                + ", ".join(sorted(missing_import_ids))
            )

        targeted_documents: list[dict[str, str]] = []
        seen_repository_paths: set[str] = set()
        pending_normalization: list[DocumentNormalizationResult] = []

        for import_id in normalized_import_ids:
            import_record = imports_by_id[import_id]
            for stored_document in self._load_import_document_records(import_record):
                repository_path = str(
                    stored_document.get("repository_path") or ""
                ).strip()
                if not repository_path or repository_path in seen_repository_paths:
                    continue
                source_abs = self.path / Path(repository_path)
                if not source_abs.is_file():
                    continue

                original_name = self._document_record_display_name(stored_document)
                content = source_abs.read_bytes()
                sha256 = str(stored_document.get("sha256") or "").strip() or hashlib.sha256(
                    content
                ).hexdigest()

                document_record = {
                    "filename": original_name,
                    "source_document_name": original_name,
                    "repository_path": repository_path,
                    "sha256": sha256,
                    "document_import_id": str(
                        stored_document.get("document_import_id") or import_id
                    ).strip()
                    or import_id,
                }
                targeted_documents.append(document_record)
                pending_normalization.append(
                    self._pending_normalization_result(
                        filename=original_name,
                        source_document_path=repository_path,
                        selected_profile_id=selected_profile_id,
                    )
                )
                seen_repository_paths.add(repository_path)

        if not targeted_documents:
            raise ValueError("No stored repository documents were found for the selected imports.")

        job_store.save_artifact(
            job_id,
            "repo_processing_context",
            {
                "reprocess_id": reprocess_id,
                "target_import_ids": normalized_import_ids,
                "documents": targeted_documents,
                "repository_path": str(self.path),
                "profile_override": selected_profile_id,
                "processing_mode": "reprocess_documents",
            },
        )
        status = job_store.get_job_status(job_id) or {}
        status["selected_profile_id"] = selected_profile_id or "auto_detect"
        status["document_normalization"] = [
            item.model_dump(mode="json") for item in pending_normalization
        ]
        status["processing_mode"] = "reprocess_documents"
        status["target_import_ids"] = normalized_import_ids
        status["repository_preprocess_state"] = "pending"
        status["repository_preprocess_message"] = "Waiting to prepare standardized markdown."
        status["repository_preprocess_updated_at"] = _utc_now_iso()
        status["repository_finalize_state"] = ""
        status["repository_finalize_message"] = ""
        status["repository_finalize_updated_at"] = ""
        job_store.save_job_status(job_id, status)

        thread = threading.Thread(
            target=self._reprocess_documents_worker,
            args=(
                job_id,
                reprocess_id,
                normalized_import_ids,
                repo_settings,
                targeted_documents,
                selected_profile_id,
            ),
            daemon=True,
        )
        thread.start()

        status_snapshot = self.get_status()
        return RepositoryReprocessDocumentsResponse(
            job_id=job_id,
            reprocess_id=reprocess_id,
            target_import_ids=normalized_import_ids,
            accepted_documents=len(targeted_documents),
            total_sources=status_snapshot.total_sources,
            total_citations=status_snapshot.total_citations,
            selected_profile_id=selected_profile_id or "auto_detect",
            document_normalization=pending_normalization,
            message=f"Queued {len(targeted_documents)} stored document(s) for reprocessing",
        )

    def _process_documents_worker(
        self,
        job_id: str,
        import_id: str,
        settings: EffectiveSettings,
        documents: list[dict[str, str]],
        profile_override: str = "",
    ) -> None:
        try:
            prepared_documents, _ = self._prepare_standardized_pipeline_documents(
                job_id=job_id,
                documents=documents,
                settings=settings,
                profile_override=profile_override,
            )
        except Exception as exc:  # noqa: BLE001
            self._set_repository_preprocess_status(
                job_id,
                state="failed",
                message=f"Standardized markdown generation failed: {type(exc).__name__}: {exc}",
                mark_job_failed=True,
            )
            return

        if not prepared_documents:
            self._set_repository_preprocess_status(
                job_id,
                state="failed",
                message="No standardized markdown files were available for extraction.",
                mark_job_failed=True,
            )
            return

        orchestrator = PipelineOrchestrator(
            job_id=job_id,
            store=self.job_store_for(job_id),
            config=ProcessingConfig(
                use_llm=settings.use_llm,
                research_purpose=settings.research_purpose,
                llm_backend=settings.llm_backend,
            ),
        )
        orchestrator.run()

        status = self.job_store_for(job_id).get_job_status(job_id) or {}
        if str(status.get("current_stage") or "") != "completed":
            self._set_repository_finalize_status(
                job_id,
                state="skipped",
                message="Repository merge was skipped because extraction did not complete.",
            )
            return

        self._set_repository_finalize_status(
            job_id,
            state="running",
            message="Merging repository results from standardized markdown extraction.",
        )
        try:
            self.merge_processing_job_results(
                job_id=job_id,
                import_id=import_id,
                documents=documents,
                settings=settings,
                profile_override=profile_override,
            )
        except Exception as exc:  # noqa: BLE001
            self._set_repository_finalize_status(
                job_id,
                state="failed",
                message=f"Repository merge failed: {type(exc).__name__}: {exc}",
                mark_job_failed=True,
            )
            return

        self._set_repository_finalize_status(
            job_id,
            state="completed",
            message="Repository merge completed.",
        )

    def _reprocess_documents_worker(
        self,
        job_id: str,
        reprocess_id: str,
        target_import_ids: list[str],
        settings: EffectiveSettings,
        documents: list[dict[str, str]],
        profile_override: str = "",
    ) -> None:
        try:
            prepared_documents, _ = self._prepare_standardized_pipeline_documents(
                job_id=job_id,
                documents=documents,
                settings=settings,
                profile_override=profile_override,
            )
        except Exception as exc:  # noqa: BLE001
            self._set_repository_preprocess_status(
                job_id,
                state="failed",
                message=f"Standardized markdown generation failed: {type(exc).__name__}: {exc}",
                mark_job_failed=True,
            )
            return

        if not prepared_documents:
            self._set_repository_preprocess_status(
                job_id,
                state="failed",
                message="No standardized markdown files were available for extraction.",
                mark_job_failed=True,
            )
            return

        orchestrator = PipelineOrchestrator(
            job_id=job_id,
            store=self.job_store_for(job_id),
            config=ProcessingConfig(
                use_llm=settings.use_llm,
                research_purpose=settings.research_purpose,
                llm_backend=settings.llm_backend,
            ),
        )
        orchestrator.run()

        status = self.job_store_for(job_id).get_job_status(job_id) or {}
        if str(status.get("current_stage") or "") != "completed":
            self._set_repository_finalize_status(
                job_id,
                state="skipped",
                message="Repository reprocessing was skipped because extraction did not complete.",
            )
            return

        self._set_repository_finalize_status(
            job_id,
            state="running",
            message="Replacing repository citation rows from standardized markdown extraction.",
        )
        try:
            self.merge_reprocessed_documents_results(
                job_id=job_id,
                reprocess_id=reprocess_id,
                target_import_ids=target_import_ids,
                documents=documents,
                settings=settings,
                profile_override=profile_override,
            )
        except Exception as exc:  # noqa: BLE001
            self._set_repository_finalize_status(
                job_id,
                state="failed",
                message=f"Repository reprocessing failed: {type(exc).__name__}: {exc}",
                mark_job_failed=True,
            )
            return

        self._set_repository_finalize_status(
            job_id,
            state="completed",
            message="Repository reprocessing completed.",
        )

    def merge_processing_job_results(
        self,
        job_id: str,
        import_id: str,
        documents: list[dict[str, str]],
        settings: EffectiveSettings | None = None,
        profile_override: str = "",
    ) -> dict[str, Any]:
        if not self.is_attached:
            raise ValueError("Attach a repository before merging document results")

        job_store = self.job_store_for(job_id)
        ingestion_raw = job_store.load_artifact(job_id, "01_ingestion") or {}
        bibliography_raw = job_store.load_artifact(job_id, "03_bibliography") or {}
        export_raw = job_store.load_artifact(job_id, "05_export") or {}
        export_rows = _load_citation_rows(export_raw.get("rows", []))
        repo_settings = self.load_effective_settings(settings)
        imported_at = _utc_now_iso()
        prepared_documents, normalization_outputs = self._resolve_document_normalization_outputs(
            job_id=job_id,
            documents=documents,
            ingested_documents=self._load_ingested_documents_by_filename(
                ingestion_raw.get("documents", [])
            ),
            bibliography_entries_by_filename=self._group_bibliography_entries_by_filename(
                bibliography_raw.get("entries", [])
            ),
            bibliography_sections_by_filename=self._load_references_sections_by_filename(
                bibliography_raw.get("sections", [])
            ),
            settings=repo_settings,
            profile_override=profile_override,
        )
        document_by_job_filename = {
            str(item.get("filename") or "").strip(): item for item in prepared_documents
        }
        doc_hash_by_filename = {
            str(item.get("filename") or ""): str(item.get("sha256") or "")[:12]
            for item in documents
        }

        with self._writer_lock():
            self._ensure_scaffold_locked()
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            imports = list(state.get("imports", []))
            meta = self._load_meta_locked()

            by_key: dict[str, SourceManifestRow] = {}
            for row in rows:
                key = repository_dedupe_key(row.original_url or row.final_url)
                if not key:
                    key = dedupe_url_key(row.original_url or row.final_url)
                if key:
                    by_key[key] = row

            next_source_id = int(meta.get("next_source_id") or _next_source_id_from_rows(rows))
            total_candidates = 0
            accepted_new = 0
            duplicates = 0
            touched_source_ids: list[str] = []

            for raw_entry in bibliography_raw.get("entries", []):
                try:
                    entry = BibliographyEntry.model_validate(raw_entry)
                except Exception:
                    continue

                document_record = document_by_job_filename.get(
                    str(entry.source_document_name or "").strip()
                )
                if document_record is not None:
                    entry.source_document_name = self._document_record_display_name(
                        document_record
                    )

                url = _entry_url(entry)
                if not url:
                    continue
                total_candidates += 1

                dedupe_key = repository_dedupe_key(url) or dedupe_url_key(url)
                if not dedupe_key:
                    continue

                existing = by_key.get(dedupe_key)
                if existing:
                    duplicates += 1
                    touched_source_ids.append(existing.id)
                    if not existing.source_document_name:
                        existing.source_document_name = entry.source_document_name
                    if not existing.citation_number and entry.ref_number:
                        existing.citation_number = str(entry.ref_number)
                    if not existing.title and entry.title:
                        existing.title = entry.title
                    continue

                source_id = f"{next_source_id:06d}"
                next_source_id += 1
                row = SourceManifestRow(
                    id=source_id,
                    repository_source_id=source_id,
                    import_type="document_process",
                    imported_at=imported_at,
                    provenance_ref=f"{import_id}:{Path(entry.source_document_name or 'document').name}",
                    source_document_name=entry.source_document_name,
                    citation_number=str(entry.ref_number or ""),
                    original_url=url,
                    title=entry.title,
                    fetch_status="queued",
                    notes="queued_for_download",
                )
                rows.append(row)
                by_key[dedupe_key] = row
                accepted_new += 1
                touched_source_ids.append(source_id)

            merged_citations = list(citations)
            for export_row in export_rows:
                row = export_row.model_copy(deep=True)
                document_record = document_by_job_filename.get(
                    str(export_row.source_document or "").strip()
                )
                if document_record is not None:
                    display_name = self._document_record_display_name(document_record)
                    row.source_document = display_name
                    row.document_import_id = str(
                        document_record.get("document_import_id") or import_id
                    ).strip()
                    row.document_repository_path = str(
                        document_record.get("repository_path") or ""
                    ).strip()
                    row.document_sha256 = str(
                        document_record.get("sha256") or ""
                    ).strip()
                    row.provenance_ref = _document_citation_provenance_for_document(
                        display_name,
                        row.document_sha256,
                    )
                row.import_type = "document_process"
                row.imported_at = imported_at
                if not row.provenance_ref:
                    row.provenance_ref = _document_citation_provenance(
                        row.source_document,
                        doc_hash_by_filename,
                    )
                url_key = repository_dedupe_key(row.cited_url) or dedupe_url_key(row.cited_url)
                if url_key and url_key in by_key:
                    row.repository_source_id = by_key[url_key].id
                merged_citations.append(row)

            standardized_markdown_files = sum(
                1 for output in normalization_outputs if output.result.standardized_markdown_path
            )
            partial_normalizations = sum(
                1 for output in normalization_outputs if output.result.status == "partial"
            )
            failed_normalizations = sum(
                1 for output in normalization_outputs if output.result.status == "failed"
            )
            self._append_document_normalization_status(
                job_id=job_id,
                normalization_outputs=normalization_outputs,
                partial_count=partial_normalizations,
                failed_count=failed_normalizations,
            )

            sorted_rows = self._sort_rows(rows)
            sorted_citations = self._sort_citations(self._dedupe_citations(merged_citations))

            imports.append(
                {
                    "import_id": import_id,
                    "import_type": "document_process",
                    "provenance": ", ".join(
                        sorted(str(item.get("filename") or "") for item in documents)
                    ),
                    "imported_at": imported_at,
                    "total_candidates": total_candidates,
                    "accepted_new": accepted_new,
                    "duplicates_skipped": duplicates,
                    "source_ids": _dedupe_strings(touched_source_ids),
                    "accepted_documents": len(documents),
                    "documents": documents,
                    "processing_job_id": job_id,
                    "standardized_markdown_files": standardized_markdown_files,
                    "partial_normalizations": partial_normalizations,
                    "failed_normalizations": failed_normalizations,
                    "selected_profile_id": profile_override or "auto_detect",
                }
            )

            self._save_state_locked(
                sources=sorted_rows,
                citations=sorted_citations,
                imports=imports,
            )
            self._save_meta_locked(
                {
                    **meta,
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": next_source_id,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(sorted_rows, sorted_citations)
            self._download_message = (
                f"Processed {len(documents)} document(s): "
                f"{accepted_new} new source URLs, {len(export_rows)} citation rows, "
                f"{standardized_markdown_files} standardized markdown files"
            )
            if partial_normalizations or failed_normalizations:
                self._download_message += (
                    f" ({partial_normalizations} partial, {failed_normalizations} failed normalization)"
                )

            return {
                "import_id": import_id,
                "accepted_documents": len(documents),
                "accepted_new": accepted_new,
                "duplicates_skipped": duplicates,
                "total_citation_rows": len(export_rows),
                "standardized_markdown_files": standardized_markdown_files,
                "selected_profile_id": profile_override or "auto_detect",
                "document_normalization": [
                    output.result.model_dump(mode="json") for output in normalization_outputs
                ],
                "message": self._download_message,
            }

    def merge_reprocessed_documents_results(
        self,
        job_id: str,
        reprocess_id: str,
        target_import_ids: list[str],
        documents: list[dict[str, str]],
        settings: EffectiveSettings | None = None,
        profile_override: str = "",
    ) -> dict[str, Any]:
        if not self.is_attached:
            raise ValueError("Attach a repository before merging document results")

        job_store = self.job_store_for(job_id)
        ingestion_raw = job_store.load_artifact(job_id, "01_ingestion") or {}
        bibliography_raw = job_store.load_artifact(job_id, "03_bibliography") or {}
        export_raw = job_store.load_artifact(job_id, "05_export") or {}
        export_rows = _load_citation_rows(export_raw.get("rows", []))
        repo_settings = self.load_effective_settings(settings)
        imported_at = _utc_now_iso()
        prepared_documents, normalization_outputs = self._resolve_document_normalization_outputs(
            job_id=job_id,
            documents=documents,
            ingested_documents=self._load_ingested_documents_by_filename(
                ingestion_raw.get("documents", [])
            ),
            bibliography_entries_by_filename=self._group_bibliography_entries_by_filename(
                bibliography_raw.get("entries", [])
            ),
            bibliography_sections_by_filename=self._load_references_sections_by_filename(
                bibliography_raw.get("sections", [])
            ),
            settings=repo_settings,
            profile_override=profile_override,
        )
        document_by_job_filename = {
            str(item.get("filename") or "").strip(): item for item in prepared_documents
        }
        normalization_by_repository_path = {
            output.result.source_document_path: output for output in normalization_outputs
        }
        standardized_markdown_files = sum(
            1 for output in normalization_outputs if output.result.standardized_markdown_path
        )
        partial_normalizations = sum(
            1 for output in normalization_outputs if output.result.status == "partial"
        )
        failed_normalizations = sum(
            1 for output in normalization_outputs if output.result.status == "failed"
        )

        with self._writer_lock():
            self._ensure_scaffold_locked()
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            imports = list(state.get("imports", []))
            meta = self._load_meta_locked()

            by_key: dict[str, SourceManifestRow] = {}
            for row in rows:
                key = repository_dedupe_key(row.original_url or row.final_url)
                if not key:
                    key = dedupe_url_key(row.original_url or row.final_url)
                if key:
                    by_key[key] = row

            next_source_id = int(meta.get("next_source_id") or _next_source_id_from_rows(rows))
            total_candidates = 0
            accepted_new = 0
            duplicates = 0
            touched_source_ids: list[str] = []

            for raw_entry in bibliography_raw.get("entries", []):
                try:
                    entry = BibliographyEntry.model_validate(raw_entry)
                except Exception:
                    continue

                document_record = document_by_job_filename.get(
                    str(entry.source_document_name or "").strip()
                )
                if document_record is None:
                    continue

                display_name = self._document_record_display_name(document_record)
                entry.source_document_name = display_name
                url = _entry_url(entry)
                if not url:
                    continue
                total_candidates += 1

                dedupe_key = repository_dedupe_key(url) or dedupe_url_key(url)
                if not dedupe_key:
                    continue

                existing = by_key.get(dedupe_key)
                if existing:
                    duplicates += 1
                    touched_source_ids.append(existing.id)
                    if not existing.source_document_name:
                        existing.source_document_name = display_name
                    if not existing.citation_number and entry.ref_number:
                        existing.citation_number = str(entry.ref_number)
                    if not existing.title and entry.title:
                        existing.title = entry.title
                    continue

                source_id = f"{next_source_id:06d}"
                next_source_id += 1
                row = SourceManifestRow(
                    id=source_id,
                    repository_source_id=source_id,
                    import_type="document_reprocess",
                    imported_at=imported_at,
                    provenance_ref=f"{reprocess_id}:{display_name}",
                    source_document_name=display_name,
                    citation_number=str(entry.ref_number or ""),
                    original_url=url,
                    title=entry.title,
                    fetch_status="queued",
                    notes="queued_for_download",
                )
                rows.append(row)
                by_key[dedupe_key] = row
                accepted_new += 1
                touched_source_ids.append(source_id)

            new_citations_by_path: dict[str, list[ExportRow]] = {
                str(item.get("repository_path") or "").strip(): [] for item in documents
            }
            for export_row in export_rows:
                document_record = document_by_job_filename.get(
                    str(export_row.source_document or "").strip()
                )
                if document_record is None:
                    continue

                row = export_row.model_copy(deep=True)
                display_name = self._document_record_display_name(document_record)
                repository_path = str(document_record.get("repository_path") or "").strip()
                row.source_document = display_name
                row.import_type = "document_reprocess"
                row.imported_at = imported_at
                row.provenance_ref = _document_citation_provenance_for_document(
                    display_name,
                    str(document_record.get("sha256") or "").strip(),
                )
                row.document_import_id = str(
                    document_record.get("document_import_id") or ""
                ).strip()
                row.document_repository_path = repository_path
                row.document_sha256 = str(document_record.get("sha256") or "").strip()

                url_key = repository_dedupe_key(row.cited_url) or dedupe_url_key(row.cited_url)
                if url_key and url_key in by_key:
                    row.repository_source_id = by_key[url_key].id
                new_citations_by_path.setdefault(repository_path, []).append(row)

            preserved_failed_documents = 0
            replaced_documents = 0
            replacement_outcomes: list[dict[str, Any]] = []
            remaining_citations = list(citations)

            for document_record in documents:
                repository_path = str(document_record.get("repository_path") or "").strip()
                if not repository_path:
                    continue
                display_name = self._document_record_display_name(document_record)
                normalization_output = normalization_by_repository_path.get(repository_path)
                normalized_status = (
                    normalization_output.result.status if normalization_output else "failed"
                )
                matching_existing = [
                    row
                    for row in citations
                    if self._citation_row_matches_document(row, document_record)
                ]
                existing_count = len(matching_existing)

                if normalized_status == "failed":
                    preserved_failed_documents += 1
                    replacement_outcomes.append(
                        {
                            "filename": display_name,
                            "repository_path": repository_path,
                            "status": "preserved_failed",
                            "replaced_existing_rows": 0,
                            "preserved_existing_rows": existing_count,
                            "new_rows": 0,
                        }
                    )
                    continue

                replaced_documents += 1
                remaining_citations = [
                    row
                    for row in remaining_citations
                    if not self._citation_row_matches_document(row, document_record)
                ]
                new_rows = new_citations_by_path.get(repository_path, [])
                remaining_citations.extend(new_rows)
                replacement_outcomes.append(
                    {
                        "filename": display_name,
                        "repository_path": repository_path,
                        "status": "replaced",
                        "replaced_existing_rows": existing_count,
                        "preserved_existing_rows": 0,
                        "new_rows": len(new_rows),
                    }
                )

            sorted_rows = self._sort_rows(rows)
            sorted_citations = self._sort_citations(
                self._dedupe_citations(remaining_citations)
            )

            imports.append(
                {
                    "import_id": reprocess_id,
                    "import_type": "document_reprocess",
                    "provenance": ", ".join(sorted(target_import_ids)),
                    "imported_at": imported_at,
                    "target_import_ids": list(target_import_ids),
                    "documents": documents,
                    "total_candidates": total_candidates,
                    "accepted_new": accepted_new,
                    "duplicates_skipped": duplicates,
                    "source_ids": _dedupe_strings(touched_source_ids),
                    "accepted_documents": len(documents),
                    "replaced_documents": replaced_documents,
                    "preserved_failed_documents": preserved_failed_documents,
                    "processing_job_id": job_id,
                    "standardized_markdown_files": standardized_markdown_files,
                    "partial_normalizations": partial_normalizations,
                    "failed_normalizations": failed_normalizations,
                    "selected_profile_id": profile_override or "auto_detect",
                }
            )

            self._save_state_locked(
                sources=sorted_rows,
                citations=sorted_citations,
                imports=imports,
            )
            self._save_meta_locked(
                {
                    **meta,
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": next_source_id,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(sorted_rows, sorted_citations)

            self._append_document_normalization_status(
                job_id=job_id,
                normalization_outputs=normalization_outputs,
                partial_count=partial_normalizations,
                failed_count=failed_normalizations,
                processing_mode="reprocess_documents",
                target_import_ids=target_import_ids,
                document_replacements=replacement_outcomes,
            )

            self._download_message = (
                f"Reprocessed {len(documents)} document(s): "
                f"{accepted_new} new source URLs, {replaced_documents} documents replaced, "
                f"{preserved_failed_documents} preserved failures"
            )
            if partial_normalizations or failed_normalizations:
                self._download_message += (
                    f" ({partial_normalizations} partial, {failed_normalizations} failed normalization)"
                )

            return {
                "reprocess_id": reprocess_id,
                "accepted_documents": len(documents),
                "accepted_new": accepted_new,
                "duplicates_skipped": duplicates,
                "replaced_documents": replaced_documents,
                "preserved_failed_documents": preserved_failed_documents,
                "total_citation_rows": len(sorted_citations),
                "standardized_markdown_files": standardized_markdown_files,
                "selected_profile_id": profile_override or "auto_detect",
                "target_import_ids": list(target_import_ids),
                "document_normalization": [
                    output.result.model_dump(mode="json") for output in normalization_outputs
                ],
                "document_replacements": replacement_outcomes,
                "message": self._download_message,
            }

    def _load_ingested_documents_by_filename(
        self,
        raw_documents: list[dict[str, Any]],
    ) -> dict[str, IngestedDocument]:
        documents_by_filename: dict[str, IngestedDocument] = {}
        for raw_document in raw_documents:
            try:
                document = IngestedDocument.model_validate(raw_document)
            except Exception:
                continue
            documents_by_filename[document.filename] = document
        return documents_by_filename

    def _group_bibliography_entries_by_filename(
        self,
        raw_entries: list[dict[str, Any]],
    ) -> dict[str, list[BibliographyEntry]]:
        grouped: dict[str, list[BibliographyEntry]] = {}
        for raw_entry in raw_entries:
            try:
                entry = BibliographyEntry.model_validate(raw_entry)
            except Exception:
                continue
            filename = str(entry.source_document_name or "").strip()
            if not filename:
                continue
            grouped.setdefault(filename, []).append(entry)
        return grouped

    def _load_references_sections_by_filename(
        self,
        raw_sections: list[dict[str, Any]],
    ) -> dict[str, ReferencesSection]:
        sections_by_filename: dict[str, ReferencesSection] = {}
        for raw_section in raw_sections:
            try:
                section = ReferencesSection.model_validate(raw_section)
            except Exception:
                continue
            sections_by_filename[section.document_filename] = section
        return sections_by_filename

    def _standardized_document_paths(
        self,
        document_record: dict[str, Any],
    ) -> tuple[str, str, str, Path, Path]:
        repository_path = str(document_record.get("repository_path") or "").strip()
        display_name = self._document_record_display_name(document_record)
        target_dir = (self.path / Path(repository_path)).parent
        markdown_name = standardized_markdown_filename(display_name)
        metadata_name = standardized_metadata_filename(display_name)
        markdown_rel = (Path(repository_path).parent / markdown_name).as_posix()
        metadata_rel = (Path(repository_path).parent / metadata_name).as_posix()
        markdown_abs = target_dir / markdown_name
        metadata_abs = target_dir / metadata_name
        return display_name, markdown_rel, metadata_rel, markdown_abs, metadata_abs

    def _load_existing_standardized_output(
        self,
        document_record: dict[str, Any],
    ) -> NormalizedDocumentOutput | None:
        repository_path = str(document_record.get("repository_path") or "").strip()
        if not repository_path:
            return None

        (
            display_name,
            markdown_rel,
            metadata_rel,
            markdown_abs,
            metadata_abs,
        ) = self._standardized_document_paths(document_record)
        if not markdown_abs.is_file():
            return None

        markdown_text = markdown_abs.read_text(encoding="utf-8")
        raw_result: dict[str, Any] = {}
        if metadata_abs.is_file():
            try:
                raw_result = json.loads(metadata_abs.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                raw_result = {}

        try:
            result = DocumentNormalizationResult.model_validate(raw_result)
        except Exception:
            result = self._pending_normalization_result(
                filename=display_name,
                source_document_path=repository_path,
                selected_profile_id="",
            )
        result.filename = display_name
        result.source_document_path = repository_path
        result.standardized_markdown_path = markdown_rel
        result.metadata_path = metadata_rel
        if result.status in {"", "pending"}:
            result.status = "normalized" if markdown_text.strip() else "failed"
        if result.status == "failed" and markdown_text.strip():
            result.status = "partial"
        reuse_warning = "Reused existing standardized markdown file."
        if reuse_warning not in result.warnings:
            result.warnings.append(reuse_warning)
        return NormalizedDocumentOutput(
            markdown_text=markdown_text,
            result=result,
            suggestion=None,
        )

    def _build_standardized_output_for_document(
        self,
        document_record: dict[str, Any],
        settings: EffectiveSettings,
        builtin_profiles: list[IngestionProfile],
        custom_profiles: list[IngestionProfile],
        profile_override: str = "",
    ) -> NormalizedDocumentOutput:
        repository_path = str(document_record.get("repository_path") or "").strip()
        if not repository_path:
            return NormalizedDocumentOutput(
                markdown_text="",
                result=DocumentNormalizationResult(
                    filename=self._document_record_display_name(document_record),
                    status="failed",
                    error_message="Repository document path missing.",
                ),
                suggestion=None,
            )

        (
            display_name,
            markdown_rel,
            metadata_rel,
            markdown_abs,
            metadata_abs,
        ) = self._standardized_document_paths(document_record)
        source_abs = self.path / Path(repository_path)
        target_dir = markdown_abs.parent
        target_dir.mkdir(parents=True, exist_ok=True)

        if not source_abs.is_file():
            failed_result = DocumentNormalizationResult(
                filename=display_name,
                source_document_path=repository_path,
                standardized_markdown_path="",
                metadata_path=metadata_rel,
                selected_profile_id=profile_override or "auto_detect",
                selected_profile_label=profile_override or "Auto-detect",
                status="failed",
                error_message="Stored repository document is missing.",
            )
            metadata_abs.write_text(
                json.dumps(failed_result.model_dump(mode="json"), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            return NormalizedDocumentOutput(
                markdown_text="",
                result=failed_result,
                suggestion=None,
            )

        with tempfile.TemporaryDirectory(prefix="repo-standardize-") as tmp:
            tmp_path = Path(tmp)
            temp_source = tmp_path / Path(display_name).name
            temp_source.write_bytes(source_abs.read_bytes())
            ingestion = run_ingestion(tmp_path)

        if not ingestion.documents:
            failed_result = DocumentNormalizationResult(
                filename=display_name,
                source_document_path=repository_path,
                standardized_markdown_path="",
                metadata_path=metadata_rel,
                selected_profile_id=profile_override or "auto_detect",
                selected_profile_label=profile_override or "Auto-detect",
                status="failed",
                error_message="No usable document content found during standardization.",
            )
            metadata_abs.write_text(
                json.dumps(failed_result.model_dump(mode="json"), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            return NormalizedDocumentOutput(
                markdown_text="",
                result=failed_result,
                suggestion=None,
            )

        document = ingestion.documents[0]
        section = detect_references_section(document)
        sections = [section] if section else []
        artifact = parse_bibliography(
            sections,
            use_llm=settings.use_llm,
            llm_backend=settings.llm_backend,
        )

        if document.inline_citation_urls:
            if artifact.entries:
                artifact.entries = merge_inline_urls_into_entries(
                    artifact.entries,
                    document.inline_citation_urls,
                )
            else:
                artifact.entries = build_entries_from_inline_urls(document.inline_citation_urls)

        for entry in artifact.entries:
            if not entry.source_document_name:
                entry.source_document_name = document.filename

        normalized_output = normalize_document_to_standardized_markdown(
            document=document,
            bibliography_entries=artifact.entries,
            references_section=section,
            builtin_profiles=builtin_profiles,
            custom_profiles=custom_profiles,
            profile_override=profile_override,
            use_llm=settings.use_llm,
            llm_backend=settings.llm_backend,
            research_purpose=settings.research_purpose,
        )
        normalized_output.result.filename = display_name
        normalized_output.result.source_document_path = repository_path
        normalized_output.result.metadata_path = metadata_rel
        if normalized_output.markdown_text:
            markdown_abs.write_text(normalized_output.markdown_text, encoding="utf-8")
            normalized_output.result.standardized_markdown_path = markdown_rel
        else:
            normalized_output.result.standardized_markdown_path = ""

        metadata_abs.write_text(
            json.dumps(
                normalized_output.result.model_dump(mode="json"),
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        return normalized_output

    def _save_document_normalization_progress(
        self,
        job_id: str,
        results: list[DocumentNormalizationResult],
    ) -> None:
        job_store = self.job_store_for(job_id)
        status = job_store.get_job_status(job_id) or {}
        if not status:
            return
        status["document_normalization"] = [
            item.model_dump(mode="json") for item in results
        ]
        job_store.save_job_status(job_id, status)

    def _prepare_standardized_pipeline_documents(
        self,
        job_id: str,
        documents: list[dict[str, str]],
        settings: EffectiveSettings,
        profile_override: str = "",
    ) -> tuple[list[dict[str, str]], list[NormalizedDocumentOutput]]:
        job_store = self.job_store_for(job_id)
        total_documents = len(documents)
        bundled_profiles = self._load_bundled_ingestion_profiles()
        custom_profiles = self._load_custom_ingestion_profiles()
        pending_suggestions = self._load_ingestion_profile_suggestions()
        prepared_documents: list[dict[str, str]] = []
        normalization_outputs: list[NormalizedDocumentOutput] = []
        progress_results = [
            self._pending_normalization_result(
                filename=self._document_record_display_name(document_record),
                source_document_path=str(document_record.get("repository_path") or "").strip(),
                selected_profile_id=profile_override,
            )
            for document_record in documents
        ]
        used_pipeline_filenames: set[str] = set()
        generated_documents = 0
        reused_documents = 0
        failed_documents = 0

        self._set_repository_preprocess_status(
            job_id,
            state="running",
            message=f"Preparing standardized markdown for 0/{total_documents} document(s).",
        )
        self._save_document_normalization_progress(job_id, progress_results)

        for index, document_record in enumerate(documents, start=1):
            display_name = self._document_record_display_name(document_record)
            self._set_repository_preprocess_status(
                job_id,
                state="running",
                message=f"Preparing standardized markdown {index}/{total_documents}: {display_name}",
            )

            existing_output = self._load_existing_standardized_output(document_record)
            if existing_output is not None:
                normalized_output = existing_output
                reused_documents += 1
            else:
                normalized_output = self._build_standardized_output_for_document(
                    document_record=document_record,
                    settings=settings,
                    builtin_profiles=bundled_profiles,
                    custom_profiles=custom_profiles,
                    profile_override=profile_override,
                )
                if normalized_output.result.standardized_markdown_path:
                    generated_documents += 1
                else:
                    failed_documents += 1

            if normalized_output.suggestion is not None:
                pending_suggestions.append(normalized_output.suggestion)

            normalization_outputs.append(normalized_output)
            progress_results[index - 1] = normalized_output.result
            self._save_document_normalization_progress(job_id, progress_results)

            markdown_path = str(normalized_output.result.standardized_markdown_path or "").strip()
            if not markdown_path or not normalized_output.markdown_text:
                continue

            pipeline_seed = standardized_markdown_filename(display_name)
            pipeline_filename = _next_unique_filename(pipeline_seed, used_pipeline_filenames)
            job_store.save_upload(
                job_id,
                pipeline_filename,
                normalized_output.markdown_text.encode("utf-8"),
            )
            prepared_documents.append(
                {
                    **document_record,
                    "filename": pipeline_filename,
                    "standardized_markdown_path": markdown_path,
                    "metadata_path": str(normalized_output.result.metadata_path or "").strip(),
                }
            )

        if pending_suggestions:
            self._save_ingestion_profile_suggestions(pending_suggestions)

        job_store.save_artifact(
            job_id,
            "00_repository_preprocess",
            {
                "prepared_documents": prepared_documents,
                "document_normalization": [
                    output.result.model_dump(mode="json") for output in normalization_outputs
                ],
                "generated_documents": generated_documents,
                "reused_documents": reused_documents,
                "failed_documents": failed_documents,
            },
        )

        prepared_count = len(prepared_documents)
        if prepared_count == 0:
            message = "No standardized markdown files were generated."
            if failed_documents:
                message += f" {failed_documents} document(s) failed preprocessing."
            self._set_repository_preprocess_status(
                job_id,
                state="failed",
                message=message,
            )
            return [], normalization_outputs

        message = (
            f"Prepared {prepared_count}/{total_documents} standardized markdown file(s)"
        )
        details: list[str] = []
        if reused_documents:
            details.append(f"{reused_documents} reused")
        if generated_documents:
            details.append(f"{generated_documents} newly generated")
        if failed_documents:
            details.append(f"{failed_documents} skipped")
        if details:
            message += f" ({', '.join(details)})"
        message += "."
        self._set_repository_preprocess_status(
            job_id,
            state="completed",
            message=message,
        )
        return prepared_documents, normalization_outputs

    def _resolve_document_normalization_outputs(
        self,
        job_id: str,
        documents: list[dict[str, str]],
        ingested_documents: dict[str, IngestedDocument],
        bibliography_entries_by_filename: dict[str, list[BibliographyEntry]],
        bibliography_sections_by_filename: dict[str, ReferencesSection],
        settings: EffectiveSettings,
        profile_override: str = "",
    ) -> tuple[list[dict[str, str]], list[NormalizedDocumentOutput]]:
        preprocess_raw = self.job_store_for(job_id).load_artifact(job_id, "00_repository_preprocess") or {}
        prepared_documents = [
            dict(item)
            for item in preprocess_raw.get("prepared_documents", [])
            if isinstance(item, dict)
        ]
        normalization_outputs: list[NormalizedDocumentOutput] = []
        for raw_result in preprocess_raw.get("document_normalization", []):
            try:
                result = DocumentNormalizationResult.model_validate(raw_result)
            except Exception:
                continue
            normalization_outputs.append(
                NormalizedDocumentOutput(
                    markdown_text="",
                    result=result,
                    suggestion=None,
                )
            )
        if prepared_documents or normalization_outputs:
            return prepared_documents, normalization_outputs

        return (
            documents,
            self._write_standardized_markdown_documents(
                documents=documents,
                ingested_documents=ingested_documents,
                bibliography_entries_by_filename=bibliography_entries_by_filename,
                bibliography_sections_by_filename=bibliography_sections_by_filename,
                settings=settings,
                profile_override=profile_override,
            ),
        )

    def _write_standardized_markdown_documents(
        self,
        documents: list[dict[str, str]],
        ingested_documents: dict[str, IngestedDocument],
        bibliography_entries_by_filename: dict[str, list[BibliographyEntry]],
        bibliography_sections_by_filename: dict[str, ReferencesSection],
        settings: EffectiveSettings,
        profile_override: str = "",
    ) -> list[NormalizedDocumentOutput]:
        outputs: list[NormalizedDocumentOutput] = []
        bundled_profiles = self._load_bundled_ingestion_profiles()
        custom_profiles = self._load_custom_ingestion_profiles()
        pending_suggestions = self._load_ingestion_profile_suggestions()
        for document_record in documents:
            filename = str(document_record.get("filename") or "").strip()
            repository_path = str(document_record.get("repository_path") or "").strip()
            display_name = self._document_record_display_name(document_record)
            if not filename or not repository_path:
                continue

            source_abs = self.path / Path(repository_path)
            target_dir = source_abs.parent
            target_dir.mkdir(parents=True, exist_ok=True)
            markdown_name = standardized_markdown_filename(display_name)
            metadata_name = standardized_metadata_filename(display_name)
            markdown_rel = (Path(repository_path).parent / markdown_name).as_posix()
            metadata_rel = (Path(repository_path).parent / metadata_name).as_posix()
            markdown_abs = target_dir / markdown_name
            metadata_abs = target_dir / metadata_name

            document = ingested_documents.get(filename)
            if document is None:
                failed_result = DocumentNormalizationResult(
                    filename=display_name,
                    source_document_path=repository_path,
                    standardized_markdown_path="",
                    metadata_path=metadata_rel,
                    selected_profile_id=profile_override or "auto_detect",
                    selected_profile_label=profile_override or "Auto-detect",
                    status="failed",
                    error_message="Ingestion output missing for stored document.",
                )
                metadata_abs.write_text(
                    json.dumps(
                        failed_result.model_dump(mode="json"),
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                outputs.append(
                    NormalizedDocumentOutput(
                        markdown_text="",
                        result=failed_result,
                        suggestion=None,
                    )
                )
                continue

            entry_list = bibliography_entries_by_filename.get(filename, [])
            section = bibliography_sections_by_filename.get(filename)
            normalized_output = normalize_document_to_standardized_markdown(
                document=document,
                bibliography_entries=entry_list or [],
                references_section=section,
                builtin_profiles=bundled_profiles,
                custom_profiles=custom_profiles,
                profile_override=profile_override,
                use_llm=settings.use_llm,
                llm_backend=settings.llm_backend,
                research_purpose=settings.research_purpose,
            )
            normalized_output.result.filename = display_name

            if normalized_output.markdown_text:
                markdown_abs.write_text(normalized_output.markdown_text, encoding="utf-8")
                normalized_output.result.standardized_markdown_path = markdown_rel
            normalized_output.result.source_document_path = repository_path
            normalized_output.result.metadata_path = metadata_rel

            metadata_abs.write_text(
                json.dumps(normalized_output.result.model_dump(mode="json"), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            if normalized_output.suggestion is not None:
                pending_suggestions.append(normalized_output.suggestion)
            outputs.append(normalized_output)

        if pending_suggestions:
            self._save_ingestion_profile_suggestions(pending_suggestions)
        return outputs

    def _append_document_normalization_status(
        self,
        job_id: str,
        normalization_outputs: list[NormalizedDocumentOutput],
        partial_count: int,
        failed_count: int,
        processing_mode: str = "",
        target_import_ids: list[str] | None = None,
        document_replacements: list[dict[str, Any]] | None = None,
    ) -> None:
        job_store = self.job_store_for(job_id)
        status = job_store.get_job_status(job_id) or {}
        if not status:
            return
        status["document_normalization"] = [
            output.result.model_dump(mode="json") for output in normalization_outputs
        ]
        status["selected_profile_id"] = (
            normalization_outputs[0].result.selected_profile_id if normalization_outputs else ""
        )
        if processing_mode:
            status["processing_mode"] = processing_mode
        if target_import_ids is not None:
            status["target_import_ids"] = list(target_import_ids)
        if document_replacements is not None:
            status["document_replacements"] = document_replacements
        stages = list(status.get("stages") or [])
        for stage in stages:
            if str(stage.get("stage") or "") != "exporting":
                continue
            warnings = list(stage.get("warnings") or [])
            errors = list(stage.get("errors") or [])
            if partial_count:
                warnings.append(f"{partial_count} document(s) had partial normalization output.")
            if failed_count:
                errors.append(f"{failed_count} document(s) failed standardized markdown normalization.")
            stage["warnings"] = _dedupe_strings(warnings)
            stage["errors"] = _dedupe_strings(errors)
            break
        status["stages"] = stages
        job_store.save_job_status(job_id, status)

    def _set_repository_preprocess_status(
        self,
        job_id: str,
        *,
        state: str,
        message: str = "",
        mark_job_failed: bool = False,
    ) -> None:
        job_store = self.job_store_for(job_id)
        status = job_store.get_job_status(job_id) or {}
        if not status:
            return

        now = _utc_now_iso()
        normalized_state = str(state or "").strip().lower() or "completed"
        status["repository_preprocess_state"] = normalized_state
        status["repository_preprocess_message"] = message
        status["repository_preprocess_updated_at"] = now

        if mark_job_failed:
            status["current_stage"] = "failed"
            status["completed_at"] = now

        job_store.save_job_status(job_id, status)

    def _set_repository_finalize_status(
        self,
        job_id: str,
        *,
        state: str,
        message: str = "",
        mark_job_failed: bool = False,
    ) -> None:
        job_store = self.job_store_for(job_id)
        status = job_store.get_job_status(job_id) or {}
        if not status:
            return

        now = _utc_now_iso()
        normalized_state = str(state or "").strip().lower() or "completed"
        status["repository_finalize_state"] = normalized_state
        status["repository_finalize_message"] = message
        status["repository_finalize_updated_at"] = now

        if mark_job_failed:
            status["current_stage"] = "failed"
            status["completed_at"] = now

        job_store.save_job_status(job_id, status)

    def _citation_row_matches_document(
        self,
        row: ExportRow,
        document_record: dict[str, Any],
    ) -> bool:
        repository_path = str(document_record.get("repository_path") or "").strip()
        if repository_path and str(row.document_repository_path or "").strip() == repository_path:
            return True

        display_name = self._document_record_display_name(document_record)
        document_import_id = str(document_record.get("document_import_id") or "").strip()
        if document_import_id and row.provenance_ref == f"{document_import_id}:{display_name}":
            return True

        sha256 = str(document_record.get("sha256") or "").strip()
        if sha256 and row.provenance_ref == _document_citation_provenance_for_document(
            display_name,
            sha256,
        ):
            return True
        return False

    def start_source_tasks(
        self,
        payload: RepositorySourceTaskRequest,
        settings: EffectiveSettings | None = None,
        live_jobs: dict[str, SourceDownloadOrchestrator] | None = None,
        live_jobs_lock: threading.Lock | None = None,
    ) -> RepositorySourceTaskResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before running source tasks")
        run_download = bool(payload.run_download or payload.force_redownload)
        run_convert = bool(payload.run_convert or payload.force_convert or (run_download and payload.include_markdown))
        run_citation_verify = bool(payload.run_citation_verify or payload.force_citation_verify)
        run_catalog = bool(payload.run_catalog or payload.force_catalog)
        run_llm_cleanup = bool(payload.run_llm_cleanup or payload.force_llm_cleanup)
        run_llm_title = bool(payload.run_llm_title or payload.force_title)
        run_llm_summary = bool(payload.run_llm_summary or payload.force_summary)
        run_llm_rating = bool(payload.run_llm_rating or payload.force_rating)
        if not (
            run_download
            or run_convert
            or run_catalog
            or run_citation_verify
            or run_llm_cleanup
            or run_llm_title
            or run_llm_summary
            or run_llm_rating
        ):
            raise ValueError("Select at least one source phase to run.")
        if run_download and not any(
            [
                payload.include_raw_file,
                payload.include_rendered_html,
                payload.include_rendered_pdf,
                payload.include_markdown,
            ]
        ):
            raise ValueError("Select at least one download output type.")

        repo_settings = self.load_effective_settings(settings)
        normalized_scope = str(payload.scope or "queued").strip().lower()
        normalized_source_ids = _normalize_source_ids(payload.source_ids)
        requested_limit = payload.limit
        selected_phases = _normalize_agent_phase_names(
            payload.selected_phases,
            run_download=run_download,
            run_convert=run_convert,
            run_cleanup=run_llm_cleanup,
            run_title=run_llm_title,
            run_catalog=run_catalog,
            run_citation_verify=run_citation_verify,
            run_rating=run_llm_rating,
            run_summary=run_llm_summary,
        )

        with self._writer_lock():
            if self._download_thread and self._download_thread.is_alive():
                raise ValueError("A repository operation is already running")

            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            imports = list(state.get("imports", []))
            selected_rows, selected_import_id, effective_scope = self._select_rows_for_task_request(
                rows=rows,
                imports=imports,
                scope=normalized_scope,
                import_id=payload.import_id,
                source_ids=normalized_source_ids,
                limit=requested_limit,
                run_download=run_download,
                run_convert=run_convert,
                run_catalog=run_catalog,
                run_citation_verify=run_citation_verify,
                run_llm_cleanup=run_llm_cleanup,
                run_llm_title=run_llm_title,
                run_llm_summary=run_llm_summary,
                run_llm_rating=run_llm_rating,
                output_options=SourceOutputOptions(
                    include_raw_file=payload.include_raw_file,
                    include_rendered_html=payload.include_rendered_html,
                    include_rendered_pdf=payload.include_rendered_pdf,
                    include_markdown=payload.include_markdown,
                ),
            )
            if not selected_rows:
                raise ValueError(f"No repository rows available for scope `{effective_scope}`.")

            project_profile_name, project_profile_yaml = self._load_project_profile_yaml(
                payload.project_profile_name,
                research_purpose=repo_settings.research_purpose,
                default_when_blank=run_llm_rating,
            )
            job_store = self.repo_job_store()
            job_id = job_store.create_job(prefix=REPO_JOB_PREFIX)
            job_store.save_artifact(
                job_id,
                "repo_source_task_context",
                {
                    "scope": effective_scope,
                    "import_id": selected_import_id,
                    "selected_ids": [row.id for row in selected_rows],
                    "selected_phases": selected_phases,
                    "repository_path": str(self.path),
                },
            )

            self._download_state = "running"
            self._download_message = (
                f"Running repository source tasks for {len(selected_rows)} row(s)"
            )

            orchestrator = SourceDownloadOrchestrator(
                job_id=job_id,
                store=job_store,
                rerun_failed_only=payload.rerun_failed_only,
                use_llm=repo_settings.use_llm,
                llm_backend=repo_settings.llm_backend,
                research_purpose=repo_settings.research_purpose,
                searxng_base_url=repo_settings.searxng_base_url,
                fetch_delay=repo_settings.fetch_delay,
                run_download=run_download,
                run_convert=run_convert,
                run_catalog=run_catalog,
                run_citation_verify=run_citation_verify,
                run_llm_cleanup=run_llm_cleanup,
                run_llm_title=run_llm_title,
                run_llm_summary=run_llm_summary,
                run_llm_rating=run_llm_rating,
                force_redownload=payload.force_redownload,
                force_convert=payload.force_convert,
                force_catalog=payload.force_catalog,
                force_citation_verify=payload.force_citation_verify,
                force_llm_cleanup=payload.force_llm_cleanup,
                force_title=payload.force_title,
                force_summary=payload.force_summary,
                force_rating=payload.force_rating,
                project_profile_name=project_profile_name,
                project_profile_yaml=project_profile_yaml,
                output_options=SourceOutputOptions(
                    include_raw_file=payload.include_raw_file,
                    include_rendered_html=payload.include_rendered_html,
                    include_rendered_pdf=payload.include_rendered_pdf,
                    include_markdown=payload.include_markdown,
                ),
                target_rows=[row.model_copy(deep=True) for row in selected_rows],
                output_dir=self.path,
                writes_to_repository=True,
                repository_path=str(self.path),
                selected_scope=effective_scope,
                selected_import_id=selected_import_id,
                selected_phases=selected_phases,
                row_persist_callback=self._persist_source_task_row,
            )

            if live_jobs is not None and live_jobs_lock is not None:
                with live_jobs_lock:
                    if job_id in live_jobs:
                        raise RuntimeError("Source download is already running")
                    live_jobs[job_id] = orchestrator

            self._download_thread = threading.Thread(
                target=self._repository_source_task_worker,
                args=(job_id, orchestrator, live_jobs, live_jobs_lock),
                daemon=True,
            )
            self._download_thread.start()

            return RepositorySourceTaskResponse(
                job_id=job_id,
                status="started",
                scope=effective_scope,
                import_id=selected_import_id,
                total_urls=len(selected_rows),
                message=self._download_message,
            )

    def _repository_source_task_worker(
        self,
        job_id: str,
        orchestrator: SourceDownloadOrchestrator,
        live_jobs: dict[str, SourceDownloadOrchestrator] | None = None,
        live_jobs_lock: threading.Lock | None = None,
    ) -> None:
        try:
            orchestrator.run()
        finally:
            if live_jobs is not None and live_jobs_lock is not None:
                with live_jobs_lock:
                    current = live_jobs.get(job_id)
                    if current is orchestrator:
                        live_jobs.pop(job_id, None)
            status = self.job_store_for(job_id).get_source_status(job_id) or {}
            with self._mutex:
                state = str(status.get("state") or "completed")
                if state == "failed":
                    self._download_state = "failed"
                elif state == "cancelled":
                    self._download_state = "cancelled"
                elif state == "cancelling":
                    self._download_state = "cancelling"
                else:
                    self._download_state = "completed"
                self._download_message = str(
                    status.get("message") or "Repository source tasks completed"
                )

    def mark_source_tasks_cancelling(self, message: str) -> None:
        with self._mutex:
            if self._download_state in {"running", "cancelling"}:
                self._download_state = "cancelling"
                self._download_message = (message or "Stop requested").strip()

    def _persist_source_task_row(self, row: SourceManifestRow) -> None:
        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            column_configs = _load_column_configs(state.get("column_configs", []))
            updated = False

            for index, existing in enumerate(rows):
                if existing.id != row.id:
                    continue
                preserved = row.model_copy(deep=True)
                preserved.import_type = existing.import_type or row.import_type
                preserved.imported_at = existing.imported_at or row.imported_at
                preserved.provenance_ref = existing.provenance_ref or row.provenance_ref
                preserved.custom_fields = dict(existing.custom_fields or {})
                rows[index] = preserved
                updated = True
                break

            if not updated:
                rows.append(row)

            rows = self._sort_rows(rows)
            self._save_state_locked(
                sources=rows,
                citations=citations,
                imports=state.get("imports", []),
                column_configs=column_configs,
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "next_source_id": _next_source_id_from_rows(rows),
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(rows, citations)

    def start_download(self, settings: EffectiveSettings) -> RepositoryActionResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before downloading")

        with self._writer_lock():
            if self._download_thread and self._download_thread.is_alive():
                raise ValueError("Repository download is already running")

            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            queued_ids = [
                row.id for row in rows if (row.fetch_status or "") in {"", "queued"}
            ]

            if not queued_ids:
                return RepositoryActionResponse(
                    status="noop",
                    message="No queued URLs to download",
                    queued_count=0,
                    total_sources=len(rows),
                    total_citations=len(state.get("citations", [])),
                )

            self._download_state = "running"
            self._download_message = f"Downloading {len(queued_ids)} queued URLs"
            self._download_thread = threading.Thread(
                target=self._download_worker,
                args=(queued_ids, settings),
                daemon=True,
            )
            self._download_thread.start()

            return RepositoryActionResponse(
                status="started",
                message=self._download_message,
                queued_count=len(queued_ids),
                total_sources=len(rows),
                total_citations=len(state.get("citations", [])),
            )

    def rebuild(self) -> RepositoryActionResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before rebuilding")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            rows = self._sort_rows(rows)
            citations = self._sort_citations(citations)
            next_source_id = _next_source_id_from_rows(rows)

            self._save_state_locked(
                sources=rows,
                citations=citations,
                imports=state.get("imports", []),
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": next_source_id,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(rows, citations)
            health = self._compute_health(rows, citations)
            self._download_message = (
                f"Rebuilt manifest/citations ({health.missing_files} missing files detected)"
            )

            queued_count = sum(
                1 for row in rows if (row.fetch_status or "") in {"", "queued"}
            )
            return RepositoryActionResponse(
                status="completed",
                message=self._download_message,
                queued_count=queued_count,
                total_sources=len(rows),
                total_citations=len(citations),
            )

    def clear_citations(self) -> RepositoryActionResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before clearing citations")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = self._sort_rows(_load_source_rows(state.get("sources", [])))
            citations = _load_citation_rows(state.get("citations", []))
            deleted_count = len(citations)

            queued_count = sum(
                1 for row in rows if (row.fetch_status or "") in {"", "queued"}
            )
            if deleted_count == 0:
                return RepositoryActionResponse(
                    status="noop",
                    message="No stored citation rows to clear.",
                    queued_count=queued_count,
                    total_sources=len(rows),
                    total_citations=0,
                )

            self._create_backup_snapshot_locked("pre_clear_citations")
            self._save_state_locked(
                sources=rows,
                citations=[],
                imports=state.get("imports", []),
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": _next_source_id_from_rows(rows),
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(rows, [])

            return RepositoryActionResponse(
                status="completed",
                message=(
                    f"Cleared {deleted_count} stored citation row(s). "
                    "Re-run extraction to regenerate citations."
                ),
                queued_count=queued_count,
                total_sources=len(rows),
                total_citations=0,
            )

    def cleanup_repository_layout(self) -> RepositoryActionResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before cleaning up")

        with self._writer_lock():
            state = self._load_state_locked()
            merged = self._merge_source_rows(_load_source_rows(state.get("sources", [])))
            rows = self._sort_rows(merged.rows)
            citations = self._sort_citations(
                self._merge_citation_rows(
                    _load_citation_rows(state.get("citations", [])),
                    rows,
                )
            )
            moved_files = self._normalize_repository_source_storage_locked(rows)
            next_source_id = _next_source_id_from_rows(rows)

            self._save_state_locked(
                sources=rows,
                citations=citations,
                imports=state.get("imports", []),
            )
            self._save_meta_locked(
                {
                    **self._load_meta_locked(),
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": next_source_id,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(rows, citations)

            queued_count = sum(
                1 for row in rows if (row.fetch_status or "") in {"", "queued"}
            )
            self._download_message = (
                f"Normalized repository layout for {len(rows)} sources and moved "
                f"{moved_files} artifact file(s) into sources/<id>/."
            )
            self._download_state = "completed"
            return RepositoryActionResponse(
                status="completed",
                message=self._download_message,
                queued_count=queued_count,
                total_sources=len(rows),
                total_citations=len(citations),
            )

    def manifest_csv_path(self) -> Path:
        if not self.is_attached:
            raise ValueError("No repository attached")
        return self.path / MANIFEST_CSV_NAME

    def manifest_xlsx_path(self) -> Path:
        if not self.is_attached:
            raise ValueError("No repository attached")
        return self.path / MANIFEST_XLSX_NAME

    def citations_csv_path(self) -> Path:
        if not self.is_attached:
            raise ValueError("No repository attached")
        return self.path / CITATIONS_CSV_NAME

    def citations_xlsx_path(self) -> Path:
        if not self.is_attached:
            raise ValueError("No repository attached")
        return self.path / CITATIONS_XLSX_NAME

    def create_export_job(self, scope: str, import_id: str = "") -> RepositoryExportJobResponse:
        if not self.is_attached:
            raise ValueError("Attach a repository before creating an export job")

        normalized_scope = str(scope or "").strip().lower()
        if normalized_scope not in {"all", "queued", "import"}:
            raise ValueError("Invalid scope. Use `all`, `queued`, or `import`.")

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            imports = list(state.get("imports", []))
            selected_rows, normalized_import_id = self._select_rows_for_scope(
                rows=rows,
                imports=imports,
                scope=normalized_scope,
                import_id=import_id,
            )
            bibliography = self._build_export_job_bibliography(selected_rows, normalized_scope)

        if not bibliography.entries:
            if normalized_scope == "import" and normalized_import_id:
                raise RuntimeError(f"No URLs available for import `{normalized_import_id}`.")
            raise RuntimeError(f"No URLs available for scope `{normalized_scope}`.")

        job_store = self.repo_job_store()
        job_id = job_store.create_job(prefix=REPO_JOB_PREFIX)
        job_store.save_artifact(job_id, "03_bibliography", bibliography.model_dump(mode="json"))

        import_suffix = f", import: {normalized_import_id}" if normalized_import_id else ""
        message = (
            f"Repository source set: {len(bibliography.entries)} URLs "
            f"(scope: {normalized_scope}{import_suffix})"
        )
        return RepositoryExportJobResponse(
            job_id=job_id,
            total_urls=len(bibliography.entries),
            scope=normalized_scope,
            import_id=normalized_import_id,
            message=message,
        )

    def seed_job_output_run(self, job_id: str) -> dict[str, int]:
        """Seed a job's 06_sources_manifest from current repository rows.

        This enables summary/rating-only runs when repository artifacts already
        exist, without forcing a new download phase.
        """
        if not self.is_attached:
            raise ValueError("Attach a repository before seeding a job output run")

        job_store = self.job_store_for(job_id)
        bibliography_raw = job_store.load_artifact(job_id, "03_bibliography")
        if not bibliography_raw:
            return {"seeded_rows": 0, "copied_files": 0}

        targets = self._build_job_targets_from_bibliography(bibliography_raw)
        if not targets:
            return {"seeded_rows": 0, "copied_files": 0}

        with self._writer_lock():
            state = self._load_state_locked()
            repo_rows = self._sort_rows(_load_source_rows(state.get("sources", [])))

        by_key: dict[str, SourceManifestRow] = {}
        for row in repo_rows:
            candidate_url = (row.original_url or row.final_url or "").strip()
            if not candidate_url:
                continue
            key = repository_dedupe_key(candidate_url) or dedupe_url_key(candidate_url)
            if not key:
                continue
            by_key.setdefault(key, row)

        output_dir = job_store.get_sources_output_dir(job_id)
        seeded_rows: list[SourceManifestRow] = []
        copied_files = 0

        for target in targets:
            key = repository_dedupe_key(target["original_url"]) or dedupe_url_key(
                target["original_url"]
            )
            if not key:
                continue
            existing = by_key.get(key)
            if not existing:
                continue

            seeded = existing.model_copy(deep=True)
            seeded.id = target["id"]
            seeded.source_document_name = (
                target["source_document_name"] or seeded.source_document_name
            )
            seeded.citation_number = target["citation_number"] or seeded.citation_number
            seeded.original_url = target["original_url"] or seeded.original_url
            if not seeded.repository_source_id:
                seeded.repository_source_id = existing.id

            copied_files += self._copy_repo_artifacts_to_job_output(
                row=seeded,
                output_dir=output_dir,
            )
            seeded_rows.append(seeded)

        if not seeded_rows:
            return {"seeded_rows": 0, "copied_files": copied_files}

        success_count = sum(1 for row in seeded_rows if (row.fetch_status or "") == "success")
        partial_count = sum(1 for row in seeded_rows if (row.fetch_status or "") == "partial")
        failed_count = len(seeded_rows) - success_count - partial_count

        artifact = SourceManifestArtifact(
            rows=seeded_rows,
            total_urls=len(seeded_rows),
            success_count=success_count,
            failed_count=failed_count,
            partial_count=partial_count,
        )
        column_configs = _load_column_configs(self._load_state_locked().get("column_configs", []))
        job_store.save_artifact(job_id, "06_sources_manifest", artifact.model_dump(mode="json"))
        job_store.save_sources_manifest_csv(
            job_id,
            build_manifest_csv(seeded_rows, base_dir=output_dir, column_configs=column_configs),
        )
        job_store.save_sources_manifest_xlsx(
            job_id,
            build_manifest_xlsx(seeded_rows, base_dir=output_dir, column_configs=column_configs),
        )

        return {"seeded_rows": len(seeded_rows), "copied_files": copied_files}

    def _build_job_targets_from_bibliography(self, bibliography_raw: dict[str, Any]) -> list[dict[str, str]]:
        targets: list[dict[str, str]] = []
        seen_keys: set[str] = set()
        row_num = 0

        for entry in bibliography_raw.get("entries", []):
            if not isinstance(entry, dict):
                continue

            url = clean_url_candidate(str(entry.get("url") or ""))
            doi = clean_url_candidate(str(entry.get("doi") or ""))
            if not url and doi:
                url = f"https://doi.org/{doi}"
            if not url:
                continue

            normalized_url, _ = normalize_url(url)
            normalized_candidate = normalized_url or url
            dedupe_key = dedupe_url_key(normalized_candidate)
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)

            row_num += 1
            targets.append(
                {
                    "id": f"{row_num:06d}",
                    "original_url": normalized_candidate,
                    "source_document_name": str(entry.get("source_document_name") or "").strip(),
                    "citation_number": str(entry.get("ref_number") or ""),
                }
            )

        return targets

    def _copy_repo_artifacts_to_job_output(
        self,
        row: SourceManifestRow,
        output_dir: Path,
    ) -> int:
        copied = 0
        for field in JOB_SEED_FILE_FIELDS:
            rel_value = str(getattr(row, field) or "").strip()
            if not rel_value:
                continue
            rel_path = Path(rel_value)
            source_path = rel_path if rel_path.is_absolute() else self.path / rel_path
            if not source_path.is_file():
                continue
            destination = output_dir / rel_path
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source_path, destination)
            copied += 1
        return copied

    def _load_project_profile_yaml(
        self,
        filename: str,
        *,
        research_purpose: str = "",
        default_when_blank: bool = False,
    ) -> tuple[str, str]:
        return resolve_project_profile_yaml(
            self.project_profiles_dir,
            filename,
            research_purpose=research_purpose,
            default_when_blank=default_when_blank,
        )

    def _select_rows_for_scope(
        self,
        rows: list[SourceManifestRow],
        imports: list[dict[str, Any]],
        scope: str,
        import_id: str,
    ) -> tuple[list[SourceManifestRow], str]:
        ordered_rows = self._sort_rows(rows)
        if scope == "all":
            return ordered_rows, ""

        if scope == "queued":
            queued = [
                row for row in ordered_rows if (row.fetch_status or "").strip() in {"", "queued"}
            ]
            return queued, ""

        normalized_import_id = str(import_id or "").strip()
        if scope == "latest_import":
            latest = sorted(
                (
                    item
                    for item in imports
                    if str(item.get("import_id") or "").strip()
                ),
                key=lambda item: str(item.get("imported_at") or ""),
                reverse=True,
            )
            if not latest:
                raise ValueError("No repository imports are available.")
            normalized_import_id = str(latest[0].get("import_id") or "").strip()
            scope = "import"

        if scope != "import":
            raise ValueError("Invalid scope. Use `all`, `queued`, `import`, `latest_import`, or `empty_only`.")
        if not normalized_import_id:
            raise ValueError("`import_id` is required when scope is `import`.")

        known_import_ids = {
            str(item.get("import_id") or "").strip()
            for item in imports
            if str(item.get("import_id") or "").strip()
        }
        if normalized_import_id not in known_import_ids:
            raise ValueError(f"Unknown import_id: {normalized_import_id}")

        import_record = next(
            (
                item
                for item in imports
                if str(item.get("import_id") or "").strip() == normalized_import_id
            ),
            None,
        )
        if import_record is not None:
            selected_ids = set(_import_source_ids(import_record))
            if selected_ids:
                selected = [row for row in ordered_rows if row.id in selected_ids]
                if selected:
                    return selected, normalized_import_id

        prefix = f"{normalized_import_id}:"
        selected = [row for row in ordered_rows if (row.provenance_ref or "").startswith(prefix)]
        return selected, normalized_import_id

    def _select_rows_for_task_request(
        self,
        *,
        rows: list[SourceManifestRow],
        imports: list[dict[str, Any]],
        scope: str,
        import_id: str,
        source_ids: set[str] | None = None,
        limit: int | None = None,
        run_download: bool = False,
        run_convert: bool = False,
        run_catalog: bool = False,
        run_citation_verify: bool = False,
        run_llm_cleanup: bool = False,
        run_llm_title: bool = False,
        run_llm_summary: bool = False,
        run_llm_rating: bool = False,
        output_options: SourceOutputOptions | None = None,
    ) -> tuple[list[SourceManifestRow], str, str]:
        normalized_source_ids = source_ids or set()
        effective_output_options = output_options or SourceOutputOptions()
        if normalized_source_ids:
            ordered_rows = self._sort_rows(rows)
            selected_rows = [row for row in ordered_rows if row.id in normalized_source_ids]
            missing = sorted(normalized_source_ids.difference({row.id for row in selected_rows}))
            if missing:
                raise ValueError(f"Unknown source_ids: {', '.join(missing[:20])}")
            if scope == "empty_only":
                selected_rows = [
                    row
                    for row in selected_rows
                    if self._row_matches_empty_only_scope(
                        row,
                        run_download=run_download,
                        run_convert=run_convert,
                        run_catalog=run_catalog,
                        run_citation_verify=run_citation_verify,
                        run_llm_cleanup=run_llm_cleanup,
                        run_llm_title=run_llm_title,
                        run_llm_summary=run_llm_summary,
                        run_llm_rating=run_llm_rating,
                        output_options=effective_output_options,
                    )
                ]
                effective_scope = "source_ids_empty_only"
            else:
                effective_scope = "source_ids"
            selected_import_id = ""
        elif scope == "empty_only":
            ordered_rows = self._sort_rows(rows)
            selected_rows = [
                row
                for row in ordered_rows
                if self._row_matches_empty_only_scope(
                    row,
                    run_download=run_download,
                    run_convert=run_convert,
                    run_catalog=run_catalog,
                    run_citation_verify=run_citation_verify,
                    run_llm_cleanup=run_llm_cleanup,
                    run_llm_title=run_llm_title,
                    run_llm_summary=run_llm_summary,
                    run_llm_rating=run_llm_rating,
                    output_options=effective_output_options,
                )
            ]
            effective_scope = "empty_only"
            selected_import_id = ""
        else:
            selected_rows, selected_import_id = self._select_rows_for_scope(
                rows=rows,
                imports=imports,
                scope=scope,
                import_id=import_id,
            )
            effective_scope = scope

        safe_limit = max(1, min(int(limit), 500)) if limit is not None else None
        if safe_limit is not None:
            selected_rows = selected_rows[:safe_limit]
        return selected_rows, selected_import_id, effective_scope

    def _row_matches_empty_only_scope(
        self,
        row: SourceManifestRow,
        *,
        run_download: bool,
        run_convert: bool,
        run_catalog: bool,
        run_citation_verify: bool,
        run_llm_cleanup: bool,
        run_llm_title: bool,
        run_llm_summary: bool,
        run_llm_rating: bool,
        output_options: SourceOutputOptions,
    ) -> bool:
        if run_download:
            if output_options.include_raw_file and not self._row_has_artifact(row, "raw_file"):
                return True
            if output_options.include_rendered_html and not self._row_has_artifact(row, "rendered_file"):
                return True
            if output_options.include_rendered_pdf and not self._row_has_artifact(row, "rendered_pdf_file"):
                return True
            if output_options.include_markdown and not self._row_has_artifact(row, "markdown_file"):
                return True
        if run_convert and output_options.include_markdown and not self._row_has_artifact(row, "markdown_file"):
            return True
        if run_catalog and not self._row_has_artifact(row, "catalog_file"):
            return True
        if run_llm_cleanup and not self._row_has_artifact(row, "llm_cleanup_file"):
            return True
        if run_llm_title and not str(row.title or "").strip():
            return True
        if run_citation_verify and not self._row_has_verified_citation(row):
            return True
        if run_llm_summary and not self._row_has_artifact(row, "summary_file"):
            return True
        if run_llm_rating and not self._row_has_artifact(row, "rating_file"):
            return True
        return False

    def _row_has_artifact(self, row: SourceManifestRow, field_name: str) -> bool:
        rel_value = str(getattr(row, field_name) or "").strip()
        if not rel_value:
            return False
        resolved = self._resolve_repository_artifact_path(row, field_name, rel_value)
        return resolved is not None and resolved.is_file()

    def _row_has_verified_citation(self, row: SourceManifestRow) -> bool:
        catalog_rel = str(row.catalog_file or "").strip()
        if not catalog_rel:
            return False
        catalog_path = self.path / catalog_rel
        if not catalog_path.is_file():
            return False
        try:
            payload = json.loads(catalog_path.read_text(encoding="utf-8"))
        except Exception:
            return False
        citation = _coerce_citation_metadata(payload.get("citation"))
        citation = _merge_citation_metadata(
            citation,
            self._row_citation_defaults_payload(row),
            overwrite_existing=False,
        )
        return bool(citation.ready_for_ris)

    def _build_export_job_bibliography(
        self,
        rows: list[SourceManifestRow],
        scope: str,
    ) -> BibliographyArtifact:
        deduped_rows: list[SourceManifestRow] = []
        seen_keys: set[str] = set()

        for row in rows:
            url_candidate = (row.original_url or row.final_url or "").strip()
            if not url_candidate:
                continue
            dedupe_key = repository_dedupe_key(url_candidate) or dedupe_url_key(url_candidate)
            if not dedupe_key:
                continue
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            deduped_rows.append(row)

        entries: list[BibliographyEntry] = []
        next_ref = 1
        for row in deduped_rows:
            url_candidate = (row.original_url or row.final_url or "").strip()
            normalized, _ = normalize_url(url_candidate)
            clean_url = normalized or url_candidate
            if not clean_url:
                continue

            parsed_ref = _parse_int(row.citation_number)
            if parsed_ref is None:
                ref_number = next_ref
                next_ref += 1
            else:
                ref_number = parsed_ref
                if parsed_ref >= next_ref:
                    next_ref = parsed_ref + 1

            entries.append(
                BibliographyEntry(
                    ref_number=ref_number,
                    raw_text=clean_url,
                    source_document_name=row.source_document_name,
                    title=row.title,
                    url=clean_url,
                    parse_confidence=1.0,
                    parse_warnings=[],
                    repair_method=f"repository_export_scope:{scope}",
                )
            )

        return BibliographyArtifact(
            sections=[],
            entries=entries,
            total_raw_entries=len(entries),
            parse_failures=0,
        )

    def merge_job_results(self, job_id: str) -> dict:
        """Merge completed job download results into the attached repository.

        New URLs get new IDs; existing URLs get updated if the new download is
        higher quality. Returns a summary dict. Silently returns if not attached
        or no results are available.
        """
        if not self.is_attached:
            return {"merged": False, "reason": "not_attached"}

        job_store = self.job_store_for(job_id)
        artifact = job_store.load_artifact(job_id, "06_sources_manifest")
        if not artifact:
            return {"merged": False, "reason": "no_manifest"}

        downloaded_rows = _load_source_rows(artifact.get("rows", []))
        if not downloaded_rows:
            return {"merged": False, "reason": "no_rows"}

        output_dir = job_store.get_sources_output_dir(job_id)

        new_sources = 0
        updated_sources = 0
        skipped = 0

        with self._writer_lock():
            state = self._load_state_locked()
            meta = self._load_meta_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))

            # Build dedupe map of existing rows
            existing_by_key: dict[str, SourceManifestRow] = {}
            for row in rows:
                url = row.original_url or row.final_url
                key = repository_dedupe_key(url)
                if not key:
                    key = dedupe_url_key(url)
                if key:
                    existing_by_key[key] = row

            next_source_id = int(meta.get("next_source_id") or 1)

            for dl_row in downloaded_rows:
                # Skip rows that were not actually fetched
                if (dl_row.fetch_status or "").strip() in {"", "queued", "failed"}:
                    skipped += 1
                    continue

                url = dl_row.original_url or dl_row.final_url
                key = repository_dedupe_key(url)
                if not key:
                    key = dedupe_url_key(url)
                if not key:
                    skipped += 1
                    continue

                existing = existing_by_key.get(key)
                if existing:
                    if _row_priority(dl_row) > _row_priority(existing):
                        self._apply_download_result(
                            target=existing, downloaded=dl_row, output_dir=output_dir
                        )
                        updated_sources += 1
                    else:
                        skipped += 1
                else:
                    # New URL — assign ID and add to repository
                    source_id = f"{next_source_id:06d}"
                    next_source_id += 1

                    new_row = SourceManifestRow(
                        id=source_id,
                        repository_source_id=source_id,
                        import_type="job_merge",
                        imported_at=_utc_now_iso(),
                        provenance_ref=f"job:{job_id}",
                        source_document_name=dl_row.source_document_name,
                        citation_number=dl_row.citation_number,
                        original_url=dl_row.original_url,
                    )
                    self._apply_download_result(
                        target=new_row, downloaded=dl_row, output_dir=output_dir
                    )
                    rows.append(new_row)
                    existing_by_key[key] = new_row
                    new_sources += 1

            merged_rows = self._sort_rows(rows)
            self._save_state_locked(
                sources=merged_rows,
                citations=citations,
                imports=state.get("imports", []),
            )
            self._save_meta_locked(
                {
                    **meta,
                    "next_source_id": _next_source_id_from_rows(merged_rows),
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(merged_rows, citations)

        return {
            "merged": True,
            "new_sources": new_sources,
            "updated_sources": updated_sources,
            "skipped": skipped,
            "total_sources": len(rows),
        }

    def start_merge(
        self,
        source_paths: list[str],
    ) -> RepositoryMergeResponse:
        """Start a merge of external repos into the currently attached repo."""
        if not self.is_attached:
            raise ValueError("Attach a repository before merging")

        with self._mutex:
            if self._download_thread and self._download_thread.is_alive():
                raise ValueError("A repository operation is already running")

            self._download_state = "running"
            self._download_message = "Merging repositories..."
            self._download_thread = threading.Thread(
                target=self._merge_worker,
                args=(source_paths,),
                daemon=True,
            )
            self._download_thread.start()

        return RepositoryMergeResponse(
            status="started",
            message="Repository merge started",
        )

    def _merge_worker(
        self,
        source_paths: list[str],
    ) -> None:
        try:
            result = self._merge_repositories(source_paths)
            with self._mutex:
                self._download_state = "completed"
                self._download_message = result.message
        except Exception as exc:  # noqa: BLE001
            with self._mutex:
                self._download_state = "failed"
                self._download_message = f"Merge failed: {type(exc).__name__}: {exc}"

    def _merge_repositories(
        self,
        source_path_strs: list[str],
    ) -> RepositoryMergeResponse:
        """Merge one or more external repositories into the currently attached repo."""
        source_paths: list[Path] = []
        for p in source_path_strs:
            resolved = Path(p).expanduser().resolve()
            if not resolved.is_dir():
                raise ValueError(f"Source repository path is not a directory: {resolved}")
            source_paths.append(resolved)

        with self._writer_lock():
            self._ensure_scaffold_locked()
            self._create_backup_snapshot_locked("pre_merge")

            primary_state = self._load_state_locked()
            primary_rows = _load_source_rows(primary_state.get("sources", []))
            primary_citations = _load_citation_rows(primary_state.get("citations", []))
            imports = list(primary_state.get("imports", []))
            meta = self._load_meta_locked()

            existing_by_key: dict[str, SourceManifestRow] = {}
            for row in primary_rows:
                key = repository_dedupe_key(row.original_url or row.final_url)
                if not key:
                    key = dedupe_url_key(row.original_url or row.final_url)
                if key:
                    existing_by_key[key] = row

            merged_rows = list(primary_rows)
            next_source_id = int(meta.get("next_source_id") or _next_source_id_from_rows(merged_rows))
            now = _utc_now_iso()
            source_id_map: dict[str, str] = {}
            external_citations: list[tuple[ExportRow, Path]] = []

            sources_merged = 0
            updated_sources = 0
            duplicates_removed = 0

            for src_path in source_paths:
                ext_state = self._load_state_from_path(src_path)
                ext_rows = _load_source_rows(ext_state.get("sources", []))
                ext_citations = _load_citation_rows(ext_state.get("citations", []))
                external_citations.extend((citation, src_path) for citation in ext_citations)
                imports.extend(ext_state.get("imports", []))

                for ext_row in ext_rows:
                    key = repository_dedupe_key(ext_row.original_url or ext_row.final_url)
                    if not key:
                        key = dedupe_url_key(ext_row.original_url or ext_row.final_url)
                    if not key:
                        continue

                    existing = existing_by_key.get(key)
                    if existing:
                        source_id_map[f"{src_path}:{ext_row.id}"] = existing.id
                        duplicates_removed += 1
                        if _row_priority(ext_row) > _row_priority(existing):
                            self._apply_repository_row_from_external(
                                target=existing,
                                source_row=ext_row,
                                origin_root=src_path,
                            )
                            updated_sources += 1
                        continue

                    source_id = f"{next_source_id:06d}"
                    next_source_id += 1
                    new_row = ext_row.model_copy(deep=True)
                    new_row.id = source_id
                    new_row.repository_source_id = source_id
                    new_row.import_type = new_row.import_type or "repository_merge"
                    new_row.imported_at = new_row.imported_at or now
                    new_row.provenance_ref = new_row.provenance_ref or f"merge:{src_path.name}"
                    self._copy_repository_artifacts_from_origin(
                        target_id=source_id,
                        source_row=new_row,
                        origin_root=src_path,
                    )
                    self._write_repository_source_metadata(new_row)
                    merged_rows.append(new_row)
                    existing_by_key[key] = new_row
                    source_id_map[f"{src_path}:{ext_row.id}"] = source_id
                    sources_merged += 1

            url_to_source_id: dict[str, str] = {}
            for row in merged_rows:
                key = repository_dedupe_key(row.original_url or row.final_url)
                if key:
                    url_to_source_id[key] = row.id

            merged_citations = list(primary_citations)
            for citation, origin_path in external_citations:
                row = citation.model_copy(deep=True)
                mapped_id = ""
                if row.repository_source_id:
                    mapped_id = source_id_map.get(f"{origin_path}:{row.repository_source_id}", "")
                if not mapped_id:
                    url_key = repository_dedupe_key(row.cited_url) or dedupe_url_key(row.cited_url)
                    mapped_id = url_to_source_id.get(url_key, "")
                if mapped_id:
                    row.repository_source_id = mapped_id
                row.import_type = row.import_type or "repository_merge"
                row.imported_at = row.imported_at or now
                row.provenance_ref = row.provenance_ref or f"merge:{origin_path.name}"
                merged_citations.append(row)

            sorted_rows = self._sort_rows(merged_rows)
            sorted_citations = self._sort_citations(self._dedupe_citations(merged_citations))
            self._save_state_locked(
                sources=sorted_rows,
                citations=sorted_citations,
                imports=imports,
            )
            self._save_meta_locked(
                {
                    **meta,
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": next_source_id,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(sorted_rows, sorted_citations)
            self._last_scan = None

        message = (
            f"Merged {sources_merged} new sources from {len(source_paths)} repo(s) "
            f"({updated_sources} updated in place, {duplicates_removed} duplicates removed, "
            f"{len(sorted_citations)} citations)"
        )
        return RepositoryMergeResponse(
            status="completed",
            message=message,
            sources_merged=sources_merged,
            duplicates_removed=duplicates_removed,
            total_merged_sources=len(sorted_rows),
            total_merged_citations=len(sorted_citations),
        )

    def _apply_repository_row_from_external(
        self,
        target: SourceManifestRow,
        source_row: SourceManifestRow,
        origin_root: Path,
    ) -> None:
        target.repository_source_id = target.id
        target.source_kind = source_row.source_kind or target.source_kind or "url"
        target.source_document_name = source_row.source_document_name or target.source_document_name
        target.citation_number = source_row.citation_number or target.citation_number
        target.original_url = source_row.original_url or target.original_url
        target.final_url = source_row.final_url
        target.fetch_status = source_row.fetch_status
        target.http_status = source_row.http_status
        target.content_type = source_row.content_type
        target.detected_type = source_row.detected_type
        target.fetch_method = source_row.fetch_method
        target.title = source_row.title
        target.title_status = source_row.title_status
        target.author_names = source_row.author_names
        target.publication_date = source_row.publication_date
        target.publication_year = source_row.publication_year
        target.document_type = source_row.document_type
        target.organization_name = source_row.organization_name
        target.organization_type = source_row.organization_type
        target.notes = source_row.notes
        target.error_message = source_row.error_message
        target.fetched_at = source_row.fetched_at or _utc_now_iso()
        target.canonical_url = source_row.canonical_url
        target.sha256 = source_row.sha256
        target.extraction_method = source_row.extraction_method
        target.markdown_char_count = source_row.markdown_char_count
        target.llm_cleanup_needed = source_row.llm_cleanup_needed
        target.llm_cleanup_status = source_row.llm_cleanup_status
        target.catalog_status = source_row.catalog_status
        target.summary_status = source_row.summary_status
        target.rating_status = source_row.rating_status
        target.tags_text = source_row.tags_text
        self._copy_repository_artifacts_from_origin(
            target_id=target.id,
            source_row=target,
            origin_root=origin_root,
            source_artifacts=source_row,
        )
        self._write_repository_source_metadata(target)

    def _copy_repository_artifacts_from_origin(
        self,
        target_id: str,
        source_row: SourceManifestRow,
        origin_root: Path,
        source_artifacts: SourceManifestRow | None = None,
    ) -> None:
        artifacts = source_artifacts or source_row
        for field in FILE_FIELDS:
            if field == "metadata_file":
                continue
            rel_value = str(getattr(artifacts, field) or "").strip()
            if not rel_value:
                setattr(source_row, field, "")
                continue
            source_path = Path(rel_value)
            source_file = source_path if source_path.is_absolute() else origin_root / source_path
            if not source_file.is_file():
                setattr(source_row, field, "")
                continue
            target_rel = _repository_source_file_path(
                source_id=target_id,
                field=field,
                source_name=source_file.name,
                source_row_id=artifacts.id,
            )
            target_abs = self.path / target_rel
            target_abs.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source_file, target_abs)
            setattr(source_row, field, target_rel.as_posix())

    def _write_repository_source_metadata(self, row: SourceManifestRow) -> None:
        metadata_rel = Path(SOURCES_DIR_NAME) / row.id / f"{row.id}_metadata.json"
        metadata_abs = self.path / metadata_rel
        metadata_abs.parent.mkdir(parents=True, exist_ok=True)
        metadata_abs.write_text(
            json.dumps(row.model_dump(mode="json"), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        row.metadata_file = metadata_rel.as_posix()

    def _apply_seed_entry_hints(
        self,
        row: SourceManifestRow,
        entry: BibliographyEntry,
    ) -> None:
        title = str(entry.title or "").strip()
        if title and not str(row.title or "").strip():
            row.title = title
            row.title_status = row.title_status or "existing"

        author_names = "; ".join(
            str(author or "").strip()
            for author in entry.authors
            if str(author or "").strip()
        )
        if author_names and not str(row.author_names or "").strip():
            row.author_names = author_names

        publication_year = str(entry.year or "").strip()
        if publication_year and not str(row.publication_year or "").strip():
            row.publication_year = publication_year

        doi = str(entry.doi or "").strip()
        if doi and not str(row.seed_doi or "").strip():
            row.seed_doi = doi

    def _normalize_repository_source_storage_locked(
        self,
        rows: list[SourceManifestRow],
    ) -> int:
        ref_counts: dict[str, int] = {}
        for row in rows:
            for field in FILE_FIELDS:
                if field == "metadata_file":
                    continue
                rel_value = str(getattr(row, field) or "").strip()
                if not rel_value or Path(rel_value).is_absolute():
                    continue
                ref_counts[Path(rel_value).as_posix()] = ref_counts.get(
                    Path(rel_value).as_posix(), 0
                ) + 1

        moved_files = 0
        for row in rows:
            for field in FILE_FIELDS:
                if field == "metadata_file":
                    continue
                rel_value = str(getattr(row, field) or "").strip()
                source_file = None
                if rel_value:
                    source_file = self._resolve_repository_artifact_path(row, field, rel_value)
                else:
                    legacy_candidate = self._legacy_repository_artifact_candidate(row, field)
                    if legacy_candidate is not None and legacy_candidate.is_file():
                        source_file = legacy_candidate
                        rel_value = _relative_or_absolute(self.path, legacy_candidate)
                if source_file is None or not source_file.is_file():
                    continue

                target_rel = _repository_source_file_path(
                    source_id=row.id,
                    field=field,
                    source_name=source_file.name,
                    source_row_id=row.id,
                )
                target_abs = self.path / target_rel
                if source_file.resolve() == target_abs.resolve():
                    setattr(row, field, target_rel.as_posix())
                    continue

                target_abs.parent.mkdir(parents=True, exist_ok=True)
                normalized_rel = Path(rel_value).as_posix()
                if (
                    not Path(rel_value).is_absolute()
                    and ref_counts.get(normalized_rel, 0) <= 1
                    and self._is_path_within_repo(source_file)
                ):
                    if target_abs.exists():
                        target_abs.unlink()
                    shutil.move(str(source_file), str(target_abs))
                    self._cleanup_empty_parent_dirs(source_file.parent)
                else:
                    shutil.copy2(source_file, target_abs)
                setattr(row, field, target_rel.as_posix())
                moved_files += 1

            self._write_repository_source_metadata(row)

        return moved_files

    def _resolve_repository_artifact_path(
        self,
        row: SourceManifestRow,
        field: str,
        rel_value: str,
    ) -> Path | None:
        candidate = Path(rel_value)
        if candidate.is_absolute():
            return candidate if candidate.is_file() else None

        direct = self.path / candidate
        if direct.is_file():
            return direct

        target_rel = _repository_source_file_path(
            source_id=row.id,
            field=field,
            source_name=candidate.name,
            source_row_id=row.id,
        )
        canonical = self.path / target_rel
        if canonical.is_file():
            return canonical

        fallback = self._legacy_repository_artifact_candidate(row, field)
        if fallback is not None and fallback.is_file():
            return fallback
        return None

    def _legacy_repository_artifact_candidate(
        self,
        row: SourceManifestRow,
        field: str,
    ) -> Path | None:
        filename = ""
        folder = ""
        if field == "raw_file":
            folder = "originals"
            filename = f"{row.id}_source"
        elif field == "rendered_file":
            folder = "rendered"
            filename = f"{row.id}_rendered.html"
        elif field == "rendered_pdf_file":
            folder = "rendered"
            filename = f"{row.id}_rendered.pdf"
        elif field == "markdown_file":
            folder = "markdown"
            filename = f"{row.id}_clean.md"
        elif field == "llm_cleanup_file":
            folder = "markdown"
            filename = f"{row.id}_llm_clean.md"
        elif field == "summary_file":
            folder = "summaries"
            filename = f"{row.id}_summary.md"
        elif field == "rating_file":
            folder = "ratings"
            filename = f"{row.id}_rating.json"
        else:
            return None

        if field == "raw_file":
            originals_dir = self.path / folder
            if originals_dir.is_dir():
                matches = sorted(originals_dir.glob(f"{row.id}_source*"))
                for match in matches:
                    if match.is_file():
                        return match
            return None
        return self.path / folder / filename

    def _cleanup_empty_parent_dirs(self, start_dir: Path) -> None:
        allowed = {
            self.path / "originals",
            self.path / "rendered",
            self.path / "markdown",
            self.path / "summaries",
            self.path / "ratings",
            self.path / "metadata",
        }
        current = start_dir
        while current in allowed:
            try:
                current.rmdir()
            except OSError:
                break
            current = current.parent

    def _resolve_source_file_path_for_kind(
        self,
        row: SourceManifestRow,
        kind: str,
    ) -> Path | None:
        normalized_kind = _normalize_source_file_kind(kind)

        if normalized_kind == "pdf":
            raw_path = self._resolve_repository_artifact_path(row, "raw_file", row.raw_file)
            if raw_path is not None and raw_path.suffix.lower() == ".pdf":
                return raw_path
            return None

        if normalized_kind == "html":
            raw_path = self._resolve_repository_artifact_path(row, "raw_file", row.raw_file)
            if raw_path is not None and raw_path.suffix.lower() in {".html", ".htm"}:
                return raw_path
            return None

        if normalized_kind == "rendered":
            rendered_pdf_path = self._resolve_repository_artifact_path(
                row,
                "rendered_pdf_file",
                row.rendered_pdf_file,
            )
            if rendered_pdf_path is not None:
                return rendered_pdf_path
            rendered_path = self._resolve_repository_artifact_path(
                row,
                "rendered_file",
                row.rendered_file,
            )
            if rendered_path is not None:
                return rendered_path
            return None

        cleanup_path = self._resolve_repository_artifact_path(
            row,
            "llm_cleanup_file",
            row.llm_cleanup_file,
        )
        if cleanup_path is not None:
            return cleanup_path
        return self._resolve_repository_artifact_path(row, "markdown_file", row.markdown_file)

    def _protected_repository_file_paths(
        self,
        rows: list[SourceManifestRow],
    ) -> set[str]:
        protected: set[str] = set()
        for row in rows:
            for field_name in FILE_FIELDS:
                rel_value = str(getattr(row, field_name) or "").strip()
                if not rel_value:
                    continue
                path = self._resolve_repository_artifact_path(row, field_name, rel_value)
                if path is None or not path.is_file():
                    continue
                if not self._is_path_within_repo(path):
                    continue
                protected.add(str(path.resolve()))
        return protected

    def _cleanup_empty_repository_dirs(self, start_dir: Path) -> None:
        stop_dirs = {
            self.path.resolve(),
            self.sources_dir.resolve(),
            self.documents_dir.resolve(),
            self.project_profiles_dir.resolve(),
            self._internal_dir().resolve(),
        }
        current = start_dir.resolve()
        while self._is_path_within_repo(current) and current not in stop_dirs:
            try:
                current.rmdir()
            except OSError:
                break
            current = current.parent

    def _is_path_within_repo(self, path: Path) -> bool:
        try:
            path.resolve().relative_to(self.path.resolve())
            return True
        except Exception:
            return False

    def export_sqlite(self) -> Path:
        """Generate a SQLite database from the attached repository's sources and citations.

        Returns the Path to the generated .db file.
        """
        if not self.is_attached:
            raise ValueError("No repository is attached")

        with self._writer_lock():
            state = self._load_state_locked()

        sources = _load_source_rows(state.get("sources", []))
        citations = _load_citation_rows(state.get("citations", []))

        if not citations:
            raise ValueError("No citations available for export")

        # Read markdown content from source files (prefer LLM cleanup, fall back to raw markdown)
        markdown_by_source_id: dict[str, str] = {}
        for src in sources:
            src_id = src.repository_source_id or src.id
            if not src_id:
                continue
            for field_name in ("llm_cleanup_file", "markdown_file"):
                rel_path = getattr(src, field_name, "") or ""
                if not rel_path:
                    continue
                full_path = self.path / rel_path
                if full_path.is_file():
                    try:
                        markdown_by_source_id[src_id] = full_path.read_text(encoding="utf-8", errors="replace")
                        break
                    except OSError:
                        continue

        db_path = self.path / INTERNAL_DIR_NAME / "wikiclaude_export.db"

        build_wikiclaude_sqlite_db(
            db_path=db_path,
            export_rows=citations,
            source_rows=sources,
            markdown_by_source_id=markdown_by_source_id or None,
        )

        # Checkpoint WAL so the .db file is self-contained for download
        import sqlite3

        with sqlite3.connect(str(db_path), timeout=10) as conn:
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")

        return db_path

    @staticmethod
    def _load_state_from_path(repo_path: Path) -> dict[str, Any]:
        """Load repository_state.json from any repo path."""
        state_path = repo_path / INTERNAL_DIR_NAME / STATE_FILE_NAME
        if not state_path.exists():
            return {"sources": [], "citations": [], "imports": []}
        try:
            data = json.loads(state_path.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                return {"sources": [], "citations": [], "imports": []}
            return {
                "sources": data.get("sources", []),
                "citations": data.get("citations", []),
                "imports": data.get("imports", []),
            }
        except Exception:
            return {"sources": [], "citations": [], "imports": []}

    def _download_worker(self, queued_ids: list[str], settings: EffectiveSettings) -> None:
        try:
            with self._writer_lock():
                state = self._load_state_locked()
                rows = _load_source_rows(state.get("sources", []))
                by_id = {row.id: row for row in rows}
                queued_rows = [by_id[row_id] for row_id in queued_ids if row_id in by_id]

            if not queued_rows:
                with self._writer_lock():
                    self._download_state = "completed"
                    self._download_message = "No queued rows were available"
                return

            job_store = self.repo_job_store()
            job_id = job_store.create_job(prefix=REPO_JOB_PREFIX)
            entries: list[BibliographyEntry] = []
            ref_counter = 1
            for row in queued_rows:
                ref_number = _parse_int(row.citation_number)
                if ref_number is None:
                    ref_number = ref_counter
                    ref_counter += 1
                entries.append(
                    BibliographyEntry(
                        ref_number=ref_number,
                        raw_text=row.original_url or row.final_url,
                        source_document_name=row.source_document_name,
                        url=row.original_url,
                        parse_confidence=1.0,
                        parse_warnings=[],
                        repair_method="repository_download",
                    )
                )

            bib = BibliographyArtifact(
                sections=[],
                entries=entries,
                total_raw_entries=len(entries),
                parse_failures=0,
            )
            job_store.save_artifact(job_id, "03_bibliography", bib.model_dump(mode="json"))

            orchestrator = SourceDownloadOrchestrator(
                job_id=job_id,
                store=job_store,
                rerun_failed_only=False,
                use_llm=settings.use_llm,
                llm_backend=settings.llm_backend,
                research_purpose=settings.research_purpose,
                searxng_base_url=settings.searxng_base_url,
            )
            orchestrator.run()

            downloaded_raw = job_store.load_artifact(job_id, "06_sources_manifest") or {}
            downloaded_rows = _load_source_rows(downloaded_raw.get("rows", []))
            output_dir = job_store.get_sources_output_dir(job_id)

            download_map = {
                repository_dedupe_key(row.original_url or row.final_url): row
                for row in downloaded_rows
                if repository_dedupe_key(row.original_url or row.final_url)
            }

            with self._writer_lock():
                state = self._load_state_locked()
                rows = _load_source_rows(state.get("sources", []))
                row_by_id = {row.id: row for row in rows}

                for row_id in queued_ids:
                    row = row_by_id.get(row_id)
                    if not row:
                        continue
                    key = repository_dedupe_key(row.original_url or row.final_url)
                    downloaded = download_map.get(key)
                    if not downloaded:
                        row.fetch_status = "failed"
                        row.error_message = "download_failure: missing_result"
                        row.fetched_at = _utc_now_iso()
                        continue

                    self._apply_download_result(
                        target=row,
                        downloaded=downloaded,
                        output_dir=output_dir,
                    )

                merged_rows = self._sort_rows(list(row_by_id.values()))
                citations = self._sort_citations(_load_citation_rows(state.get("citations", [])))
                self._save_state_locked(
                    sources=merged_rows,
                    citations=citations,
                    imports=state.get("imports", []),
                )
                self._save_meta_locked(
                    {
                        **self._load_meta_locked(),
                        "next_source_id": _next_source_id_from_rows(merged_rows),
                        "updated_at": _utc_now_iso(),
                    }
                )
                self._rebuild_outputs_locked(merged_rows, citations)
                self._download_state = "completed"
                self._download_message = (
                    f"Completed repository download for {len(queued_ids)} queued URLs"
                )
        except Exception as exc:  # noqa: BLE001
            with self._writer_lock():
                self._download_state = "failed"
                self._download_message = f"Repository download failed: {type(exc).__name__}: {exc}"

    def _apply_download_result(
        self,
        target: SourceManifestRow,
        downloaded: SourceManifestRow,
        output_dir: Path,
    ) -> None:
        target.repository_source_id = target.id
        target.source_kind = downloaded.source_kind or target.source_kind or "url"
        target.final_url = downloaded.final_url
        target.fetch_status = downloaded.fetch_status
        target.http_status = downloaded.http_status
        target.content_type = downloaded.content_type
        target.detected_type = downloaded.detected_type
        target.fetch_method = downloaded.fetch_method
        target.title = downloaded.title
        target.title_status = downloaded.title_status
        target.author_names = downloaded.author_names
        target.publication_date = downloaded.publication_date
        target.publication_year = downloaded.publication_year
        target.document_type = downloaded.document_type
        target.organization_name = downloaded.organization_name
        target.organization_type = downloaded.organization_type
        target.notes = downloaded.notes
        target.error_message = downloaded.error_message
        target.fetched_at = downloaded.fetched_at or _utc_now_iso()
        target.canonical_url = downloaded.canonical_url
        target.sha256 = downloaded.sha256
        target.extraction_method = downloaded.extraction_method
        target.markdown_char_count = downloaded.markdown_char_count
        target.llm_cleanup_needed = downloaded.llm_cleanup_needed
        target.llm_cleanup_status = downloaded.llm_cleanup_status
        target.catalog_status = downloaded.catalog_status
        target.summary_status = downloaded.summary_status
        target.rating_status = downloaded.rating_status
        target.tags_text = downloaded.tags_text

        for field in FILE_FIELDS:
            if field == "metadata_file":
                continue
            rel_value = getattr(downloaded, field)
            if not rel_value:
                setattr(target, field, "")
                continue
            src = output_dir / rel_value
            if not src.exists():
                setattr(target, field, "")
                continue
            dest_rel = _repository_source_file_path(
                source_id=target.id,
                field=field,
                source_name=src.name,
                source_row_id=downloaded.id,
            )
            dest = self.path / dest_rel
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dest)
            setattr(target, field, dest_rel.as_posix())

        self._write_repository_source_metadata(target)

    def _import_entries(
        self,
        entries: list[BibliographyEntry],
        import_type: str,
        provenance_label: str,
        default_source_document: str,
        *,
        write_placeholder_citations: bool = True,
        source_kind: str = "url",
    ) -> RepositoryImportResponse:
        import_id = uuid.uuid4().hex[:12]
        imported_at = _utc_now_iso()

        with self._writer_lock():
            state = self._load_state_locked()
            rows = _load_source_rows(state.get("sources", []))
            citations = _load_citation_rows(state.get("citations", []))
            imports = list(state.get("imports", []))
            meta = self._load_meta_locked()
            next_source_id = int(meta.get("next_source_id") or _next_source_id_from_rows(rows))

            by_key: dict[str, SourceManifestRow] = {}
            for row in rows:
                key = repository_dedupe_key(row.original_url or row.final_url)
                if key:
                    by_key[key] = row

            total_candidates = 0
            accepted_new = 0
            duplicates = 0
            touched_source_ids: list[str] = []

            for entry in entries:
                url = _entry_url(entry)
                if not url:
                    continue
                total_candidates += 1

                dedupe_key = repository_dedupe_key(url)
                if not dedupe_key:
                    dedupe_key = dedupe_url_key(url)

                existing = by_key.get(dedupe_key)
                if existing:
                    duplicates += 1
                    touched_source_ids.append(existing.id)
                    self._apply_seed_entry_hints(existing, entry)
                    if write_placeholder_citations:
                        citations.append(
                            self._placeholder_citation_row(
                                entry=entry,
                                source_id=existing.id,
                                import_type=import_type,
                                imported_at=imported_at,
                                provenance_ref=f"{import_id}:{provenance_label}",
                                default_source_document=default_source_document,
                            )
                        )
                    continue

                source_id = f"{next_source_id:06d}"
                next_source_id += 1

                row = SourceManifestRow(
                    id=source_id,
                    repository_source_id=source_id,
                    source_kind=source_kind,
                    import_type=import_type,
                    imported_at=imported_at,
                    provenance_ref=f"{import_id}:{provenance_label}",
                    source_document_name=entry.source_document_name or default_source_document,
                    citation_number=str(entry.ref_number or ""),
                    original_url=url,
                    fetch_status="queued",
                    notes="queued_for_download",
                )
                self._apply_seed_entry_hints(row, entry)
                rows.append(row)
                by_key[dedupe_key] = row
                accepted_new += 1
                touched_source_ids.append(source_id)

                if write_placeholder_citations:
                    citations.append(
                        self._placeholder_citation_row(
                            entry=entry,
                            source_id=source_id,
                            import_type=import_type,
                            imported_at=imported_at,
                            provenance_ref=f"{import_id}:{provenance_label}",
                            default_source_document=default_source_document,
                        )
                    )

            deduped_citations = self._dedupe_citations(citations)
            sorted_rows = self._sort_rows(rows)
            sorted_citations = self._sort_citations(deduped_citations)

            imports.append(
                {
                    "import_id": import_id,
                    "import_type": import_type,
                    "provenance": provenance_label,
                    "imported_at": imported_at,
                    "total_candidates": total_candidates,
                    "accepted_new": accepted_new,
                    "duplicates_skipped": duplicates,
                    "source_ids": _dedupe_strings(touched_source_ids),
                }
            )

            self._save_state_locked(
                sources=sorted_rows,
                citations=sorted_citations,
                imports=imports,
            )
            self._save_meta_locked(
                {
                    **meta,
                    "schema_version": SCHEMA_VERSION,
                    "next_source_id": next_source_id,
                    "updated_at": _utc_now_iso(),
                }
            )
            self._rebuild_outputs_locked(sorted_rows, sorted_citations)
            self._download_message = (
                f"Imported {accepted_new} new URLs ({duplicates} duplicates skipped)"
            )

            queued_count = sum(
                1 for row in sorted_rows if (row.fetch_status or "") in {"", "queued"}
            )
            return RepositoryImportResponse(
                import_id=import_id,
                import_type=import_type,
                total_candidates=total_candidates,
                accepted_new=accepted_new,
                duplicates_skipped=duplicates,
                total_sources=len(sorted_rows),
                queued_count=queued_count,
                message=self._download_message,
            )

    def _placeholder_citation_row(
        self,
        entry: BibliographyEntry,
        source_id: str,
        import_type: str,
        imported_at: str,
        provenance_ref: str,
        default_source_document: str,
    ) -> ExportRow:
        return ExportRow(
            repository_source_id=source_id,
            import_type=import_type,
            imported_at=imported_at,
            provenance_ref=provenance_ref,
            source_document=entry.source_document_name or default_source_document,
            citation_ref_numbers=str(entry.ref_number or ""),
            cited_authors="; ".join(entry.authors),
            cited_title=entry.title,
            cited_year=entry.year,
            cited_source=entry.journal_or_source,
            cited_doi=entry.doi,
            cited_url=_entry_url(entry),
            cited_raw_entry=entry.raw_text,
            match_method="repository_placeholder",
            warnings="placeholder_row",
        )

    def _dedupe_citations(self, rows: list[ExportRow]) -> list[ExportRow]:
        deduped: list[ExportRow] = []
        seen: set[tuple[str, ...]] = set()
        for row in rows:
            key = (
                row.repository_source_id,
                row.import_type,
                row.provenance_ref,
                row.citation_ref_numbers,
                row.cited_url,
                row.citation_raw,
                row.citing_sentence,
                row.citing_paragraph,
                row.cited_raw_entry,
            )
            if key in seen:
                continue
            seen.add(key)
            deduped.append(row)
        return deduped

    def _scan_and_merge_locked(self) -> tuple[_MergedScan, RepositoryScanSummary]:
        state = self._load_state_locked()
        sources = _load_source_rows(state.get("sources", []))
        citations = _load_citation_rows(state.get("citations", []))

        manifests_scanned = 0
        artifacts_scanned = 0
        citations_scanned = 0

        manifest_paths = sorted(self._iter_paths_named(MANIFEST_CSV_NAME))
        for manifest_path in manifest_paths:
            manifests_scanned += 1
            provenance = _relative_or_absolute(self.path, manifest_path)
            sources.extend(self._read_manifest_csv(manifest_path, provenance))

        artifact_paths = sorted(self._iter_paths_named("06_sources_manifest.json"))
        for artifact_path in artifact_paths:
            artifacts_scanned += 1
            provenance = _relative_or_absolute(self.path, artifact_path)
            sources.extend(self._read_sources_artifact_json(artifact_path, provenance))

        citation_paths = sorted(self._iter_paths_named(CITATIONS_CSV_NAME))
        for citation_path in citation_paths:
            citations_scanned += 1
            provenance = _relative_or_absolute(self.path, citation_path)
            citations.extend(self._read_citations_csv(citation_path, provenance))

        citation_xlsx_paths = sorted(self._iter_paths_named(CITATIONS_XLSX_NAME))
        for citation_xlsx_path in citation_xlsx_paths:
            citations_scanned += 1
            provenance = _relative_or_absolute(self.path, citation_xlsx_path)
            citations.extend(self._read_citations_xlsx(citation_xlsx_path, provenance))

        merged = self._merge_source_rows(sources)
        merged_citations = self._merge_citation_rows(citations, merged.rows)
        merged = _MergedScan(
            rows=merged.rows,
            citations=merged_citations,
            next_source_id=merged.next_source_id,
            duplicate_urls_removed=merged.duplicate_urls_removed,
        )

        scan = RepositoryScanSummary(
            scanned_at=_utc_now_iso(),
            total_sources=len(merged.rows),
            total_citations=len(merged.citations),
            next_source_id=merged.next_source_id,
            manifests_scanned=manifests_scanned,
            artifacts_scanned=artifacts_scanned,
            citations_scanned=citations_scanned,
            duplicate_urls_removed=merged.duplicate_urls_removed,
        )
        return merged, scan

    def _merge_source_rows(self, rows: list[SourceManifestRow]) -> _MergedScan:
        best_by_key: dict[str, SourceManifestRow] = {}
        duplicate_urls_removed = 0

        for row in rows:
            if not row.id:
                row.id = row.repository_source_id or ""
            if not row.repository_source_id:
                row.repository_source_id = row.id
            if not row.import_type:
                row.import_type = "legacy_scan"
            if not row.imported_at:
                row.imported_at = row.fetched_at or _utc_now_iso()
            if not row.provenance_ref:
                row.provenance_ref = "legacy_scan"
            if not row.source_kind:
                row.source_kind = "url"

            key = _source_row_identity_key(row)
            if not key:
                continue

            existing = best_by_key.get(key)
            if not existing:
                best_by_key[key] = row
                continue

            if _row_priority(row) > _row_priority(existing):
                best_by_key[key] = row
            duplicate_urls_removed += 1

        merged_rows = list(best_by_key.values())

        used_ids: set[int] = set()
        pending: list[SourceManifestRow] = []
        for row in merged_rows:
            parsed = _parse_numeric_id(row.id or row.repository_source_id)
            if parsed is None or parsed in used_ids:
                pending.append(row)
                continue
            used_ids.add(parsed)
            row.id = f"{parsed:06d}"
            row.repository_source_id = row.id

        next_id = (max(used_ids) + 1) if used_ids else 1
        for row in pending:
            while next_id in used_ids:
                next_id += 1
            used_ids.add(next_id)
            row.id = f"{next_id:06d}"
            row.repository_source_id = row.id
            next_id += 1

        merged_rows = self._sort_rows(merged_rows)
        return _MergedScan(
            rows=merged_rows,
            citations=[],
            next_source_id=_next_source_id_from_rows(merged_rows),
            duplicate_urls_removed=duplicate_urls_removed,
        )

    def _merge_citation_rows(
        self,
        rows: list[ExportRow],
        sources: list[SourceManifestRow],
    ) -> list[ExportRow]:
        source_by_key: dict[str, str] = {}
        source_ids = {source.id for source in sources}

        for source in sources:
            key = repository_dedupe_key(source.original_url or source.final_url)
            if key:
                source_by_key[key] = source.id

        normalized: list[ExportRow] = []
        for row in rows:
            if not row.repository_source_id:
                key = repository_dedupe_key(row.cited_url)
                if key and key in source_by_key:
                    row.repository_source_id = source_by_key[key]
            if not row.import_type:
                row.import_type = "legacy_scan"
            if not row.imported_at:
                row.imported_at = _utc_now_iso()
            if not row.provenance_ref:
                row.provenance_ref = "legacy_scan"
            if row.repository_source_id and row.repository_source_id not in source_ids:
                key = repository_dedupe_key(row.cited_url)
                if key and key in source_by_key:
                    row.repository_source_id = source_by_key[key]
            normalized.append(row)

        return self._sort_citations(self._dedupe_citations(normalized))

    def _rebuild_outputs_locked(
        self,
        sources: list[SourceManifestRow],
        citations: list[ExportRow],
    ) -> None:
        column_configs = _load_column_configs(self._load_state_locked().get("column_configs", []))
        self.manifest_csv_path().write_text(
            build_manifest_csv(sources, base_dir=self.path, column_configs=column_configs),
            encoding="utf-8-sig",
        )
        self.manifest_xlsx_path().write_bytes(
            build_manifest_xlsx(sources, base_dir=self.path, column_configs=column_configs)
        )

        citation_artifact = ExportArtifact(
            rows=citations,
            total_citations_found=len(citations),
            total_bib_entries=len(citations),
            matched_count=0,
            unmatched_count=0,
        )
        self.citations_csv_path().write_text(
            write_csv(citation_artifact),
            encoding="utf-8-sig",
        )

    def _compute_health(
        self,
        rows: list[SourceManifestRow],
        citations: list[ExportRow],
    ) -> RepositoryHealth:
        missing_files = 0
        source_ids = {row.id for row in rows}
        orphaned_citations = 0

        for row in rows:
            for field in FILE_FIELDS:
                rel = getattr(row, field)
                if not rel:
                    continue
                if not (self.path / rel).exists():
                    missing_files += 1

        for citation in citations:
            if citation.repository_source_id and citation.repository_source_id not in source_ids:
                orphaned_citations += 1

        return RepositoryHealth(
            missing_files=missing_files,
            orphaned_citation_rows=orphaned_citations,
        )

    def _read_manifest_csv(self, path: Path, provenance_ref: str) -> list[SourceManifestRow]:
        try:
            text = path.read_text(encoding="utf-8-sig")
        except Exception:
            text = path.read_text(encoding="utf-8", errors="replace")

        reader = csv.DictReader(io.StringIO(text))
        rows: list[SourceManifestRow] = []
        for raw in reader:
            if not raw:
                continue
            payload = dict(raw)
            payload.setdefault("id", str(payload.get("repository_source_id") or ""))
            payload.setdefault(
                "repository_source_id",
                str(payload.get("id") or payload.get("repository_source_id") or ""),
            )
            payload.setdefault("import_type", "legacy_scan")
            payload.setdefault("imported_at", str(payload.get("fetched_at") or _utc_now_iso()))
            payload.setdefault("provenance_ref", provenance_ref)
            row = _safe_manifest_row(payload)
            if row:
                rows.append(row)
        return rows

    def _read_sources_artifact_json(
        self,
        path: Path,
        provenance_ref: str,
    ) -> list[SourceManifestRow]:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return []

        rows: list[SourceManifestRow] = []
        for raw_row in payload.get("rows", []):
            if not isinstance(raw_row, dict):
                continue
            candidate = dict(raw_row)
            candidate.setdefault("repository_source_id", str(candidate.get("id") or ""))
            candidate.setdefault("import_type", "legacy_scan")
            candidate.setdefault("imported_at", str(candidate.get("fetched_at") or _utc_now_iso()))
            candidate.setdefault("provenance_ref", provenance_ref)
            row = _safe_manifest_row(candidate)
            if row:
                rows.append(row)
        return rows

    def _read_citations_csv(self, path: Path, provenance_ref: str) -> list[ExportRow]:
        try:
            text = path.read_text(encoding="utf-8-sig")
        except Exception:
            text = path.read_text(encoding="utf-8", errors="replace")

        reader = csv.DictReader(io.StringIO(text))
        rows: list[ExportRow] = []
        for raw in reader:
            if not raw:
                continue
            payload = dict(raw)
            payload.setdefault("import_type", "legacy_scan")
            payload.setdefault("imported_at", _utc_now_iso())
            payload.setdefault("provenance_ref", provenance_ref)
            row = _safe_export_row(payload)
            if row:
                rows.append(row)
        return rows

    def _read_citations_xlsx(self, path: Path, provenance_ref: str) -> list[ExportRow]:
        try:
            from openpyxl import load_workbook
        except Exception:
            return []

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return []

        rows: list[ExportRow] = []
        try:
            worksheet = workbook.active
            iterator = worksheet.iter_rows(values_only=True)
            headers_raw = next(iterator, None)
            if not headers_raw:
                return rows
            headers = [str(item or "").strip() for item in headers_raw]

            for values in iterator:
                if not values:
                    continue
                payload: dict[str, Any] = {}
                has_values = False
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[idx] if idx < len(values) else None
                    if value is None:
                        continue
                    text = str(value).strip()
                    if not text:
                        continue
                    has_values = True
                    payload[header] = text
                if not has_values:
                    continue
                payload.setdefault("import_type", "legacy_scan")
                payload.setdefault("imported_at", _utc_now_iso())
                payload.setdefault("provenance_ref", provenance_ref)
                row = _safe_export_row(payload)
                if row:
                    rows.append(row)
        finally:
            workbook.close()

        return rows

    def _iter_paths_named(self, filename: str):
        for path in self.path.rglob(filename):
            if self._is_internal_path(path):
                continue
            if path.is_file():
                yield path

    def _is_internal_path(self, path: Path) -> bool:
        try:
            return path.resolve().is_relative_to(self._internal_dir().resolve())
        except Exception:
            return False

    def _resolve_path(self, value: str) -> Path:
        candidate = Path((value or "").strip()).expanduser()
        if not str(candidate):
            raise ValueError("Repository path is required")
        if not candidate.is_absolute():
            raise ValueError("Repository path must be absolute")
        if candidate.exists() and not candidate.is_dir():
            raise ValueError("Repository path must point to a directory")
        candidate.mkdir(parents=True, exist_ok=True)
        test_file = candidate / ".ra_repo_write_test"
        try:
            test_file.write_text("ok", encoding="utf-8")
            test_file.unlink(missing_ok=True)
        except Exception as exc:
            raise ValueError(f"Repository path is not writable: {exc}") from exc
        return candidate

    def _ensure_internal_dirs(self) -> None:
        self._internal_dir().mkdir(parents=True, exist_ok=True)
        self._backups_dir().mkdir(parents=True, exist_ok=True)
        self._repo_jobs_dir().mkdir(parents=True, exist_ok=True)
        self._lock_path().touch(exist_ok=True)
        (self.path / PROJECT_PROFILES_DIR_NAME).mkdir(parents=True, exist_ok=True)

    def _ensure_scaffold_locked(self) -> None:
        self._internal_dir().mkdir(parents=True, exist_ok=True)
        self._backups_dir().mkdir(parents=True, exist_ok=True)
        self._repo_jobs_dir().mkdir(parents=True, exist_ok=True)
        self._lock_path().touch(exist_ok=True)
        self.project_profiles_dir.mkdir(parents=True, exist_ok=True)
        self.documents_dir.mkdir(parents=True, exist_ok=True)
        self.sources_dir.mkdir(parents=True, exist_ok=True)

        if not self._meta_path().exists():
            self._save_meta_locked(self._default_meta())

        if not self._state_path().exists():
            self._save_state_locked(sources=[], citations=[], imports=[], column_configs=[])

        settings_path = self._internal_dir() / REPO_SETTINGS_FILE_NAME
        if not settings_path.exists():
            settings = RepoSettings()
            settings_path.write_text(
                json.dumps(settings.model_dump(mode="json"), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        self._sync_bundled_ingestion_profiles()
        self.store.sync_project_profiles_to(self.project_profiles_dir)
        profiles_path = self._internal_dir() / INGESTION_PROFILES_FILE_NAME
        if not profiles_path.exists():
            profiles_path.write_text("[]\n", encoding="utf-8")
        suggestions_path = self._internal_dir() / INGESTION_PROFILE_SUGGESTIONS_FILE_NAME
        if not suggestions_path.exists():
            suggestions_path.write_text("[]\n", encoding="utf-8")
        if not self._agent_idempotency_path().exists():
            self._agent_idempotency_path().write_text("{}\n", encoding="utf-8")
        self._load_agent_tokens_locked()
        if not self._agent_resources_path().exists():
            self._agent_resources_path().write_text("[]\n", encoding="utf-8")

        if not self.manifest_csv_path().exists():
            self.manifest_csv_path().write_text(
                build_manifest_csv([]),
                encoding="utf-8-sig",
            )
        if not self.manifest_xlsx_path().exists():
            self.manifest_xlsx_path().write_bytes(build_manifest_xlsx([]))
        if not self.citations_csv_path().exists():
            empty_citations = ExportArtifact(
                rows=[],
                total_citations_found=0,
                total_bib_entries=0,
                matched_count=0,
                unmatched_count=0,
            )
            self.citations_csv_path().write_text(
                write_csv(empty_citations),
                encoding="utf-8-sig",
            )
        self._refresh_agent_resource_index_locked()

    def _default_meta(self) -> dict[str, Any]:
        now = _utc_now_iso()
        return {
            "schema_version": SCHEMA_VERSION,
            "created_at": now,
            "updated_at": now,
            "last_scan_at": "",
            "next_source_id": 1,
        }

    def _load_meta_locked(self) -> dict[str, Any]:
        path = self._meta_path()
        if not path.exists():
            return self._default_meta()
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                return self._default_meta()
            merged = self._default_meta()
            merged.update(data)
            return merged
        except Exception:
            return self._default_meta()

    def _save_meta_locked(self, data: dict[str, Any]) -> None:
        payload = self._default_meta()
        payload.update(data)
        self._meta_path().write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _load_state_locked(self) -> dict[str, Any]:
        path = self._state_path()
        if not path.exists():
            return {"sources": [], "citations": [], "imports": [], "column_configs": []}
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                return {"sources": [], "citations": [], "imports": [], "column_configs": []}
            return {
                "sources": data.get("sources", []),
                "citations": data.get("citations", []),
                "imports": data.get("imports", []),
                "column_configs": data.get("column_configs", []),
            }
        except Exception:
            return {"sources": [], "citations": [], "imports": [], "column_configs": []}

    def _save_state_locked(
        self,
        sources: list[SourceManifestRow],
        citations: list[ExportRow],
        imports: list[dict[str, Any]],
        column_configs: list[RepositoryColumnConfig] | None = None,
    ) -> None:
        normalized_column_configs = (
            column_configs
            if column_configs is not None
            else _load_column_configs(self._load_state_locked().get("column_configs", []))
        )
        payload = {
            "sources": [row.model_dump(mode="json") for row in sources],
            "citations": [row.model_dump(mode="json") for row in citations],
            "imports": imports,
            "column_configs": [
                config.model_dump(mode="json") for config in normalized_column_configs
            ],
        }
        self._state_path().write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _sort_rows(self, rows: list[SourceManifestRow]) -> list[SourceManifestRow]:
        return sorted(
            rows,
            key=lambda row: (
                _parse_numeric_id(row.id) if _parse_numeric_id(row.id) is not None else 10**9,
                row.id,
                row.original_url,
            ),
        )

    def _sort_citations(self, rows: list[ExportRow]) -> list[ExportRow]:
        return sorted(
            rows,
            key=lambda row: (
                _parse_numeric_id(row.repository_source_id)
                if _parse_numeric_id(row.repository_source_id) is not None
                else 10**9,
                row.repository_source_id,
                row.imported_at,
                row.cited_url,
            ),
        )

    def _create_backup_snapshot_locked(self, reason: str) -> None:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        backup_dir = self._backups_dir() / f"{timestamp}_{reason}"
        backup_dir.mkdir(parents=True, exist_ok=True)
        candidates = [
            self.path / MANIFEST_CSV_NAME,
            self.path / MANIFEST_XLSX_NAME,
            self.path / CITATIONS_CSV_NAME,
            self._meta_path(),
            self._state_path(),
        ]
        for src in candidates:
            if src.exists():
                shutil.copy2(src, backup_dir / src.name)

    def _internal_dir(self) -> Path:
        return self.path / INTERNAL_DIR_NAME

    def _meta_path(self) -> Path:
        return self._internal_dir() / META_FILE_NAME

    def _state_path(self) -> Path:
        return self._internal_dir() / STATE_FILE_NAME

    def _lock_path(self) -> Path:
        return self._internal_dir() / LOCK_FILE_NAME

    def _backups_dir(self) -> Path:
        return self._internal_dir() / "backups"

    def _agent_resources_path(self) -> Path:
        return self._internal_dir() / AGENT_RESOURCES_FILE_NAME

    def _agent_tokens_path(self) -> Path:
        return self._internal_dir() / AGENT_TOKENS_FILE_NAME

    def _agent_idempotency_path(self) -> Path:
        return self._internal_dir() / AGENT_IDEMPOTENCY_FILE_NAME

    def _agent_audit_path(self) -> Path:
        return self._internal_dir() / AGENT_AUDIT_FILE_NAME

    def _repo_jobs_dir(self) -> Path:
        return self._internal_dir() / REPO_JOBS_DIR_NAME

    def _load_agent_tokens_locked(self) -> dict[str, str]:
        path = self._agent_tokens_path()
        raw: dict[str, Any] = {}
        if path.exists():
            try:
                parsed = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(parsed, dict):
                    raw = parsed
            except Exception:
                raw = {}

        read_token = str(raw.get("read_token") or "").strip()
        write_token = str(raw.get("write_token") or "").strip()
        updated = False
        if not read_token:
            read_token = f"ra-read-{uuid.uuid4().hex}"
            updated = True
        if not write_token:
            write_token = f"ra-write-{uuid.uuid4().hex}"
            updated = True
        if updated or not path.exists():
            payload = {
                "read_token": read_token,
                "write_token": write_token,
            }
            path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        return {"read_token": read_token, "write_token": write_token}

    def load_agent_tokens(self) -> dict[str, str]:
        if not self.is_attached:
            raise ValueError("No repository attached")
        with self._writer_lock():
            self._ensure_scaffold_locked()
            return self._load_agent_tokens_locked()

    def _load_agent_idempotency_locked(self) -> dict[str, dict[str, str]]:
        path = self._agent_idempotency_path()
        if not path.exists():
            return {}
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return {}
        if not isinstance(raw, dict):
            return {}
        data: dict[str, dict[str, str]] = {}
        for key, value in raw.items():
            if not isinstance(value, dict):
                continue
            data[str(key)] = {
                "request_fingerprint": str(value.get("request_fingerprint") or "").strip(),
                "run_id": str(value.get("run_id") or "").strip(),
            }
        return data

    def _save_agent_idempotency_locked(self, payload: dict[str, dict[str, str]]) -> None:
        self._agent_idempotency_path().write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def resolve_agent_idempotency(
        self,
        idempotency_key: str,
        request_fingerprint: str,
    ) -> str:
        if not self.is_attached:
            raise ValueError("No repository attached")
        normalized_key = str(idempotency_key or "").strip()
        if not normalized_key:
            return ""
        with self._writer_lock():
            self._ensure_scaffold_locked()
            stored = self._load_agent_idempotency_locked()
            entry = stored.get(normalized_key)
            if not entry:
                return ""
            stored_fingerprint = str(entry.get("request_fingerprint") or "").strip()
            if stored_fingerprint and stored_fingerprint != request_fingerprint:
                raise ValueError(
                    "Idempotency key already exists for a different request payload."
                )
            return str(entry.get("run_id") or "").strip()

    def remember_agent_idempotency(
        self,
        idempotency_key: str,
        request_fingerprint: str,
        run_id: str,
    ) -> None:
        normalized_key = str(idempotency_key or "").strip()
        normalized_run_id = str(run_id or "").strip()
        if not normalized_key or not normalized_run_id:
            return
        with self._writer_lock():
            self._ensure_scaffold_locked()
            stored = self._load_agent_idempotency_locked()
            stored[normalized_key] = {
                "request_fingerprint": request_fingerprint,
                "run_id": normalized_run_id,
            }
            self._save_agent_idempotency_locked(stored)

    def append_agent_audit_record(self, payload: dict[str, Any]) -> None:
        if not self.is_attached:
            raise ValueError("No repository attached")
        with self._writer_lock():
            self._ensure_scaffold_locked()
            with self._agent_audit_path().open("a", encoding="utf-8") as handle:
                handle.write(json.dumps(payload, ensure_ascii=False) + "\n")

    def _save_agent_resources_locked(self, resources: list[AgentResourceRecord]) -> None:
        self._agent_resources_path().write_text(
            json.dumps(
                [resource.model_dump(mode="json") for resource in resources],
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

    def _load_agent_resources_locked(self) -> list[AgentResourceRecord]:
        path = self._agent_resources_path()
        if not path.exists():
            return []
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return []
        if not isinstance(raw, list):
            return []
        resources: list[AgentResourceRecord] = []
        for item in raw:
            try:
                resources.append(AgentResourceRecord.model_validate(item))
            except Exception:
                continue
        return resources

    def refresh_agent_resource_index(self) -> list[AgentResourceRecord]:
        if not self.is_attached:
            raise ValueError("No repository attached")
        with self._writer_lock():
            self._ensure_scaffold_locked()
            return self._refresh_agent_resource_index_locked()

    def _refresh_agent_resource_index_locked(self) -> list[AgentResourceRecord]:
        if not self.is_attached:
            return []

        candidates: list[tuple[str, Path]] = []
        root_memory = self.path / "CLAUDE.md"
        if root_memory.is_file():
            candidates.append(("memory", root_memory))

        candidates.extend(
            ("skill", path)
            for path in sorted((self.path / ".claude" / "agents").glob("*.md"))
            if path.is_file()
        )
        candidates.extend(
            ("memory", path)
            for path in sorted((self.path / ".researchassistant" / "memory").glob("*.md"))
            if path.is_file()
        )
        candidates.extend(
            ("skill", path)
            for path in sorted((self.path / ".researchassistant" / "skills").glob("*.md"))
            if path.is_file()
        )
        candidates.extend(
            ("rubric", path)
            for path in sorted(self.project_profiles_dir.glob("*.y*ml"))
            if path.is_file()
        )

        resources: list[AgentResourceRecord] = []
        for kind, source_path in candidates:
            try:
                content = source_path.read_text(encoding="utf-8", errors="replace")
            except OSError:
                continue
            relative_path = source_path.relative_to(self.path).as_posix()
            title, tags, description = _derive_agent_resource_metadata(
                source_path=source_path,
                content=content,
                kind=kind,
            )
            content_hash = hashlib.sha256(content.encode("utf-8")).hexdigest()
            last_modified = _path_mtime_iso(source_path)
            stable_key = f"{kind}:{relative_path}"
            resource_id = hashlib.sha1(stable_key.encode("utf-8")).hexdigest()[:16]
            resources.append(
                AgentResourceRecord(
                    resource_id=resource_id,
                    kind=kind,
                    path=relative_path,
                    title=title,
                    tags=tags,
                    last_modified_at=last_modified,
                    short_description=description,
                    content_hash=content_hash,
                    mime_type=(
                        "application/yaml"
                        if source_path.suffix.lower() in {".yaml", ".yml"}
                        else "text/markdown"
                    ),
                )
            )

        resources = sorted(
            resources,
            key=lambda item: (item.kind, item.title.lower(), item.path.lower()),
        )
        self._save_agent_resources_locked(resources)
        return resources

    def list_agent_resources(self) -> list[AgentResourceRecord]:
        if not self.is_attached:
            raise ValueError("No repository attached")
        with self._writer_lock():
            self._ensure_scaffold_locked()
            return self._refresh_agent_resource_index_locked()

    def get_agent_resource(self, resource_id: str) -> AgentResourceContent:
        if not self.is_attached:
            raise ValueError("No repository attached")
        normalized_id = str(resource_id or "").strip()
        if not normalized_id:
            raise ValueError("resource_id is required")

        with self._writer_lock():
            self._ensure_scaffold_locked()
            resources = self._refresh_agent_resource_index_locked()
            resource = next(
                (item for item in resources if item.resource_id == normalized_id),
                None,
            )
        if resource is None:
            raise ValueError(f"Unknown resource_id: {normalized_id}")

        abs_path = self.path / Path(resource.path)
        if not abs_path.is_file():
            raise ValueError(f"Resource file not found: {resource.path}")
        content = abs_path.read_text(encoding="utf-8", errors="replace")
        return AgentResourceContent(resource=resource, content=content)

    @contextmanager
    def _writer_lock(self):
        with self._mutex:
            lock_file = self._lock_path() if self.is_attached else None
            handle = None
            try:
                if lock_file is not None:
                    lock_file.parent.mkdir(parents=True, exist_ok=True)
                    handle = lock_file.open("a+")
                    if fcntl is not None:
                        fcntl.flock(handle.fileno(), fcntl.LOCK_EX)
                yield
            finally:
                if handle is not None:
                    if fcntl is not None:
                        try:
                            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
                        except Exception:
                            pass
                    handle.close()


def _normalize_source_ids(source_ids: list[str]) -> set[str]:
    normalized: set[str] = set()
    for item in source_ids:
        value = str(item or "").strip()
        if value:
            normalized.add(value)
    return normalized


def _publication_year_from_date_text(value: str) -> str:
    match = re.search(r"\b(19|20)\d{2}\b", str(value or ""))
    return match.group(0) if match else ""


def _normalize_rating_score(value: Any) -> float | None:
    if value in {"", None}:
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    bounded = max(0.0, min(1.0, numeric))
    return round(round(bounded / 0.05) * 0.05, 2)


def _split_relevant_sections_text(value: Any) -> list[str]:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    items = [
        re.sub(r"^[\-\*\d\.\)\s]+", "", line).strip()
        for line in text.split("\n")
    ]
    return [item for item in items if item]


def _normalize_source_file_kind(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if normalized not in {"pdf", "html", "rendered", "md"}:
        raise ValueError("Invalid file kind. Use `pdf`, `html`, `rendered`, or `md`.")
    return normalized


def _normalize_source_file_kinds(values: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for item in values:
        kind = _normalize_source_file_kind(item)
        if kind in seen:
            continue
        seen.add(kind)
        normalized.append(kind)
    return normalized


def _resolve_export_destination_path(value: str) -> Path:
    candidate = Path((value or "").strip()).expanduser()
    if not str(candidate):
        raise ValueError("Export destination path is required")
    if not candidate.is_absolute():
        raise ValueError("Export destination path must be absolute")
    resolved = candidate.resolve()
    if not resolved.exists() or not resolved.is_dir():
        raise ValueError("Export destination must point to an existing directory")

    test_file = resolved / ".ra_export_write_test"
    try:
        test_file.write_text("ok", encoding="utf-8")
        test_file.unlink(missing_ok=True)
    except Exception as exc:
        raise ValueError(f"Export destination is not writable: {exc}") from exc
    return resolved


def _sanitize_export_title(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', " ", str(value or ""))
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    if len(cleaned) > 120:
        cleaned = cleaned[:120].rstrip(" .")
    return cleaned or "Untitled"


def _build_flat_export_filename(
    *,
    source_id: str,
    title: str,
    extension: str,
    used_names: set[str],
) -> str:
    normalized_ext = str(extension or "").strip()
    if normalized_ext and not normalized_ext.startswith("."):
        normalized_ext = f".{normalized_ext}"
    base_name = f"{source_id} - {_sanitize_export_title(title)}"
    candidate = f"{base_name}{normalized_ext}"
    counter = 2
    while candidate.lower() in used_names:
        candidate = f"{base_name} ({counter}){normalized_ext}"
        counter += 1
    used_names.add(candidate.lower())
    return candidate


def _normalize_repository_bundle_file_kinds(values: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for item in values:
        kind = str(item or "").strip().lower()
        if kind not in REPOSITORY_BUNDLE_FILE_KINDS:
            raise ValueError("Invalid bundle file kind. Use `pdf`, `rendered`, `html`, or `md`.")
        if kind in seen:
            continue
        seen.add(kind)
        normalized.append(kind)
    return normalized


def _normalize_repository_bundle_base_url(value: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        raise ValueError("Cloud exports require a Base URL for the uploaded storage files.")
    if re.search(r"\s", normalized):
        raise ValueError("Base URL cannot contain whitespace.")

    if re.match(r"^https?://", normalized, flags=re.IGNORECASE):
        parsed = urlsplit(normalized)
        if not parsed.scheme or not parsed.netloc:
            raise ValueError("Base URL must be a valid http(s) URL or a relative preview path like ./files/.")
        return normalized.rstrip("/") + "/"

    if normalized.startswith("/") or normalized.startswith("//") or "://" in normalized:
        raise ValueError("Base URL must be a valid http(s) URL or a relative preview path like ./files/.")
    if not re.fullmatch(r"(?:\.\.?/)?[A-Za-z0-9._~!$&'()*+,;=:@%-]+(?:/[A-Za-z0-9._~!$&'()*+,;=:@%-]+)*/?", normalized):
        raise ValueError("Base URL must be a valid http(s) URL or a relative preview path like ./files/.")
    return normalized.rstrip("/") + "/"


def _repository_bundle_metadata_value(record: dict[str, Any], key: str) -> str:
    if key == "export_url":
        return str(
            record.get("citation_url")
            or record.get("final_url")
            or record.get("original_url")
            or ""
        )
    if key == "export_overall_rating":
        value = record.get("rating_overall_relevance")
        if value in {"", None}:
            value = record.get("rating_overall")
        return "" if value in {"", None} else str(value)
    value = record.get(key, "")
    return "" if value is None else str(value)


def _dedupe_export_header(label: str, used_headers: set[str]) -> str:
    candidate = str(label or "").strip() or "Column"
    if candidate not in used_headers:
        used_headers.add(candidate)
        return candidate
    counter = 2
    while f"{candidate} ({counter})" in used_headers:
        counter += 1
    deduped = f"{candidate} ({counter})"
    used_headers.add(deduped)
    return deduped


def _repository_bundle_csv_layout(
    column_configs: list[RepositoryColumnConfig] | None = None,
) -> tuple[list[RepositoryColumnConfig], list[str], list[str], list[str]]:
    custom_configs = [config for config in column_configs or [] if config.kind == "custom"]
    used_headers: set[str] = set()
    base_headers = [
        _dedupe_export_header(label, used_headers)
        for _field_name, label in REPOSITORY_BUNDLE_RESEARCH_COLUMNS
    ]
    custom_headers = [
        _dedupe_export_header(
            _manifest_column_label(config.id, column_configs),
            used_headers,
        )
        for config in custom_configs
    ]
    return custom_configs, base_headers, custom_headers, base_headers + custom_headers


def _build_repository_bundle_csv_record(
    record: dict[str, Any],
    *,
    custom_configs: list[RepositoryColumnConfig],
    base_headers: list[str],
    custom_headers: list[str],
) -> dict[str, str]:
    export_record: dict[str, str] = {}
    for (field_name, _label), header in zip(REPOSITORY_BUNDLE_RESEARCH_COLUMNS, base_headers):
        export_record[header] = _repository_bundle_metadata_value(record, field_name)
    for config, header in zip(custom_configs, custom_headers):
        export_record[header] = _repository_bundle_metadata_value(record, config.id)
    return export_record


def _build_repository_bundle_csv_rows(
    rows: list[SourceManifestRow],
    *,
    base_dir: Path | None = None,
    column_configs: list[RepositoryColumnConfig] | None = None,
    manifest_records: dict[str, dict[str, Any]] | None = None,
) -> tuple[list[str], list[dict[str, str]]]:
    custom_configs, base_headers, custom_headers, fieldnames = _repository_bundle_csv_layout(column_configs)
    export_records: list[dict[str, str]] = []
    for row in rows:
        record = (
            manifest_records.get(row.id)
            if manifest_records is not None and row.id in manifest_records
            else build_manifest_record(row, base_dir=base_dir, column_configs=column_configs)
        )
        export_records.append(
            _build_repository_bundle_csv_record(
                record,
                custom_configs=custom_configs,
                base_headers=base_headers,
                custom_headers=custom_headers,
            )
        )
    return fieldnames, export_records


def _render_repository_bundle_csv(
    fieldnames: list[str],
    export_records: list[dict[str, str]],
) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for export_record in export_records:
        writer.writerow(export_record)
    return output.getvalue()


def _repository_bundle_author_label(record: dict[str, Any]) -> str:
    raw = str(
        record.get("citation_authors")
        or record.get("author_names")
        or record.get("citation_publisher")
        or record.get("organization_name")
        or ""
    ).strip()
    if not raw:
        return "Unknown Author"
    authors = [chunk.strip() for chunk in raw.split(";") if chunk.strip()]
    if not authors:
        return _sanitize_export_title(raw) or "Unknown Author"
    first = authors[0]
    if len(authors) > 1:
        first = f"{first} et al"
    return _sanitize_export_title(first) or "Unknown Author"


def _repository_bundle_date_label(record: dict[str, Any]) -> str:
    raw = str(
        record.get("citation_issued")
        or record.get("publication_date")
        or record.get("publication_year")
        or ""
    ).strip()
    return _sanitize_export_title(raw) or "Undated"


def _repository_bundle_title_label(row: SourceManifestRow, record: dict[str, Any]) -> str:
    raw = str(record.get("citation_title") or record.get("title") or "").strip()
    if not raw:
        raw = f"Source {row.id}"
    return _sanitize_export_title(raw) or f"Source {row.id}"


def _build_repository_bundle_filename_base(
    row: SourceManifestRow,
    record: dict[str, Any],
) -> str:
    return " - ".join(
        [
            _repository_bundle_author_label(record),
            _repository_bundle_date_label(record),
            _repository_bundle_title_label(row, record),
        ]
    )


def _build_repository_bundle_filename(
    *,
    row: SourceManifestRow,
    record: dict[str, Any],
    extension: str,
    used_names: set[str],
) -> str:
    normalized_ext = str(extension or "").strip()
    if normalized_ext and not normalized_ext.startswith("."):
        normalized_ext = f".{normalized_ext}"
    base_name = _build_repository_bundle_filename_base(row, record)
    candidate = f"{base_name}{normalized_ext}"
    counter = 2
    while candidate.lower() in used_names:
        candidate = f"{base_name} ({counter}){normalized_ext}"
        counter += 1
    used_names.add(candidate.lower())
    return candidate


def _normalize_storage_slug(value: str, *, fallback: str = "file", max_length: int = 80) -> str:
    normalized = normalize_unicode(str(value or ""))
    normalized = unicodedata.normalize("NFKD", normalized)
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized)
    normalized = normalized.strip("-")
    if len(normalized) > max_length:
        normalized = normalized[:max_length].rstrip("-")
    return normalized or fallback


def _build_repository_bundle_storage_filename(
    *,
    row: SourceManifestRow,
    kind: str,
    display_name: str,
    extension: str,
) -> str:
    normalized_ext = str(extension or "").strip().lower()
    if normalized_ext and not normalized_ext.startswith("."):
        normalized_ext = f".{normalized_ext}"
    if not normalized_ext:
        normalized_ext = _default_extension_for_source_kind(kind)

    display_path = Path(str(display_name or "").strip() or f"{row.id}-{kind}")
    stem = display_path.stem if display_path.suffix else display_path.name
    normalized_stem = _normalize_storage_slug(stem, fallback="file")
    normalized_id = _normalize_storage_slug(row.id, fallback="source", max_length=24)
    normalized_kind = _normalize_storage_slug(kind, fallback="file", max_length=24)
    return f"{normalized_stem}-{normalized_id}-{normalized_kind}{normalized_ext}"


def _resolve_repository_bundle_file_path(
    service: AttachedRepositoryService,
    row: SourceManifestRow,
    kind: str,
) -> Path | None:
    normalized_kind = str(kind or "").strip().lower()
    if normalized_kind == "pdf":
        raw_path = service._resolve_repository_artifact_path(row, "raw_file", row.raw_file)
        if raw_path is not None and raw_path.suffix.lower() == ".pdf":
            return raw_path
        return None
    if normalized_kind == "rendered":
        rendered_pdf_path = service._resolve_repository_artifact_path(
            row,
            "rendered_pdf_file",
            row.rendered_pdf_file,
        )
        if rendered_pdf_path is not None:
            return rendered_pdf_path
        return service._resolve_repository_artifact_path(row, "rendered_file", row.rendered_file)
    if normalized_kind == "html":
        raw_path = service._resolve_repository_artifact_path(row, "raw_file", row.raw_file)
        if raw_path is not None and raw_path.suffix.lower() in {".html", ".htm"}:
            return raw_path
        return service._resolve_repository_artifact_path(row, "rendered_file", row.rendered_file)
    if normalized_kind == "md":
        cleanup_path = service._resolve_repository_artifact_path(
            row,
            "llm_cleanup_file",
            row.llm_cleanup_file,
        )
        if cleanup_path is not None:
            return cleanup_path
        return service._resolve_repository_artifact_path(row, "markdown_file", row.markdown_file)
    raise ValueError("Invalid bundle file kind. Use `pdf`, `rendered`, `html`, or `md`.")


def _default_extension_for_source_kind(kind: str) -> str:
    normalized_kind = _normalize_source_file_kind(kind)
    if normalized_kind in {"pdf", "rendered"}:
        return ".pdf"
    if normalized_kind == "html":
        return ".html"
    return ".md"


def _media_type_for_repository_source_path(path: Path, kind: str) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return "application/pdf"
    if suffix in {".html", ".htm"}:
        return "text/html; charset=utf-8"
    if suffix in {".md", ".markdown", ".txt"}:
        return "text/plain; charset=utf-8"

    guessed, _ = mimetypes.guess_type(path.name)
    if guessed:
        return guessed
    return {
        "pdf": "application/pdf",
        "rendered": "application/pdf",
        "html": "text/html; charset=utf-8",
        "md": "text/plain; charset=utf-8",
    }[_normalize_source_file_kind(kind)]


def _repository_source_file_headers(path: Path) -> dict[str, str]:
    headers = {
        "Referrer-Policy": "no-referrer",
        "X-Content-Type-Options": "nosniff",
    }
    if path.suffix.lower() in {".html", ".htm"}:
        headers["Content-Security-Policy"] = (
            "sandbox; default-src 'none'; img-src data: blob:; "
            "style-src 'unsafe-inline'; font-src data:; media-src data: blob:;"
        )
    return headers


PROCESSABLE_BUILTIN_COLUMNS = {
    "title",
    "author_names",
    "publication_date",
    "publication_year",
    "document_type",
    "organization_name",
    "organization_type",
    "tags_text",
    "notes",
    "summary_text",
    "rating_rationale",
    "relevant_sections",
    "citation_title",
    "citation_authors",
    "citation_issued",
    "citation_type",
    "citation_url",
    "citation_publisher",
    "citation_container_title",
    "citation_volume",
    "citation_issue",
    "citation_pages",
    "citation_doi",
    "citation_report_number",
    "citation_standard_number",
    "citation_language",
    "citation_accessed",
    "citation_ready",
}

NON_LLM_PROCESSABLE_BUILTIN_COLUMNS = {
    "citation_ready",
}

DEFAULT_BUILTIN_COLUMN_PROMPTS: dict[str, str] = {
    "title": (
        "Resolve the source title using document front matter, headings, or clearly labeled metadata. "
        "Prefer the actual document title. If no clear title exists, return a concise descriptive title "
        "of 10 words or fewer supported by the source. Output only the title."
    ),
    "author_names": (
        "Extract the source authors from bylines, front matter, citation metadata, or clearly labeled "
        "metadata. Output a semicolon-separated list of author names. If no individual author is supported "
        "but the producing or publishing organization is clear, output that organization name as the sole author. "
        "If neither is supported, return blank."
    ),
    "publication_date": (
        "Determine the source publication date from visible publication info, front matter, or citation metadata. "
        "Output the most specific supported date using YYYY-MM-DD, YYYY-MM, or YYYY. If no supported date exists, return blank."
    ),
    "publication_year": (
        "Determine the source publication year from visible publication info, publication date, or citation metadata. "
        "Output only the four-digit year as an integer. If no supported year exists, return blank."
    ),
    "document_type": (
        "Classify the source into one short phrase such as report, journal article, web page, standard, "
        "white paper, memo, or presentation. Output only the document type."
    ),
    "organization_name": (
        "Identify the main producing or publishing organization for the source. Output only the organization name. "
        "If none is clearly supported, return blank."
    ),
    "organization_type": (
        "Classify the producing organization into one short phrase such as government, university, nonprofit, "
        "company, utility, standards body, media, or blog. Output only the organization type."
    ),
    "summary_text": (
        "Write one concise paragraph summarizing the source for the stated research purpose. "
        "Output a single paragraph of exactly 3 or 4 sentences with no bullets or headings."
    ),
    "citation_title": (
        "Determine the citation title for this source using supported citation metadata, front matter, headings, "
        "or publication information. Output only the citation title. If unsupported, return blank."
    ),
    "citation_authors": (
        "Determine the citation authors for this source using supported citation metadata, bylines, front matter, "
        "or publication information. Output a semicolon-separated list of author names. If unsupported, return blank."
    ),
    "citation_issued": (
        "Determine the citation issued date for this source using supported publication information. "
        "Output the most specific supported date using YYYY-MM-DD, YYYY-MM, or YYYY. If unsupported, return blank."
    ),
    "citation_type": (
        "Determine the citation item type for this source. Output one short phrase such as report, article, web page, "
        "book, standard, or presentation."
    ),
    "citation_url": (
        "Determine the best supported citation URL for this source. Output only one URL. If unsupported, return blank."
    ),
    "citation_publisher": (
        "Determine the citation publisher for this source. Output only the publisher name. If unsupported, return blank."
    ),
    "citation_container_title": (
        "Determine the citation container title for this source if it is part of a journal, periodical, or collection. "
        "Output only the container title. If unsupported, return blank."
    ),
    "citation_volume": "Determine the citation volume for this source. Output only the volume value. If unsupported, return blank.",
    "citation_issue": "Determine the citation issue for this source. Output only the issue value. If unsupported, return blank.",
    "citation_pages": "Determine the citation pages for this source. Output only the pages value. If unsupported, return blank.",
    "citation_doi": "Determine the citation DOI for this source. Output only the DOI. If unsupported, return blank.",
    "citation_report_number": (
        "Determine the citation report number for this source. Output only the report number. If unsupported, return blank."
    ),
    "citation_standard_number": (
        "Determine the citation standard number for this source. Output only the standard number. If unsupported, return blank."
    ),
    "citation_language": (
        "Determine the citation language for this source. Output only one short phrase naming the language. If unsupported, return blank."
    ),
    "citation_accessed": (
        "Determine the citation accessed date when it is explicitly supported by the repository metadata. "
        "Output the most specific supported date using YYYY-MM-DD, YYYY-MM, or YYYY. If unsupported, return blank."
    ),
}

DATE_SORT_FIELDS = {
    "imported_at",
    "fetched_at",
    "publication_date",
    "citation_verified_at",
}


def _column_config_lookup(
    column_configs: list[RepositoryColumnConfig] | None,
) -> dict[str, RepositoryColumnConfig]:
    lookup: dict[str, RepositoryColumnConfig] = {}
    for config in column_configs or []:
        if config.kind == "custom":
            lookup[config.id] = config
            continue
        builtin_key = str(config.builtin_key or config.id or "").strip()
        if builtin_key:
            lookup[builtin_key] = config
    return lookup


def _default_builtin_column_prompt(field_name: str) -> str:
    return str(DEFAULT_BUILTIN_COLUMN_PROMPTS.get(str(field_name or "").strip(), "")).strip()


def _effective_column_instruction_prompt(
    config: RepositoryColumnConfig | None = None,
    *,
    field_name: str = "",
) -> str:
    if config is not None:
        saved_prompt = str(config.instruction_prompt or "").strip()
        if saved_prompt:
            return saved_prompt
        if config.kind == "builtin":
            return _default_builtin_column_prompt(str(config.builtin_key or config.id or field_name).strip())
        return ""
    return _default_builtin_column_prompt(field_name)


def _effective_column_output_constraint(
    config: RepositoryColumnConfig | None = None,
    *,
    field_name: str = "",
) -> RepositoryColumnOutputConstraint | None:
    if config is not None and config.output_constraint is not None:
        return config.output_constraint
    prompt_text = _effective_column_instruction_prompt(config, field_name=field_name)
    if not prompt_text:
        return None
    return _infer_column_output_constraint(prompt_text)


def _effective_column_include_row_context(
    config: RepositoryColumnConfig | None = None,
) -> bool:
    if config is None:
        return False
    return bool(config.include_row_context)


def _effective_column_include_source_text(
    config: RepositoryColumnConfig | None = None,
) -> bool:
    if config is None:
        return True
    return bool(config.include_source_text)


def _column_requires_llm(
    config: RepositoryColumnConfig | None = None,
    *,
    field_name: str = "",
) -> bool:
    if config is not None and config.kind == "custom":
        return True
    target_key = str(field_name or "").strip()
    if config is not None:
        target_key = str(config.builtin_key or config.id or target_key).strip()
    return target_key not in NON_LLM_PROCESSABLE_BUILTIN_COLUMNS


def _manifest_column_label(
    value: str,
    column_configs: list[RepositoryColumnConfig] | None = None,
) -> str:
    config = _column_config_lookup(column_configs).get(value)
    if config is not None and str(config.label or "").strip():
        return str(config.label or "").strip()

    overrides = {
        "id": "ID",
        "source_kind": "Source Kind",
        "original_url": "Original URL",
        "final_url": "Final URL",
        "canonical_url": "Canonical URL",
        "http_status": "HTTP Status",
        "markdown_char_count": "Markdown Char Count",
        "llm_cleanup_needed": "LLM Cleanup Needed",
        "llm_cleanup_file": "LLM Cleanup File",
        "llm_cleanup_status": "LLM Cleanup Status",
        "catalog_file": "Catalog File",
        "catalog_status": "Catalog Status",
        "author_names": "Authors",
        "publication_date": "Publication Date",
        "publication_year": "Publication Year",
        "document_type": "Document Type",
        "organization_name": "Organization",
        "organization_type": "Organization Type",
        "tags_text": "Tags",
        "sha256": "SHA256",
    }
    if value in overrides:
        return overrides[value]

    parts: list[str] = []
    for piece in str(value or "").split("_"):
        if not piece:
            continue
        lowered = piece.lower()
        if lowered in {"id", "url", "html", "md", "pdf", "llm"}:
            parts.append(lowered.upper())
        else:
            parts.append(lowered.capitalize())
    return " ".join(parts) or value


def _manifest_field_is_numeric(field_name: str) -> bool:
    lowered = str(field_name or "").strip().lower()
    return (
        lowered in {"http_status", "markdown_char_count"}
        or lowered.startswith("rating_")
        or lowered.startswith("flag_")
        or lowered.endswith("_count")
        or lowered.endswith("_score")
        or lowered.endswith("_confidence")
    )


def _manifest_column_sort_type(
    field_name: str,
    column_configs: list[RepositoryColumnConfig] | None = None,
) -> str:
    config = _column_config_lookup(column_configs).get(field_name)
    effective_constraint = _effective_column_output_constraint(config, field_name=field_name)
    if effective_constraint is not None:
        kind = str(effective_constraint.kind or "").strip().lower()
        if kind in {"integer", "number"}:
            return "number"
        if kind == "date":
            return "date"
    if field_name in DATE_SORT_FIELDS:
        return "date"
    if _manifest_field_is_numeric(field_name):
        return "number"
    return "text"


def _manifest_column_processable(
    field_name: str,
    column_configs: list[RepositoryColumnConfig] | None = None,
) -> bool:
    config = _column_config_lookup(column_configs).get(field_name)
    if config is not None and config.kind == "custom":
        return True
    return field_name in PROCESSABLE_BUILTIN_COLUMNS


def _manifest_record_has_column_value(
    record: dict[str, str | int | float | bool],
    field_name: str,
) -> bool:
    normalized_field = str(field_name or "").strip()
    if normalized_field == "citation_ready":
        return bool(record.get(normalized_field))
    return bool(str(record.get(normalized_field) or "").strip())


def _build_manifest_column_metadata(
    records: list[dict[str, str | int | float | bool]],
    column_configs: list[RepositoryColumnConfig] | None = None,
) -> list[dict[str, Any]]:
    ordered_fields = list(SOURCE_MANIFEST_COLUMNS) + list(MANIFEST_DERIVED_COLUMNS)
    seen = set(ordered_fields)
    custom_field_ids: list[str] = []
    for config in column_configs or []:
        if config.kind != "custom":
            continue
        if config.id in seen:
            continue
        seen.add(config.id)
        custom_field_ids.append(config.id)
    dynamic_fields: list[str] = []
    for record in records:
        for key in record.keys():
            if key in seen:
                continue
            seen.add(key)
            dynamic_fields.append(key)

    config_lookup = _column_config_lookup(column_configs)
    custom_field_id_set = set(custom_field_ids)
    all_fields = ordered_fields + custom_field_ids + sorted(dynamic_fields)
    return [
        {
            "key": field_name,
            "label": _manifest_column_label(field_name, column_configs),
            "sortable": True,
            "type": (
                "number"
                if _manifest_column_sort_type(field_name, column_configs) == "number"
                else "text"
            ),
            "kind": "custom" if field_name in custom_field_id_set else "builtin",
            "renamable": field_name in custom_field_id_set,
            "processable": _manifest_column_processable(field_name, column_configs),
            "requires_llm": _column_requires_llm(
                config_lookup.get(field_name),
                field_name=field_name,
            ),
            "sort_type": _manifest_column_sort_type(field_name, column_configs),
            "instruction_prompt": _effective_column_instruction_prompt(
                config_lookup.get(field_name),
                field_name=field_name,
            ),
            "output_constraint": (
                effective_constraint.model_dump(mode="json")
                if (
                    (effective_constraint := _effective_column_output_constraint(
                        config_lookup.get(field_name),
                        field_name=field_name,
                    ))
                    is not None
                )
                else None
            ),
            "include_row_context": _effective_column_include_row_context(
                config_lookup.get(field_name)
            ),
            "include_source_text": _effective_column_include_source_text(
                config_lookup.get(field_name)
            ),
            "last_run_at": str(config_lookup.get(field_name).last_run_at or "")
            if field_name in config_lookup
            else "",
            "last_run_status": str(config_lookup.get(field_name).last_run_status or "")
            if field_name in config_lookup
            else "",
        }
        for field_name in all_fields
    ]


def _manifest_record_float(value: Any) -> float | None:
    if value in {"", None}:
        return None
    if isinstance(value, bool):
        return float(int(value))
    try:
        return float(value)
    except Exception:
        return None


def _manifest_record_matches_thresholds(
    record: dict[str, str | int | float | bool],
    thresholds: dict[str, tuple[float | None, float | None]],
) -> bool:
    for field_name, (min_value, max_value) in thresholds.items():
        if min_value is None and max_value is None:
            continue
        numeric_value = _manifest_record_float(record.get(field_name))
        if numeric_value is None:
            return False
        if min_value is not None and numeric_value < float(min_value):
            return False
        if max_value is not None and numeric_value > float(max_value):
            return False
    return True


def _manifest_record_has_sort_value(
    record: dict[str, str | int | float | bool],
    sort_by: str,
) -> bool:
    value = record.get(sort_by)
    return value not in {"", None}


def _manifest_record_date_value(value: Any) -> float | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).timestamp()
    except Exception:
        pass
    full_match = re.match(r"^(\d{4})[-/](\d{2})[-/](\d{2})", raw)
    if full_match:
        try:
            return datetime(
                int(full_match.group(1)),
                int(full_match.group(2)),
                int(full_match.group(3)),
                tzinfo=timezone.utc,
            ).timestamp()
        except Exception:
            return None
    year_month_match = re.match(r"^(\d{4})[-/](\d{2})$", raw)
    if year_month_match:
        try:
            return datetime(
                int(year_month_match.group(1)),
                int(year_month_match.group(2)),
                1,
                tzinfo=timezone.utc,
            ).timestamp()
        except Exception:
            return None
    year_match = re.match(r"^(\d{4})$", raw)
    if year_match:
        try:
            return datetime(int(year_match.group(1)), 1, 1, tzinfo=timezone.utc).timestamp()
        except Exception:
            return None
    return None


def _manifest_record_sort_value(
    record: dict[str, str | int | float | bool],
    sort_by: str,
    *,
    sort_type: str = "text",
) -> tuple[int, int | float | str]:
    if sort_by == "id":
        raw = str(record.get("id") or "").strip()
        return (0, int(raw)) if raw.isdigit() else (1, raw.lower())

    value = record.get(sort_by)
    if isinstance(value, bool):
        return (0, int(value))
    if isinstance(value, (int, float)):
        return (0, value)
    if sort_type == "date":
        date_value = _manifest_record_date_value(value)
        if date_value is not None:
            return (0, date_value)
        return (1, str(value or "").lower())
    if sort_type == "number":
        numeric_value = _manifest_record_float(value)
        if numeric_value is not None:
            return (0, numeric_value)
    return (1, str(value or "").lower())


def _sort_manifest_records(
    records: list[dict[str, str | int | float | bool]],
    *,
    sort_by: str,
    reverse: bool,
    sort_type: str = "text",
) -> list[dict[str, str | int | float | bool]]:
    if not sort_by:
        return records
    present: list[dict[str, str | int | float | bool]] = []
    missing: list[dict[str, str | int | float | bool]] = []
    for record in records:
        if _manifest_record_has_sort_value(record, sort_by):
            present.append(record)
        else:
            missing.append(record)

    present.sort(
        key=lambda item: _manifest_record_sort_value(item, sort_by, sort_type=sort_type),
        reverse=reverse,
    )
    return present + missing


MANIFEST_EXPORT_SYNTHETIC_FILE_COLUMNS: dict[str, str] = {
    "file_pdf": "pdf",
    "file_html": "html",
    "file_rendered": "rendered",
    "file_md": "md",
}


def _manifest_export_file_value(row: SourceManifestRow, synthetic_key: str) -> str:
    normalized_key = str(synthetic_key or "").strip().lower()
    if normalized_key == "file_pdf":
        raw_file = str(row.raw_file or "").strip()
        return raw_file if raw_file.lower().endswith(".pdf") else ""
    if normalized_key == "file_html":
        raw_file = str(row.raw_file or "").strip()
        return raw_file if raw_file.lower().endswith((".html", ".htm")) else ""
    if normalized_key == "file_rendered":
        return str(row.rendered_file or row.rendered_pdf_file or "").strip()
    if normalized_key == "file_md":
        return str(row.llm_cleanup_file or row.markdown_file or "").strip()
    return ""


def _build_manifest_export_rows(
    rows: list[SourceManifestRow],
    *,
    base_dir: Path | None = None,
    column_configs: list[RepositoryColumnConfig] | None = None,
    column_scope: str = "all",
    column_keys: list[str] | None = None,
) -> tuple[list[str], list[dict[str, str | int | float | bool]]]:
    records: list[dict[str, str | int | float | bool]] = []
    for row in rows:
        record = build_manifest_record(row, base_dir=base_dir, column_configs=column_configs)
        for synthetic_key in MANIFEST_EXPORT_SYNTHETIC_FILE_COLUMNS:
            record[synthetic_key] = _manifest_export_file_value(row, synthetic_key)
        records.append(record)

    manifest_columns = _build_manifest_column_metadata(records, column_configs)
    all_fieldnames = [str(column.get("key") or "") for column in manifest_columns if str(column.get("key") or "").strip()]
    all_fieldnames.extend(
        field_name
        for field_name in MANIFEST_EXPORT_SYNTHETIC_FILE_COLUMNS
        if field_name not in all_fieldnames
    )

    normalized_column_scope = str(column_scope or "all").strip().lower() or "all"
    if normalized_column_scope == "visible":
        requested: list[str] = []
        seen: set[str] = set()
        available = set(all_fieldnames)
        for key in column_keys or []:
            normalized_key = str(key or "").strip()
            if not normalized_key or normalized_key in seen or normalized_key not in available:
                continue
            seen.add(normalized_key)
            requested.append(normalized_key)
        if not requested:
            raise ValueError("Select at least one visible column for spreadsheet export.")
        fieldnames = requested
    else:
        fieldnames = all_fieldnames

    filtered_records = [
        {field_name: record.get(field_name, "") for field_name in fieldnames}
        for record in records
    ]
    return fieldnames, filtered_records


def _build_manifest_export_csv(
    fieldnames: list[str],
    records: list[dict[str, str | int | float | bool]],
) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for record in records:
        writer.writerow(record)
    return output.getvalue()


def _build_repository_bundle_csv(
    rows: list[SourceManifestRow],
    *,
    base_dir: Path | None = None,
    column_configs: list[RepositoryColumnConfig] | None = None,
    manifest_records: dict[str, dict[str, Any]] | None = None,
) -> str:
    fieldnames, export_records = _build_repository_bundle_csv_rows(
        rows,
        base_dir=base_dir,
        column_configs=column_configs,
        manifest_records=manifest_records,
    )
    return _render_repository_bundle_csv(fieldnames, export_records)


def _read_repository_bundle_markdown_text(
    service: AttachedRepositoryService,
    row: SourceManifestRow,
) -> str:
    markdown_path = _resolve_repository_bundle_file_path(service, row, "md")
    if markdown_path is None or not markdown_path.is_file():
        return ""
    try:
        return markdown_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return markdown_path.read_text(encoding="utf-8", errors="replace")


def _build_repository_bundle_viewer_row(
    service: AttachedRepositoryService,
    row: SourceManifestRow,
    *,
    record: dict[str, Any],
    csv_record: dict[str, str],
    custom_configs: list[RepositoryColumnConfig],
    custom_headers: list[str],
    artifact_entries: dict[str, dict[str, str]],
) -> dict[str, Any]:
    citation_title = str(record.get("citation_title") or "").strip()
    source_title = str(record.get("title") or "").strip()
    citation_authors = str(record.get("citation_authors") or "").strip()
    source_authors = str(record.get("author_names") or "").strip()
    citation_issued = str(record.get("citation_issued") or "").strip()
    publication_date = str(record.get("publication_date") or "").strip()
    publication_year = str(record.get("publication_year") or "").strip()
    organization_name = str(record.get("organization_name") or "").strip()
    citation_publisher = str(record.get("citation_publisher") or "").strip()

    ris_text, _ris_exported, _ris_skipped = build_ris_records([row], base_dir=service.path)

    return {
        "id": row.id,
        "sourceTitle": source_title,
        "title": citation_title or source_title or f"Source {row.id}",
        "sourceAuthors": source_authors,
        "authors": citation_authors or source_authors,
        "publicationDate": citation_issued or publication_date or publication_year,
        "publicationYear": publication_year,
        "organization": organization_name or citation_publisher,
        "organizationType": str(record.get("organization_type") or "").strip(),
        "overallRating": _repository_bundle_metadata_value(record, "export_overall_rating"),
        "summary": str(record.get("summary_text") or "").strip(),
        "ratingRationale": str(record.get("rating_rationale") or "").strip(),
        "relevantSections": str(record.get("relevant_sections") or "").strip(),
        "citationType": str(record.get("citation_type") or "").strip(),
        "reportNumber": str(record.get("citation_report_number") or "").strip(),
        "citationUrl": str(record.get("citation_url") or "").strip(),
        "sourceUrl": str(record.get("final_url") or record.get("original_url") or "").strip(),
        "exportUrl": _repository_bundle_metadata_value(record, "export_url"),
        "documentType": str(record.get("document_type") or "").strip(),
        "markdownCharCount": int(row.markdown_char_count or 0),
        "customFields": [
            {
                "key": config.id,
                "label": header,
                "value": _repository_bundle_metadata_value(record, config.id),
            }
            for config, header in zip(custom_configs, custom_headers)
        ],
        "files": artifact_entries,
        "csvRecord": csv_record,
        "ris": ris_text,
        "markdown": _read_repository_bundle_markdown_text(service, row),
    }


def _build_manifest_export_xlsx(
    fieldnames: list[str],
    records: list[dict[str, str | int | float | bool]],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "manifest"
    worksheet.append(fieldnames)
    worksheet.freeze_panes = "A2"

    hyperlink_columns = {name for name in fieldnames if name.endswith("_url") or name.startswith("file_")}
    column_index_by_name = {name: index + 1 for index, name in enumerate(fieldnames)}

    for record in records:
        worksheet.append([record.get(column_name, "") for column_name in fieldnames])
        row_index = worksheet.max_row
        for hyperlink_column in hyperlink_columns:
            cell_value = record.get(hyperlink_column, "")
            if not cell_value:
                continue
            cell = worksheet.cell(row=row_index, column=column_index_by_name[hyperlink_column])
            cell.value = cell_value
            cell.hyperlink = str(cell_value)
            cell.style = "Hyperlink"

    for index, column_name in enumerate(fieldnames, start=1):
        if column_name.endswith("_url") or column_name.startswith("file_"):
            width = 40
        elif column_name in {"summary_text", "rating_rationale", "relevant_sections", "rating_raw_json"}:
            width = 56
        else:
            width = 24
        worksheet.column_dimensions[get_column_letter(index)].width = width

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _parse_json_object(raw_value: str) -> dict[str, Any]:
    try:
        payload = json.loads(raw_value)
    except Exception as exc:
        raise ValueError(f"Invalid JSON response from LLM: {exc}") from exc
    if not isinstance(payload, dict):
        raise ValueError("LLM response must be a JSON object.")
    return payload


def _serialize_column_output_constraint(
    constraint: RepositoryColumnOutputConstraint | None,
) -> dict[str, Any]:
    if constraint is None:
        return {
            "kind": "text",
            "allowed_values": [],
            "max_words": None,
            "fallback_value": "",
            "format_hint": "",
        }
    return constraint.model_dump(mode="json")


def _coerce_column_output_constraint_payload(
    payload: Any,
) -> RepositoryColumnOutputConstraint | None:
    if payload in {None, ""}:
        return None
    try:
        return (
            payload
            if isinstance(payload, RepositoryColumnOutputConstraint)
            else RepositoryColumnOutputConstraint.model_validate(payload)
        )
    except Exception:
        return None


def _infer_column_output_constraint(
    prompt_text: str,
    *,
    existing: RepositoryColumnOutputConstraint | None = None,
) -> RepositoryColumnOutputConstraint:
    normalized_prompt = str(prompt_text or "").strip()
    lowered = normalized_prompt.lower()
    if existing is not None and not normalized_prompt:
        return existing

    if any(phrase in lowered for phrase in {"yes or no", "yes/no", "only yes or no"}):
        return RepositoryColumnOutputConstraint(
            kind="yes_no",
            allowed_values=["yes", "no"],
            fallback_value="",
            format_hint="Lowercase yes or no only.",
        )
    if any(phrase in lowered for phrase in {"four-digit year", "year as an integer", "year only"}):
        return RepositoryColumnOutputConstraint(kind="integer", fallback_value="")
    if "yyyy-mm-dd" in lowered or re.search(r"\boutput (an? )?date\b", lowered):
        return RepositoryColumnOutputConstraint(
            kind="date",
            fallback_value="",
            format_hint="YYYY-MM-DD",
        )
    if any(phrase in lowered for phrase in {"integer", "whole number", "count only"}):
        return RepositoryColumnOutputConstraint(kind="integer", fallback_value="")
    if any(phrase in lowered for phrase in {"numeric", "number only", "decimal"}):
        return RepositoryColumnOutputConstraint(kind="number", fallback_value="")
    if any(phrase in lowered for phrase in {"one short phrase", "short phrase", "single phrase"}):
        return RepositoryColumnOutputConstraint(kind="text", max_words=8, fallback_value="")
    return existing or RepositoryColumnOutputConstraint(kind="text", fallback_value="")


def _truncate_column_run_text(text: str, max_chars: int) -> str:
    normalized = str(text or "").strip()
    if not normalized:
        return ""
    limit = int(max_chars or 0)
    if limit <= 0:
        limit = 12000
    if len(normalized) <= limit:
        return normalized
    head = max(1000, limit // 2)
    tail = max(500, limit - head - 32)
    return f"{normalized[:head].rstrip()}\n\n[... truncated ...]\n\n{normalized[-tail:].lstrip()}"


def _cell_string_value(value: Any) -> str:
    if value in {None, ""}:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    return str(value).strip()


def _normalize_numeric_output(value: Any, *, integer_only: bool) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    try:
        parsed = float(normalized)
    except Exception:
        match = re.search(r"-?\d+(?:\.\d+)?", normalized)
        if not match:
            return ""
        parsed = float(match.group(0))
    if integer_only:
        return str(int(round(parsed)))
    if float(parsed).is_integer():
        return str(int(parsed))
    return f"{parsed:.6f}".rstrip("0").rstrip(".")


def _normalize_date_output(value: Any) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    iso_datetime_match = re.match(r"^(\d{4}-\d{2}-\d{2})", normalized)
    if iso_datetime_match:
        return iso_datetime_match.group(1)
    full_match = re.match(r"^(\d{4})[-/](\d{2})[-/](\d{2})", normalized)
    if full_match:
        return f"{full_match.group(1)}-{full_match.group(2)}-{full_match.group(3)}"
    year_month_match = re.match(r"^(\d{4})[-/](\d{2})$", normalized)
    if year_month_match:
        return f"{year_month_match.group(1)}-{year_month_match.group(2)}"
    year_match = re.match(r"^(\d{4})$", normalized)
    if year_match:
        return year_match.group(1)
    return ""


def _coerce_column_output_value(
    value: Any,
    constraint: RepositoryColumnOutputConstraint,
) -> str:
    normalized = str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not normalized:
        return str(constraint.fallback_value or "")

    if constraint.kind == "yes_no":
        lowered = normalized.lower()
        if lowered.startswith("yes"):
            return "yes"
        if lowered.startswith("no"):
            return "no"
        return str(constraint.fallback_value or "")

    if constraint.kind == "integer":
        return _normalize_numeric_output(normalized, integer_only=True)

    if constraint.kind == "number":
        return _normalize_numeric_output(normalized, integer_only=False)

    if constraint.kind == "date":
        coerced = _normalize_date_output(normalized)
        return coerced or str(constraint.fallback_value or "")

    single_line = " ".join(part.strip() for part in normalized.split("\n") if part.strip())
    if constraint.allowed_values:
        lowered_lookup = {item.lower(): item for item in constraint.allowed_values}
        resolved = lowered_lookup.get(single_line.lower())
        if resolved is not None:
            single_line = resolved
        else:
            return str(constraint.fallback_value or "")
    if constraint.max_words is not None and constraint.max_words > 0:
        single_line = " ".join(single_line.split()[: constraint.max_words])
    return single_line.strip()


def _column_run_author_fallback(
    *,
    config: RepositoryColumnConfig,
    value: str,
    row: SourceManifestRow,
    manifest_record: dict[str, str | int | float | bool],
) -> str:
    target_key = str(config.builtin_key or config.id).strip()
    if target_key not in {"author_names", "citation_authors"}:
        return value
    normalized_value = str(value or "").strip()
    if normalized_value:
        return normalized_value
    for candidate in (
        row.organization_name,
        manifest_record.get("organization_name"),
        manifest_record.get("citation_publisher"),
    ):
        normalized_candidate = _cell_string_value(candidate)
        if normalized_candidate:
            return normalized_candidate
    return ""


def _column_value_patch_for_field(
    config: RepositoryColumnConfig,
    value: str,
) -> dict[str, Any]:
    if config.kind == "custom":
        return {"custom_fields": {config.id: value}}
    target_key = str(config.builtin_key or config.id).strip()
    patch: dict[str, Any] = {target_key: value}
    citation_override_field = ""
    citation_patch_key = ""
    if target_key == "title":
        citation_override_field = "title"
        citation_patch_key = "citation_title"
    elif target_key == "author_names":
        citation_override_field = "authors"
        citation_patch_key = "citation_authors"
    elif target_key in {"publication_date", "publication_year"}:
        citation_override_field = "issued"
        citation_patch_key = "citation_issued"
    elif target_key == "document_type":
        citation_override_field = "item_type"
        citation_patch_key = "citation_type"
    elif target_key == "organization_name":
        citation_override_field = "publisher"
        citation_patch_key = "citation_publisher"
    if citation_patch_key:
        patch[citation_patch_key] = value
        patch["citation_override_fields"] = [citation_override_field]
    return patch


def repository_dedupe_key(url: str) -> str:
    normalized, err = normalize_url(url)
    candidate = normalized or (url or "").strip()
    if not candidate:
        return ""
    if err and "://" not in candidate:
        candidate = f"https://{candidate}"

    try:
        parsed = urlsplit(candidate)
    except Exception:
        return candidate.lower()

    if not parsed.netloc:
        return candidate.lower()

    filtered_params: list[tuple[str, str]] = []
    for key, value in parse_qsl(parsed.query, keep_blank_values=True):
        key_lower = key.lower()
        if key_lower in TRACKING_PARAM_EXACT:
            continue
        if any(key_lower.startswith(prefix) for prefix in TRACKING_PARAM_PREFIXES):
            continue
        filtered_params.append((key, value))

    filtered_params.sort(key=lambda item: (item[0].lower(), item[1]))
    query = urlencode(filtered_params, doseq=True)

    return urlunsplit(
        (
            parsed.scheme.lower() or "https",
            parsed.netloc.lower(),
            parsed.path or "/",
            query,
            "",  # strip fragments
        )
    )


def _dedupe_strings(values: list[str]) -> list[str]:
    deduped: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        deduped.append(text)
        seen.add(text)
    return deduped


def _import_source_ids(import_record: dict[str, Any]) -> list[str]:
    ids: list[str] = []
    explicit_ids = import_record.get("source_ids") or []
    if isinstance(explicit_ids, list):
        ids.extend(str(item or "").strip() for item in explicit_ids)

    documents = import_record.get("documents") or []
    if isinstance(documents, list):
        for item in documents:
            if not isinstance(item, dict):
                continue
            ids.append(str(item.get("source_id") or "").strip())

    return _dedupe_strings(ids)


def _entry_url(entry: BibliographyEntry) -> str:
    url = (entry.url or "").strip()
    if url:
        normalized, _ = normalize_url(url)
        return normalized or url
    doi = (entry.doi or "").strip()
    if not doi:
        return ""
    normalized, _ = normalize_url(f"https://doi.org/{doi}")
    return normalized or f"https://doi.org/{doi}"


def _row_priority(row: SourceManifestRow) -> tuple[int, int, int, int, int, int, int]:
    status_rank = {
        "success": 3,
        "partial": 2,
        "queued": 1,
        "": 1,
        "failed": 0,
    }.get((row.fetch_status or "").strip().lower(), 0)
    return (
        status_rank,
        1 if row.markdown_file else 0,
        1 if row.llm_cleanup_file else 0,
        1 if row.summary_file else 0,
        1 if row.rating_file else 0,
        1 if row.rendered_pdf_file else 0,
        1 if row.raw_file else 0,
    )


@dataclass(frozen=True)
class _DuplicateCandidateContext:
    row: SourceManifestRow
    manifest: dict[str, Any]
    quality_score: int
    doi_key: str
    url_keys: tuple[str, ...]
    sha_key: str
    year: str
    title_normalized: str
    title_tokens: tuple[str, ...]
    title_token_set: frozenset[str]
    author_signature: str
    organization_signature: str


def _next_unique_filename(filename: str, used_filenames: set[str]) -> str:
    base_name = Path(filename or "").name or "upload"
    stem = Path(base_name).stem or "upload"
    suffix = Path(base_name).suffix

    candidate = base_name
    counter = 2
    while candidate.lower() in used_filenames:
        candidate = f"{stem}_{counter}{suffix}"
        counter += 1

    used_filenames.add(candidate.lower())
    return candidate


def _is_original_document_import_path(path: Path) -> bool:
    if not path.is_file():
        return False
    lower_name = path.name.lower()
    if lower_name.endswith(".standardized.md") or lower_name.endswith(".standardized.json"):
        return False
    return path.suffix.lower() in SUPPORTED_DOCUMENT_IMPORT_EXTENSIONS


def _document_citation_provenance_for_document(
    filename: str,
    sha256: str,
) -> str:
    safe_name = Path(filename or "").name or "document"
    hash_part = str(sha256 or "").strip()[:12]
    if hash_part:
        return f"document:{hash_part}:{safe_name}"
    return f"document:{safe_name}"


def _document_citation_provenance(
    filename: str,
    doc_hash_by_filename: dict[str, str],
) -> str:
    return _document_citation_provenance_for_document(
        filename,
        doc_hash_by_filename.get(filename, ""),
    )


def _extract_seed_entries_from_markdown(
    *,
    filename: str,
    text: str,
) -> list[BibliographyEntry]:
    source_name = Path(filename or "seed.md").name
    front_matter = _parse_simple_front_matter(text)
    default_title = str(front_matter.get("title") or "").strip()
    entries: list[BibliographyEntry] = []
    seen_keys: set[str] = set()
    next_ref = 1
    lines = text.splitlines()

    markdown_link_pattern = re.compile(r"\[([^\]]+)\]\((https?://[^)\s]+)\)")
    bare_url_pattern = re.compile(r"(?P<url>https?://[^\s<>()]+|www\.[^\s<>()]+)")

    for line in lines:
        link_matches = list(markdown_link_pattern.finditer(line))
        for match in link_matches:
            raw_url = match.group(2)
            clean_url = clean_url_candidate(raw_url)
            if not clean_url:
                continue
            dedupe_key = repository_dedupe_key(clean_url) or dedupe_url_key(clean_url) or clean_url
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            link_text = re.sub(r"\s+", " ", match.group(1)).strip()
            title = link_text if link_text and link_text != clean_url else ""
            entries.append(
                BibliographyEntry(
                    ref_number=next_ref,
                    raw_text=line.strip() or clean_url,
                    source_document_name=source_name,
                    title=title or default_title,
                    url=clean_url,
                    parse_confidence=1.0,
                    parse_warnings=[],
                    repair_method="seed_markdown",
                )
            )
            next_ref += 1

        stripped_for_bare = markdown_link_pattern.sub(" ", line)
        for match in bare_url_pattern.finditer(stripped_for_bare):
            raw_url = match.group("url")
            clean_url = clean_url_candidate(raw_url)
            if not clean_url:
                continue
            dedupe_key = repository_dedupe_key(clean_url) or dedupe_url_key(clean_url) or clean_url
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            title = _derive_seed_entry_title_from_line(line=line, url=clean_url)
            entries.append(
                BibliographyEntry(
                    ref_number=next_ref,
                    raw_text=line.strip() or clean_url,
                    source_document_name=source_name,
                    title=title or default_title,
                    url=clean_url,
                    parse_confidence=1.0,
                    parse_warnings=[],
                    repair_method="seed_markdown",
                )
            )
            next_ref += 1

    return entries


def _extract_seed_entries_from_document(
    *,
    filename: str,
    document: IngestedDocument,
) -> list[BibliographyEntry]:
    source_name = Path(filename or document.filename or "seed").name
    lines = [block.text.strip() for block in document.blocks if str(block.text or "").strip()]
    text = document.full_text or "\n".join(lines)
    entries: list[BibliographyEntry] = []
    seen_keys: set[str] = set()
    next_ref = 1

    for raw_url in document.inline_citation_urls.values():
        clean_url = clean_url_candidate(raw_url)
        if not clean_url:
            continue
        dedupe_key = repository_dedupe_key(clean_url) or dedupe_url_key(clean_url) or clean_url
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)
        title = _derive_seed_entry_title_from_context(lines=lines, url=clean_url)
        entries.append(
            BibliographyEntry(
                ref_number=next_ref,
                raw_text=clean_url,
                source_document_name=source_name,
                title=title,
                url=clean_url,
                parse_confidence=1.0,
                parse_warnings=[],
                repair_method="seed_document",
            )
        )
        next_ref += 1

    bare_url_pattern = re.compile(r"(?P<url>https?://[^\s<>()]+|www\.[^\s<>()]+)")
    for line in text.splitlines():
        for match in bare_url_pattern.finditer(line):
            raw_url = match.group("url")
            clean_url = clean_url_candidate(raw_url)
            if not clean_url:
                continue
            dedupe_key = repository_dedupe_key(clean_url) or dedupe_url_key(clean_url) or clean_url
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            entries.append(
                BibliographyEntry(
                    ref_number=next_ref,
                    raw_text=line.strip() or clean_url,
                    source_document_name=source_name,
                    title=_derive_seed_entry_title_from_line(line=line, url=clean_url),
                    url=clean_url,
                    parse_confidence=1.0,
                    parse_warnings=[],
                    repair_method="seed_document",
                )
            )
            next_ref += 1

    return entries


def _derive_seed_entry_title_from_context(
    *,
    lines: list[str],
    url: str,
) -> str:
    for line in lines:
        if url not in line:
            continue
        title = _derive_seed_entry_title_from_line(line=line, url=url)
        if title:
            return title
    return ""


def _derive_seed_entry_title_from_line(*, line: str, url: str) -> str:
    stripped = str(line or "").strip()
    if not stripped:
        return ""
    candidate = stripped.replace(url, " ")
    candidate = re.sub(r"\[[^\]]+\]\((?:https?://[^)\s]+)\)", " ", candidate)
    candidate = re.sub(r"\s+", " ", candidate)
    candidate = candidate.strip(" -:|,.;[]()")
    if not candidate:
        return ""
    if len(candidate) > 240:
        candidate = candidate[:240].rstrip(" ,.;:-")
    if re.fullmatch(r"https?://.+", candidate):
        return ""
    return candidate


def _local_document_detected_type(ext: str) -> str:
    normalized = str(ext or "").strip().lower()
    if normalized == ".pdf":
        return "pdf"
    if normalized in {".html", ".htm"}:
        return "html"
    if normalized in {".doc", ".docx", ".md", ".rtf", ".txt"}:
        return "document"
    return "unsupported"


def _extract_markdown_seed_title(text: str) -> str:
    front_matter = _parse_simple_front_matter(text)
    title = str(front_matter.get("title") or "").strip()
    if title:
        return title
    return _extract_markdown_title(text)


def _source_row_identity_key(row: SourceManifestRow) -> str:
    source_kind = str(row.source_kind or "").strip().lower() or "url"
    if source_kind == "uploaded_document":
        sha256 = str(row.sha256 or "").strip().lower()
        if sha256:
            return f"uploaded_document:sha256:{sha256}"
        source_id = str(row.repository_source_id or row.id or "").strip()
        if source_id:
            return f"uploaded_document:id:{source_id}"
        raw_file = str(row.raw_file or "").strip().lower()
        if raw_file:
            return f"uploaded_document:file:{raw_file}"
        return ""

    candidate_url = row.original_url or row.final_url
    dedupe_key = repository_dedupe_key(candidate_url)
    if dedupe_key:
        return f"url:{dedupe_key}"
    fallback = dedupe_url_key(candidate_url)
    if fallback:
        return f"url:{fallback}"
    source_id = str(row.repository_source_id or row.id or "").strip()
    return f"url:id:{source_id}" if source_id else ""


def _sortable_source_id(value: str) -> tuple[int, str]:
    raw = str(value or "").strip()
    return (0, f"{int(raw):09d}") if raw.isdigit() else (1, raw)


def _normalize_duplicate_text(value: Any) -> str:
    normalized = re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower())
    return re.sub(r"\s+", " ", normalized).strip()


def _duplicate_title_tokens(value: Any) -> tuple[str, ...]:
    normalized = _normalize_duplicate_text(value)
    if not normalized:
        return ()
    tokens: list[str] = []
    seen: set[str] = set()
    for token in normalized.split():
        if token in DUPLICATE_TITLE_STOP_WORDS:
            continue
        if token.isdigit():
            continue
        if len(token) < 3 and token.isalpha():
            continue
        if token in seen:
            continue
        seen.add(token)
        tokens.append(token)
    return tuple(tokens)


def _duplicate_signature(value: Any) -> str:
    tokens = sorted(_duplicate_title_tokens(value))
    if len(tokens) < 2:
        return ""
    return " ".join(tokens)


def _normalize_duplicate_doi(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", text)
    text = re.sub(r"^doi:\s*", "", text)
    return text.strip(" /")


def _duplicate_doi_candidates(manifest: dict[str, Any], row: SourceManifestRow) -> list[str]:
    return [
        str(manifest.get("citation_doi") or ""),
        str(manifest.get("citation_url") or ""),
        row.original_url,
        row.final_url,
    ]


def _duplicate_year(manifest: dict[str, Any], row: SourceManifestRow) -> str:
    for candidate in (
        manifest.get("citation_issued"),
        row.publication_year,
        row.publication_date,
    ):
        year = _publication_year_from_date_text(str(candidate or ""))
        if year:
            return year
    return ""


def _duplicate_title_value(manifest: dict[str, Any], row: SourceManifestRow) -> str:
    return str(manifest.get("citation_title") or row.title or "").strip()


def _duplicate_author_value(manifest: dict[str, Any], row: SourceManifestRow) -> str:
    return str(manifest.get("citation_authors") or row.author_names or "").strip()


def _duplicate_organization_value(manifest: dict[str, Any], row: SourceManifestRow) -> str:
    return str(manifest.get("citation_publisher") or row.organization_name or "").strip()


def _duplicate_url_keys(manifest: dict[str, Any], row: SourceManifestRow) -> tuple[str, ...]:
    keys: list[str] = []
    for candidate in (
        str(manifest.get("citation_url") or ""),
        row.canonical_url,
        row.final_url,
        row.original_url,
    ):
        key = repository_dedupe_key(candidate) or dedupe_url_key(candidate)
        if key:
            keys.append(key)
    return tuple(_dedupe_strings(keys))


def _duplicate_quality_score(row: SourceManifestRow, manifest: dict[str, Any]) -> int:
    fetch_status = str(row.fetch_status or "").strip().lower()
    score = {
        "success": 40,
        "partial": 28,
        "not_applicable": 22,
        "queued": 10,
        "pending": 8,
        "failed": 0,
    }.get(fetch_status, 4)
    score += 8 if row.llm_cleanup_file else 0
    score += 6 if row.markdown_file else 0
    score += 6 if row.raw_file else 0
    score += 4 if row.summary_file else 0
    score += 4 if row.rating_file else 0
    score += 4 if row.catalog_file else 0
    score += 3 if row.metadata_file else 0
    score += 3 if bool(manifest.get("citation_ready")) else 0
    score += 2 if _duplicate_title_value(manifest, row) else 0
    score += 2 if _duplicate_author_value(manifest, row) else 0
    score += 2 if _duplicate_organization_value(manifest, row) else 0
    score += 1 if _duplicate_year(manifest, row) else 0
    return score


def _build_duplicate_candidate_context(
    row: SourceManifestRow,
    *,
    base_dir: Path,
) -> _DuplicateCandidateContext:
    manifest = build_manifest_record(row, base_dir=base_dir)
    title_value = _duplicate_title_value(manifest, row)
    title_tokens = _duplicate_title_tokens(title_value)
    doi_key = ""
    for candidate in _duplicate_doi_candidates(manifest, row):
        doi_key = _normalize_duplicate_doi(candidate)
        if doi_key:
            break
    sha_key = ""
    raw_sha = str(row.sha256 or "").strip().lower()
    if raw_sha and (
        str(row.source_kind or "").strip().lower() == "uploaded_document"
        or str(row.fetch_status or "").strip().lower() in {"success", "partial", "not_applicable"}
    ):
        sha_key = raw_sha
    return _DuplicateCandidateContext(
        row=row,
        manifest=manifest,
        quality_score=_duplicate_quality_score(row, manifest),
        doi_key=doi_key,
        url_keys=_duplicate_url_keys(manifest, row),
        sha_key=sha_key,
        year=_duplicate_year(manifest, row),
        title_normalized=_normalize_duplicate_text(title_value),
        title_tokens=title_tokens,
        title_token_set=frozenset(title_tokens),
        author_signature=_duplicate_signature(_duplicate_author_value(manifest, row)),
        organization_signature=_duplicate_signature(_duplicate_organization_value(manifest, row)),
    )


def _duplicate_contexts_by_bucket_size(
    buckets: dict[str, list[_DuplicateCandidateContext]],
) -> list[list[_DuplicateCandidateContext]]:
    ordered = []
    for items in buckets.values():
        unique: list[_DuplicateCandidateContext] = []
        seen: set[str] = set()
        for context in items:
            if context.row.id in seen:
                continue
            seen.add(context.row.id)
            unique.append(context)
        if len(unique) >= 2:
            ordered.append(unique)
    ordered.sort(
        key=lambda items: (
            -len(items),
            min((_sortable_source_id(item.row.id) for item in items), default=(1, "")),
        )
    )
    return ordered


def _duplicate_title_similarity(
    left: _DuplicateCandidateContext,
    right: _DuplicateCandidateContext,
) -> tuple[float, float, float]:
    if not left.title_token_set or not right.title_token_set:
        return 0.0, 0.0, 0.0
    intersection = len(left.title_token_set & right.title_token_set)
    union = len(left.title_token_set | right.title_token_set)
    smallest = min(len(left.title_token_set), len(right.title_token_set))
    if intersection <= 0 or union <= 0 or smallest <= 0:
        return 0.0, 0.0, 0.0
    jaccard = intersection / union
    overlap = intersection / smallest
    sequence_ratio = SequenceMatcher(
        None,
        left.title_normalized,
        right.title_normalized,
    ).ratio()
    return jaccard, overlap, sequence_ratio


def _duplicate_titles_match(
    left: _DuplicateCandidateContext,
    right: _DuplicateCandidateContext,
) -> bool:
    if not left.title_normalized or not right.title_normalized:
        return False
    if left.title_normalized == right.title_normalized:
        return True
    if len(left.title_token_set) < 3 or len(right.title_token_set) < 3:
        return False
    jaccard, overlap, sequence_ratio = _duplicate_title_similarity(left, right)
    intersection = len(left.title_token_set & right.title_token_set)
    if overlap >= 0.9 and intersection >= 4:
        return True
    return overlap >= 0.8 and jaccard >= 0.65 and sequence_ratio >= 0.74


def _cluster_duplicate_title_contexts(
    contexts: list[_DuplicateCandidateContext],
) -> list[list[_DuplicateCandidateContext]]:
    unique = sorted(
        {context.row.id: context for context in contexts}.values(),
        key=lambda context: _sortable_source_id(context.row.id),
    )
    clusters: list[list[_DuplicateCandidateContext]] = []
    visited: set[int] = set()
    for index, context in enumerate(unique):
        if index in visited:
            continue
        stack = [index]
        component: list[_DuplicateCandidateContext] = []
        while stack:
            current = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            component.append(unique[current])
            for neighbor, candidate in enumerate(unique):
                if neighbor == current or neighbor in visited:
                    continue
                if _duplicate_titles_match(unique[current], candidate):
                    stack.append(neighbor)
        if len(component) >= 2:
            clusters.append(component)
    clusters.sort(
        key=lambda items: (
            -len(items),
            min((_sortable_source_id(item.row.id) for item in items), default=(1, "")),
        )
    )
    return clusters


def _duplicate_keep_sort_key(context: _DuplicateCandidateContext) -> tuple[int, ...]:
    priority = _row_priority(context.row)
    source_id = str(context.row.id or "").strip()
    return (
        context.quality_score,
        priority[0],
        priority[1],
        priority[2],
        priority[3],
        priority[4],
        priority[5],
        priority[6],
        -int(source_id) if source_id.isdigit() else 0,
    )


def _duplicate_candidate_row(context: _DuplicateCandidateContext) -> RepositoryDuplicateCandidateRow:
    manifest = context.manifest
    return RepositoryDuplicateCandidateRow(
        id=context.row.id,
        title=_duplicate_title_value(manifest, context.row),
        author_names=_duplicate_author_value(manifest, context.row),
        publication_date=str(manifest.get("citation_issued") or context.row.publication_date or ""),
        publication_year=context.year,
        organization_name=_duplicate_organization_value(manifest, context.row),
        document_type=str(context.row.document_type or ""),
        source_kind=str(context.row.source_kind or ""),
        fetch_status=str(context.row.fetch_status or ""),
        original_url=str(context.row.original_url or ""),
        final_url=str(context.row.final_url or ""),
        citation_url=str(manifest.get("citation_url") or ""),
        citation_doi=str(manifest.get("citation_doi") or ""),
        imported_at=str(context.row.imported_at or ""),
        quality_score=context.quality_score,
    )


def _build_duplicate_candidate_group(
    contexts: list[_DuplicateCandidateContext],
    *,
    group_id: str,
    match_reason: str,
    confidence: str,
) -> RepositoryDuplicateCandidateGroup:
    ordered = sorted(
        contexts,
        key=lambda context: _duplicate_keep_sort_key(context),
        reverse=True,
    )
    suggested_keep_id = ordered[0].row.id if ordered else ""
    return RepositoryDuplicateCandidateGroup(
        group_id=group_id,
        match_reason=match_reason,
        confidence=confidence,
        suggested_keep_id=suggested_keep_id,
        suggested_delete_ids=[context.row.id for context in ordered[1:]],
        rows=[_duplicate_candidate_row(context) for context in ordered],
    )


def _repository_source_file_path(
    source_id: str,
    field: str,
    source_name: str,
    source_row_id: str = "",
) -> Path:
    original_name = Path(source_name or "").name or f"{source_id}_{field}"
    if source_row_id and original_name.startswith(source_row_id):
        normalized_name = source_id + original_name[len(source_row_id) :]
    elif original_name.startswith(source_id):
        normalized_name = original_name
    else:
        normalized_name = f"{source_id}_{original_name}"
    return Path(SOURCES_DIR_NAME) / source_id / normalized_name


def _load_source_rows(payload: list[Any]) -> list[SourceManifestRow]:
    rows: list[SourceManifestRow] = []
    for item in payload:
        if isinstance(item, SourceManifestRow):
            rows.append(item)
            continue
        if isinstance(item, dict):
            row = _safe_manifest_row(item)
            if row:
                rows.append(row)
    return rows


def _load_column_configs(payload: list[Any]) -> list[RepositoryColumnConfig]:
    configs: list[RepositoryColumnConfig] = []
    seen: set[str] = set()
    for item in payload:
        try:
            config = (
                item
                if isinstance(item, RepositoryColumnConfig)
                else RepositoryColumnConfig.model_validate(item)
            )
        except Exception:
            continue
        normalized_id = str(config.id or "").strip()
        if not normalized_id or normalized_id in seen:
            continue
        seen.add(normalized_id)
        configs.append(config.model_copy(update={"id": normalized_id}))
    return configs


def _load_citation_rows(payload: list[Any]) -> list[ExportRow]:
    rows: list[ExportRow] = []
    for item in payload:
        if isinstance(item, ExportRow):
            rows.append(item)
            continue
        if isinstance(item, dict):
            row = _safe_export_row(item)
            if row:
                rows.append(row)
    return rows


def _safe_manifest_row(payload: dict[str, Any]) -> SourceManifestRow | None:
    try:
        if not str(payload.get("source_kind") or "").strip():
            payload["source_kind"] = "url"
        if payload.get("http_status") in {"", None}:
            payload["http_status"] = None
        else:
            payload["http_status"] = _parse_int(str(payload.get("http_status")))
        llm_cleanup_needed = payload.get("llm_cleanup_needed")
        if llm_cleanup_needed in {"", None}:
            payload["llm_cleanup_needed"] = False
        elif isinstance(llm_cleanup_needed, str):
            payload["llm_cleanup_needed"] = llm_cleanup_needed.strip().lower() in {
                "1",
                "true",
                "yes",
                "y",
            }
        payload["markdown_char_count"] = _parse_int(str(payload.get("markdown_char_count") or "0")) or 0
        custom_fields = payload.get("custom_fields")
        if not isinstance(custom_fields, dict):
            payload["custom_fields"] = {}
        else:
            payload["custom_fields"] = {
                str(key): "" if value in {None, ""} else str(value)
                for key, value in custom_fields.items()
                if str(key or "").strip()
            }
        return _normalize_row_phase_metadata(SourceManifestRow.model_validate(payload))
    except Exception:
        return None


def _safe_export_row(payload: dict[str, Any]) -> ExportRow | None:
    try:
        confidence_val = payload.get("match_confidence")
        if confidence_val in {"", None}:
            payload["match_confidence"] = 0.0
        else:
            payload["match_confidence"] = float(confidence_val)
        return ExportRow.model_validate(payload)
    except Exception:
        return None


def _parse_numeric_id(value: str | None) -> int | None:
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        return int(text)
    return None


def _parse_int(value: str) -> int | None:
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _safe_float(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def _extract_ref_numbers(value: str | None) -> list[int]:
    text = str(value or "").strip()
    if not text:
        return []

    refs: list[int] = []
    for match in re.findall(r"\d+", text):
        try:
            refs.append(int(match))
        except Exception:
            continue
    deduped: list[int] = []
    seen: set[int] = set()
    for item in refs:
        if item in seen:
            continue
        seen.add(item)
        deduped.append(item)
    return deduped


def _next_source_id_from_rows(rows: list[SourceManifestRow]) -> int:
    numeric_ids = [_parse_numeric_id(row.id) for row in rows]
    valid = [item for item in numeric_ids if item is not None]
    if not valid:
        return 1
    return max(valid) + 1


def _latest_stage_timestamp(stages: Any) -> str:
    if not isinstance(stages, list):
        return ""
    candidates: list[str] = []
    for stage in stages:
        if not isinstance(stage, dict):
            continue
        completed = str(stage.get("completed_at") or "").strip()
        started = str(stage.get("started_at") or "").strip()
        if completed:
            candidates.append(completed)
        elif started:
            candidates.append(started)
    if not candidates:
        return ""
    return max(candidates)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _relative_or_absolute(root: Path, path: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except Exception:
        return str(path)


def _normalize_agent_phase_names(
    values: list[str] | tuple[str, ...] | set[str] | Any,
    *,
    run_download: bool = False,
    run_convert: bool = False,
    run_cleanup: bool = False,
    run_title: bool = False,
    run_catalog: bool = False,
    run_citation_verify: bool = False,
    run_rating: bool = False,
    run_summary: bool = False,
) -> list[str]:
    allowed = {
        PHASE_FETCH,
        PHASE_CONVERT,
        PHASE_CLEANUP,
        PHASE_TITLE,
        PHASE_CATALOG,
        PHASE_CITATION_VERIFY,
        PHASE_SUMMARY,
        PHASE_RATING,
    }
    normalized: list[str] = []
    seen: set[str] = set()
    if isinstance(values, (list, tuple, set)):
        for item in values:
            phase = _normalize_phase_name(item)
            if phase not in allowed or phase in seen:
                continue
            seen.add(phase)
            normalized.append(phase)
    if normalized:
        return normalized

    defaults: list[str] = []
    if run_download:
        defaults.append(PHASE_FETCH)
    if run_convert:
        defaults.append(PHASE_CONVERT)
    if run_cleanup:
        defaults.append(PHASE_CLEANUP)
    if run_title:
        defaults.append(PHASE_TITLE)
    if run_catalog:
        defaults.append(PHASE_CATALOG)
    if run_citation_verify:
        defaults.append(PHASE_CITATION_VERIFY)
    if run_summary:
        defaults.append(PHASE_SUMMARY)
    if run_rating:
        defaults.append(PHASE_RATING)
    return defaults


def _encode_agent_offset_cursor(offset: int) -> str:
    payload = json.dumps({"offset": max(0, int(offset))}, separators=(",", ":"))
    return base64.urlsafe_b64encode(payload.encode("utf-8")).decode("ascii")


def _decode_agent_offset_cursor(cursor: str) -> int:
    normalized = str(cursor or "").strip()
    if not normalized:
        return 0
    try:
        padding = "=" * (-len(normalized) % 4)
        raw = base64.urlsafe_b64decode((normalized + padding).encode("ascii")).decode(
            "utf-8"
        )
        parsed = json.loads(raw)
    except Exception as exc:
        raise ValueError("Invalid cursor.") from exc
    if not isinstance(parsed, dict):
        raise ValueError("Invalid cursor.")
    offset = parsed.get("offset")
    if not isinstance(offset, int) or offset < 0:
        raise ValueError("Invalid cursor.")
    return offset


def _coerce_optional_float(value: Any) -> float | None:
    if value in {"", None}:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _file_sha256(path: Path | None) -> str:
    if path is None or not path.is_file():
        return ""
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _phase_is_stale(metadata: SourcePhaseMetadata | None, current_digest: str) -> bool:
    if metadata is None:
        return False
    if metadata.stale:
        return True
    reference_digest = str(metadata.content_digest or "").strip()
    return bool(reference_digest and current_digest and reference_digest != current_digest)


def _agent_source_artifact_uri(source_id: str, kind: str) -> str:
    return f"repo://sources/{source_id}/{kind}"


def _row_import_id(row: SourceManifestRow) -> str:
    prefix = str(row.provenance_ref or "").split(":", 1)[0].strip()
    if not prefix:
        return ""
    if prefix.lower() in {"merge", "document", "repository", "scan", "manual"}:
        return ""
    if "/" in prefix or prefix.startswith("http"):
        return ""
    return prefix


def _row_updated_at(
    row: SourceManifestRow,
    base_dir: Path,
    *,
    imports: list[dict[str, Any]] | None = None,
) -> str:
    candidates: list[str] = [str(row.fetched_at or "").strip(), str(row.imported_at or "").strip()]
    for field_name in (
        "metadata_file",
        "rating_file",
        "summary_file",
        "llm_cleanup_file",
        "markdown_file",
        "rendered_file",
        "rendered_pdf_file",
        "raw_file",
    ):
        rel_value = str(getattr(row, field_name) or "").strip()
        if not rel_value:
            continue
        path = Path(rel_value)
        full_path = path if path.is_absolute() else base_dir / path
        if full_path.is_file():
            candidates.append(_path_mtime_iso(full_path))

    normalized_candidates = [item for item in candidates if item]
    if imports:
        import_id = _row_import_id(row)
        if import_id:
            matching = next(
                (
                    item
                    for item in imports
                    if str(item.get("import_id") or "").strip() == import_id
                ),
                None,
            )
            if matching is not None:
                imported_at = str(matching.get("imported_at") or "").strip()
                if imported_at:
                    normalized_candidates.append(imported_at)
    return max(normalized_candidates) if normalized_candidates else ""


def _agent_fetch_status(row: SourceManifestRow) -> str:
    metadata = row.phase_metadata.get("fetch")
    if metadata is not None and metadata.status:
        return metadata.status
    status = str(row.fetch_status or "").strip().lower()
    if status in {"success", "completed"}:
        return "completed"
    if status in {"partial", "failed", "queued"}:
        return status
    return "pending"


def _agent_phase_status(row: SourceManifestRow, phase: str) -> str:
    normalized_phase = _normalize_phase_name(phase)
    metadata = row.phase_metadata.get(normalized_phase)
    if metadata is not None and metadata.status:
        return metadata.status

    if normalized_phase == PHASE_CONVERT:
        cleanup_status = str(row.llm_cleanup_status or "").strip().lower()
        if cleanup_status == "failed":
            return "failed"
        if row.llm_cleanup_file or row.markdown_file:
            return "completed"
        return "pending"

    if normalized_phase == PHASE_CLEANUP:
        status = str(row.llm_cleanup_status or "").strip().lower()
        if status in {"cleaned", "existing", "not_needed", "completed"}:
            return "completed"
        if status in {"failed", "missing_markdown"}:
            return "failed"
        if status == "stale":
            return "stale"
        if status == "not_requested" or status.startswith("skipped"):
            return "skipped"
        return "pending"

    if normalized_phase == PHASE_TITLE:
        status = str(row.title_status or "").strip().lower()
        if status in {"generated", "existing", "completed", "extracted"}:
            return "completed"
        if status in {"failed", "missing_markdown"}:
            return "failed"
        if status == "stale":
            return "stale"
        if status == "not_requested" or status.startswith("skipped"):
            return "skipped"
        return "pending"

    if normalized_phase == PHASE_CATALOG:
        status = str(row.catalog_status or "").strip().lower()
        if status in {"generated", "existing", "completed"}:
            return "completed"
        if status in {"failed", "missing_markdown"}:
            return "failed"
        if status == "stale":
            return "stale"
        if status in {"skipped", "not_applicable"} or status.startswith("skipped"):
            return "skipped"
        return "pending"

    if normalized_phase == PHASE_CITATION_VERIFY:
        status = str(_row_citation_verification_status(row) or "").strip().lower()
        if status in {"verified", "candidate", "completed"}:
            return "completed"
        if status in {"failed", "missing_markdown"}:
            return "failed"
        if status == "stale":
            return "stale"
        if status in {"not_requested", "not_applicable"} or status.startswith("skipped"):
            return "skipped"
        return "pending"

    if normalized_phase == PHASE_SUMMARY:
        status = str(row.summary_status or "").strip().lower()
        if status in {"generated", "existing", "completed"}:
            return "completed"
        if status in {"failed", "missing_markdown"}:
            return "failed"
        if status == "stale":
            return "stale"
        if status == "skipped" or status.startswith("skipped"):
            return "skipped"
        return "pending"

    if normalized_phase == PHASE_RATING:
        status = str(row.rating_status or "").strip().lower()
        if status in {"generated", "existing", "completed"}:
            return "completed"
        if status in {"failed", "missing_markdown"}:
            return "failed"
        if status == "stale":
            return "stale"
        if status == "skipped" or status.startswith("skipped"):
            return "skipped"
        return "pending"

    return _agent_fetch_status(row)


def _sort_agent_source_records(
    records: list[tuple[AgentSourceRecord, dict[str, Any]]],
    *,
    sort_by: str,
    sort_dir: str,
) -> list[tuple[AgentSourceRecord, dict[str, Any]]]:
    def sort_value(item: tuple[AgentSourceRecord, dict[str, Any]]) -> Any:
        record = item[0]
        if sort_by == "rating_overall":
            return record.rating_overall
        if sort_by == "updated_at":
            return record.updated_at
        if sort_by == "title":
            return record.title.lower()
        return record.source_id.lower()

    present: list[tuple[AgentSourceRecord, dict[str, Any]]] = []
    missing: list[tuple[AgentSourceRecord, dict[str, Any]]] = []
    for item in records:
        value = sort_value(item)
        if value in {"", None}:
            missing.append(item)
        else:
            present.append(item)

    present = sorted(present, key=lambda item: item[0].source_id.lower())
    missing = sorted(missing, key=lambda item: item[0].source_id.lower())
    present = sorted(
        present,
        key=sort_value,
        reverse=(sort_dir == "desc"),
    )
    return present + missing


def _normalize_phase_outcome(status: str) -> str:
    normalized = str(status or "").strip().lower()
    if normalized in {"completed", "success", "generated", "existing"}:
        return "success"
    if normalized in {"failed", "missing_markdown"}:
        return "failed"
    if normalized in {"partial", "stale"}:
        return "partial"
    if normalized in {"skipped", "not_requested"} or normalized.startswith("skipped"):
        return "skipped"
    return "pending"


def _build_agent_run_counts(
    *,
    rows: list[SourceManifestRow],
    selected_phases: list[str],
    fallback_status: SourceDownloadStatus | None,
) -> AgentRunCounts:
    total = len(rows) if rows else int(getattr(fallback_status, "total_urls", 0) or 0)
    if not rows:
        return AgentRunCounts(
            total=total,
            processed=int(getattr(fallback_status, "processed_urls", 0) or 0),
            success=int(getattr(fallback_status, "success_count", 0) or 0),
            failed=int(getattr(fallback_status, "failed_count", 0) or 0),
            partial=int(getattr(fallback_status, "partial_count", 0) or 0),
            skipped=int(getattr(fallback_status, "skipped_count", 0) or 0),
        )

    success = 0
    failed = 0
    partial = 0
    skipped = 0
    processed = 0
    phases = selected_phases or ["fetch"]

    for row in rows:
        outcomes = [
            _normalize_phase_outcome(
                _agent_fetch_status(row) if phase == "fetch" else _agent_phase_status(row, phase)
            )
            for phase in phases
        ]
        if any(item == "failed" for item in outcomes):
            failed += 1
            processed += 1
            continue
        if any(item == "partial" for item in outcomes):
            partial += 1
            processed += 1
            continue
        if all(item == "skipped" for item in outcomes):
            skipped += 1
            processed += 1
            continue
        if all(item == "success" for item in outcomes):
            success += 1
            processed += 1
            continue

    if fallback_status is not None:
        processed = max(processed, int(fallback_status.processed_urls or 0))
        total = max(total, int(fallback_status.total_urls or 0))
        skipped = max(skipped, int(fallback_status.skipped_count or 0))
    return AgentRunCounts(
        total=total,
        processed=processed,
        success=success,
        failed=failed,
        partial=partial,
        skipped=skipped,
    )


def _build_pending_source_status(
    *,
    job_id: str,
    store: FileStore,
    orchestrator: SourceDownloadOrchestrator,
) -> SourceDownloadStatus:
    bibliography = store.load_artifact(job_id, "03_bibliography") or {}
    if getattr(orchestrator, "target_rows", None):
        pending_total = len(getattr(orchestrator, "target_rows", []))
    else:
        pending_total = len(
            [entry for entry in bibliography.get("entries", []) if isinstance(entry, dict)]
        )
    phase_states = {
        key: value.model_copy(deep=True)
        for key, value in getattr(orchestrator, "_phase_states", {}).items()
    }
    return SourceDownloadStatus(
        job_id=job_id,
        state="cancelling" if getattr(orchestrator, "cancel_requested", False) else "running",
        total_urls=pending_total,
        processed_urls=0,
        cancel_requested=bool(getattr(orchestrator, "cancel_requested", False)),
        cancel_requested_at=None,
        stop_after_current_item=False,
        message=(
            "Stop requested | stopping before the next item"
            if getattr(orchestrator, "cancel_requested", False)
            else "Preparing source task run..."
        ),
        run_download=bool(getattr(orchestrator, "run_download", False)),
        run_convert=bool(getattr(orchestrator, "run_convert", False)),
        run_catalog=bool(getattr(orchestrator, "run_catalog", False)),
        run_citation_verify=bool(getattr(orchestrator, "run_citation_verify", False)),
        run_llm_cleanup=bool(getattr(orchestrator, "run_llm_cleanup", False)),
        run_llm_title=bool(getattr(orchestrator, "run_llm_title", False)),
        run_llm_summary=bool(getattr(orchestrator, "run_llm_summary", False)),
        run_llm_rating=bool(getattr(orchestrator, "run_llm_rating", False)),
        force_redownload=bool(getattr(orchestrator, "force_redownload", False)),
        force_convert=bool(getattr(orchestrator, "force_convert", False)),
        force_catalog=bool(getattr(orchestrator, "force_catalog", False)),
        force_citation_verify=bool(getattr(orchestrator, "force_citation_verify", False)),
        force_llm_cleanup=bool(getattr(orchestrator, "force_llm_cleanup", False)),
        force_title=bool(getattr(orchestrator, "force_title", False)),
        force_summary=bool(getattr(orchestrator, "force_summary", False)),
        force_rating=bool(getattr(orchestrator, "force_rating", False)),
        output_dir=str(getattr(orchestrator, "status_output_dir", "output_run")),
        manifest_csv=str(getattr(orchestrator, "status_manifest_csv", "output_run/manifest.csv")),
        manifest_xlsx=str(
            getattr(orchestrator, "status_manifest_xlsx", "output_run/manifest.xlsx")
        ),
        bundle_file=str(getattr(orchestrator, "status_bundle_file", "output_run.zip")),
        writes_to_repository=bool(getattr(orchestrator, "writes_to_repository", False)),
        repository_path=str(getattr(orchestrator, "repository_path", "")),
        selected_scope=str(getattr(orchestrator, "selected_scope", "")),
        selected_import_id=str(getattr(orchestrator, "selected_import_id", "")),
        selected_phases=list(getattr(orchestrator, "selected_phases", [])),
        phase_states=phase_states,
        items=[],
    )


def _parse_simple_front_matter(content: str) -> dict[str, Any]:
    lines = content.splitlines()
    if not lines or lines[0].strip() != "---":
        return {}

    data: dict[str, Any] = {}
    current_key = ""
    for line in lines[1:]:
        stripped = line.strip()
        if stripped == "---":
            break
        match = re.match(r"^([A-Za-z0-9_-]+)\s*:\s*(.*)$", line)
        if match:
            current_key = match.group(1).strip().lower()
            raw_value = match.group(2).strip()
            if not raw_value:
                data[current_key] = []
                continue
            if raw_value.startswith("[") and raw_value.endswith("]"):
                pieces = [
                    item.strip().strip("'\"")
                    for item in raw_value[1:-1].split(",")
                    if item.strip()
                ]
                data[current_key] = pieces
            else:
                data[current_key] = raw_value.strip("'\"")
            continue
        if current_key and re.match(r"^\s*-\s+.+$", line):
            data.setdefault(current_key, [])
            if isinstance(data[current_key], list):
                data[current_key].append(line.split("-", 1)[1].strip().strip("'\""))
    return data


def _extract_markdown_title(content: str) -> str:
    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            return stripped.lstrip("#").strip()
    for line in content.splitlines():
        stripped = line.strip()
        if stripped:
            return stripped[:200]
    return ""


def _extract_markdown_tags(content: str) -> list[str]:
    for line in content.splitlines():
        stripped = line.strip()
        if not stripped.lower().startswith("tags:"):
            continue
        _, raw_tags = stripped.split(":", 1)
        return [
            item.strip().strip("'\"")
            for item in raw_tags.split(",")
            if item.strip()
        ]
    return []


def _extract_markdown_description(content: str) -> str:
    paragraph: list[str] = []
    for line in content.splitlines():
        stripped = line.strip()
        if not stripped:
            if paragraph:
                break
            continue
        if stripped.startswith("#") or stripped.lower().startswith("tags:"):
            continue
        paragraph.append(stripped)
        if len(" ".join(paragraph)) >= 220:
            break
    return " ".join(paragraph)[:240].strip()


def _extract_simple_yaml_value(content: str, key: str) -> str:
    pattern = rf"(?m)^{re.escape(key)}\s*:\s*(.+?)\s*$"
    match = re.search(pattern, content)
    if not match:
        return ""
    return match.group(1).strip().strip("'\"")


def _derive_agent_resource_metadata(
    *,
    source_path: Path,
    content: str,
    kind: str,
) -> tuple[str, list[str], str]:
    front_matter = _parse_simple_front_matter(content)
    title = ""
    tags: list[str] = []
    description = ""

    if kind == "rubric":
        title = (
            str(front_matter.get("title") or "").strip()
            or _extract_simple_yaml_value(content, "name")
            or _extract_simple_yaml_value(content, "title")
            or source_path.stem
        )
        raw_tags = front_matter.get("tags") or []
        if isinstance(raw_tags, list):
            tags = [str(item).strip() for item in raw_tags if str(item).strip()]
        description = (
            str(front_matter.get("description") or "").strip()
            or _extract_simple_yaml_value(content, "description")
            or _extract_simple_yaml_value(content, "summary")
        )
    else:
        title = (
            str(front_matter.get("title") or "").strip()
            or _extract_markdown_title(content)
            or source_path.stem
        )
        raw_tags = front_matter.get("tags") or _extract_markdown_tags(content)
        if isinstance(raw_tags, list):
            tags = [str(item).strip() for item in raw_tags if str(item).strip()]
        elif isinstance(raw_tags, str) and raw_tags.strip():
            tags = [
                item.strip()
                for item in raw_tags.split(",")
                if item.strip()
            ]
        description = (
            str(front_matter.get("description") or "").strip()
            or str(front_matter.get("summary") or "").strip()
            or _extract_markdown_description(content)
        )

    deduped_tags: list[str] = []
    seen: set[str] = set()
    for item in tags:
        normalized = item.lower()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        deduped_tags.append(item)
    return title[:200].strip(), deduped_tags, description[:280].strip()


def _path_mtime_iso(path: Path) -> str:
    try:
        return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()
    except Exception:
        return ""
