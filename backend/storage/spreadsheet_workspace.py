"""Spreadsheet workspace service for generic tabular data editing."""

from __future__ import annotations

import csv
import io
import json
import shutil
import sqlite3
import tempfile
import threading
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Literal

from openpyxl import Workbook, load_workbook

from backend.llm.client import UnifiedLLMClient
from backend.llm.prompts import COLUMN_PROMPT_FIX_SYSTEM, COLUMN_PROMPT_FIX_USER
from backend.models.repository import RepositoryColumnOutputConstraint
from backend.models.spreadsheets import (
    SpreadsheetColumnConfig,
    SpreadsheetColumnCreateRequest,
    SpreadsheetColumnPromptFixRequest,
    SpreadsheetColumnPromptFixResponse,
    SpreadsheetColumnRunRequest,
    SpreadsheetColumnRunRowError,
    SpreadsheetColumnRunStartResponse,
    SpreadsheetColumnRunStatus,
    SpreadsheetColumnUpdateRequest,
    SpreadsheetDataType,
    SpreadsheetExportRequest,
    SpreadsheetFilterRequest,
    SpreadsheetManifestResponse,
    SpreadsheetRowPatchRequest,
    SpreadsheetSessionResponse,
    SpreadsheetSessionSummary,
    SpreadsheetSessionTargetSelectRequest,
    SpreadsheetSessionUploadResponse,
    SpreadsheetSourceFormat,
    SpreadsheetTargetDescriptor,
    SpreadsheetWorkspaceStatusResponse,
)
from backend.models.settings import EffectiveSettings
from backend.storage.attached_repository import AttachedRepositoryService
from backend.storage.file_store import FileStore


SPREADSHEET_WORKSPACE_DIR_NAME = "spreadsheets"
SPREADSHEET_SESSIONS_DIR_NAME = "sessions"
SPREADSHEET_INDEX_FILE_NAME = "workspace_index.json"
SPREADSHEET_SESSION_METADATA_FILE_NAME = "session.json"
SPREADSHEET_SESSION_DB_FILE_NAME = "data.sqlite"
SPREADSHEET_SOURCE_FILE_NAME = "source"
SPREADSHEET_JOB_PREFIX = "sheet"

SUPPORTED_SPREADSHEET_EXTENSIONS: dict[str, SpreadsheetSourceFormat] = {
    ".csv": "csv",
    ".xlsx": "xlsx",
    ".json": "json",
    ".jsonl": "jsonl",
    ".ndjson": "ndjson",
    ".parquet": "parquet",
    ".sqlite": "sqlite",
    ".db": "sqlite",
}

SPREADSHEET_RUN_SYSTEM = """You generate exactly one spreadsheet cell value.

Hard rules:
- Return JSON only.
- Output only the value for the target cell, never an explanation.
- No markdown.
- Keep the value single-cell-safe.
- Follow the requested output constraint exactly.
- Use only the provided selected input column values.
- If evidence is insufficient, return status `insufficient_evidence` and the fallback cell value.

Return JSON only in exactly this shape:
{
  "value": "",
  "status": "ok" | "insufficient_evidence"
}"""

SPREADSHEET_RUN_USER = """Research purpose:
{research_purpose}

Column label:
{column_label}

Column instructions:
{column_prompt}

Output constraint:
{output_constraint_json}

Application hard rules:
{hard_rules}

Current cell value:
{current_value}

Selected input columns:
{selected_columns_json}

Return JSON only."""


@dataclass
class _ImportedRow:
    ordinal: int
    current_values: dict[str, Any]
    original_payload: dict[str, Any]


@dataclass
class _ImportedTarget:
    descriptor: SpreadsheetTargetDescriptor
    columns: list[SpreadsheetColumnConfig]
    rows: list[_ImportedRow]


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def _normalize_multiline_text(value: Any) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    return "\n".join(line.rstrip() for line in text.split("\n")).strip()


def _slugify_label(value: str, fallback: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in {" ", "_", "-"} else " " for ch in str(value or ""))
    normalized = " ".join(cleaned.split()).strip()
    return normalized or fallback


def _row_has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _quote_sqlite_identifier(value: str) -> str:
    return '"' + str(value or "").replace('"', '""') + '"'


def _coerce_output_constraint(payload: Any) -> RepositoryColumnOutputConstraint | None:
    if payload is None or payload == "":
        return None
    if isinstance(payload, RepositoryColumnOutputConstraint):
        return payload
    try:
        return RepositoryColumnOutputConstraint.model_validate(payload)
    except Exception:
        return None


def _serialize_output_constraint(
    value: RepositoryColumnOutputConstraint | None,
) -> dict[str, Any] | None:
    return value.model_dump(mode="json") if value is not None else None


def _infer_output_constraint(
    prompt: str,
    existing: RepositoryColumnOutputConstraint | None = None,
) -> RepositoryColumnOutputConstraint | None:
    normalized = str(prompt or "").strip().lower()
    if not normalized:
        return existing
    kind: Literal["text", "yes_no", "integer", "number", "date"] = "text"
    allowed_values: list[str] = []
    max_words: int | None = existing.max_words if existing else None
    if "yes/no" in normalized or "yes or no" in normalized or normalized.startswith("is "):
        kind = "yes_no"
        allowed_values = ["yes", "no"]
        max_words = 1
    elif "integer" in normalized or "whole number" in normalized:
        kind = "integer"
    elif "number" in normalized or "numeric" in normalized or "score" in normalized:
        kind = "number"
    elif "date" in normalized:
        kind = "date"
    elif "one word" in normalized:
        max_words = 1
    elif "short" in normalized and max_words is None:
        max_words = 8
    fallback_value = existing.fallback_value if existing else ""
    format_hint = existing.format_hint if existing else ""
    return RepositoryColumnOutputConstraint(
        kind=kind,
        allowed_values=allowed_values or (existing.allowed_values if existing else []),
        max_words=max_words,
        fallback_value=fallback_value,
        format_hint=format_hint,
    )


def _parse_json_object(value: str) -> dict[str, Any]:
    try:
        payload = json.loads(value or "{}")
    except json.JSONDecodeError as exc:
        raise ValueError("LLM returned invalid JSON.") from exc
    if not isinstance(payload, dict):
        raise ValueError("LLM returned an invalid payload.")
    return payload


def _coerce_run_value(
    value: Any,
    constraint: RepositoryColumnOutputConstraint,
) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        return str(constraint.fallback_value or "")
    if constraint.kind == "yes_no":
        lowered = text.lower()
        if lowered in {"yes", "true", "1"}:
            return "yes"
        if lowered in {"no", "false", "0"}:
            return "no"
        return str(constraint.fallback_value or "")
    if constraint.kind == "integer":
        try:
            return str(int(float(text)))
        except Exception:
            return str(constraint.fallback_value or "")
    if constraint.kind == "number":
        try:
            number = float(text)
        except Exception:
            return str(constraint.fallback_value or "")
        return str(int(number)) if float(number).is_integer() else str(number)
    if constraint.kind == "date":
        return text[:10]
    words = text.split()
    if constraint.max_words and len(words) > constraint.max_words:
        text = " ".join(words[: constraint.max_words])
    return text


def _infer_data_type(values: Iterable[Any]) -> SpreadsheetDataType:
    kinds: set[str] = set()
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool):
            kinds.add("boolean")
            continue
        if isinstance(value, int) and not isinstance(value, bool):
            kinds.add("integer")
            continue
        if isinstance(value, float):
            kinds.add("number")
            continue
        kinds.add("string")
    if not kinds:
        return "string"
    if len(kinds) > 1:
        if kinds == {"integer", "number"}:
            return "number"
        return "mixed"
    only = next(iter(kinds))
    return "number" if only == "number" else only  # type: ignore[return-value]


def _sort_key_for_value(data_type: SpreadsheetDataType, value: Any) -> tuple[int, Any]:
    if value in {"", None}:
        return (1, "")
    if data_type == "boolean":
        return (0, int(bool(value)))
    if data_type == "integer":
        try:
            return (0, int(value))
        except Exception:
            return (1, str(value).lower())
    if data_type == "number":
        try:
            return (0, float(value))
        except Exception:
            return (1, str(value).lower())
    return (0, str(value).lower())


def _json_path_label(tokens: list[Any]) -> str:
    if not tokens:
        return "$"
    chunks = ["$"]
    for token in tokens:
        if isinstance(token, int):
            chunks.append(f"[{token}]")
        else:
            chunks.append(f".{token}")
    return "".join(chunks)


def _is_record_array(value: Any) -> bool:
    if not isinstance(value, list) or not value:
        return False
    has_record = False
    for item in value:
        if item is None:
            continue
        if not isinstance(item, dict):
            return False
        has_record = True
    return has_record


def _discover_json_arrays(value: Any, path: list[Any] | None = None) -> list[list[Any]]:
    current_path = list(path or [])
    matches: list[list[Any]] = []
    if _is_record_array(value):
        matches.append(current_path)
    if isinstance(value, dict):
        for key, child in value.items():
            matches.extend(_discover_json_arrays(child, [*current_path, str(key)]))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            if isinstance(child, (dict, list)):
                matches.extend(_discover_json_arrays(child, [*current_path, index]))
    return matches


def _resolve_json_path(root: Any, tokens: list[Any]) -> Any:
    current = root
    for token in tokens:
        if isinstance(token, int):
            if not isinstance(current, list) or token >= len(current):
                raise ValueError("Spreadsheet JSON path no longer exists.")
            current = current[token]
            continue
        if not isinstance(current, dict) or token not in current:
            raise ValueError("Spreadsheet JSON path no longer exists.")
        current = current[token]
    return current


class SpreadsheetWorkspaceService:
    def __init__(self, store: FileStore, repository_service: AttachedRepositoryService):
        self.store = store
        self.repository_service = repository_service
        self._mutex = threading.RLock()
        self._run_threads: dict[str, threading.Thread] = {}

    @property
    def is_available(self) -> bool:
        return self.repository_service.is_attached

    def get_status(self) -> SpreadsheetWorkspaceStatusResponse:
        with self._mutex:
            if not self.is_available:
                return SpreadsheetWorkspaceStatusResponse(available=False)
            self._ensure_workspace_dirs()
            index = self._load_index()
            current_session_id = str(index.get("current_session_id") or "")
            sessions = [
                SpreadsheetSessionSummary.model_validate(item)
                for item in index.get("sessions", [])
                if isinstance(item, dict)
            ]
            current_session = None
            if current_session_id:
                try:
                    current_session = self.get_session(current_session_id)
                except ValueError:
                    current_session = None
            return SpreadsheetWorkspaceStatusResponse(
                available=True,
                current_session_id=current_session_id,
                sessions=sessions,
                current_session=current_session,
            )

    def upload_session(self, filename: str, content: bytes) -> SpreadsheetSessionUploadResponse:
        if not self.is_available:
            raise ValueError("Attach a repository before opening spreadsheet files.")
        normalized_name = Path(filename or "").name
        extension = Path(normalized_name).suffix.lower()
        source_format = SUPPORTED_SPREADSHEET_EXTENSIONS.get(extension)
        if not source_format:
            raise ValueError(
                "Unsupported spreadsheet file. Use .csv, .xlsx, .json, .jsonl, .ndjson, .parquet, .sqlite, or .db."
            )

        with self._mutex:
            self._ensure_workspace_dirs()
            session_id = uuid.uuid4().hex[:12]
            session_dir = self._session_dir(session_id)
            session_dir.mkdir(parents=True, exist_ok=True)
            source_path = session_dir / f"{SPREADSHEET_SOURCE_FILE_NAME}{extension}"
            source_path.write_bytes(content)

            targets = self._load_targets_from_source(
                session_id=session_id,
                source_path=source_path,
                source_format=source_format,
            )
            if not targets:
                raise ValueError("No usable row-oriented data was found in the uploaded file.")

            self._initialize_cache(session_id)
            for target in targets:
                self._store_target(session_id, target)

            created_at = _utc_now_iso()
            metadata = {
                "session_id": session_id,
                "filename": normalized_name,
                "original_filename": normalized_name,
                "source_format": source_format,
                "created_at": created_at,
                "updated_at": created_at,
                "active_target_id": targets[0].descriptor.id,
                "targets": [target.descriptor.model_dump(mode="json") for target in targets],
            }
            self._save_session_metadata(session_id, metadata)
            self._save_session_summary(
                SpreadsheetSessionSummary(
                    session_id=session_id,
                    filename=normalized_name,
                    original_filename=normalized_name,
                    source_format=source_format,
                    created_at=created_at,
                    updated_at=created_at,
                    active_target_id=targets[0].descriptor.id,
                    target_count=len(targets),
                )
            )
            self._set_current_session(session_id)
            return SpreadsheetSessionUploadResponse(
                session=self.get_session(session_id),
                message=f"Opened {normalized_name} in the Spreadsheets workspace.",
            )

    def get_session(self, session_id: str) -> SpreadsheetSessionResponse:
        if not self.is_available:
            raise ValueError("Attach a repository before using the spreadsheet workspace.")
        with self._mutex:
            metadata = self._load_session_metadata(session_id)
            targets = [
                SpreadsheetTargetDescriptor.model_validate(item)
                for item in metadata.get("targets", [])
                if isinstance(item, dict)
            ]
            active_target_id = str(metadata.get("active_target_id") or "")
            active_target = next((item for item in targets if item.id == active_target_id), None)
            columns = self._load_columns(session_id, active_target_id) if active_target_id else []
            summary = SpreadsheetSessionSummary(
                session_id=str(metadata.get("session_id") or session_id),
                filename=str(metadata.get("filename") or ""),
                original_filename=str(metadata.get("original_filename") or metadata.get("filename") or ""),
                source_format=str(metadata.get("source_format") or "csv"),  # type: ignore[arg-type]
                created_at=str(metadata.get("created_at") or ""),
                updated_at=str(metadata.get("updated_at") or ""),
                active_target_id=active_target_id,
                target_count=len(targets),
            )
            return SpreadsheetSessionResponse(
                session=summary,
                targets=targets,
                active_target=active_target,
                columns=columns,
            )

    def select_target(
        self,
        session_id: str,
        payload: SpreadsheetSessionTargetSelectRequest,
    ) -> SpreadsheetSessionResponse:
        with self._mutex:
            metadata = self._load_session_metadata(session_id)
            target_id = str(payload.target_id or "").strip()
            if target_id not in {
                str(item.get("id") or "")
                for item in metadata.get("targets", [])
                if isinstance(item, dict)
            }:
                raise ValueError(f"Unknown spreadsheet target: {target_id}")
            metadata["active_target_id"] = target_id
            metadata["updated_at"] = _utc_now_iso()
            self._save_session_metadata(session_id, metadata)
            self._update_session_summary(
                session_id,
                active_target_id=target_id,
                updated_at=str(metadata["updated_at"]),
            )
            self._set_current_session(session_id)
            return self.get_session(session_id)

    def list_manifest(
        self,
        session_id: str,
        *,
        q: str = "",
        sort_by: str = "",
        sort_dir: str = "",
        limit: int = 50,
        offset: int = 0,
    ) -> SpreadsheetManifestResponse:
        with self._mutex:
            metadata = self._load_session_metadata(session_id)
            target_id = str(metadata.get("active_target_id") or "")
            if not target_id:
                raise ValueError("No active spreadsheet target is selected.")
            columns = self._load_columns(session_id, target_id)
            rows = self._load_rows(session_id, target_id)
            filtered = self._filter_rows(rows, columns, SpreadsheetFilterRequest(q=q))
            normalized_sort_dir = str(sort_dir or "").strip().lower()
            if normalized_sort_dir and normalized_sort_dir not in {"asc", "desc"}:
                raise ValueError("Invalid sort_dir. Use `asc` or `desc`.")
            if sort_by:
                column = next((item for item in columns if item.id == sort_by), None)
                if column is None:
                    raise ValueError(f"Unknown spreadsheet sort column: {sort_by}")
                filtered = sorted(
                    filtered,
                    key=lambda item: _sort_key_for_value(column.data_type, item.get(sort_by)),
                    reverse=normalized_sort_dir == "desc",
                )
            safe_limit = max(1, min(int(limit), 500))
            safe_offset = max(0, int(offset))
            return SpreadsheetManifestResponse(
                rows=filtered[safe_offset : safe_offset + safe_limit],
                total=len(filtered),
                limit=safe_limit,
                offset=safe_offset,
                sort_by=str(sort_by or ""),
                sort_dir=normalized_sort_dir if sort_by else "",
                columns=columns,
                filters=SpreadsheetFilterRequest(q=q),
            )

    def patch_row(
        self,
        session_id: str,
        row_id: str,
        payload: SpreadsheetRowPatchRequest,
    ) -> dict[str, Any]:
        if not payload.values:
            raise ValueError("At least one spreadsheet value is required.")
        with self._mutex:
            target_id = self._active_target_id(session_id)
            columns = {item.id: item for item in self._load_columns(session_id, target_id)}
            row = self._load_row_payload(session_id, target_id, row_id)
            current_values = dict(row["data"])
            for column_id, raw_value in payload.values.items():
                column = columns.get(str(column_id or ""))
                if column is None:
                    raise ValueError(f"Unknown spreadsheet column: {column_id}")
                current_values[column.id] = self._coerce_value_for_column(column, raw_value)
            self._save_row_payload(session_id, target_id, row_id, current_values, row["original"])
            self._touch_session(session_id)
            return self._row_dict(row_id, current_values)

    def create_column(
        self,
        session_id: str,
        payload: SpreadsheetColumnCreateRequest,
    ) -> SpreadsheetColumnConfig:
        with self._mutex:
            target_id = self._active_target_id(session_id)
            columns = self._load_columns(session_id, target_id)
            existing_ids = {column.id for column in columns}
            column_id = ""
            while not column_id or column_id in existing_ids:
                column_id = f"custom_{uuid.uuid4().hex[:8]}"
            label = _slugify_label(payload.label, "New Column")
            column = SpreadsheetColumnConfig(
                id=column_id,
                source_key=label,
                label=label,
                kind="custom",
                data_type="string",
                ordinal=len(columns),
                input_column_ids=[item.id for item in columns if item.kind == "source"],
            )
            self._upsert_columns(session_id, target_id, [*columns, column])
            self._touch_session(session_id)
            return column

    def update_column(
        self,
        session_id: str,
        column_id: str,
        payload: SpreadsheetColumnUpdateRequest,
    ) -> SpreadsheetColumnConfig:
        with self._mutex:
            target_id = self._active_target_id(session_id)
            columns = self._load_columns(session_id, target_id)
            updated: list[SpreadsheetColumnConfig] = []
            target_column: SpreadsheetColumnConfig | None = None
            for column in columns:
                if column.id != column_id:
                    updated.append(column)
                    continue
                next_column = column.model_copy(deep=True)
                if payload.label is not None:
                    if next_column.kind != "custom":
                        raise ValueError("Only custom spreadsheet columns can be renamed.")
                    next_label = _slugify_label(payload.label, "New Column")
                    next_column.label = next_label
                    next_column.source_key = next_label
                if payload.instruction_prompt is not None:
                    next_column.instruction_prompt = _normalize_multiline_text(payload.instruction_prompt)
                if payload.output_constraint is not None:
                    next_column.output_constraint = _coerce_output_constraint(payload.output_constraint)
                elif payload.instruction_prompt is not None:
                    next_column.output_constraint = _infer_output_constraint(
                        next_column.instruction_prompt,
                        existing=next_column.output_constraint,
                    )
                if payload.input_column_ids is not None:
                    allowed_ids = {item.id for item in columns if item.id != next_column.id}
                    next_column.input_column_ids = [
                        item
                        for item in payload.input_column_ids
                        if str(item or "").strip() in allowed_ids
                    ]
                target_column = next_column
                updated.append(next_column)
            if target_column is None:
                raise ValueError(f"Unknown spreadsheet column: {column_id}")
            self._upsert_columns(session_id, target_id, updated)
            self._touch_session(session_id)
            return target_column

    def fix_column_prompt(
        self,
        session_id: str,
        column_id: str,
        payload: SpreadsheetColumnPromptFixRequest,
    ) -> SpreadsheetColumnPromptFixResponse:
        normalized_prompt = _normalize_multiline_text(payload.draft_prompt)
        if not normalized_prompt:
            raise ValueError("draft_prompt is required")
        settings = self._effective_settings()
        if not settings.use_llm or not settings.llm_backend.model.strip():
            raise ValueError("Prompt fix requires an enabled chat-capable LLM backend.")
        column = self._require_column(session_id, self._active_target_id(session_id), column_id)
        client = UnifiedLLMClient(settings.llm_backend)
        try:
            raw_response = client.sync_chat_completion(
                system_prompt=COLUMN_PROMPT_FIX_SYSTEM,
                user_prompt=COLUMN_PROMPT_FIX_USER.format(
                    column_label=column.label,
                    current_prompt=column.instruction_prompt,
                    current_constraint_json=json.dumps(
                        _serialize_output_constraint(column.output_constraint),
                        ensure_ascii=False,
                    ),
                    draft_prompt=normalized_prompt,
                ),
                response_format="json",
            )
        finally:
            client.sync_close()
        parsed = _parse_json_object(raw_response)
        rewritten_prompt = _normalize_multiline_text(parsed.get("prompt") or normalized_prompt)
        output_constraint = _coerce_output_constraint(parsed.get("output_constraint")) or _infer_output_constraint(
            rewritten_prompt,
            existing=column.output_constraint,
        )
        notes = [
            str(item).strip()
            for item in parsed.get("notes", [])
            if str(item or "").strip()
        ] if isinstance(parsed.get("notes"), list) else []
        return SpreadsheetColumnPromptFixResponse(
            status="completed",
            column_id=column.id,
            prompt=rewritten_prompt,
            output_constraint=output_constraint,
            notes=notes,
        )

    def start_column_run(
        self,
        session_id: str,
        column_id: str,
        payload: SpreadsheetColumnRunRequest,
    ) -> SpreadsheetColumnRunStartResponse:
        with self._mutex:
            settings = self._effective_settings()
            if not settings.use_llm or not settings.llm_backend.model.strip():
                raise ValueError("Spreadsheet column runs require an enabled chat-capable LLM backend.")
            target_id = self._active_target_id(session_id)
            columns = self._load_columns(session_id, target_id)
            column = self._require_column(session_id, target_id, column_id)
            if not column.instruction_prompt.strip():
                raise ValueError("Save instructions for this spreadsheet column before running it.")
            if column.output_constraint is None:
                column.output_constraint = _infer_output_constraint(column.instruction_prompt)
            rows = self._load_rows(session_id, target_id)
            filtered_rows = self._rows_for_scope(rows, columns, column, payload)
            if not filtered_rows:
                if payload.scope == "selected":
                    raise ValueError("No selected spreadsheet rows are available for this column run.")
                if payload.scope == "empty_only":
                    raise ValueError(f"No blank {column.label} cells are available for this column run.")
                raise ValueError("No spreadsheet rows match the current filters.")
            populated_rows = sum(1 for row in filtered_rows if _row_has_value(row.get(column.id)))
            if populated_rows > 0 and not payload.confirm_overwrite:
                return SpreadsheetColumnRunStartResponse(
                    job_id="",
                    status="confirmation_required",
                    column_id=column.id,
                    total_rows=len(filtered_rows),
                    populated_rows=populated_rows,
                    message=(
                        f"{populated_rows} matching row(s) already have values in {column.label}. "
                        "Confirm overwrite to continue."
                    ),
                )

            job_store = self.repository_service.repo_job_store()
            job_id = job_store.create_job(prefix=SPREADSHEET_JOB_PREFIX)
            status = SpreadsheetColumnRunStatus(
                job_id=job_id,
                session_id=session_id,
                target_id=target_id,
                column_id=column.id,
                column_label=column.label,
                state="pending",
                total_rows=len(filtered_rows),
                message=f"Queued {len(filtered_rows)} row(s) for {column.label}.",
            )
            job_store.save_artifact(
                job_id,
                "spreadsheet_column_run_context",
                {
                    "session_id": session_id,
                    "target_id": target_id,
                    "column_id": column.id,
                    "row_ids": [str(row.get("id") or "") for row in filtered_rows],
                    "filters": payload.filters.model_dump(mode="json"),
                },
            )
            job_store.save_source_status(job_id, status.model_dump(mode="json"))
            column.last_run_status = "running"
            self._upsert_columns(
                session_id,
                target_id,
                [column if item.id == column.id else item for item in columns],
            )

            thread = threading.Thread(
                target=self._run_column_job,
                kwargs={
                    "session_id": session_id,
                    "target_id": target_id,
                    "column_id": column.id,
                    "job_id": job_id,
                    "row_ids": [str(row.get("id") or "") for row in filtered_rows],
                    "settings": settings,
                },
                daemon=True,
            )
            self._run_threads[job_id] = thread
            thread.start()
            return SpreadsheetColumnRunStartResponse(
                job_id=job_id,
                status="started",
                column_id=column.id,
                total_rows=len(filtered_rows),
                populated_rows=populated_rows,
                message=f"Started spreadsheet run for {column.label}.",
            )

    def get_column_run_status(
        self,
        session_id: str,
        job_id: str,
    ) -> SpreadsheetColumnRunStatus:
        if not self.is_available:
            raise ValueError("Attach a repository before using spreadsheet runs.")
        status_payload = self.repository_service.repo_job_store().get_source_status(job_id)
        if not status_payload:
            raise ValueError("Spreadsheet column run status not found.")
        status = SpreadsheetColumnRunStatus.model_validate(status_payload)
        if status.session_id != session_id:
            raise ValueError("Spreadsheet column run status not found.")
        return status

    def export_session(
        self,
        payload: SpreadsheetExportRequest,
    ) -> tuple[bytes, dict[str, str], str]:
        session_id = str(payload.session_id or "").strip()
        if not session_id:
            raise ValueError("session_id is required")
        with self._mutex:
            metadata = self._load_session_metadata(session_id)
            target_id = str(metadata.get("active_target_id") or "")
            if not target_id:
                raise ValueError("No active spreadsheet target is selected.")
            source_format = str(metadata.get("source_format") or "").strip().lower()
            source_path = self._source_path(session_id)
            if not source_path.exists():
                raise ValueError("Spreadsheet source file not found.")
            columns = self._load_columns(session_id, target_id)
            row_payloads = self._load_row_payloads(session_id, target_id)
            filename = Path(str(metadata.get("filename") or "spreadsheet")).name

        if source_format == "csv":
            content = self._export_csv(columns, row_payloads).encode("utf-8-sig")
            media_type = "text/csv; charset=utf-8"
        elif source_format == "xlsx":
            content = self._export_xlsx(source_path, metadata, target_id, columns, row_payloads)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif source_format == "json":
            content = self._export_json(source_path, metadata, target_id, columns, row_payloads)
            media_type = "application/json"
        elif source_format in {"jsonl", "ndjson"}:
            content = self._export_jsonl(columns, row_payloads).encode("utf-8")
            media_type = "application/x-ndjson"
        elif source_format == "sqlite":
            content = self._export_sqlite(source_path, metadata, target_id, columns, row_payloads)
            media_type = "application/vnd.sqlite3"
        elif source_format == "parquet":
            content = self._export_parquet(columns, row_payloads)
            media_type = "application/octet-stream"
        else:
            raise ValueError(f"Unsupported spreadsheet format: {source_format}")

        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-ResearchAssistant-Requested-Count": str(len(row_payloads)),
            "X-ResearchAssistant-Exported-Count": str(len(row_payloads)),
            "X-ResearchAssistant-Skipped-Count": "0",
        }
        return content, headers, media_type

    def _run_column_job(
        self,
        *,
        session_id: str,
        target_id: str,
        column_id: str,
        job_id: str,
        row_ids: list[str],
        settings: EffectiveSettings,
    ) -> None:
        job_store = self.repository_service.repo_job_store()
        row_errors: list[SpreadsheetColumnRunRowError] = []
        started_at = _utc_now_iso()
        try:
            with self._mutex:
                columns = self._load_columns(session_id, target_id)
                column = self._require_column(session_id, target_id, column_id)
                rows_by_id = {row["id"]: row for row in self._load_rows(session_id, target_id)}
            for index, row_id in enumerate(row_ids, start=1):
                current_status = SpreadsheetColumnRunStatus(
                    job_id=job_id,
                    session_id=session_id,
                    target_id=target_id,
                    column_id=column.id,
                    column_label=column.label,
                    state="running",
                    total_rows=len(row_ids),
                    processed_rows=index - 1,
                    succeeded_rows=(index - 1) - len(row_errors),
                    failed_rows=len(row_errors),
                    current_row_id=row_id,
                    message=f"Processing row {index}/{len(row_ids)}",
                    started_at=started_at,
                    row_errors=row_errors,
                )
                job_store.save_source_status(job_id, current_status.model_dump(mode="json"))
                row = rows_by_id.get(row_id)
                if row is None:
                    row_errors.append(SpreadsheetColumnRunRowError(row_id=row_id, message="Row not found"))
                    continue
                try:
                    value = self._generate_column_value_for_row(
                        settings=settings,
                        column=column,
                        columns=columns,
                        row=row,
                    )
                    with self._mutex:
                        payload = self._load_row_payload(session_id, target_id, row_id)
                        current_values = dict(payload["data"])
                        current_values[column.id] = self._coerce_value_for_column(column, value)
                        self._save_row_payload(session_id, target_id, row_id, current_values, payload["original"])
                except Exception as exc:
                    row_errors.append(
                        SpreadsheetColumnRunRowError(row_id=row_id, message=str(exc))
                    )
                    continue

            completed_at = _utc_now_iso()
            with self._mutex:
                columns = self._load_columns(session_id, target_id)
                for item in columns:
                    if item.id == column_id:
                        item.last_run_at = completed_at
                        item.last_run_status = "failed" if row_errors else "completed"
                self._upsert_columns(session_id, target_id, columns)
                self._touch_session(session_id)
            final_status = SpreadsheetColumnRunStatus(
                job_id=job_id,
                session_id=session_id,
                target_id=target_id,
                column_id=column_id,
                column_label=column.label,
                state="failed" if row_errors else "completed",
                total_rows=len(row_ids),
                processed_rows=len(row_ids),
                succeeded_rows=len(row_ids) - len(row_errors),
                failed_rows=len(row_errors),
                message=(
                    f"Completed spreadsheet run for {column.label}."
                    if not row_errors
                    else f"Completed spreadsheet run for {column.label} with {len(row_errors)} row error(s)."
                ),
                started_at=started_at,
                completed_at=completed_at,
                row_errors=row_errors,
            )
            job_store.save_source_status(job_id, final_status.model_dump(mode="json"))
        finally:
            self._run_threads.pop(job_id, None)

    def _generate_column_value_for_row(
        self,
        *,
        settings: EffectiveSettings,
        column: SpreadsheetColumnConfig,
        columns: list[SpreadsheetColumnConfig],
        row: dict[str, Any],
    ) -> str:
        constraint = column.output_constraint or RepositoryColumnOutputConstraint(
            kind="text",
            fallback_value="",
        )
        selected_ids = column.input_column_ids or [
            item.id for item in columns if item.kind == "source" and item.id != column.id
        ]
        selected_values = {
            item.label: row.get(item.id)
            for item in columns
            if item.id in selected_ids
        }
        hard_rules = "\n".join(
            [
                "- Output only the target cell value.",
                "- No markdown.",
                "- No explanations.",
                "- Keep the value single-cell-safe.",
                f"- Use `{constraint.fallback_value}` when evidence is insufficient."
                if constraint.fallback_value
                else "- Use a blank value when evidence is insufficient.",
            ]
        )
        client = UnifiedLLMClient(settings.llm_backend)
        try:
            raw_response = client.sync_chat_completion(
                system_prompt=SPREADSHEET_RUN_SYSTEM,
                user_prompt=SPREADSHEET_RUN_USER.format(
                    research_purpose=settings.research_purpose or "",
                    column_label=column.label,
                    column_prompt=column.instruction_prompt,
                    output_constraint_json=json.dumps(
                        _serialize_output_constraint(constraint),
                        ensure_ascii=False,
                    ),
                    hard_rules=hard_rules,
                    current_value=str(row.get(column.id) or ""),
                    selected_columns_json=json.dumps(selected_values, ensure_ascii=False, indent=2),
                ),
                response_format="json",
            )
        finally:
            client.sync_close()
        parsed = _parse_json_object(raw_response)
        normalized_status = str(parsed.get("status") or "").strip().lower()
        normalized_value = _coerce_run_value(parsed.get("value"), constraint)
        if normalized_status == "insufficient_evidence" and not normalized_value:
            return str(constraint.fallback_value or "")
        return normalized_value

    def _effective_settings(self) -> EffectiveSettings:
        return self.repository_service.load_effective_settings()

    def _filter_rows(
        self,
        rows: list[dict[str, Any]],
        columns: list[SpreadsheetColumnConfig],
        filters: SpreadsheetFilterRequest,
    ) -> list[dict[str, Any]]:
        query = str(filters.q or "").strip().lower()
        if not query:
            return rows
        filtered: list[dict[str, Any]] = []
        for row in rows:
            haystack = " ".join(str(row.get(column.id) or "") for column in columns).lower()
            if query in haystack:
                filtered.append(row)
        return filtered

    def _rows_for_scope(
        self,
        rows: list[dict[str, Any]],
        columns: list[SpreadsheetColumnConfig],
        column: SpreadsheetColumnConfig,
        payload: SpreadsheetColumnRunRequest,
    ) -> list[dict[str, Any]]:
        normalized_scope = str(payload.scope or "filtered").strip().lower()
        if normalized_scope == "all":
            base = rows
        elif normalized_scope == "selected":
            wanted = {str(item or "").strip() for item in payload.row_ids if str(item or "").strip()}
            base = [row for row in rows if str(row.get("id") or "") in wanted]
        elif normalized_scope in {"filtered", "empty_only"}:
            base = self._filter_rows(rows, columns, payload.filters)
        else:
            raise ValueError(f"Unsupported spreadsheet column run scope: {normalized_scope}")
        if normalized_scope == "empty_only":
            return [row for row in base if not _row_has_value(row.get(column.id))]
        return base

    def _coerce_value_for_column(self, column: SpreadsheetColumnConfig, raw_value: Any) -> Any:
        if raw_value is None:
            return None if column.data_type in {"integer", "number", "boolean", "null"} else ""
        if column.kind == "custom":
            return str(raw_value).strip()
        if column.data_type == "string":
            return str(raw_value)
        if column.data_type == "boolean":
            text = str(raw_value).strip().lower()
            if text in {"", "null"}:
                return None
            if text in {"1", "true", "yes", "y"}:
                return True
            if text in {"0", "false", "no", "n"}:
                return False
            raise ValueError(f"Invalid boolean value for {column.label}: {raw_value}")
        if column.data_type == "integer":
            text = str(raw_value).strip()
            if not text:
                return None
            try:
                return int(text)
            except Exception as exc:
                raise ValueError(f"Invalid integer value for {column.label}: {raw_value}") from exc
        if column.data_type == "number":
            text = str(raw_value).strip()
            if not text:
                return None
            try:
                return float(text)
            except Exception as exc:
                raise ValueError(f"Invalid numeric value for {column.label}: {raw_value}") from exc
        if column.data_type == "null":
            return None
        return str(raw_value)

    def _export_fieldnames(self, columns: list[SpreadsheetColumnConfig]) -> list[str]:
        return [column.source_key for column in sorted(columns, key=lambda item: item.ordinal)]

    def _export_row_object(
        self,
        columns: list[SpreadsheetColumnConfig],
        payload: dict[str, Any],
    ) -> dict[str, Any]:
        original_payload = payload.get("original") or {}
        original_row = dict(original_payload.get("row") or {})
        current_values = dict(payload.get("data") or {})
        exported = dict(original_row)
        for column in sorted(columns, key=lambda item: item.ordinal):
            exported[column.source_key] = current_values.get(column.id)
        return exported

    def _export_csv(
        self,
        columns: list[SpreadsheetColumnConfig],
        row_payloads: list[dict[str, Any]],
    ) -> str:
        output = io.StringIO()
        fieldnames = self._export_fieldnames(columns)
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for payload in row_payloads:
            row = self._export_row_object(columns, payload)
            writer.writerow({field: _csv_value(row.get(field)) for field in fieldnames})
        return output.getvalue()

    def _export_xlsx(
        self,
        source_path: Path,
        metadata: dict[str, Any],
        target_id: str,
        columns: list[SpreadsheetColumnConfig],
        row_payloads: list[dict[str, Any]],
    ) -> bytes:
        workbook = load_workbook(source_path)
        try:
            target = next(
                SpreadsheetTargetDescriptor.model_validate(item)
                for item in metadata.get("targets", [])
                if isinstance(item, dict) and str(item.get("id") or "") == target_id
            )
            sheet_name = str(target.selector.get("sheet_name") or target.label)
            worksheet = workbook[sheet_name]
            if worksheet.max_row > 0:
                worksheet.delete_rows(1, worksheet.max_row)
            fieldnames = self._export_fieldnames(columns)
            worksheet.append(fieldnames)
            for payload in row_payloads:
                exported = self._export_row_object(columns, payload)
                worksheet.append([exported.get(field) for field in fieldnames])
            stream = io.BytesIO()
            workbook.save(stream)
            return stream.getvalue()
        finally:
            workbook.close()

    def _export_json(
        self,
        source_path: Path,
        metadata: dict[str, Any],
        target_id: str,
        columns: list[SpreadsheetColumnConfig],
        row_payloads: list[dict[str, Any]],
    ) -> bytes:
        root = json.loads(source_path.read_text(encoding="utf-8"))
        target = next(
            SpreadsheetTargetDescriptor.model_validate(item)
            for item in metadata.get("targets", [])
            if isinstance(item, dict) and str(item.get("id") or "") == target_id
        )
        tokens = list(target.selector.get("path_tokens") or [])
        array_ref = _resolve_json_path(root, tokens)
        if not isinstance(array_ref, list):
            raise ValueError("Spreadsheet JSON path no longer points to a list.")
        updated_rows = [self._export_row_object(columns, payload) for payload in row_payloads]
        array_ref[:] = updated_rows
        return json.dumps(root, ensure_ascii=False, indent=2).encode("utf-8")

    def _export_jsonl(
        self,
        columns: list[SpreadsheetColumnConfig],
        row_payloads: list[dict[str, Any]],
    ) -> str:
        return "\n".join(
            json.dumps(self._export_row_object(columns, payload), ensure_ascii=False)
            for payload in row_payloads
        ) + "\n"

    def _export_sqlite(
        self,
        source_path: Path,
        metadata: dict[str, Any],
        target_id: str,
        columns: list[SpreadsheetColumnConfig],
        row_payloads: list[dict[str, Any]],
    ) -> bytes:
        target = next(
            SpreadsheetTargetDescriptor.model_validate(item)
            for item in metadata.get("targets", [])
            if isinstance(item, dict) and str(item.get("id") or "") == target_id
        )
        table_name = str(target.selector.get("table_name") or target.label)
        quoted_table_name = _quote_sqlite_identifier(table_name)
        with tempfile.TemporaryDirectory(prefix="sheet-export-sqlite-") as tmp:
            temp_path = Path(tmp) / Path(source_path.name or "spreadsheet.sqlite").name
            shutil.copy2(source_path, temp_path)
            with sqlite3.connect(temp_path) as export_conn:
                table_columns = {
                    str(row[1])
                    for row in export_conn.execute(f"PRAGMA table_info({quoted_table_name})")
                }
                for column in columns:
                    if column.source_key not in table_columns:
                        export_conn.execute(
                            f"ALTER TABLE {quoted_table_name} ADD COLUMN {_quote_sqlite_identifier(column.source_key)} TEXT"
                        )
                assignments = ", ".join(
                    f"{_quote_sqlite_identifier(column.source_key)} = ?"
                    for column in columns
                )
                for payload in row_payloads:
                    original = dict(payload.get("original") or {})
                    rowid = original.get("rowid")
                    if rowid is None:
                        raise ValueError("SQLite export requires rowid-backed tables.")
                    exported = self._export_row_object(columns, payload)
                    export_conn.execute(
                        f"UPDATE {quoted_table_name} SET {assignments} WHERE rowid = ?",
                        [exported.get(column.source_key) for column in columns] + [rowid],
                    )
                export_conn.commit()
            return temp_path.read_bytes()

    def _export_parquet(
        self,
        columns: list[SpreadsheetColumnConfig],
        row_payloads: list[dict[str, Any]],
    ) -> bytes:
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except Exception as exc:
            raise ValueError("Parquet support requires pyarrow to be installed.") from exc
        rows = [self._export_row_object(columns, payload) for payload in row_payloads]
        table = pa.Table.from_pylist(rows)
        sink = io.BytesIO()
        pq.write_table(table, sink)
        return sink.getvalue()

    def _load_targets_from_source(
        self,
        *,
        session_id: str,
        source_path: Path,
        source_format: SpreadsheetSourceFormat,
    ) -> list[_ImportedTarget]:
        if source_format == "csv":
            rows = self._read_csv_rows(source_path)
            return [self._build_imported_target("target_001", "Rows", "file", {}, rows)]
        if source_format == "xlsx":
            return self._read_xlsx_targets(source_path)
        if source_format == "json":
            return self._read_json_targets(source_path)
        if source_format in {"jsonl", "ndjson"}:
            rows = self._read_jsonl_rows(source_path)
            return [self._build_imported_target("target_001", "Rows", "file", {}, rows)]
        if source_format == "sqlite":
            return self._read_sqlite_targets(source_path)
        if source_format == "parquet":
            rows = self._read_parquet_rows(source_path)
            return [self._build_imported_target("target_001", "Rows", "file", {}, rows)]
        raise ValueError(f"Unsupported spreadsheet source format: {source_format}")

    def _read_csv_rows(self, path: Path) -> list[dict[str, Any]]:
        text = path.read_text(encoding="utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows: list[dict[str, Any]] = []
        for raw in reader:
            if not raw:
                continue
            row = {str(key): value for key, value in raw.items() if key is not None}
            if any(str(value or "").strip() for value in row.values()):
                rows.append(row)
        return rows

    def _read_xlsx_targets(self, path: Path) -> list[_ImportedTarget]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        targets: list[_ImportedTarget] = []
        try:
            for index, worksheet in enumerate(workbook.worksheets, start=1):
                rows_iter = worksheet.iter_rows(values_only=True)
                headers: list[str] = []
                for header_row in rows_iter:
                    header_values = [str(value).strip() if value is not None else "" for value in header_row]
                    if any(header_values):
                        headers = [value or f"Column {idx + 1}" for idx, value in enumerate(header_values)]
                        break
                if not headers:
                    continue
                rows: list[dict[str, Any]] = []
                for values in rows_iter:
                    row = {
                        headers[idx]: values[idx] if idx < len(headers) else None
                        for idx in range(len(headers))
                    }
                    if any(item not in {"", None} for item in row.values()):
                        rows.append(row)
                targets.append(
                    self._build_imported_target(
                        f"target_{index:03d}",
                        worksheet.title,
                        "sheet",
                        {"sheet_name": worksheet.title},
                        rows,
                    )
                )
            return targets
        finally:
            workbook.close()

    def _read_json_targets(self, path: Path) -> list[_ImportedTarget]:
        root = json.loads(path.read_text(encoding="utf-8"))
        arrays = _discover_json_arrays(root)
        if not arrays:
            raise ValueError("JSON spreadsheets require a top-level or nested array of objects.")
        targets: list[_ImportedTarget] = []
        for index, tokens in enumerate(arrays, start=1):
            rows = _resolve_json_path(root, tokens)
            if not isinstance(rows, list):
                continue
            normalized_rows = [dict(item) for item in rows if isinstance(item, dict)]
            if not normalized_rows:
                continue
            targets.append(
                self._build_imported_target(
                    f"target_{index:03d}",
                    _json_path_label(tokens),
                    "json_path",
                    {"path_tokens": tokens},
                    normalized_rows,
                )
            )
        return targets

    def _read_jsonl_rows(self, path: Path) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for line in path.read_text(encoding="utf-8").splitlines():
            text = line.strip()
            if not text:
                continue
            payload = json.loads(text)
            if isinstance(payload, dict):
                rows.append(payload)
        return rows

    def _read_parquet_rows(self, path: Path) -> list[dict[str, Any]]:
        try:
            import pyarrow.parquet as pq
        except Exception as exc:
            raise ValueError("Parquet support requires pyarrow to be installed.") from exc
        table = pq.read_table(path)
        return table.to_pylist()

    def _read_sqlite_targets(self, path: Path) -> list[_ImportedTarget]:
        targets: list[_ImportedTarget] = []
        with sqlite3.connect(path) as conn:
            rows = conn.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
            ).fetchall()
            for index, (table_name,) in enumerate(rows, start=1):
                try:
                    result = conn.execute(f"SELECT rowid, * FROM {_quote_sqlite_identifier(str(table_name))}")
                except sqlite3.DatabaseError:
                    continue
                headers = [item[0] for item in result.description]
                records: list[dict[str, Any]] = []
                for values in result.fetchall():
                    row = {headers[idx]: values[idx] for idx in range(len(headers))}
                    records.append(row)
                if not records:
                    schema_rows = conn.execute(
                        f"PRAGMA table_info({_quote_sqlite_identifier(str(table_name))})"
                    ).fetchall()
                    records = [
                        {
                            str(column[1]): None
                            for column in schema_rows
                        }
                    ] if schema_rows else []
                imported = self._build_imported_target(
                    f"target_{index:03d}",
                    str(table_name),
                    "table",
                    {"table_name": str(table_name)},
                    records,
                    sqlite_rowid_field="rowid",
                )
                targets.append(imported)
        return targets

    def _build_imported_target(
        self,
        target_id: str,
        label: str,
        kind: Literal["sheet", "table", "json_path", "file"],
        selector: dict[str, Any],
        rows: list[dict[str, Any]],
        *,
        sqlite_rowid_field: str = "",
    ) -> _ImportedTarget:
        column_keys: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row.keys():
                if key == sqlite_rowid_field:
                    continue
                normalized = str(key)
                if normalized in seen:
                    continue
                seen.add(normalized)
                column_keys.append(normalized)

        columns: list[SpreadsheetColumnConfig] = []
        for index, key in enumerate(column_keys, start=1):
            column_values = [row.get(key) for row in rows]
            columns.append(
                SpreadsheetColumnConfig(
                    id=f"source_{index:03d}",
                    source_key=key,
                    label=key,
                    kind="source",
                    data_type=_infer_data_type(column_values),
                    ordinal=index - 1,
                )
            )
        source_ids = [column.id for column in columns]
        for column in columns:
            column.input_column_ids = [item for item in source_ids if item != column.id]

        imported_rows: list[_ImportedRow] = []
        for index, row in enumerate(rows, start=1):
            current_values = {
                column.id: row.get(column.source_key)
                for column in columns
            }
            original_row = {
                key: value
                for key, value in row.items()
                if key != sqlite_rowid_field
            }
            original_payload = {"row": original_row}
            if sqlite_rowid_field:
                original_payload["rowid"] = row.get(sqlite_rowid_field)
            imported_rows.append(
                _ImportedRow(
                    ordinal=index - 1,
                    current_values=current_values,
                    original_payload=original_payload,
                )
            )

        descriptor = SpreadsheetTargetDescriptor(
            id=target_id,
            label=label,
            kind=kind,
            selector=selector,
            row_count=len(imported_rows),
            column_count=len(columns),
        )
        return _ImportedTarget(descriptor=descriptor, columns=columns, rows=imported_rows)

    def _store_target(self, session_id: str, target: _ImportedTarget) -> None:
        with self._connect_cache(session_id) as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO targets(target_id, label, kind, selector_json, row_count, column_count)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    target.descriptor.id,
                    target.descriptor.label,
                    target.descriptor.kind,
                    json.dumps(target.descriptor.selector, ensure_ascii=False),
                    target.descriptor.row_count,
                    target.descriptor.column_count,
                ),
            )
            conn.executemany(
                """
                INSERT OR REPLACE INTO columns(
                    target_id,
                    column_id,
                    source_key,
                    label,
                    kind,
                    data_type,
                    ordinal,
                    instruction_prompt,
                    output_constraint_json,
                    input_column_ids_json,
                    last_run_at,
                    last_run_status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        target.descriptor.id,
                        column.id,
                        column.source_key,
                        column.label,
                        column.kind,
                        column.data_type,
                        column.ordinal,
                        column.instruction_prompt,
                        json.dumps(_serialize_output_constraint(column.output_constraint), ensure_ascii=False),
                        json.dumps(column.input_column_ids, ensure_ascii=False),
                        column.last_run_at,
                        column.last_run_status,
                    )
                    for column in target.columns
                ],
            )
            conn.executemany(
                """
                INSERT OR REPLACE INTO rows(target_id, row_id, ordinal, original_json, data_json)
                VALUES (?, ?, ?, ?, ?)
                """,
                [
                    (
                        target.descriptor.id,
                        f"row_{index + 1:06d}",
                        row.ordinal,
                        json.dumps(row.original_payload, ensure_ascii=False),
                        json.dumps(row.current_values, ensure_ascii=False),
                    )
                    for index, row in enumerate(target.rows)
                ],
            )
            conn.commit()

    def _initialize_cache(self, session_id: str) -> None:
        with self._connect_cache(session_id) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS targets(
                    target_id TEXT PRIMARY KEY,
                    label TEXT NOT NULL,
                    kind TEXT NOT NULL,
                    selector_json TEXT NOT NULL DEFAULT '{}',
                    row_count INTEGER NOT NULL DEFAULT 0,
                    column_count INTEGER NOT NULL DEFAULT 0
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS columns(
                    target_id TEXT NOT NULL,
                    column_id TEXT NOT NULL,
                    source_key TEXT NOT NULL,
                    label TEXT NOT NULL,
                    kind TEXT NOT NULL,
                    data_type TEXT NOT NULL,
                    ordinal INTEGER NOT NULL DEFAULT 0,
                    instruction_prompt TEXT NOT NULL DEFAULT '',
                    output_constraint_json TEXT NOT NULL DEFAULT 'null',
                    input_column_ids_json TEXT NOT NULL DEFAULT '[]',
                    last_run_at TEXT NOT NULL DEFAULT '',
                    last_run_status TEXT NOT NULL DEFAULT '',
                    PRIMARY KEY(target_id, column_id)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS rows(
                    target_id TEXT NOT NULL,
                    row_id TEXT NOT NULL,
                    ordinal INTEGER NOT NULL DEFAULT 0,
                    original_json TEXT NOT NULL DEFAULT '{}',
                    data_json TEXT NOT NULL DEFAULT '{}',
                    PRIMARY KEY(target_id, row_id)
                )
                """
            )
            conn.commit()

    def _row_dict(self, row_id: str, data: dict[str, Any]) -> dict[str, Any]:
        return {"id": row_id, **data}

    def _load_rows(self, session_id: str, target_id: str) -> list[dict[str, Any]]:
        with self._connect_cache(session_id) as conn:
            rows = conn.execute(
                "SELECT row_id, data_json FROM rows WHERE target_id = ? ORDER BY ordinal, row_id",
                (target_id,),
            ).fetchall()
        return [
            self._row_dict(str(row_id), json.loads(data_json or "{}"))
            for row_id, data_json in rows
        ]

    def _load_row_payload(self, session_id: str, target_id: str, row_id: str) -> dict[str, Any]:
        with self._connect_cache(session_id) as conn:
            row = conn.execute(
                "SELECT original_json, data_json FROM rows WHERE target_id = ? AND row_id = ?",
                (target_id, row_id),
            ).fetchone()
        if row is None:
            raise ValueError(f"Unknown spreadsheet row: {row_id}")
        original_json, data_json = row
        return {
            "original": json.loads(original_json or "{}"),
            "data": json.loads(data_json or "{}"),
        }

    def _load_row_payloads(self, session_id: str, target_id: str) -> list[dict[str, Any]]:
        with self._connect_cache(session_id) as conn:
            rows = conn.execute(
                "SELECT row_id, original_json, data_json FROM rows WHERE target_id = ? ORDER BY ordinal, row_id",
                (target_id,),
            ).fetchall()
        return [
            {
                "id": row_id,
                "original": json.loads(original_json or "{}"),
                "data": json.loads(data_json or "{}"),
            }
            for row_id, original_json, data_json in rows
        ]

    def _save_row_payload(
        self,
        session_id: str,
        target_id: str,
        row_id: str,
        current_values: dict[str, Any],
        original_payload: dict[str, Any],
    ) -> None:
        with self._connect_cache(session_id) as conn:
            conn.execute(
                """
                UPDATE rows
                SET data_json = ?, original_json = ?
                WHERE target_id = ? AND row_id = ?
                """,
                (
                    json.dumps(current_values, ensure_ascii=False),
                    json.dumps(original_payload, ensure_ascii=False),
                    target_id,
                    row_id,
                ),
            )
            conn.commit()

    def _load_columns(self, session_id: str, target_id: str) -> list[SpreadsheetColumnConfig]:
        with self._connect_cache(session_id) as conn:
            rows = conn.execute(
                """
                SELECT
                    column_id,
                    source_key,
                    label,
                    kind,
                    data_type,
                    ordinal,
                    instruction_prompt,
                    output_constraint_json,
                    input_column_ids_json,
                    last_run_at,
                    last_run_status
                FROM columns
                WHERE target_id = ?
                ORDER BY ordinal, column_id
                """,
                (target_id,),
            ).fetchall()
        return [
            SpreadsheetColumnConfig(
                id=str(column_id),
                source_key=str(source_key),
                label=str(label),
                kind=str(kind),  # type: ignore[arg-type]
                data_type=str(data_type),  # type: ignore[arg-type]
                ordinal=int(ordinal or 0),
                instruction_prompt=str(instruction_prompt or ""),
                output_constraint=_coerce_output_constraint(json.loads(output_constraint_json or "null")),
                input_column_ids=[
                    str(item)
                    for item in json.loads(input_column_ids_json or "[]")
                    if str(item or "").strip()
                ],
                last_run_at=str(last_run_at or ""),
                last_run_status=str(last_run_status or ""),
            )
            for (
                column_id,
                source_key,
                label,
                kind,
                data_type,
                ordinal,
                instruction_prompt,
                output_constraint_json,
                input_column_ids_json,
                last_run_at,
                last_run_status,
            ) in rows
        ]

    def _upsert_columns(
        self,
        session_id: str,
        target_id: str,
        columns: list[SpreadsheetColumnConfig],
    ) -> None:
        with self._connect_cache(session_id) as conn:
            conn.execute("DELETE FROM columns WHERE target_id = ?", (target_id,))
            conn.executemany(
                """
                INSERT INTO columns(
                    target_id,
                    column_id,
                    source_key,
                    label,
                    kind,
                    data_type,
                    ordinal,
                    instruction_prompt,
                    output_constraint_json,
                    input_column_ids_json,
                    last_run_at,
                    last_run_status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        target_id,
                        column.id,
                        column.source_key,
                        column.label,
                        column.kind,
                        column.data_type,
                        index,
                        column.instruction_prompt,
                        json.dumps(_serialize_output_constraint(column.output_constraint), ensure_ascii=False),
                        json.dumps(column.input_column_ids, ensure_ascii=False),
                        column.last_run_at,
                        column.last_run_status,
                    )
                    for index, column in enumerate(columns)
                ],
            )
            conn.commit()

    def _require_column(
        self,
        session_id: str,
        target_id: str,
        column_id: str,
    ) -> SpreadsheetColumnConfig:
        column = next(
            (item for item in self._load_columns(session_id, target_id) if item.id == column_id),
            None,
        )
        if column is None:
            raise ValueError(f"Unknown spreadsheet column: {column_id}")
        return column

    def _connect_cache(self, session_id: str) -> sqlite3.Connection:
        connection = sqlite3.connect(self._cache_path(session_id))
        connection.row_factory = sqlite3.Row
        return connection

    def _ensure_workspace_dirs(self) -> None:
        self._workspace_dir().mkdir(parents=True, exist_ok=True)
        self._sessions_dir().mkdir(parents=True, exist_ok=True)
        if not self._index_path().exists():
            self._index_path().write_text(
                json.dumps({"current_session_id": "", "sessions": []}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

    def _workspace_dir(self) -> Path:
        if not self.is_available:
            raise ValueError("Attach a repository before using the spreadsheet workspace.")
        return self.repository_service.path / ".ra_repo" / SPREADSHEET_WORKSPACE_DIR_NAME

    def _sessions_dir(self) -> Path:
        return self._workspace_dir() / SPREADSHEET_SESSIONS_DIR_NAME

    def _index_path(self) -> Path:
        return self._workspace_dir() / SPREADSHEET_INDEX_FILE_NAME

    def _session_dir(self, session_id: str) -> Path:
        return self._sessions_dir() / str(session_id or "").strip()

    def _session_metadata_path(self, session_id: str) -> Path:
        return self._session_dir(session_id) / SPREADSHEET_SESSION_METADATA_FILE_NAME

    def _cache_path(self, session_id: str) -> Path:
        return self._session_dir(session_id) / SPREADSHEET_SESSION_DB_FILE_NAME

    def _source_path(self, session_id: str) -> Path:
        session_dir = self._session_dir(session_id)
        for path in session_dir.glob(f"{SPREADSHEET_SOURCE_FILE_NAME}.*"):
            return path
        return session_dir / SPREADSHEET_SOURCE_FILE_NAME

    def _load_index(self) -> dict[str, Any]:
        try:
            return json.loads(self._index_path().read_text(encoding="utf-8"))
        except Exception:
            return {"current_session_id": "", "sessions": []}

    def _save_index(self, payload: dict[str, Any]) -> None:
        self._index_path().write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _load_session_metadata(self, session_id: str) -> dict[str, Any]:
        path = self._session_metadata_path(session_id)
        if not path.exists():
            raise ValueError("Spreadsheet session not found.")
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            raise ValueError("Spreadsheet session metadata is invalid.") from exc
        if not isinstance(payload, dict):
            raise ValueError("Spreadsheet session metadata is invalid.")
        return payload

    def _save_session_metadata(self, session_id: str, payload: dict[str, Any]) -> None:
        self._session_metadata_path(session_id).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _save_session_summary(self, summary: SpreadsheetSessionSummary) -> None:
        index = self._load_index()
        sessions = [
            item
            for item in index.get("sessions", [])
            if not isinstance(item, dict) or str(item.get("session_id") or "") != summary.session_id
        ]
        sessions.insert(0, summary.model_dump(mode="json"))
        index["sessions"] = sessions
        self._save_index(index)

    def _update_session_summary(self, session_id: str, **patch: Any) -> None:
        index = self._load_index()
        sessions: list[dict[str, Any]] = []
        found = False
        for item in index.get("sessions", []):
            if not isinstance(item, dict):
                continue
            if str(item.get("session_id") or "") != session_id:
                sessions.append(item)
                continue
            next_item = dict(item)
            next_item.update(patch)
            sessions.append(next_item)
            found = True
        if found:
            index["sessions"] = sessions
            self._save_index(index)

    def _set_current_session(self, session_id: str) -> None:
        index = self._load_index()
        index["current_session_id"] = session_id
        self._save_index(index)

    def _touch_session(self, session_id: str) -> None:
        metadata = self._load_session_metadata(session_id)
        updated_at = _utc_now_iso()
        metadata["updated_at"] = updated_at
        self._save_session_metadata(session_id, metadata)
        self._update_session_summary(session_id, updated_at=updated_at)
        self._set_current_session(session_id)

    def _active_target_id(self, session_id: str) -> str:
        metadata = self._load_session_metadata(session_id)
        target_id = str(metadata.get("active_target_id") or "")
        if not target_id:
            raise ValueError("No active spreadsheet target is selected.")
        return target_id


def _csv_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return value
