"""Read a planning spreadsheet into sources and columns.

Planning spreadsheets are written for people. A typical one has a merged banner
row above the real header, a `Prompts:` row between the header and the data
holding one LLM instruction per column, and data rows where only an id and a URL
are filled in.

`csv.DictReader` and openpyxl's header inference both assume row 0 is the
header, which is the precise thing that is untrue here. So this module reads a
raw rectangular grid of strings and locates the rows by scoring them.

Two rules govern the whole module:

* **Never drop anything silently.** A row with an unusable id still appears in
  the output with a blank id and an anomaly attached, so the repository
  allocates one and the user can see why.
* **Always show the working.** Candidate rows and their scores are returned, and
  every automatic choice can be overridden, so a wrong guess is correctable
  rather than mysterious.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from .encoding import repair_grid
from .models import (
    RowCandidate,
    RowLayout,
    SheetAnomaly,
    SheetColumn,
    SheetDocument,
    SheetPlan,
    SheetProvidedColumn,
    SheetSource,
)

MAX_HEADER_SEARCH_ROWS = 20
HEADER_FILL_RATIO = 0.6
PROMPT_MIN_MEAN_LENGTH = 80
PROMPT_MAX_DATA_RATIO = 0.3
ID_COLUMN_MIN_RATIO = 0.8
ID_COLUMN_NAMED_MIN_RATIO = 0.5
URL_COLUMN_MIN_RATIO = 0.6

_PROMPTS_LABEL = re.compile(r"^\s*prompts?\s*:?\s*$", re.IGNORECASE)
_ID_HEADER = re.compile(r"\b(id|no\.?|number|#)\b|^id#?$", re.IGNORECASE)
_URL_HEADER = re.compile(r"^\s*(url|link|source\s*url|address)\s*$", re.IGNORECASE)
_INTEGER = re.compile(r"^\d{1,7}$")


class SheetReadError(ValueError):
    """The file could not be read as a grid at all."""


# ---------------------------------------------------------------------------
# reading
# ---------------------------------------------------------------------------


def read_grid(path: Path) -> tuple[list[list[str]], str]:
    """Read a spreadsheet into a rectangular grid of trimmed strings."""
    path = Path(path)
    if not path.is_file():
        raise SheetReadError(f"Not a file: {path}")

    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return _read_delimited(path, "\t" if suffix == ".tsv" else ","), "csv"
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path), "xlsx"
    raise SheetReadError(
        f"Unsupported spreadsheet type {suffix or '(none)'}. Save the sheet as .csv or .xlsx."
    )


def _read_delimited(path: Path, delimiter: str) -> list[list[str]]:
    # utf-8-sig: the app writes its own CSVs with a BOM and Excel usually does
    # too; without it the first header cell arrives with a stray marker glued on.
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = [[(cell or "").strip() for cell in row] for row in csv.reader(row for row in handle)]
    return _rectangular(rows)


def _read_xlsx(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is in requirements
        raise SheetReadError(f"openpyxl is required to read .xlsx files: {exc}") from exc

    # data_only: formula cells yield their computed value rather than "=A1".
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows: list[list[str]] = []
        for row in sheet.iter_rows():
            rows.append(["" if cell.value is None else str(cell.value).strip() for cell in row])
    finally:
        workbook.close()
    return _rectangular(rows)


def _rectangular(rows: list[list[str]]) -> list[list[str]]:
    width = max((len(row) for row in rows), default=0)
    return [row + [""] * (width - len(row)) for row in rows]


# ---------------------------------------------------------------------------
# locating
# ---------------------------------------------------------------------------


def _is_url(value: str) -> bool:
    text = (value or "").strip().lower()
    return text.startswith("http://") or text.startswith("https://")


def _looks_like_url_cell(value: str) -> bool:
    text = (value or "").strip()
    if _is_url(text):
        return True
    # A bare domain is still a URL cell; `create_sources` normalizes it.
    return bool(re.match(r"^[a-z0-9][a-z0-9.-]*\.[a-z]{2,}(/|$)", text, re.IGNORECASE))


def locate_rows(
    grid: list[list[str]],
    *,
    header_row: int | None = None,
    prompts_row: int | None = None,
    no_prompts_row: bool = False,
) -> tuple[RowLayout, list[SheetAnomaly]]:
    layout = RowLayout()
    anomalies: list[SheetAnomaly] = []

    if not grid:
        anomalies.append(SheetAnomaly(code="no_header_row_found", message="The sheet is empty."))
        return layout, anomalies

    # --- header -----------------------------------------------------------
    search_limit = min(len(grid), MAX_HEADER_SEARCH_ROWS)
    fills = [sum(1 for cell in grid[i] if cell.strip()) for i in range(search_limit)]
    max_fill = max(fills, default=0)

    candidates: list[RowCandidate] = []
    for index in range(search_limit):
        if max_fill <= 0:
            break
        score = fills[index] / max_fill
        has_url_header = any(_URL_HEADER.match(cell) for cell in grid[index])
        if has_url_header:
            score += 0.5
        candidates.append(
            RowCandidate(
                row=index,
                score=round(score, 3),
                reason=f"{fills[index]} filled cells" + (", names a URL column" if has_url_header else ""),
            )
        )
    layout.header_candidates = sorted(candidates, key=lambda c: (-c.score, c.row))[:5]

    if header_row is not None:
        layout.header_row = header_row
        layout.detection["header"] = "explicit override"
    else:
        chosen = next(
            (c.row for c in candidates if c.score >= HEADER_FILL_RATIO),
            -1,
        )
        # A row that names a URL column beats a merely wide row above it.
        named = next((c.row for c in sorted(candidates, key=lambda c: c.row) if c.score >= 1.0), -1)
        layout.header_row = named if named >= 0 else chosen
        layout.detection["header"] = (
            "first row filling >=60% of the widest row" if layout.header_row >= 0 else "not found"
        )

    if layout.header_row < 0 or layout.header_row >= len(grid):
        anomalies.append(
            SheetAnomaly(
                code="no_header_row_found",
                message=(
                    "Could not identify a header row. Re-run with --header-row N "
                    "(0-based) naming the row that holds the column names."
                ),
            )
        )
        return layout, anomalies

    header = grid[layout.header_row]

    # --- prompts ----------------------------------------------------------
    if no_prompts_row:
        layout.prompts_row = -1
        layout.detection["prompts"] = "disabled by caller"
    elif prompts_row is not None:
        layout.prompts_row = prompts_row
        layout.detection["prompts"] = "explicit override"
    else:
        layout.prompts_row, layout.detection["prompts"] = _find_prompts_row(grid, layout.header_row)
        if layout.prompts_row < 0:
            anomalies.append(
                SheetAnomaly(
                    code="no_prompts_row_found",
                    message=(
                        "No prompts row found, so no columns will be created. If the sheet "
                        "has one, re-run with --prompts-row N."
                    ),
                )
            )

    layout.first_data_row = (
        max(layout.header_row, layout.prompts_row) + 1
        if layout.prompts_row >= 0
        else layout.header_row + 1
    )

    data_rows = [row for row in grid[layout.first_data_row :] if any(cell.strip() for cell in row)]

    # --- id and url columns ------------------------------------------------
    layout.url_column = _find_url_column(header, data_rows)
    if layout.url_column < 0:
        anomalies.append(
            SheetAnomaly(
                code="no_url_column_found",
                message="No column holds URLs. There is nothing to create sources from.",
            )
        )
    layout.id_column = _find_id_column(header, data_rows, exclude=layout.url_column)
    layout.detection["id_column"] = (
        header[layout.id_column] if 0 <= layout.id_column < len(header) else "not found"
    )
    layout.detection["url_column"] = (
        header[layout.url_column] if 0 <= layout.url_column < len(header) else "not found"
    )

    return layout, anomalies


def _find_prompts_row(grid: list[list[str]], header_row: int) -> tuple[int, str]:
    # 1. An explicit label in the first cell is unambiguous.
    for index in range(header_row + 1, min(len(grid), header_row + 4)):
        first = grid[index][0] if grid[index] else ""
        if _PROMPTS_LABEL.match(first):
            return index, "labelled `Prompts:` in the first cell"

    # 2. Otherwise, a row of long prose where the data rows hold ids and URLs.
    for index in range(header_row + 1, min(len(grid), header_row + 4)):
        cells = [cell for cell in grid[index] if cell.strip()]
        if len(cells) < 2:
            continue
        mean_length = sum(len(cell) for cell in cells) / len(cells)
        data_like = sum(1 for cell in cells if _looks_like_url_cell(cell) or _INTEGER.match(cell))
        if mean_length > PROMPT_MIN_MEAN_LENGTH and data_like / len(cells) < PROMPT_MAX_DATA_RATIO:
            return index, f"long prose row (mean {int(mean_length)} chars, no ids or URLs)"

    return -1, "not found"


def _find_url_column(header: list[str], data_rows: list[list[str]]) -> int:
    for index, cell in enumerate(header):
        if _URL_HEADER.match(cell):
            return index
    best, best_ratio = -1, 0.0
    for index in range(len(header)):
        values = [row[index] for row in data_rows if index < len(row) and row[index].strip()]
        if not values:
            continue
        ratio = sum(1 for value in values if _looks_like_url_cell(value)) / len(values)
        if ratio >= URL_COLUMN_MIN_RATIO and ratio > best_ratio:
            best, best_ratio = index, ratio
    return best


def _find_id_column(header: list[str], data_rows: list[list[str]], *, exclude: int) -> int:
    named: list[int] = []
    numeric: list[int] = []
    for index in range(len(header)):
        if index == exclude:
            continue
        values = [row[index] for row in data_rows if index < len(row) and row[index].strip()]
        if not values:
            continue
        ratio = sum(1 for value in values if _INTEGER.match(value)) / len(values)
        # A column the sheet itself calls an id keeps that job even with a few
        # placeholder cells in it -- one "TBD" should not cost the whole column,
        # since the rows it does number still carry the user's intent.
        if index < len(header) and _ID_HEADER.search(header[index] or ""):
            if ratio >= ID_COLUMN_NAMED_MIN_RATIO:
                named.append(index)
        elif ratio >= ID_COLUMN_MIN_RATIO:
            numeric.append(index)

    if named:
        return named[0]
    return numeric[0] if numeric else -1


# ---------------------------------------------------------------------------
# extraction
# ---------------------------------------------------------------------------


def parse_planning_sheet(
    path: Path | str,
    *,
    header_row: int | None = None,
    prompts_row: int | None = None,
    no_prompts_row: bool = False,
    repair_encoding: str = "auto",
    merge_duplicate_urls: bool = False,
) -> SheetPlan:
    """Read a planning spreadsheet into a reviewable plan."""
    path = Path(path)
    grid, source_format = read_grid(path)
    grid, encoding_report = repair_grid(grid, mode=repair_encoding)

    plan = SheetPlan(
        path=str(path),
        source_format=source_format,
        encoding_repair=encoding_report,
    )

    layout, anomalies = locate_rows(
        grid,
        header_row=header_row,
        prompts_row=prompts_row,
        no_prompts_row=no_prompts_row,
    )
    plan.layout = layout
    plan.anomalies.extend(anomalies)

    if encoding_report.applied:
        plan.anomalies.append(
            SheetAnomaly(
                code="mojibake_repaired",
                message=(
                    f"Repaired {encoding_report.cells_changed} cell(s) that were written as "
                    f"UTF-8 and read as {encoding_report.codec}. The source file still contains "
                    "the original characters."
                ),
                subject=encoding_report.codec,
            )
        )
    elif encoding_report.refused_reason and encoding_report.suspicious_before:
        plan.anomalies.append(
            SheetAnomaly(
                code="mojibake_unrepairable",
                message=(
                    f"Found {encoding_report.suspicious_before} damaged character sequence(s) "
                    f"that could not be repaired automatically: {encoding_report.refused_reason}."
                ),
            )
        )

    if layout.header_row < 0:
        plan.summary = "Could not read the sheet: no header row found."
        return plan

    header = grid[layout.header_row]
    prompts = grid[layout.prompts_row] if 0 <= layout.prompts_row < len(grid) else []

    _extract_columns(plan, header, prompts, layout)
    _extract_sources(plan, grid, layout)
    if merge_duplicate_urls:
        _merge_duplicate_urls(plan)
    _check_sources(plan)
    _extract_provided_columns(plan, grid, header, prompts, layout)

    plan.summary = _summarize(plan)
    return plan


def _extract_columns(
    plan: SheetPlan,
    header: list[str],
    prompts: list[str],
    layout: RowLayout,
) -> None:
    for index, label in enumerate(header):
        label = (label or "").strip()
        prompt = (prompts[index] if index < len(prompts) else "").strip()
        if index in {layout.id_column, layout.url_column}:
            continue
        if _PROMPTS_LABEL.match(prompt):
            # The row label itself, not a prompt.
            continue
        column = SheetColumn(index=index, label=label, prompt=prompt)
        if label and prompt:
            plan.columns.append(column)
        elif label:
            plan.skipped_columns.append(column)
            plan.anomalies.append(
                SheetAnomaly(
                    code="column_without_prompt",
                    message=f"Column {label!r} has no prompt and will not be created.",
                    subject=label,
                )
            )
        elif prompt:
            plan.anomalies.append(
                SheetAnomaly(
                    code="prompt_without_header",
                    message=(
                        f"Column {index} has a prompt but no heading, so it cannot be named. "
                        "Check for a merged cell in the header row."
                    ),
                    subject=str(index),
                )
            )


def _extract_sources(plan: SheetPlan, grid: list[list[str]], layout: RowLayout) -> None:
    if layout.url_column < 0:
        return
    for row_index in range(layout.first_data_row, len(grid)):
        row = grid[row_index]
        if not any(cell.strip() for cell in row):
            continue
        url = row[layout.url_column].strip() if layout.url_column < len(row) else ""
        raw_id = (
            row[layout.id_column].strip()
            if 0 <= layout.id_column < len(row)
            else ""
        )

        if not url:
            plan.anomalies.append(
                SheetAnomaly(
                    code="row_missing_url",
                    message=f"Row {row_index} has no URL and will be skipped.",
                    row=row_index,
                )
            )
            continue

        source_id = raw_id
        if raw_id and not _INTEGER.match(raw_id):
            plan.anomalies.append(
                SheetAnomaly(
                    code="non_numeric_id",
                    message=(
                        f"Row {row_index}: id {raw_id!r} is not a number, so an id will be "
                        "allocated automatically."
                    ),
                    subject=raw_id,
                    row=row_index,
                )
            )
            source_id = ""
        elif not raw_id:
            plan.anomalies.append(
                SheetAnomaly(
                    code="empty_id_cell",
                    message=f"Row {row_index} has no id; one will be allocated automatically.",
                    row=row_index,
                )
            )

        if not _looks_like_url_cell(url):
            # There is no address to fetch, so this cannot become a URL source:
            # handing it to `create_sources` could only ever produce
            # `url_invalid`, and one such row blocks the whole import. It keeps
            # its id and is reported, because the row is real -- the document
            # just arrives by hand instead of over the network.
            plan.documents.append(SheetDocument(row=row_index, id=source_id, label=url))
            plan.anomalies.append(
                SheetAnomaly(
                    code="document_row",
                    message=(
                        f"Row {row_index}: {url[:60]!r} is not a web address, so nothing will be "
                        f"fetched for id {source_id or '(unnumbered)'}. Attach the document by "
                        "hand, or correct the URL cell."
                    ),
                    subject=url,
                    row=row_index,
                )
            )
            continue

        plan.sources.append(SheetSource(row=row_index, id=source_id, url=url))


def _dedupe_key(url: str) -> str:
    """The repository's own idea of "the same URL", so a merge here matches it.

    Imported lazily: `attached_repository` is the app's largest module and
    importing it at module scope would drag the whole storage layer into a
    parser that otherwise only needs the standard library.
    """
    from backend.storage.attached_repository import repository_dedupe_key

    return repository_dedupe_key(url) or url.strip().lower().rstrip("/")


def _merge_duplicate_urls(plan: SheetPlan) -> None:
    """Collapse rows sharing a URL into the one with the lowest id.

    The rows are not discarded: the survivor records which ids it now stands
    for, so a value imported from a merged row can still be traced to the sheet
    row it came from, and `_extract_provided_columns` can gather what those rows
    said rather than only what the survivor said.
    """
    groups: dict[str, list[SheetSource]] = {}
    for source in plan.sources:
        groups.setdefault(_dedupe_key(source.url), []).append(source)

    survivors: list[SheetSource] = []
    merged_total = 0
    for members in groups.values():
        # Lowest numeric id wins; an unnumbered row sorts last so it never
        # displaces a row the user numbered deliberately.
        members.sort(key=lambda s: (int(s.id) if s.id.isdigit() else 10**9, s.row))
        keeper, rest = members[0], members[1:]
        if rest:
            keeper.merged_ids = [item.id for item in rest if item.id]
            keeper.merged_rows = [item.row for item in rest]
            merged_total += len(rest)
        survivors.append(keeper)

    if not merged_total:
        return

    survivors.sort(key=lambda s: s.row)
    plan.sources = survivors
    plan.anomalies.append(
        SheetAnomaly(
            code="duplicate_urls_merged",
            message=(
                f"{merged_total} row(s) shared a URL with an earlier row and were merged into "
                f"it, leaving {len(survivors)} source(s). Each survivor records the ids it "
                "stands for."
            ),
        )
    )


def _extract_provided_columns(
    plan: SheetPlan,
    grid: list[list[str]],
    header: list[str],
    prompts: list[str],
    layout: RowLayout,
) -> None:
    """Collect columns the user filled in themselves.

    A column with a heading, no prompt and no data is simply not a column. One
    with a heading, no prompt and *data* is the user's own work -- a collection
    date, the channel a link came from -- and until now the only thing said
    about it was that it would not be created, after which its contents were
    dropped on the floor.
    """
    if layout.url_column < 0:
        return

    row_ids: dict[int, str] = {}
    for source in plan.sources:
        row_ids[source.row] = source.id
        for row in source.merged_rows:
            row_ids[row] = source.id
    for document in plan.documents:
        row_ids[document.row] = document.id

    seen_labels: dict[str, int] = {}
    for index, raw_label in enumerate(header):
        label = (raw_label or "").strip()
        if index in {layout.id_column, layout.url_column} or not label:
            continue
        if (prompts[index] if index < len(prompts) else "").strip():
            continue

        # Values are gathered per source rather than per row, so a merged group
        # contributes everything its rows said, in the order the sheet says it.
        gathered: dict[str, list[str]] = {}
        for row_index in range(layout.first_data_row, len(grid)):
            source_id = row_ids.get(row_index)
            if not source_id:
                continue
            row = grid[row_index]
            value = row[index].strip() if index < len(row) else ""
            if not value:
                continue
            bucket = gathered.setdefault(source_id, [])
            if value not in bucket:
                bucket.append(value)

        if not gathered:
            continue

        key = label.casefold()
        if key in seen_labels:
            plan.anomalies.append(
                SheetAnomaly(
                    code="duplicate_provided_column",
                    message=(
                        f"Column {label!r} appears at positions {seen_labels[key]} and {index}, "
                        "both with data and no prompt. Only the first will be imported."
                    ),
                    subject=label,
                )
            )
            continue
        seen_labels[key] = index

        plan.provided_columns.append(
            SheetProvidedColumn(
                index=index,
                label=label,
                values={sid: "; ".join(values) for sid, values in gathered.items()},
            )
        )

    # A column that is about to be imported is not a column that "will not be
    # created", and printing both leaves the user to work out which is true.
    imported = {column.label for column in plan.provided_columns}
    plan.skipped_columns = [
        column for column in plan.skipped_columns if column.label not in imported
    ]
    plan.anomalies = [
        anomaly
        for anomaly in plan.anomalies
        if not (anomaly.code == "column_without_prompt" and anomaly.subject in imported)
    ]

    _add_merged_ids_column(plan)


def _add_merged_ids_column(plan: SheetPlan) -> None:
    """Record which sheet rows each merged source stands for, as a column.

    Without it a merged source's provided values name several channels with no
    way back to the rows they came from, and the sheet can no longer be joined
    to the repository on id alone.
    """
    values = {
        source.id: ", ".join(source.merged_ids)
        for source in plan.sources
        if source.id and source.merged_ids
    }
    if not values:
        return
    plan.provided_columns.append(
        SheetProvidedColumn(index=-1, label="Merged ID#s", values=values)
    )


def _check_sources(plan: SheetPlan) -> None:
    if not plan.sources:
        plan.anomalies.append(
            SheetAnomaly(
                code="no_sources_found",
                message="No rows had a URL, so there is nothing to create.",
            )
        )
        return

    seen_ids: dict[str, int] = {}
    for source in plan.sources:
        if not source.id:
            continue
        if source.id in seen_ids:
            plan.anomalies.append(
                SheetAnomaly(
                    code="duplicate_source_id",
                    message=(
                        f"Id {source.id} is used by rows {seen_ids[source.id]} and {source.row}. "
                        "Fix the sheet, or one of them will be refused."
                    ),
                    subject=source.id,
                    row=source.row,
                )
            )
        else:
            seen_ids[source.id] = source.row

    seen_urls: dict[str, int] = {}
    for source in plan.sources:
        key = source.url.strip().lower().rstrip("/")
        if key in seen_urls:
            plan.anomalies.append(
                SheetAnomaly(
                    code="duplicate_url",
                    message=f"Rows {seen_urls[key]} and {source.row} share the same URL.",
                    subject=source.url,
                    row=source.row,
                )
            )
        else:
            seen_urls[key] = source.row

    numeric = sorted(int(s.id) for s in plan.sources if s.id)
    if numeric and numeric[-1] - numeric[0] + 1 != len(numeric):
        missing = sorted(set(range(numeric[0], numeric[-1] + 1)) - set(numeric))
        plan.anomalies.append(
            SheetAnomaly(
                code="non_contiguous_ids",
                message=(
                    f"Ids run {numeric[0]}-{numeric[-1]} with {len(missing)} gap(s): "
                    f"{', '.join(str(m) for m in missing[:10])}"
                    f"{'...' if len(missing) > 10 else ''}. This is fine, just worth knowing."
                ),
            )
        )


def _summarize(plan: SheetPlan) -> str:
    numeric = sorted(int(s.id) for s in plan.sources if s.id)
    id_range = f" (ids {numeric[0]}-{numeric[-1]})" if numeric else ""
    parts = [f"{len(plan.sources)} source(s){id_range}"]
    merged = sum(len(source.merged_ids) for source in plan.sources)
    if merged:
        parts.append(f"{merged} duplicate row(s) merged in")
    if plan.documents:
        parts.append(f"{len(plan.documents)} document(s) to attach by hand")
    parts.append(f"{len(plan.columns)} column(s) with prompts")
    if plan.provided_columns:
        parts.append(f"{len(plan.provided_columns)} column(s) of provided data")
    if plan.skipped_columns:
        parts.append(f"{len(plan.skipped_columns)} column(s) skipped (no prompt)")
    blocking = plan.blocking_anomalies
    if blocking:
        parts.append(f"{len(blocking)} problem(s) needing attention")
    elif plan.anomalies:
        parts.append(f"{len(plan.anomalies)} note(s)")
    return "Found " + ", ".join(parts) + "."


def sheet_plan_to_create_sources_params(plan: SheetPlan) -> dict[str, Any]:
    return {
        "sources": [{"url": s.url, "id": s.id} for s in plan.sources],
        "skip_existing": True,
    }


def sheet_plan_to_create_columns_params(plan: SheetPlan) -> dict[str, Any]:
    columns: list[dict[str, Any]] = [
        {
            "label": column.label,
            "instruction_prompt": column.prompt,
            "include_source_text": True,
        }
        for column in plan.columns
    ]
    # A provided column holds the user's own data and is never run, so it needs
    # no prompt, no source text and no constraint -- only somewhere to live.
    columns.extend(
        {
            "label": column.label,
            "instruction_prompt": "",
            "include_source_text": False,
            "provided": True,
        }
        for column in plan.provided_columns
    )
    return {"columns": columns, "skip_existing": True}


def sheet_plan_to_set_values_params(plan: SheetPlan) -> list[dict[str, Any]]:
    """One `set_column_values` request per provided column, keyed by label.

    Labels rather than column ids because the columns do not exist yet when the
    plan is written; the operation resolves the label against the repository.
    """
    return [
        {
            "column_label": column.label,
            "values": {f"{int(sid):06d}": value for sid, value in column.values.items() if sid},
            "overwrite": False,
        }
        for column in plan.provided_columns
    ]
