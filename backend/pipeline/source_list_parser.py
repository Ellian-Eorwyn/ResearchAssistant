"""Parse uploaded source URL spreadsheets into bibliography entries."""

from __future__ import annotations

import csv
import html
import io
import re
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import parse_qsl, quote, unquote, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook

from backend.models.bibliography import BibliographyEntry

SUPPORTED_SOURCE_LIST_EXTENSIONS = {".csv", ".xlsx"}

URL_COLUMN_CANDIDATES = {
    "url",
    "sourceurl",
    "originalurl",
    "finalurl",
    "link",
    "uri",
    "website",
}
REF_COLUMN_CANDIDATES = {"refnumber", "reference", "citationnumber", "citation", "ref"}
SOURCE_DOC_COLUMN_CANDIDATES = {
    "sourcedocument",
    "sourcedocumentname",
    "document",
    "documentname",
    "sourcefile",
    "filename",
}
AUTHORS_COLUMN_CANDIDATES = {"authors", "author"}
TITLE_COLUMN_CANDIDATES = {"title", "citedtitle", "name"}
YEAR_COLUMN_CANDIDATES = {"year", "publicationyear"}
DOI_COLUMN_CANDIDATES = {"doi"}
RAW_TEXT_COLUMN_CANDIDATES = {"rawtext", "entry", "referenceentry"}
TRACKING_PARAM_EXACT = {"gclid", "fbclid", "msclkid"}
TRACKING_PARAM_PREFIXES = ("utm_",)


@dataclass
class SourceListParseResult:
    entries: list[BibliographyEntry]
    total_rows: int
    accepted_rows: int
    missing_url_rows: int
    estimated_duplicate_urls: int


def parse_source_list_upload(filename: str, content: bytes) -> SourceListParseResult:
    ext = Path(filename or "").suffix.lower()
    if ext not in SUPPORTED_SOURCE_LIST_EXTENSIONS:
        raise ValueError("Unsupported file type. Use .csv or .xlsx.")

    rows = _read_rows(ext=ext, content=content)
    if not rows:
        raise ValueError("No data rows found in uploaded file.")

    normalized_headers = {_normalize_header(k): k for k in rows[0].keys()}
    url_header = _pick_header(normalized_headers, URL_COLUMN_CANDIDATES)
    if not url_header:
        raise ValueError("Spreadsheet must include a 'URL' column.")

    ref_header = _pick_header(normalized_headers, REF_COLUMN_CANDIDATES)
    source_doc_header = _pick_header(normalized_headers, SOURCE_DOC_COLUMN_CANDIDATES)
    authors_header = _pick_header(normalized_headers, AUTHORS_COLUMN_CANDIDATES)
    title_header = _pick_header(normalized_headers, TITLE_COLUMN_CANDIDATES)
    year_header = _pick_header(normalized_headers, YEAR_COLUMN_CANDIDATES)
    doi_header = _pick_header(normalized_headers, DOI_COLUMN_CANDIDATES)
    raw_text_header = _pick_header(normalized_headers, RAW_TEXT_COLUMN_CANDIDATES)

    entries: list[BibliographyEntry] = []
    missing_url_rows = 0
    next_ref_number = 1

    for row in rows:
        url = _clean_url_candidate(_cell_text(row.get(url_header)))
        if not url:
            missing_url_rows += 1
            continue

        ref_number = _parse_int(_cell_text(row.get(ref_header)))
        if ref_number is None:
            ref_number = next_ref_number
            next_ref_number += 1
        else:
            next_ref_number = max(next_ref_number, ref_number + 1)

        title = _cell_text(row.get(title_header))
        doi = _cell_text(row.get(doi_header))
        source_document_name = _cell_text(row.get(source_doc_header))
        raw_text = _cell_text(row.get(raw_text_header)) or title or url
        authors = _parse_authors(_cell_text(row.get(authors_header)))

        entries.append(
            BibliographyEntry(
                ref_number=ref_number,
                raw_text=raw_text,
                source_document_name=source_document_name,
                authors=authors,
                title=title,
                year=_cell_text(row.get(year_header)),
                doi=doi,
                url=url,
                parse_confidence=1.0,
                parse_warnings=[],
                repair_method="source_upload",
            )
        )

    if not entries:
        raise ValueError("No usable rows found. At least one row must include a URL.")

    return SourceListParseResult(
        entries=entries,
        total_rows=len(rows),
        accepted_rows=len(entries),
        missing_url_rows=missing_url_rows,
        estimated_duplicate_urls=_estimate_duplicate_urls(entries),
    )


def _read_rows(ext: str, content: bytes) -> list[dict[str, str]]:
    if ext == ".csv":
        return _read_csv_rows(content)
    return _read_xlsx_rows(content)


def _read_csv_rows(content: bytes) -> list[dict[str, str]]:
    text = _decode_text(content)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    rows: list[dict[str, str]] = []
    for raw_row in reader:
        row = {
            str(k).strip(): _cell_text(v)
            for k, v in (raw_row or {}).items()
            if k is not None
        }
        if any(v for v in row.values()):
            rows.append(row)
    return rows


def _read_xlsx_rows(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []

        rows_iter = ws.iter_rows(values_only=True)
        headers: list[str] = []
        for candidate in rows_iter:
            candidate_vals = [_cell_text(v) for v in (candidate or ())]
            if any(candidate_vals):
                headers = [v or f"column_{idx+1}" for idx, v in enumerate(candidate_vals)]
                break
        if not headers:
            return []

        data_rows: list[dict[str, str]] = []
        for values in rows_iter:
            values_list = [_cell_text(v) for v in (values or ())]
            if not any(values_list):
                continue
            row: dict[str, str] = {}
            for idx, header in enumerate(headers):
                row[header] = values_list[idx] if idx < len(values_list) else ""
            data_rows.append(row)
        return data_rows
    finally:
        wb.close()


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except Exception:
            continue
    return content.decode("utf-8", errors="replace")


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_int(value: str) -> int | None:
    if not value:
        return None
    try:
        return int(value)
    except Exception:
        return None


def _parse_authors(value: str) -> list[str]:
    if not value:
        return []
    tokens = [a.strip() for a in re.split(r"[;|]", value) if a.strip()]
    return tokens


def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (header or "").strip().lower())


def _pick_header(headers_by_normalized: dict[str, str], candidates: set[str]) -> str:
    for candidate in candidates:
        found = headers_by_normalized.get(candidate)
        if found:
            return found
    return ""


def _estimate_duplicate_urls(entries: list[BibliographyEntry]) -> int:
    seen: set[str] = set()
    duplicates = 0
    for entry in entries:
        key = _normalize_url_for_dedupe(entry.url)
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
    return duplicates


def _normalize_url_for_dedupe(url: str) -> str:
    candidate = _clean_url_candidate(url)
    if not candidate:
        return ""
    if "://" not in candidate:
        candidate = f"https://{candidate}"
    try:
        parsed = urlsplit(candidate)
    except Exception:
        return candidate.lower()

    if not parsed.netloc:
        return candidate.lower()

    filtered_params: list[tuple[str, str]] = []
    for key, value in parse_qsl(parsed.query, keep_blank_values=True):
        key_lower = key.lower()
        if key_lower in TRACKING_PARAM_EXACT:
            continue
        if any(key_lower.startswith(prefix) for prefix in TRACKING_PARAM_PREFIXES):
            continue
        filtered_params.append((key, value))
    filtered_params.sort(key=lambda item: (item[0].lower(), item[1]))
    query = urlencode(filtered_params, doseq=True)
    canonical_path = quote(unquote(parsed.path or "/"), safe="/:@!$&'()*+,;=-._~")

    normalized = urlunsplit(
        (
            parsed.scheme.lower() or "https",
            parsed.netloc.lower(),
            canonical_path,
            query,
            "",
        )
    )
    return normalized


def _clean_url_candidate(value: str) -> str:
    candidate = html.unescape((value or "").strip())
    if not candidate:
        return ""
    if candidate.startswith("<") and candidate.endswith(">"):
        candidate = candidate[1:-1].strip()

    candidate = candidate.strip("\"'`")
    candidate = _trim_unbalanced_trailing_closers(candidate)
    return candidate


def _trim_unbalanced_trailing_closers(candidate: str) -> str:
    value = (candidate or "").strip()
    if not value:
        return ""

    matching_openers = {")": "(", "]": "[", "}": "{", ">": "<"}
    while value:
        last = value[-1]
        if last in "\"'`":
            value = value[:-1].rstrip()
            continue
        opener = matching_openers.get(last)
        if not opener:
            break
        if value.count(last) <= value.count(opener):
            break
        value = value[:-1].rstrip()
    return value
