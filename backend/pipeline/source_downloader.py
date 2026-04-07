"""Download and capture citation source URLs into local output bundles."""

from __future__ import annotations

import asyncio
import csv
import html
import hashlib
import io
import json
import logging
import re
import shutil
import subprocess
import tempfile
import threading
from collections.abc import Awaitable, Callable, Iterator, Sequence
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, TypeVar
from urllib.parse import parse_qsl, quote, unquote, urlencode, urljoin, urlsplit, urlunsplit

import httpx
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.llm.client import UnifiedLLMClient
from backend.llm.prompts import (
    SOURCE_CATALOG_SYSTEM,
    SOURCE_CATALOG_USER,
    SOURCE_CITATION_VERIFY_SYSTEM,
    SOURCE_CITATION_VERIFY_USER,
    SOURCE_MARKDOWN_CLEANUP_SYSTEM,
    SOURCE_MARKDOWN_CLEANUP_USER,
    SOURCE_RATING_SYSTEM,
    SOURCE_RATING_USER,
    SOURCE_SUMMARY_SYSTEM,
    SOURCE_SUMMARY_USER,
    SOURCE_TITLE_SYSTEM,
    SOURCE_TITLE_USER,
)
from backend.models.citation_metadata import CitationAuthor, CitationFieldEvidence, CitationMetadata
from backend.models.repository import RepositoryColumnConfig
from backend.models.settings import LLMBackendConfig
from backend.models.sources import (
    SOURCE_MANIFEST_COLUMNS,
    SourceDownloadStatus,
    SourceOutputOptions,
    SourceOutputSummary,
    SourceItemStatus,
    SourceManifestArtifact,
    SourceManifestRow,
    SourcePhaseMetadata,
)
from backend.pipeline.standardized_markdown import extract_markdown_title_candidate
from backend.storage.file_store import FileStore

try:
    import trafilatura
except ImportError:  # pragma: no cover - optional at runtime
    trafilatura = None


logger = logging.getLogger(__name__)
T = TypeVar("T")
_MUPDF_MESSAGE_LOCK = threading.Lock()

HTTP_TIMEOUT_SECONDS = 25.0
PLAYWRIGHT_TIMEOUT_MS = 20000
PLAYWRIGHT_VIEWPORT_WIDTH = 1366
PLAYWRIGHT_VIEWPORT_HEIGHT = 1200
MAX_VISUAL_CAPTURE_SEGMENTS = 40
MIN_MARKDOWN_SCORE = 180
MIN_FALLBACK_MARKDOWN_SCORE = 20
MAX_CLEANUP_SOURCE_CHARS = 24000  # legacy fallback
MAX_SUMMARY_SOURCE_CHARS = 20000  # legacy fallback

NOTE_RUNTIME_MISSING_TRAFILATURA = "runtime_missing_trafilatura"
NOTE_RUNTIME_MISSING_PLAYWRIGHT = "runtime_missing_playwright"
NOTE_RUNTIME_MISSING_TEXTUTIL = "runtime_missing_textutil"
NOTE_RUNTIME_MISSING_TESSERACT = "runtime_missing_tesseract"
NOTE_RUNTIME_MISSING_LLM_VISION = "runtime_missing_llm_vision"
NOTE_BLOCKED_REQUEST = "blocked_request"
NOTE_EXTRACTION_FAILURE = "extraction_failure"
NOTE_OCR_LOCAL_USED = "ocr_local_used"
NOTE_OCR_LLM_FALLBACK_USED = "ocr_llm_fallback_used"
NOTE_DOC_CONVERSION_FAILED = "doc_conversion_failed"
NOTE_LLM_CLEANUP_FAILED = "llm_cleanup_failed"
NOTE_LLM_CLEANUP_SKIPPED_LLM_NOT_CONFIGURED = "llm_cleanup_skipped_llm_not_configured"
NOTE_TITLE_GENERATION_FAILED = "title_generation_failed"
NOTE_TITLE_SKIPPED_LLM_NOT_CONFIGURED = "title_skipped_llm_not_configured"
NOTE_SUMMARY_GENERATION_FAILED = "summary_generation_failed"
NOTE_SUMMARY_SKIPPED_LLM_NOT_CONFIGURED = "summary_skipped_llm_not_configured"
NOTE_RATING_GENERATION_FAILED = "rating_generation_failed"
NOTE_RATING_SKIPPED_LLM_NOT_CONFIGURED = "rating_skipped_llm_not_configured"
NOTE_VISUAL_CAPTURE_FAILED = "visual_capture_failed"
NOTE_VISUAL_CAPTURE_SEGMENTED = "visual_capture_segmented"
PHASE_FETCH = "fetch"
PHASE_CONVERT = "convert"
PHASE_CLEANUP = "cleanup"
PHASE_TITLE = "title"
PHASE_CATALOG = "catalog"
PHASE_CITATION_VERIFY = "citation_verify"
PHASE_SUMMARY = "summary"
PHASE_RATING = "rating"

LEGACY_PHASE_ALIASES = {
    "summarize": PHASE_SUMMARY,
    "tag": PHASE_RATING,
}
PHASE_METADATA_ALIASES = {
    PHASE_SUMMARY: ("summarize",),
    PHASE_RATING: ("tag",),
}

PROMPT_VERSION_CLEANUP = "source_markdown_cleanup.v1"
PROMPT_VERSION_CATALOG = "source_catalog.v1"
PROMPT_VERSION_CITATION_VERIFY = "source_citation_verify.v1"
PROMPT_VERSION_TITLE = "source_title.v1"
PROMPT_VERSION_SUMMARY = "source_summary.v1"
PROMPT_VERSION_RATING = "source_rating.v1"

INSTALL_BOOTSTRAP_COMMAND = "./scripts/bootstrap_venv.sh"
INSTALL_REQUIREMENTS_COMMAND = ".venv/bin/python -m pip install -r requirements.txt"
INSTALL_PLAYWRIGHT_BROWSER_COMMAND = ".venv/bin/python -m playwright install chromium"

PDF_NATIVE_PAGE_MIN_CHARS = 120
PDF_NATIVE_DOC_MIN_AVG_CHARS = 180
PDF_TEXT_ALPHA_MIN_RATIO = 0.55
PDF_OCR_MIN_CHARS = 80

DOCUMENT_CONTENT_TYPE_EXT = {
    "application/msword": ".doc",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.template": ".dotx",
    "application/rtf": ".rtf",
    "text/markdown": ".md",
    "text/x-markdown": ".md",
    "application/markdown": ".md",
    "text/plain": ".txt",
}

DOCUMENT_EXTENSIONS = {
    ".doc",
    ".docx",
    ".dot",
    ".dotx",
    ".md",
    ".txt",
    ".rtf",
}

BLOCKED_PAGE_PATTERNS = [
    re.compile(r"\bjust a moment\b", re.IGNORECASE),
    re.compile(r"checking (?:if|your browser)", re.IGNORECASE),
    re.compile(r"cf[-_ ]?(?:ray|challenge|chl|turnstile)", re.IGNORECASE),
    re.compile(r"\bcaptcha\b", re.IGNORECASE),
    re.compile(r"\baccess denied\b", re.IGNORECASE),
    re.compile(r"\brequest blocked\b", re.IGNORECASE),
    re.compile(r"\bsecurity check\b", re.IGNORECASE),
]

TRACKING_PARAM_EXACT = {"gclid", "fbclid", "msclkid"}
TRACKING_PARAM_PREFIXES = ("utm_",)

MANIFEST_DERIVED_COLUMNS = [
    "summary_text",
    "rating_overall",
    "rating_confidence",
    "rating_rationale",
    "relevant_sections",
    "rating_dimensions_json",
    "flag_scores_json",
    "rating_raw_json",
    "citation_title",
    "citation_authors",
    "citation_issued",
    "citation_url",
    "citation_publisher",
    "citation_container_title",
    "citation_volume",
    "citation_issue",
    "citation_pages",
    "citation_language",
    "citation_accessed",
    "citation_type",
    "citation_doi",
    "citation_report_number",
    "citation_standard_number",
    "citation_verification_status",
    "citation_blocked_reasons",
    "citation_manual_override_fields",
    "citation_field_evidence_json",
    "citation_verified_at",
    "citation_ready",
    "citation_missing_fields",
    "citation_confidence",
]

RATING_DIMENSION_CONTAINER_KEYS = ("ratings", "scores", "dimensions")
RATING_RESERVED_KEYS = {
    "confidence",
    "rationale",
    "relevant_sections",
    "relevant_section",
    "sections",
    "flags",
    "ratings",
    "scores",
    "dimensions",
    "overall",
    "overall_rating",
    "overall_score",
    "summary",
}

CITATION_REQUIRED_FIELDS = ("title", "authors", "issued", "url")
CITATION_REQUIRED_FIELD_LABELS = {
    "title": "title",
    "authors": "authors",
    "issued": "publication_year",
    "url": "url",
}
CITATION_VERIFIABLE_FIELDS = (
    "item_type",
    "title",
    "authors",
    "issued",
    "publisher",
    "container_title",
    "volume",
    "issue",
    "pages",
    "doi",
    "url",
    "report_number",
    "standard_number",
    "language",
    "accessed",
)
CITATION_CONFIDENCE_INCREMENT = 0.05
DOI_CSL_ACCEPT_HEADER = "application/vnd.citationstyles.csl+json"
DOI_RIS_ACCEPT_HEADER = "application/x-research-info-systems"
DOI_PATTERN = re.compile(r"\b(10\.\d{4,9}/[-._;()/:A-Z0-9]+)\b", re.IGNORECASE)
REPORT_NUMBER_PATTERN = re.compile(
    r"\b(?:report|publication|document|working paper|technical report|report no\.?|publication no\.?)\s*[:#]?\s*([A-Z0-9][A-Z0-9._/-]{2,})",
    re.IGNORECASE,
)
STANDARD_NUMBER_PATTERN = re.compile(
    r"\b(?:ANSI|ASHRAE|ASTM|IEC|IEEE|ISO|NIST|UL)\s*[A-Z0-9._/-]{1,20}\b",
    re.IGNORECASE,
)
HTML_META_TAG_PATTERN = re.compile(
    r"""<meta\s+[^>]*(?:name|property)\s*=\s*["']([^"']+)["'][^>]*content\s*=\s*["']([^"']*)["'][^>]*>""",
    re.IGNORECASE,
)
HTML_TITLE_PATTERN = re.compile(r"(?is)<title[^>]*>(.*?)</title>")
HTML_JSON_LD_PATTERN = re.compile(
    r'(?is)<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>'
)
CORPORATE_AUTHOR_KEYWORDS = {
    "agency",
    "association",
    "board",
    "bureau",
    "center",
    "centre",
    "commission",
    "committee",
    "company",
    "corporation",
    "council",
    "department",
    "district",
    "foundation",
    "group",
    "institute",
    "lab",
    "laboratory",
    "ministry",
    "office",
    "organization",
    "programme",
    "program",
    "society",
    "team",
    "university",
}
CANONICAL_CITATION_TYPE_MAP = {
    "article-journal": "journal article",
    "journalarticle": "journal article",
    "journal article": "journal article",
    "journal-article": "journal article",
    "journal_article": "journal article",
    "article": "journal article",
    "article-magazine": "magazine article",
    "magazine article": "magazine article",
    "article-newspaper": "newspaper article",
    "newspaper article": "newspaper article",
    "book": "book",
    "book-chapter": "book chapter",
    "book chapter": "book chapter",
    "chapter": "book chapter",
    "report": "report",
    "working-paper": "report",
    "working paper": "report",
    "technical report": "report",
    "webpage": "web page",
    "website": "web page",
    "web page": "web page",
    "blog post": "web page",
    "dataset": "dataset",
    "dissertation": "thesis",
    "thesis": "thesis",
    "conference-paper": "conference paper",
    "conference paper": "conference paper",
    "proceedings-article": "conference paper",
    "standard": "standard",
}
RIS_TYPE_MAP = {
    "journal article": "JOUR",
    "magazine article": "MGZN",
    "newspaper article": "NEWS",
    "book": "BOOK",
    "book chapter": "CHAP",
    "report": "RPRT",
    "web page": "ELEC",
    "thesis": "THES",
    "conference paper": "CONF",
    "standard": "RPRT",
    "dataset": "DATA",
}


@dataclass
class SourceTarget:
    id: str
    source_document_name: str
    citation_number: str
    original_url: str


@dataclass
class RuntimeCapabilities:
    trafilatura_available: bool
    playwright_python_available: bool
    playwright_browser_available: bool
    textutil_available: bool
    tesseract_available: bool
    llm_vision_enabled: bool
    runtime_notes: list[str]
    runtime_guidance: list[dict[str, str]]


@contextmanager
def suppress_mupdf_messages() -> Iterator[None]:
    """Temporarily silence noisy MuPDF parser messages for malformed PDFs."""
    with _MUPDF_MESSAGE_LOCK:
        enabled = False
        prev_errors = True
        prev_warnings = True
        try:
            prev_errors = bool(fitz.TOOLS.mupdf_display_errors())
            prev_warnings = bool(fitz.TOOLS.mupdf_display_warnings())
            fitz.TOOLS.mupdf_display_errors(False)
            fitz.TOOLS.mupdf_display_warnings(False)
            enabled = True
        except Exception:
            enabled = False

        try:
            yield
        finally:
            if enabled:
                fitz.TOOLS.mupdf_display_errors(prev_errors)
                fitz.TOOLS.mupdf_display_warnings(prev_warnings)


def run_async_in_sync(
    async_fn: Callable[..., Awaitable[T]],
    *args: Any,
    **kwargs: Any,
) -> T:
    """
    Execute an async function from synchronous code in both loop/no-loop contexts.

    If a loop is already running on this thread, run the coroutine in a helper
    thread to avoid `asyncio.run()` reentrancy errors.
    """
    try:
        asyncio.get_running_loop()
    except RuntimeError:
        return asyncio.run(async_fn(*args, **kwargs))

    result: dict[str, Any] = {}
    done = threading.Event()

    def _runner() -> None:
        try:
            result["value"] = asyncio.run(async_fn(*args, **kwargs))
        except BaseException as exc:  # noqa: BLE001
            result["error"] = exc
        finally:
            done.set()

    worker = threading.Thread(target=_runner, daemon=True)
    worker.start()
    done.wait()

    error = result.get("error")
    if isinstance(error, BaseException):
        raise error
    return result["value"]


def _normalize_phase_name(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return ""
    return LEGACY_PHASE_ALIASES.get(normalized, normalized)


def _get_phase_metadata(
    row: SourceManifestRow,
    phase: str,
) -> SourcePhaseMetadata | None:
    canonical = _normalize_phase_name(phase)
    metadata = row.phase_metadata.get(canonical)
    if metadata is not None:
        if metadata.phase != canonical:
            metadata.phase = canonical
            row.phase_metadata[canonical] = metadata
        return metadata
    for alias in PHASE_METADATA_ALIASES.get(canonical, ()):
        metadata = row.phase_metadata.get(alias)
        if metadata is None:
            continue
        metadata.phase = canonical
        row.phase_metadata[canonical] = metadata
        return metadata
    return None


def _set_phase_metadata(
    row: SourceManifestRow,
    phase: str,
    metadata: SourcePhaseMetadata,
) -> None:
    canonical = _normalize_phase_name(phase)
    metadata.phase = canonical
    row.phase_metadata[canonical] = metadata


def _normalize_row_phase_metadata(row: SourceManifestRow) -> SourceManifestRow:
    if not row.phase_metadata:
        return row
    updates: dict[str, SourcePhaseMetadata] = {}
    removals: list[str] = []
    for key, metadata in row.phase_metadata.items():
        canonical = _normalize_phase_name(key)
        if canonical == key:
            if metadata.phase != canonical:
                metadata.phase = canonical
            continue
        if canonical in row.phase_metadata or canonical in updates:
            removals.append(key)
            continue
        metadata.phase = canonical
        updates[canonical] = metadata
        removals.append(key)
    for key in removals:
        row.phase_metadata.pop(key, None)
    row.phase_metadata.update(updates)
    return row


class PlaywrightRenderer:
    """Lazily initialized Playwright HTML renderer."""

    def __init__(
        self,
        timeout_ms: int = PLAYWRIGHT_TIMEOUT_MS,
        startup_error: str = "",
    ):
        self.timeout_ms = timeout_ms
        self._playwright = None
        self._browser = None
        self._startup_error: str = startup_error

    def render(self, url: str) -> tuple[str, str]:
        if self._startup_error:
            return "", self._startup_error
        try:
            self._ensure_started()
        except Exception as exc:  # pragma: no cover - runtime dependency
            self._startup_error = _normalize_playwright_error(exc)
            return "", self._startup_error

        page = self._browser.new_page()
        try:
            page.goto(url, wait_until="networkidle", timeout=self.timeout_ms)
            return page.content(), ""
        except Exception as exc:  # pragma: no cover - runtime/browser failures
            return "", _normalize_playwright_error(exc)
        finally:
            page.close()

    def capture_visual_pdf(self, url: str) -> tuple[bytes, str, list[str]]:
        if self._startup_error:
            return b"", self._startup_error, []
        try:
            self._ensure_started()
        except Exception as exc:  # pragma: no cover - runtime dependency
            self._startup_error = _normalize_playwright_error(exc)
            return b"", self._startup_error, []

        page = self._browser.new_page(
            viewport={
                "width": PLAYWRIGHT_VIEWPORT_WIDTH,
                "height": PLAYWRIGHT_VIEWPORT_HEIGHT,
            }
        )
        notes: list[str] = []
        try:
            page.goto(url, wait_until="networkidle", timeout=self.timeout_ms)
            page.wait_for_timeout(200)
            return self._page_to_visual_pdf(page, notes), "", notes
        except Exception as exc:  # pragma: no cover - runtime/browser failures
            return b"", _normalize_playwright_error(exc), notes
        finally:
            page.close()

    def _page_to_visual_pdf(self, page, notes: list[str]) -> bytes:
        full_page_error: Exception | None = None
        try:
            full_page_png = page.screenshot(type="png", full_page=True)
            return png_images_to_pdf_bytes([full_page_png])
        except Exception as exc:
            full_page_error = exc

        segmented_pngs = self._capture_segmented_screenshots(page)
        if segmented_pngs:
            notes.append(NOTE_VISUAL_CAPTURE_SEGMENTED)
            return png_images_to_pdf_bytes(segmented_pngs)
        if full_page_error:
            raise full_page_error
        raise RuntimeError("visual_capture_failed: no screenshot data")

    def _capture_segmented_screenshots(self, page) -> list[bytes]:
        viewport = page.viewport_size or {}
        viewport_height = int(viewport.get("height") or PLAYWRIGHT_VIEWPORT_HEIGHT)
        if viewport_height <= 0:
            viewport_height = PLAYWRIGHT_VIEWPORT_HEIGHT

        scroll_height = page.evaluate(
            """() => Math.max(
                document.documentElement?.scrollHeight || 0,
                document.body?.scrollHeight || 0,
                document.documentElement?.offsetHeight || 0,
                document.body?.offsetHeight || 0,
                window.innerHeight || 0
            )"""
        )
        total_height = int(scroll_height or viewport_height)
        if total_height <= 0:
            total_height = viewport_height

        max_scroll = max(total_height - viewport_height, 0)
        captures: list[bytes] = []
        last_scroll = -1
        y = 0

        for _ in range(MAX_VISUAL_CAPTURE_SEGMENTS):
            scroll_y = min(y, max_scroll)
            if scroll_y == last_scroll and captures:
                break
            page.evaluate("(targetY) => window.scrollTo(0, targetY)", scroll_y)
            page.wait_for_timeout(80)
            captures.append(page.screenshot(type="png", full_page=False))
            last_scroll = scroll_y
            if scroll_y >= max_scroll:
                break
            y += viewport_height

        page.evaluate("() => window.scrollTo(0, 0)")
        return captures

    def _ensure_started(self) -> None:
        if self._browser is not None:
            return
        from playwright.sync_api import sync_playwright

        self._playwright = sync_playwright().start()
        self._browser = self._playwright.chromium.launch(headless=True)

    def close(self) -> None:
        try:
            if self._browser is not None:
                self._browser.close()
        finally:
            if self._playwright is not None:
                self._playwright.stop()


class SourceDownloadOrchestrator:
    """Runs source download/capture pipeline for a completed extraction job."""

    def __init__(
        self,
        job_id: str,
        store: FileStore,
        rerun_failed_only: bool = False,
        use_llm: bool = False,
        llm_backend: LLMBackendConfig | None = None,
        research_purpose: str = "",
        fetch_delay: float = 2.0,
        run_download: bool = True,
        run_convert: bool = False,
        run_catalog: bool = False,
        run_citation_verify: bool = False,
        run_llm_cleanup: bool = False,
        run_llm_title: bool = False,
        run_llm_summary: bool = True,
        run_llm_rating: bool = False,
        force_redownload: bool = False,
        force_convert: bool = False,
        force_catalog: bool = False,
        force_citation_verify: bool = False,
        force_llm_cleanup: bool = False,
        force_title: bool = False,
        force_summary: bool = False,
        force_rating: bool = False,
        project_profile_name: str = "",
        project_profile_yaml: str = "",
        output_options: SourceOutputOptions | None = None,
        target_rows: list[SourceManifestRow] | None = None,
        output_dir: Path | None = None,
        writes_to_repository: bool = False,
        repository_path: str = "",
        selected_scope: str = "",
        selected_import_id: str = "",
        selected_phases: list[str] | None = None,
        row_persist_callback: Callable[[SourceManifestRow], None] | None = None,
    ):
        self.job_id = job_id
        self.store = store
        self.rerun_failed_only = rerun_failed_only
        self.use_llm = use_llm
        self.llm_backend = llm_backend or LLMBackendConfig()
        self.research_purpose = (research_purpose or "").strip()
        self.fetch_delay = max(1.0, min(10.0, fetch_delay))
        self.force_redownload = bool(force_redownload)
        self.force_convert = bool(force_convert)
        self.force_catalog = bool(force_catalog)
        self.force_citation_verify = bool(force_citation_verify)
        self.force_llm_cleanup = bool(force_llm_cleanup)
        self.force_title = bool(force_title)
        self.force_summary = bool(force_summary)
        self.force_rating = bool(force_rating)
        self.output_options = output_options or SourceOutputOptions()
        self.run_download = bool(run_download or self.force_redownload)
        self.run_convert = bool(run_convert or (self.run_download and self.output_options.include_markdown))
        self.run_citation_verify = bool(run_citation_verify or self.force_citation_verify)
        self.run_catalog = bool(run_catalog or self.force_catalog)
        self.run_llm_cleanup = bool(run_llm_cleanup or self.force_llm_cleanup)
        self.run_llm_title = bool(run_llm_title or self.force_title)
        self.run_llm_summary = bool(run_llm_summary or self.force_summary)
        self.run_llm_rating = bool(run_llm_rating or self.force_rating)
        self.project_profile_name = (project_profile_name or "").strip()
        self.project_profile_yaml = (project_profile_yaml or "").strip()
        self.target_rows = [
            _normalize_row_phase_metadata(row.model_copy(deep=True))
            for row in (target_rows or [])
        ]
        self.execution_output_dir = self.store.get_sources_output_dir(job_id)
        self.output_dir = output_dir or self.execution_output_dir
        self.writes_to_repository = bool(writes_to_repository)
        self.repository_path = (repository_path or "").strip()
        self.selected_scope = (selected_scope or "").strip()
        self.selected_import_id = (selected_import_id or "").strip()
        self.selected_phases = [
            normalized_phase
            for phase in (selected_phases or [])
            if (normalized_phase := _normalize_phase_name(phase))
            in {
                PHASE_FETCH,
                PHASE_CONVERT,
                PHASE_CLEANUP,
                PHASE_TITLE,
                PHASE_CATALOG,
                PHASE_CITATION_VERIFY,
                PHASE_SUMMARY,
                PHASE_RATING,
            }
        ]
        self.row_persist_callback = row_persist_callback
        self.logs_dir = self.execution_output_dir / "logs"
        self.log_file = self.logs_dir / "source_download.jsonl"
        self.status: SourceDownloadStatus | None = None
        self._status_items: dict[str, SourceItemStatus] = {}
        self.duplicate_urls_removed = 0
        self._cancel_event = threading.Event()
        self._llm_client: UnifiedLLMClient | None = None
        self._incremental_save_counter: int = 0
        self._phase_states: dict[str, SourcePhaseMetadata] = {}
        self.runtime_capabilities = RuntimeCapabilities(
            trafilatura_available=trafilatura is not None,
            playwright_python_available=False,
            playwright_browser_available=False,
            textutil_available=check_textutil_available(),
            tesseract_available=check_tesseract_available(),
            llm_vision_enabled=False,
            runtime_notes=[],
            runtime_guidance=[],
        )
        if self.writes_to_repository:
            self.status_output_dir = "repository"
            self.status_manifest_csv = "manifest.csv"
            self.status_manifest_xlsx = "manifest.xlsx"
            self.status_bundle_file = ""
        else:
            self.status_output_dir = "output_run"
            self.status_manifest_csv = "output_run/manifest.csv"
            self.status_manifest_xlsx = "output_run/manifest.xlsx"
            self.status_bundle_file = "output_run.zip"

    def request_cancel(self) -> None:
        self._cancel_event.set()
        self._mark_status_cancelling()

    @property
    def cancel_requested(self) -> bool:
        return self._cancel_event.is_set()

    def _has_running_item(self) -> bool:
        return any(item.status == "running" for item in self._status_items.values())

    def _requested_phase_names(self) -> list[str]:
        if self.selected_phases:
            return list(self.selected_phases)
        phases: list[str] = []
        if self.run_download:
            phases.append(PHASE_FETCH)
        if self.run_convert:
            phases.append(PHASE_CONVERT)
        if self.run_llm_cleanup:
            phases.append(PHASE_CLEANUP)
        if self.run_llm_title:
            phases.append(PHASE_TITLE)
        if self.run_catalog:
            phases.append(PHASE_CATALOG)
        if self.run_citation_verify:
            phases.append(PHASE_CITATION_VERIFY)
        if self.run_llm_summary:
            phases.append(PHASE_SUMMARY)
        if self.run_llm_rating:
            phases.append(PHASE_RATING)
        return phases

    def _initial_phase_states(self) -> dict[str, SourcePhaseMetadata]:
        states: dict[str, SourcePhaseMetadata] = {}
        for phase in self._requested_phase_names():
            states[phase] = SourcePhaseMetadata(phase=phase, status="pending")
        self._phase_states = states
        return {key: value.model_copy(deep=True) for key, value in states.items()}

    def _set_phase_state(self, phase: str, status: str) -> None:
        if phase not in self._phase_states:
            self._phase_states[phase] = SourcePhaseMetadata(phase=phase, status=status)
        else:
            self._phase_states[phase].status = status
        if self.status is not None:
            self.status.phase_states = {
                key: value.model_copy(deep=True)
                for key, value in self._phase_states.items()
            }
            self._save_status()

    def _mark_status_cancelling(self) -> None:
        if not self.status:
            return
        if self.status.state in {"completed", "failed", "cancelled"}:
            return
        self.status.state = "cancelling"
        self.status.cancel_requested = True
        if not self.status.cancel_requested_at:
            self.status.cancel_requested_at = _utc_now_iso()
        self.status.stop_after_current_item = self._has_running_item()
        self.status.message = self._compose_status_message("Stop requested")
        self._save_status()

    def _effective_max_source_chars(self) -> int:
        """Return the max source chars to send to the LLM, auto-scaled or manual."""
        manual = getattr(self.llm_backend, "max_source_chars", 0)
        if manual > 0:
            return manual
        num_ctx = getattr(self.llm_backend, "num_ctx", 8192)
        # Reserve ~1500 tokens for prompts + response, ~3 chars per token
        return min((num_ctx - 1500) * 3, 60000)

    def _truncate_for_llm(self, text: str, max_chars: int) -> str:
        """Truncate text keeping start and end sections for better LLM context."""
        if len(text) <= max_chars:
            return text
        head_chars = int(max_chars * 0.6)
        tail_chars = int(max_chars * 0.3)
        separator = "\n\n[... middle section truncated for length ...]\n\n"
        return text[:head_chars] + separator + text[-tail_chars:]

    def _maybe_incremental_save(
        self,
        rows_by_id: dict[str, SourceManifestRow],
        targets: list,
    ) -> None:
        """Save an incremental manifest snapshot every 5 processed sources."""
        self._incremental_save_counter += 1
        if self._incremental_save_counter % 5 == 0:
            self._save_incremental_manifest(rows_by_id, targets)

    def _save_incremental_manifest(
        self,
        rows_by_id: dict[str, SourceManifestRow],
        targets: list,
    ) -> None:
        """Save a JSON artifact snapshot of current progress."""
        rows = [rows_by_id[t.id] for t in targets if t.id in rows_by_id]
        if not rows:
            return
        counts = _count_fetch_outcomes(rows)
        artifact = SourceManifestArtifact(
            rows=rows,
            total_urls=len(rows),
            success_count=counts["success"],
            failed_count=counts["failed"],
            partial_count=counts["partial"],
        )
        self.store.save_artifact(
            self.job_id, "06_sources_manifest", artifact.model_dump()
        )

    def run(self) -> None:
        """Execute source downloads sequentially with per-URL fault isolation."""
        try:
            if not (
                self.run_download
                or self.run_convert
                or self.run_catalog
                or self.run_citation_verify
                or self.run_llm_cleanup
                or self.run_llm_title
                or self.run_llm_summary
                or self.run_llm_rating
            ):
                raise RuntimeError("Select at least one phase to run")
            if self.run_download and not any(
                [
                    self.output_options.include_raw_file,
                    self.output_options.include_rendered_html,
                    self.output_options.include_rendered_pdf,
                    self.output_options.include_markdown,
                ]
            ):
                raise RuntimeError("Select at least one download output type")

            targets = self._build_targets()
            previous_rows = self._load_previous_rows()
            if not self.run_download and not previous_rows:
                raise RuntimeError(
                    "No existing downloaded sources found. Run download phase first."
                )
            rows_by_id = {r.id: r for r in previous_rows}
            self._ensure_output_dirs()
            self.runtime_capabilities = detect_runtime_capabilities(
                use_llm=self.use_llm,
                llm_backend=self.llm_backend,
            )
            self._initialize_status(
                targets=targets,
                runtime_capabilities=self.runtime_capabilities,
                existing_rows=previous_rows,
            )
            self._append_log(
                {
                    "event": "started",
                    "job_id": self.job_id,
                    "total_urls": len(targets),
                    "rerun_failed_only": self.rerun_failed_only,
                    "run_download": self.run_download,
                    "run_convert": self.run_convert,
                    "run_catalog": self.run_catalog,
                    "run_citation_verify": self.run_citation_verify,
                    "run_llm_cleanup": self.run_llm_cleanup,
                    "run_llm_title": self.run_llm_title,
                    "run_llm_summary": self.run_llm_summary,
                    "run_llm_rating": self.run_llm_rating,
                    "force_redownload": self.force_redownload,
                    "force_convert": self.force_convert,
                    "force_catalog": self.force_catalog,
                    "force_citation_verify": self.force_citation_verify,
                    "force_llm_cleanup": self.force_llm_cleanup,
                    "force_title": self.force_title,
                    "force_summary": self.force_summary,
                    "force_rating": self.force_rating,
                    "project_profile_name": self.project_profile_name,
                    "selected_phases": self._requested_phase_names(),
                    "output_options": self.output_options.model_dump(mode="json"),
                    "runtime_notes": self.runtime_capabilities.runtime_notes,
                    "runtime_guidance": self.runtime_capabilities.runtime_guidance,
                    "timestamp": _utc_now_iso(),
                }
            )

            # Create a shared LLM client for all LLM calls in this run
            if self.use_llm and llm_backend_ready_for_chat(self.llm_backend):
                self._llm_client = UnifiedLLMClient(self.llm_backend)

            timeout = httpx.Timeout(HTTP_TIMEOUT_SECONDS, connect=12.0)
            renderer_startup_error = ""
            if not self.runtime_capabilities.playwright_browser_available:
                renderer_startup_error = (
                    f"playwright_not_installed: run `{INSTALL_PLAYWRIGHT_BROWSER_COMMAND}`"
                )
                if not self.runtime_capabilities.playwright_python_available:
                    renderer_startup_error = (
                        f"playwright_not_installed: run `{INSTALL_BOOTSTRAP_COMMAND}`"
                    )
            renderer = PlaywrightRenderer(startup_error=renderer_startup_error)

            cancelled = False
            try:
                with httpx.Client(
                    timeout=timeout,
                    follow_redirects=True,
                    headers={
                        "User-Agent": "ResearchAssistant/0.1 (+local source downloader)",
                        "Accept": "*/*",
                    },
                ) as client:
                    last_idx = len(targets) - 1
                    for idx, target in enumerate(targets):
                        if self._cancel_event.is_set():
                            cancelled = True
                            break
                        existing_row = rows_by_id.get(target.id)

                        if self.rerun_failed_only and self._should_skip_successful_target(
                            target, rows_by_id
                        ):
                            if self._should_run_llm_postprocess(existing_row):
                                self._mark_item_running(target)
                                row = self._process_existing_row(target, existing_row)
                                rows_by_id[row.id] = row
                                self._mark_item_finished(row)
                                self._maybe_incremental_save(rows_by_id, targets)
                            else:
                                self._mark_item_skipped(target.id, existing_row)
                            continue

                        if self.run_download:
                            if self._should_skip_download_target(existing_row):
                                if self._should_run_llm_postprocess(existing_row):
                                    self._mark_item_running(target)
                                    row = self._process_existing_row(target, existing_row)
                                    rows_by_id[row.id] = row
                                    self._mark_item_finished(row)
                                    self._maybe_incremental_save(rows_by_id, targets)
                                else:
                                    self._mark_item_skipped(target.id, existing_row)
                                continue

                            self._mark_item_running(target)
                            row = self._process_target(
                                target=target,
                                client=client,
                                renderer=renderer,
                                existing_row=existing_row,
                            )
                            rows_by_id[row.id] = row
                            self._persist_sink_row(row)
                            self._mark_item_finished(row)
                            self._maybe_incremental_save(rows_by_id, targets)
                        else:
                            if existing_row is None:
                                self._mark_item_skipped(target.id, existing_row)
                                continue
                            self._mark_item_running(target)
                            row = self._process_existing_row(target, existing_row)
                            rows_by_id[row.id] = row
                            self._persist_sink_row(row)
                            self._mark_item_finished(row)
                            self._maybe_incremental_save(rows_by_id, targets)

                        if idx < last_idx and not self._cancel_event.is_set():
                            self._cancel_event.wait(self.fetch_delay)
            finally:
                renderer.close()
                if self._llm_client is not None:
                    self._llm_client.sync_close()
                    self._llm_client = None

            final_rows = [rows_by_id[t.id] for t in targets if t.id in rows_by_id]
            counts = _count_fetch_outcomes(final_rows)
            artifact = SourceManifestArtifact(
                rows=final_rows,
                total_urls=len(final_rows),
                success_count=counts["success"],
                failed_count=counts["failed"],
                partial_count=counts["partial"],
            )
            output_summary = summarize_output_rows(final_rows)

            csv_content = build_manifest_csv(final_rows, base_dir=self.output_dir)
            xlsx_bytes = build_manifest_xlsx(final_rows, base_dir=self.output_dir)
            self.store.save_sources_manifest_csv(self.job_id, csv_content)
            self.store.save_sources_manifest_xlsx(self.job_id, xlsx_bytes)
            self.store.save_artifact(
                self.job_id, "06_sources_manifest", artifact.model_dump()
            )
            bundle_path = (
                self.store.build_sources_bundle(self.job_id)
                if not self.writes_to_repository
                else None
            )

            if cancelled:
                self._mark_status_cancelled(artifact, bundle_path, output_summary)
                self._append_log(
                    {
                        "event": "cancelled",
                        "job_id": self.job_id,
                        "total_urls": artifact.total_urls,
                        "success_count": artifact.success_count,
                        "failed_count": artifact.failed_count,
                        "partial_count": artifact.partial_count,
                        "output_summary": output_summary.model_dump(mode="json"),
                        "bundle_file": str(bundle_path) if bundle_path else "",
                        "timestamp": _utc_now_iso(),
                    }
                )
            else:
                self._mark_status_completed(artifact, bundle_path, output_summary)
                self._append_log(
                    {
                        "event": "completed",
                        "job_id": self.job_id,
                        "total_urls": artifact.total_urls,
                        "success_count": artifact.success_count,
                        "failed_count": artifact.failed_count,
                        "partial_count": artifact.partial_count,
                        "output_summary": output_summary.model_dump(mode="json"),
                        "bundle_file": str(bundle_path) if bundle_path else "",
                        "timestamp": _utc_now_iso(),
                    }
                )
        except Exception as exc:
            logger.exception("Source download failed for job %s", self.job_id)
            self._mark_status_failed(f"{type(exc).__name__}: {exc}")
            self._append_log(
                {
                    "event": "failed",
                    "job_id": self.job_id,
                    "error": f"{type(exc).__name__}: {exc}",
                    "timestamp": _utc_now_iso(),
                }
            )

    def _build_targets(self) -> list[SourceTarget]:
        if self.target_rows:
            return [
                SourceTarget(
                    id=row.id,
                    source_document_name=row.source_document_name,
                    citation_number=row.citation_number,
                    original_url=row.original_url or row.final_url,
                )
                for row in self.target_rows
            ]

        bib = self.store.load_artifact(self.job_id, "03_bibliography")
        if bib is None:
            raise RuntimeError("Bibliography artifact not found")

        ingestion = self.store.load_artifact(self.job_id, "01_ingestion") or {}
        docs = ingestion.get("documents", [])
        source_doc_name = docs[0].get("filename", "") if docs else ""

        deduped_targets: list[SourceTarget] = []
        seen_keys: set[str] = set()
        duplicate_count = 0
        row_num = 0
        for entry in bib.get("entries", []):
            url = clean_url_candidate(str(entry.get("url") or ""))
            doi = clean_url_candidate(str(entry.get("doi") or ""))
            if not url and doi:
                url = f"https://doi.org/{doi}"
            if not url:
                continue

            normalized_url, _ = normalize_url(url)
            url = normalized_url or url

            dedupe_key = dedupe_url_key(url)
            if dedupe_key in seen_keys:
                duplicate_count += 1
                continue
            seen_keys.add(dedupe_key)

            row_num += 1
            entry_source_doc = str(entry.get("source_document_name") or "").strip()
            deduped_targets.append(
                SourceTarget(
                    id=f"{row_num:06d}",
                    source_document_name=entry_source_doc or source_doc_name,
                    citation_number=str(entry.get("ref_number") or ""),
                    original_url=url,
                )
            )

        self.duplicate_urls_removed = duplicate_count

        if not deduped_targets:
            raise RuntimeError("No citation URLs found in bibliography entries")
        return deduped_targets

    def _load_previous_rows(self) -> list[SourceManifestRow]:
        if self.target_rows:
            return [row.model_copy(deep=True) for row in self.target_rows]
        previous = self.store.load_artifact(self.job_id, "06_sources_manifest")
        if not previous:
            return []
        rows: list[SourceManifestRow] = []
        for raw in previous.get("rows", []):
            try:
                rows.append(_normalize_row_phase_metadata(SourceManifestRow.model_validate(raw)))
            except Exception:
                continue
        return rows

    def _initialize_status(
        self,
        targets: list[SourceTarget],
        runtime_capabilities: RuntimeCapabilities,
        existing_rows: list[SourceManifestRow],
    ) -> None:
        existing_by_id = {row.id: row for row in existing_rows}
        items = [
            SourceItemStatus(
                id=t.id,
                original_url=t.original_url,
                citation_number=t.citation_number,
                source_kind=existing_by_id.get(t.id, SourceManifestRow(id=t.id)).source_kind,
                citation_verification_status=_row_citation_verification_status(
                    existing_by_id.get(t.id),
                    self.output_dir,
                ),
            )
            for t in targets
        ]
        phase_bits: list[str] = []
        if self.run_download:
            phase_bits.append("download")
        if self.run_convert:
            phase_bits.append("convert")
        if self.run_llm_cleanup:
            phase_bits.append("cleanup")
        if self.run_llm_title:
            phase_bits.append("title")
        if self.run_catalog:
            phase_bits.append("catalog")
        if self.run_citation_verify:
            phase_bits.append("citation verify")
        if self.run_llm_summary:
            phase_bits.append("summary")
        if self.run_llm_rating:
            phase_bits.append("rating")
        phase_text = ", ".join(phase_bits) if phase_bits else "none"
        cancel_requested = self._cancel_event.is_set()
        self._status_items = {i.id: i for i in items}
        self.status = SourceDownloadStatus(
            job_id=self.job_id,
            state="cancelling" if cancel_requested else "running",
            total_urls=len(targets),
            processed_urls=0,
            success_count=0,
            failed_count=0,
            partial_count=0,
            skipped_count=0,
            duplicate_urls_removed=self.duplicate_urls_removed,
            started_at=_utc_now_iso(),
            completed_at=None,
            cancel_requested=cancel_requested,
            cancel_requested_at=_utc_now_iso() if cancel_requested else None,
            stop_after_current_item=False,
            current_item_id="",
            current_url="",
            message="",
            runtime_notes=runtime_capabilities.runtime_notes,
            runtime_guidance=runtime_capabilities.runtime_guidance,
            rerun_failed_only=self.rerun_failed_only,
            run_download=self.run_download,
            run_convert=self.run_convert,
            run_catalog=self.run_catalog,
            run_citation_verify=self.run_citation_verify,
            run_llm_cleanup=self.run_llm_cleanup,
            run_llm_title=self.run_llm_title,
            run_llm_summary=self.run_llm_summary,
            run_llm_rating=self.run_llm_rating,
            force_redownload=self.force_redownload,
            force_convert=self.force_convert,
            force_catalog=self.force_catalog,
            force_citation_verify=self.force_citation_verify,
            force_llm_cleanup=self.force_llm_cleanup,
            force_title=self.force_title,
            force_summary=self.force_summary,
            force_rating=self.force_rating,
            project_profile_name=self.project_profile_name,
            output_options=self.output_options,
            output_summary=summarize_output_rows(existing_rows),
            output_dir=self.status_output_dir,
            manifest_csv=self.status_manifest_csv,
            manifest_xlsx=self.status_manifest_xlsx,
            bundle_file=self.status_bundle_file,
            writes_to_repository=self.writes_to_repository,
            repository_path=self.repository_path,
            selected_scope=self.selected_scope,
            selected_import_id=self.selected_import_id,
            selected_phases=self._requested_phase_names(),
            items=items,
        )
        self.status.phase_states = self._initial_phase_states()
        self.status.message = self._compose_status_message(
            "Stop requested" if cancel_requested else f"Running phases: {phase_text}"
        )
        self._save_status()

    def _ensure_output_dirs(self) -> None:
        self.execution_output_dir.mkdir(parents=True, exist_ok=True)
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        if self.writes_to_repository:
            return
        for sub in ["originals", "rendered", "markdown", "summaries", "ratings", "metadata", "logs"]:
            (self.output_dir / sub).mkdir(parents=True, exist_ok=True)

    def _compose_status_message(self, message: str) -> str:
        details: list[str] = []
        if (
            self.status is not None
            and self.status.cancel_requested
            and self.status.state == "cancelling"
        ):
            if self.status.stop_after_current_item:
                details.append("finishing current item before stopping")
            else:
                details.append("stopping before the next item")
        if self.duplicate_urls_removed > 0:
            details.append(f"removed {self.duplicate_urls_removed} duplicate URLs")
        if self.runtime_capabilities.runtime_notes:
            details.append(
                f"{len(self.runtime_capabilities.runtime_notes)} runtime warnings"
            )
        if not details:
            return message
        return f"{message} | {' | '.join(details)}"

    def _save_status(self) -> None:
        if not self.status:
            return
        self.store.save_source_status(self.job_id, self.status.model_dump(mode="json"))

    def _append_log(self, event: dict) -> None:
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        with self.log_file.open("a", encoding="utf-8") as f:
            f.write(json.dumps(event, ensure_ascii=False) + "\n")

    def _persist_sink_row(self, row: SourceManifestRow) -> None:
        if self.row_persist_callback is None:
            return
        self.row_persist_callback(row.model_copy(deep=True))

    def _begin_row_phase(
        self,
        row: SourceManifestRow,
        phase: str,
        *,
        model: str = "",
        profile_name: str = "",
        prompt_version: str = "",
    ) -> None:
        canonical = _normalize_phase_name(phase)
        metadata = _get_phase_metadata(row, canonical) or SourcePhaseMetadata(phase=canonical)
        metadata.phase = canonical
        metadata.status = "running"
        metadata.error = ""
        metadata.error_code = ""
        metadata.started_at = _utc_now_iso()
        metadata.completed_at = ""
        if model:
            metadata.model = model
        if profile_name:
            metadata.profile_name = profile_name
        if prompt_version:
            metadata.prompt_version = prompt_version
        _set_phase_metadata(row, canonical, metadata)

    def _complete_row_phase(
        self,
        row: SourceManifestRow,
        phase: str,
        *,
        status: str,
        content_digest: str = "",
        error: str = "",
        error_code: str = "",
        stale: bool = False,
        model: str = "",
        profile_name: str = "",
        prompt_version: str = "",
    ) -> None:
        canonical = _normalize_phase_name(phase)
        metadata = _get_phase_metadata(row, canonical) or SourcePhaseMetadata(phase=canonical)
        metadata.phase = canonical
        metadata.status = status
        metadata.error = error
        metadata.error_code = error_code
        if not metadata.started_at:
            metadata.started_at = _utc_now_iso()
        metadata.completed_at = _utc_now_iso()
        metadata.stale = bool(stale)
        if content_digest:
            metadata.content_digest = content_digest
        if model:
            metadata.model = model
        if profile_name:
            metadata.profile_name = profile_name
        if prompt_version:
            metadata.prompt_version = prompt_version
        _set_phase_metadata(row, canonical, metadata)

    def _mark_downstream_stale(self, row: SourceManifestRow, markdown_digest: str) -> None:
        for phase, status_field in (
            (PHASE_CLEANUP, "llm_cleanup_status"),
            (PHASE_TITLE, "title_status"),
            (PHASE_CATALOG, "catalog_status"),
            (PHASE_CITATION_VERIFY, ""),
            (PHASE_SUMMARY, "summary_status"),
            (PHASE_RATING, "rating_status"),
        ):
            metadata = _get_phase_metadata(row, phase)
            if metadata is None:
                continue
            if not metadata.content_digest or metadata.content_digest == markdown_digest:
                metadata.stale = False
                _set_phase_metadata(row, phase, metadata)
                continue
            metadata.status = "stale"
            metadata.stale = True
            metadata.completed_at = _utc_now_iso()
            _set_phase_metadata(row, phase, metadata)
            if status_field:
                setattr(row, status_field, "stale")

    def _effective_markdown_rel_path(self, row: SourceManifestRow) -> str:
        cleanup_metadata = _get_phase_metadata(row, PHASE_CLEANUP)
        cleanup_is_current = not (
            cleanup_metadata is not None
            and (cleanup_metadata.stale or str(cleanup_metadata.status or "").strip().lower() == "stale")
        )
        if (
            cleanup_is_current
            and row.llm_cleanup_file
            and _has_output_file(self.output_dir, row.llm_cleanup_file)
        ):
            return row.llm_cleanup_file
        return row.markdown_file

    def _effective_markdown_digest(self, row: SourceManifestRow) -> str:
        rel_path = self._effective_markdown_rel_path(row)
        if not rel_path:
            return ""
        text = self._read_text(Path(rel_path))
        if not text.strip():
            return ""
        return hashlib.sha256(text.encode("utf-8")).hexdigest()

    def _source_file_rel(self, row: SourceManifestRow, filename: str, subdir: str) -> Path:
        if self.writes_to_repository:
            return Path("sources") / row.id / filename
        return Path(subdir) / filename

    def _raw_file_rel(self, row: SourceManifestRow, suffix: str) -> Path:
        return self._source_file_rel(row, f"{row.id}_source{suffix}", "originals")

    def _rendered_html_rel(self, row: SourceManifestRow) -> Path:
        return self._source_file_rel(row, f"{row.id}_rendered.html", "rendered")

    def _rendered_pdf_rel(self, row: SourceManifestRow) -> Path:
        return self._source_file_rel(row, f"{row.id}_rendered.pdf", "rendered")

    def _markdown_rel(self, row: SourceManifestRow) -> Path:
        return self._source_file_rel(row, f"{row.id}_clean.md", "markdown")

    def _llm_cleanup_rel(self, row: SourceManifestRow) -> Path:
        return self._source_file_rel(row, f"{row.id}_llm_clean.md", "markdown")

    def _catalog_rel(self, row: SourceManifestRow) -> Path:
        return self._source_file_rel(row, f"{row.id}_catalog.json", "metadata")

    def _summary_rel(self, row: SourceManifestRow) -> Path:
        return self._source_file_rel(row, f"{row.id}_summary.md", "summaries")

    def _rating_rel(self, row: SourceManifestRow) -> Path:
        return self._source_file_rel(row, f"{row.id}_rating.json", "ratings")

    def _metadata_rel(self, row: SourceManifestRow) -> Path:
        if self.writes_to_repository:
            return Path("sources") / row.id / f"{row.id}_metadata.json"
        return Path("metadata") / f"{row.id}.json"

    def _should_skip_successful_target(
        self, target: SourceTarget, rows_by_id: dict[str, SourceManifestRow]
    ) -> bool:
        previous = rows_by_id.get(target.id)
        return bool(previous and previous.fetch_status == "success")

    def _should_skip_download_target(self, existing_row: SourceManifestRow | None) -> bool:
        if not self.run_download:
            return True
        if existing_row is None:
            return False
        if self.force_redownload:
            return False
        if existing_row.fetch_status == "failed":
            return False
        if existing_row.fetch_status in {"", "queued"}:
            return False

        checks: list[bool] = []
        if self.output_options.include_raw_file:
            checks.append(bool(existing_row.raw_file))
        if self.output_options.include_rendered_html:
            checks.append(bool(existing_row.rendered_file))
        if self.output_options.include_rendered_pdf:
            checks.append(bool(existing_row.rendered_pdf_file))
        if self.run_convert:
            checks.append(bool(existing_row.markdown_file))

        if not checks:
            return True
        return all(checks)

    def _should_run_llm_postprocess(self, existing_row: SourceManifestRow | None) -> bool:
        if existing_row is None:
            return False
        return (
            self.run_convert
            or self.run_catalog
            or self.run_citation_verify
            or self.run_llm_cleanup
            or self.run_llm_title
            or self.run_llm_summary
            or self.run_llm_rating
        )

    def _process_existing_row(
        self,
        target: SourceTarget,
        existing_row: SourceManifestRow,
    ) -> SourceManifestRow:
        row = existing_row.model_copy(deep=True)
        row.id = target.id
        row.source_document_name = target.source_document_name or row.source_document_name
        row.citation_number = target.citation_number or row.citation_number
        row.original_url = target.original_url or row.original_url
        notes = parse_notes(row.notes)
        event: dict = {
            "event": "source_postprocess",
            "job_id": self.job_id,
            "id": row.id,
            "original_url": row.original_url,
            "started_at": _utc_now_iso(),
        }
        if self.run_convert:
            self._set_phase_state(PHASE_CONVERT, "running")
        if self.run_llm_cleanup:
            self._set_phase_state(PHASE_CLEANUP, "running")
        if self.run_llm_title:
            self._set_phase_state(PHASE_TITLE, "running")
        if self.run_catalog:
            self._set_phase_state(PHASE_CATALOG, "running")
        if self.run_citation_verify:
            self._set_phase_state(PHASE_CITATION_VERIFY, "running")
        if self.run_llm_summary:
            self._set_phase_state(PHASE_SUMMARY, "running")
        if self.run_llm_rating:
            self._set_phase_state(PHASE_RATING, "running")
        return self._finalize_row(
            row=row,
            notes=notes,
            event=event,
            update_fetched_at=False,
        )

    def _mark_item_skipped(
        self,
        item_id: str,
        existing_row: SourceManifestRow | None,
    ) -> None:
        if not self.status:
            return
        item = self._status_items.get(item_id)
        if item:
            item.status = "skipped"
            item.fetch_status = existing_row.fetch_status if existing_row else "skipped"
            item.catalog_status = (
                existing_row.catalog_status if existing_row else item.catalog_status
            )
            item.citation_verification_status = (
                _row_citation_verification_status(existing_row, self.output_dir)
                if existing_row
                else item.citation_verification_status
            )
            item.title_status = (
                existing_row.title_status if existing_row else item.title_status
            )
            item.llm_cleanup_status = (
                existing_row.llm_cleanup_status if existing_row else item.llm_cleanup_status
            )
            item.summary_status = (
                existing_row.summary_status if existing_row else item.summary_status
            )
            item.rating_status = (
                existing_row.rating_status if existing_row else item.rating_status
            )
        self.status.skipped_count += 1
        self.status.processed_urls += 1
        self.status.current_item_id = item_id
        self.status.current_url = item.original_url if item else ""
        self.status.message = self._compose_status_message("Skipped (already complete)")
        self._save_status()

    def _mark_item_running(self, target: SourceTarget) -> None:
        if not self.status:
            return
        item = self._status_items.get(target.id)
        if item:
            item.status = "running"
            item.error_message = ""
        self.status.current_item_id = target.id
        self.status.current_url = target.original_url
        self.status.message = self._compose_status_message(f"Processing {target.id}")
        self._save_status()

    def _mark_item_finished(self, row: SourceManifestRow) -> None:
        if not self.status:
            return
        item = self._status_items.get(row.id)
        if item:
            item.fetch_status = row.fetch_status
            item.catalog_status = row.catalog_status
            item.citation_verification_status = _row_citation_verification_status(
                row, self.output_dir
            )
            item.title_status = row.title_status
            item.llm_cleanup_status = row.llm_cleanup_status
            item.summary_status = row.summary_status
            item.rating_status = row.rating_status
            item.error_message = row.error_message
            overall = _row_task_outcome(row, self._requested_phase_names())
            if overall == "failed":
                item.status = "failed"
            elif overall == "partial":
                item.status = "completed"
            elif overall == "skipped":
                item.status = "skipped"
            else:
                item.status = "completed"

        self.status.processed_urls += 1
        overall = _row_task_outcome(row, self._requested_phase_names())
        if overall == "success":
            self.status.success_count += 1
        elif overall == "partial":
            self.status.partial_count += 1
        elif overall == "skipped":
            self.status.skipped_count += 1
        else:
            self.status.failed_count += 1
        self.status.message = self._compose_status_message(
            f"Processed {self.status.processed_urls}/{self.status.total_urls} URLs"
        )
        self._save_status()

    def _mark_status_failed(self, error_message: str) -> None:
        if not self.status:
            self.status = SourceDownloadStatus(
                job_id=self.job_id,
                state="failed",
                message=self._compose_status_message(error_message),
                runtime_notes=self.runtime_capabilities.runtime_notes,
                runtime_guidance=self.runtime_capabilities.runtime_guidance,
                completed_at=_utc_now_iso(),
                project_profile_name=self.project_profile_name,
                output_dir=self.status_output_dir,
                manifest_csv=self.status_manifest_csv,
                manifest_xlsx=self.status_manifest_xlsx,
                bundle_file=self.status_bundle_file,
                writes_to_repository=self.writes_to_repository,
                repository_path=self.repository_path,
                selected_scope=self.selected_scope,
                selected_import_id=self.selected_import_id,
                selected_phases=self._requested_phase_names(),
            )
        else:
            self.status.state = "failed"
            self.status.message = self._compose_status_message(error_message)
            self.status.completed_at = _utc_now_iso()
            self.status.stop_after_current_item = False
        for phase in self._requested_phase_names():
            self._phase_states.setdefault(phase, SourcePhaseMetadata(phase=phase))
            if self._phase_states[phase].status in {"pending", "running"}:
                self._phase_states[phase].status = "failed"
        self.status.phase_states = {
            key: value.model_copy(deep=True)
            for key, value in self._phase_states.items()
        }
        self._save_status()

    def _mark_status_completed(
        self,
        artifact: SourceManifestArtifact,
        bundle_path: Path | None,
        output_summary: SourceOutputSummary,
    ) -> None:
        if not self.status:
            return
        self.status.state = "completed"
        self.status.processed_urls = artifact.total_urls
        self.status.success_count = artifact.success_count
        self.status.failed_count = artifact.failed_count
        self.status.partial_count = artifact.partial_count
        self.status.completed_at = _utc_now_iso()
        self.status.stop_after_current_item = False
        self.status.current_item_id = ""
        self.status.current_url = ""
        self.status.message = self._compose_status_message(
            f"Completed: {artifact.success_count} success, "
            f"{artifact.partial_count} partial, {artifact.failed_count} failed"
        )
        self.status.bundle_file = bundle_path.name if bundle_path else ""
        self.status.output_summary = output_summary
        for phase in self._requested_phase_names():
            self._phase_states.setdefault(phase, SourcePhaseMetadata(phase=phase))
            if self._phase_states[phase].status in {"pending", "running"}:
                self._phase_states[phase].status = "completed"
        self.status.phase_states = {
            key: value.model_copy(deep=True)
            for key, value in self._phase_states.items()
        }
        self._save_status()

    def _mark_status_cancelled(
        self,
        artifact: SourceManifestArtifact,
        bundle_path: Path | None,
        output_summary: SourceOutputSummary,
    ) -> None:
        if not self.status:
            return

        for item in self.status.items:
            if item.status in {"pending", "running"}:
                item.status = "cancelled"
                if not item.fetch_status:
                    item.fetch_status = "cancelled"

        self.status.state = "cancelled"
        self.status.cancel_requested = True
        if not self.status.cancel_requested_at:
            self.status.cancel_requested_at = _utc_now_iso()
        self.status.completed_at = _utc_now_iso()
        self.status.current_item_id = ""
        self.status.current_url = ""
        self.status.bundle_file = bundle_path.name if bundle_path else ""
        self.status.success_count = artifact.success_count
        self.status.failed_count = artifact.failed_count
        self.status.partial_count = artifact.partial_count
        self.status.processed_urls = min(self.status.processed_urls, self.status.total_urls)
        detail = (
            "Stopped after the current item"
            if self.status.stop_after_current_item
            else "Stopped before the next item"
        )
        self.status.message = self._compose_status_message(
            f"{detail}: {self.status.processed_urls}/{self.status.total_urls} URLs processed"
        )
        self.status.stop_after_current_item = False
        self.status.output_summary = output_summary
        for phase in self._requested_phase_names():
            self._phase_states.setdefault(phase, SourcePhaseMetadata(phase=phase))
            if self._phase_states[phase].status in {"pending", "running"}:
                self._phase_states[phase].status = "cancelled"
        self.status.phase_states = {
            key: value.model_copy(deep=True)
            for key, value in self._phase_states.items()
        }
        self._save_status()

    def _process_target(
        self,
        target: SourceTarget,
        client: httpx.Client,
        renderer: PlaywrightRenderer,
        existing_row: SourceManifestRow | None = None,
    ) -> SourceManifestRow:
        if existing_row is not None:
            row = existing_row.model_copy(deep=True)
            row.id = target.id
            row.source_document_name = target.source_document_name or row.source_document_name
            row.citation_number = target.citation_number or row.citation_number
            row.original_url = target.original_url
        else:
            row = SourceManifestRow(
                id=target.id,
                source_document_name=target.source_document_name,
                citation_number=target.citation_number,
                original_url=target.original_url,
            )
        notes: list[str] = parse_notes(row.notes)
        self._set_phase_state(PHASE_FETCH, "running")
        self._begin_row_phase(row, PHASE_FETCH)
        event: dict = {
            "event": "source_processed",
            "job_id": self.job_id,
            "id": target.id,
            "original_url": target.original_url,
            "started_at": _utc_now_iso(),
        }

        if existing_row is not None and existing_row.source_kind == "uploaded_document":
            row.fetch_status = "not_applicable"
            row.error_message = ""
            if not row.notes:
                notes.append("local_document")
            self._complete_row_phase(
                row,
                PHASE_FETCH,
                status="skipped",
                content_digest=row.sha256,
                error="not_applicable: uploaded repository document",
                error_code="not_applicable",
            )
            if self.run_convert:
                self._set_phase_state(PHASE_CONVERT, "running")
            if self.run_llm_cleanup:
                self._set_phase_state(PHASE_CLEANUP, "running")
            if self.run_llm_title:
                self._set_phase_state(PHASE_TITLE, "running")
            if self.run_catalog:
                self._set_phase_state(PHASE_CATALOG, "running")
            if self.run_citation_verify:
                self._set_phase_state(PHASE_CITATION_VERIFY, "running")
            if self.run_llm_summary:
                self._set_phase_state(PHASE_SUMMARY, "running")
            if self.run_llm_rating:
                self._set_phase_state(PHASE_RATING, "running")
            return self._finalize_row(row, notes, event, update_fetched_at=False)

        normalized_url, url_error = normalize_url(target.original_url)
        if url_error:
            row.fetch_status = "failed"
            row.error_message = f"invalid_url: {url_error}"
            notes.append("invalid_url")
            self._complete_row_phase(
                row,
                PHASE_FETCH,
                status="failed",
                error=row.error_message,
                error_code="invalid_url",
            )
            return self._finalize_row(row, notes, event)

        try:
            response = client.get(normalized_url)
        except httpx.TimeoutException as exc:
            row.fetch_status = "failed"
            row.final_url = normalized_url
            row.error_message = f"timeout: {exc}"
            notes.append("timeout")
            self._complete_row_phase(
                row,
                PHASE_FETCH,
                status="failed",
                error=row.error_message,
                error_code="timeout",
            )
            return self._finalize_row(row, notes, event)
        except httpx.RequestError as exc:
            row.fetch_status = "failed"
            row.final_url = normalized_url
            row.error_message = f"network_failure: {type(exc).__name__}: {exc}"
            notes.append("network_failure")
            self._complete_row_phase(
                row,
                PHASE_FETCH,
                status="failed",
                error=row.error_message,
                error_code="network_failure",
            )
            return self._finalize_row(row, notes, event)

        row.final_url = str(response.url)
        row.http_status = response.status_code
        row.content_type = response.headers.get("content-type", "")
        row.detected_type = detect_source_type(
            content_type=row.content_type,
            final_url=row.final_url,
            body=response.content,
        )
        row.fetch_method = "http"
        row.error_message = ""

        if row.detected_type == "pdf":
            self._handle_pdf_response(row, response, notes)
        elif row.detected_type == "html":
            self._handle_html_response(row, response, normalized_url, renderer, notes)
        elif row.detected_type == "document":
            self._handle_document_response(row, response, notes)
        else:
            self._handle_unsupported_response(row, response, notes)

        fetch_error_code = _phase_error_code(row.error_message)
        fetch_status = "completed"
        fetch_error = ""
        if (
            (row.http_status or 0) >= 400
            or fetch_error_code in {
                "invalid_url",
                "timeout",
                "network_failure",
                "blocked_request",
                "unsupported_content",
            }
        ):
            fetch_status = "failed"
            fetch_error = row.error_message
        self._complete_row_phase(
            row,
            PHASE_FETCH,
            status=fetch_status,
            content_digest=row.sha256,
            error=fetch_error,
            error_code=fetch_error_code if fetch_status == "failed" else "",
        )
        if self.run_convert:
            self._set_phase_state(PHASE_CONVERT, "running")
        if self.run_llm_cleanup:
            self._set_phase_state(PHASE_CLEANUP, "running")
        if self.run_llm_title:
            self._set_phase_state(PHASE_TITLE, "running")
        if self.run_catalog:
            self._set_phase_state(PHASE_CATALOG, "running")
        if self.run_citation_verify:
            self._set_phase_state(PHASE_CITATION_VERIFY, "running")
        if self.run_llm_summary:
            self._set_phase_state(PHASE_SUMMARY, "running")
        if self.run_llm_rating:
            self._set_phase_state(PHASE_RATING, "running")

        return self._finalize_row(row, notes, event)

    def _handle_pdf_response(
        self,
        row: SourceManifestRow,
        response: httpx.Response,
        notes: list[str],
    ) -> None:
        if self.output_options.include_raw_file:
            rel_path = self._raw_file_rel(row, ".pdf")
            row.raw_file = rel_path.as_posix()
            self._write_binary(rel_path, response.content)
        row.sha256 = hashlib.sha256(response.content).hexdigest()
        if response.status_code >= 400:
            row.fetch_status = "failed"
            reason = classify_http_status(response.status_code)
            notes.append(reason)
            row.error_message = f"{reason}: http_status_{response.status_code}"
        else:
            if not self.run_convert:
                row.fetch_status = "success"
                return

            markdown_text, extraction_method, conversion_notes = self._convert_pdf_to_markdown(
                response.content
            )
            notes.extend(conversion_notes)

            if markdown_text:
                markdown_rel = self._markdown_rel(row)
                self._write_text(markdown_rel, markdown_text)
                row.markdown_file = markdown_rel.as_posix()
                row.markdown_char_count = len(markdown_text)
                row.extraction_method = extraction_method
                row.fetch_status = "success"
            else:
                row.fetch_status = "partial"
                row.error_message = "extraction_failure: markdown not generated"
                notes.append(NOTE_EXTRACTION_FAILURE)

    def _handle_html_response(
        self,
        row: SourceManifestRow,
        response: httpx.Response,
        normalized_url: str,
        renderer: PlaywrightRenderer,
        notes: list[str],
    ) -> None:
        raw_html = decode_html(response)
        if raw_html:
            if self.output_options.include_raw_file:
                raw_rel = self._raw_file_rel(row, ".html")
                self._write_text(raw_rel, raw_html)
                row.raw_file = raw_rel.as_posix()
            row.sha256 = hashlib.sha256(raw_html.encode("utf-8")).hexdigest()
            row.title = extract_title(raw_html)
            row.canonical_url = extract_canonical_url(raw_html)

        if self.output_options.include_rendered_pdf:
            self._cancel_event.wait(self.fetch_delay)
            rendered_pdf, rendered_pdf_error, rendered_pdf_notes = renderer.capture_visual_pdf(
                normalized_url
            )
            notes.extend(rendered_pdf_notes)
            if rendered_pdf:
                rendered_pdf_rel = self._rendered_pdf_rel(row)
                self._write_binary(rendered_pdf_rel, rendered_pdf)
                row.rendered_pdf_file = rendered_pdf_rel.as_posix()
            elif rendered_pdf_error:
                notes.append(NOTE_VISUAL_CAPTURE_FAILED)
                notes.append(normalize_render_error_note(rendered_pdf_error))

        blocked_by_challenge = detect_blocked_page(
            html_text=raw_html,
            title=row.title,
            final_url=row.final_url,
        )

        raw_markdown = ""
        raw_used_fallback = False
        raw_score = 0
        if self.run_convert:
            raw_markdown, raw_used_fallback, raw_notes = extract_markdown_with_fallback(
                raw_html,
                self.runtime_capabilities,
            )
            notes.extend(raw_notes)
            raw_score = markdown_score(raw_markdown)

        rendered_html = ""
        rendered_markdown = ""
        rendered_used_fallback = False
        rendered_score = 0

        should_render = False
        if not blocked_by_challenge:
            should_render = self.output_options.include_rendered_html or (
                self.run_convert
                and (response.status_code >= 400 or raw_score < MIN_MARKDOWN_SCORE)
            )
        if should_render:
            self._cancel_event.wait(self.fetch_delay)
            rendered_html, render_error = renderer.render(normalized_url)
            if render_error:
                notes.append(normalize_render_error_note(render_error))
            elif rendered_html:
                if self.output_options.include_rendered_html:
                    rendered_rel = self._rendered_html_rel(row)
                    self._write_text(rendered_rel, rendered_html)
                    row.rendered_file = rendered_rel.as_posix()

                if not row.title:
                    row.title = extract_title(rendered_html)
                if not row.canonical_url:
                    row.canonical_url = extract_canonical_url(rendered_html)

                if self.run_convert:
                    rendered_markdown, rendered_used_fallback, rendered_notes = (
                        extract_markdown_with_fallback(
                            rendered_html,
                            self.runtime_capabilities,
                        )
                    )
                    notes.extend(rendered_notes)
                    rendered_score = markdown_score(rendered_markdown)

        if self.run_convert:
            markdown_to_write = raw_markdown
            used_fallback = raw_used_fallback
            extraction_method = "raw_html" if raw_markdown else ""
            if rendered_score > raw_score:
                markdown_to_write = rendered_markdown
                used_fallback = rendered_used_fallback
                extraction_method = "rendered_html"
                row.fetch_method = "playwright"

            if markdown_to_write:
                markdown_rel = self._markdown_rel(row)
                self._write_text(markdown_rel, markdown_to_write)
                row.markdown_file = markdown_rel.as_posix()
                row.markdown_char_count = len(markdown_to_write)
                row.extraction_method = (
                    f"{extraction_method}_fallback" if used_fallback else extraction_method
                )

        if blocked_by_challenge:
            notes.append(NOTE_BLOCKED_REQUEST)
            row.fetch_status = "failed"
            row.error_message = blocked_error_message(response.status_code)
            return

        if response.status_code >= 400:
            reason = classify_http_status(response.status_code)
            notes.append(reason)
            row.fetch_status = "failed"
            row.error_message = f"{reason}: http_status_{response.status_code}"
            return

        if not self.run_convert:
            row.fetch_status = "success"
            return

        if row.markdown_file:
            row.fetch_status = "success"
            return

        if row.raw_file:
            row.fetch_status = "partial"
            notes.append(NOTE_EXTRACTION_FAILURE)
            row.error_message = "extraction_failure: markdown not generated"
            return

        row.fetch_status = "failed"
        row.error_message = "network_failure: empty_html_response"

    def _handle_document_response(
        self,
        row: SourceManifestRow,
        response: httpx.Response,
        notes: list[str],
    ) -> None:
        extension = infer_document_extension(
            final_url=row.final_url,
            content_type=row.content_type,
            content_disposition=response.headers.get("content-disposition", ""),
        )
        if self.output_options.include_raw_file:
            rel_path = self._raw_file_rel(row, extension)
            row.raw_file = rel_path.as_posix()
            self._write_binary(rel_path, response.content)
        row.sha256 = hashlib.sha256(response.content).hexdigest()

        if response.status_code >= 400:
            row.fetch_status = "failed"
            reason = classify_http_status(response.status_code)
            notes.append(reason)
            row.error_message = f"{reason}: http_status_{response.status_code}"
            return

        notes.append("download_only")
        if not self.run_convert:
            row.fetch_status = "success"
            return

        markdown_text, extraction_method, conversion_notes = self._convert_document_to_markdown(
            extension=extension,
            binary_content=response.content,
        )
        notes.extend(conversion_notes)

        if markdown_text:
            markdown_rel = self._markdown_rel(row)
            self._write_text(markdown_rel, markdown_text)
            row.markdown_file = markdown_rel.as_posix()
            row.markdown_char_count = len(markdown_text)
            row.extraction_method = extraction_method
            row.title = row.title or first_nonempty_line(markdown_text)[:500]
            row.fetch_status = "success"
            return

        if extension in {".txt", ".md", ".docx", ".doc", ".dot", ".dotx"}:
            row.fetch_status = "partial"
            row.error_message = "extraction_failure: markdown not generated"
            notes.append(NOTE_EXTRACTION_FAILURE)
            return

        row.fetch_status = "success"

    def _handle_unsupported_response(
        self,
        row: SourceManifestRow,
        response: httpx.Response,
        notes: list[str],
    ) -> None:
        if self.output_options.include_raw_file:
            rel_path = self._raw_file_rel(row, ".bin")
            self._write_binary(rel_path, response.content)
            row.raw_file = rel_path.as_posix()
        row.sha256 = hashlib.sha256(response.content).hexdigest()
        row.fetch_status = "failed"
        notes.append("unsupported_content")
        row.error_message = (
            f"unsupported_content: content_type={row.content_type or 'unknown'}"
        )

    def _generate_markdown_from_existing_artifacts(
        self,
        row: SourceManifestRow,
        notes: list[str],
    ) -> None:
        if row.markdown_file and not self.force_convert and _has_output_file(self.output_dir, row.markdown_file):
            return

        raw_path = Path(row.raw_file) if row.raw_file else None
        raw_exists = bool(raw_path and _has_output_file(self.output_dir, row.raw_file))
        rendered_html_path = (
            Path(row.rendered_file)
            if row.rendered_file and _has_output_file(self.output_dir, row.rendered_file)
            else None
        )

        if row.detected_type == "pdf":
            if not raw_exists or raw_path is None:
                row.error_message = "convert_missing_prerequisite: raw_file_not_found"
                return
            markdown_text, extraction_method, conversion_notes = self._convert_pdf_to_markdown(
                self._read_binary(raw_path)
            )
            notes.extend(conversion_notes)
            if not markdown_text:
                row.error_message = "extraction_failure: markdown not generated"
                return
            markdown_rel = self._markdown_rel(row)
            self._write_text(markdown_rel, markdown_text)
            row.markdown_file = markdown_rel.as_posix()
            row.markdown_char_count = len(markdown_text)
            row.extraction_method = extraction_method
            return

        if row.detected_type == "document":
            if not raw_exists or raw_path is None:
                row.error_message = "convert_missing_prerequisite: raw_file_not_found"
                return
            markdown_text, extraction_method, conversion_notes = self._convert_document_to_markdown(
                extension=raw_path.suffix.lower(),
                binary_content=self._read_binary(raw_path),
            )
            notes.extend(conversion_notes)
            if not markdown_text:
                row.error_message = "extraction_failure: markdown not generated"
                return
            markdown_rel = self._markdown_rel(row)
            self._write_text(markdown_rel, markdown_text)
            row.markdown_file = markdown_rel.as_posix()
            row.markdown_char_count = len(markdown_text)
            row.extraction_method = extraction_method
            if not row.title:
                row.title = first_nonempty_line(markdown_text)[:500]
            return

        if row.detected_type == "html":
            raw_markdown = ""
            raw_used_fallback = False
            raw_score = 0
            if raw_exists and raw_path is not None:
                raw_html = self._read_text(raw_path)
                raw_markdown, raw_used_fallback, raw_notes = extract_markdown_with_fallback(
                    raw_html,
                    self.runtime_capabilities,
                )
                notes.extend(raw_notes)
                raw_score = markdown_score(raw_markdown)

            rendered_markdown = ""
            rendered_used_fallback = False
            rendered_score = 0
            if rendered_html_path is not None:
                rendered_html = self._read_text(rendered_html_path)
                rendered_markdown, rendered_used_fallback, rendered_notes = (
                    extract_markdown_with_fallback(
                        rendered_html,
                        self.runtime_capabilities,
                    )
                )
                notes.extend(rendered_notes)
                rendered_score = markdown_score(rendered_markdown)

            markdown_to_write = raw_markdown
            used_fallback = raw_used_fallback
            extraction_method = "raw_html" if raw_markdown else ""
            if rendered_score > raw_score:
                markdown_to_write = rendered_markdown
                used_fallback = rendered_used_fallback
                extraction_method = "rendered_html"

            if not markdown_to_write:
                row.error_message = "convert_missing_prerequisite: markdown_source_not_found"
                return

            markdown_rel = self._markdown_rel(row)
            self._write_text(markdown_rel, markdown_to_write)
            row.markdown_file = markdown_rel.as_posix()
            row.markdown_char_count = len(markdown_to_write)
            row.extraction_method = (
                f"{extraction_method}_fallback" if used_fallback else extraction_method
            )
            return

        if row.markdown_file and _has_output_file(self.output_dir, row.markdown_file):
            return
        row.error_message = "convert_missing_prerequisite: unsupported_or_missing_artifact"

    def _finalize_row(
        self,
        row: SourceManifestRow,
        notes: list[str],
        event: dict,
        update_fetched_at: bool = True,
    ) -> SourceManifestRow:
        if update_fetched_at or not row.fetched_at:
            row.fetched_at = _utc_now_iso()

        previous_convert_digest = ""
        convert_metadata = _get_phase_metadata(row, PHASE_CONVERT)
        if convert_metadata is not None:
            previous_convert_digest = convert_metadata.content_digest

        if self.run_convert:
            self._begin_row_phase(row, PHASE_CONVERT, prompt_version="convert.pipeline.v1")
            self._generate_markdown_from_existing_artifacts(row, notes)
            current_markdown_digest = ""
            if row.markdown_file:
                markdown_text = self._read_text(Path(row.markdown_file))
                if markdown_text.strip():
                    current_markdown_digest = hashlib.sha256(
                        markdown_text.encode("utf-8")
                    ).hexdigest()
            convert_error = ""
            convert_error_code = ""
            convert_status = "completed"
            if not current_markdown_digest:
                convert_status = "failed"
                convert_error = row.error_message or "convert_missing_prerequisite: no markdown available"
                convert_error_code = _phase_error_code(convert_error) or "convert_missing_prerequisite"
            self._complete_row_phase(
                row,
                PHASE_CONVERT,
                status=convert_status,
                content_digest=current_markdown_digest,
                error=convert_error,
                error_code=convert_error_code,
                prompt_version="convert.pipeline.v1",
            )
            if current_markdown_digest:
                self._mark_downstream_stale(row, current_markdown_digest)
            if previous_convert_digest and current_markdown_digest and previous_convert_digest == current_markdown_digest:
                current_metadata = _get_phase_metadata(row, PHASE_CONVERT)
                if current_metadata is not None:
                    current_metadata.stale = False
                    _set_phase_metadata(row, PHASE_CONVERT, current_metadata)

        if self.run_llm_cleanup:
            self._begin_row_phase(
                row,
                PHASE_CLEANUP,
                model=self.llm_backend.model,
                prompt_version=PROMPT_VERSION_CLEANUP,
            )
        self._generate_markdown_cleanup(row, notes)
        if self.run_llm_cleanup:
            cleanup_digest = self._effective_markdown_digest(row)
            cleanup_status = _phase_completion_status(
                row.llm_cleanup_status,
                success={"cleaned", "not_needed", "existing"},
                failed={"failed", "missing_markdown"},
                skipped={"not_requested"},
            )
            cleanup_error = ""
            cleanup_error_code = ""
            if cleanup_status == "failed":
                cleanup_error = (
                    row.error_message
                    or f"{row.llm_cleanup_status or 'llm_cleanup_failed'}: cleanup failed"
                )
                cleanup_error_code = _phase_error_code(cleanup_error)
            elif cleanup_status == "skipped":
                cleanup_digest = cleanup_digest or self._effective_markdown_digest(row)
            self._complete_row_phase(
                row,
                PHASE_CLEANUP,
                status=cleanup_status,
                content_digest=cleanup_digest,
                error=cleanup_error,
                error_code=cleanup_error_code,
                model=self.llm_backend.model if cleanup_status == "completed" else "",
                prompt_version=PROMPT_VERSION_CLEANUP,
            )

        if self.run_llm_title:
            self._begin_row_phase(
                row,
                PHASE_TITLE,
                model=self.llm_backend.model,
                prompt_version=PROMPT_VERSION_TITLE,
            )
        self._generate_source_title(row, notes)
        if self.run_llm_title:
            title_digest = self._effective_markdown_digest(row)
            title_phase_status = _phase_completion_status(
                row.title_status,
                success={"existing", "extracted", "generated"},
                failed={"failed", "missing_markdown"},
                skipped={"not_requested"},
            )
            title_error = ""
            title_error_code = ""
            if title_phase_status == "failed":
                title_error = row.error_message or f"{row.title_status or 'title_generation_failed'}: title generation failed"
                title_error_code = _phase_error_code(title_error)
            self._complete_row_phase(
                row,
                PHASE_TITLE,
                status=title_phase_status,
                content_digest=title_digest,
                error=title_error,
                error_code=title_error_code,
                model=self.llm_backend.model if row.title_status == "generated" else "",
                prompt_version=PROMPT_VERSION_TITLE,
            )

        self._generate_source_catalog(row, notes)
        self._generate_source_citation_verification(row, notes)
        self._generate_source_summary(row, notes)
        self._generate_source_rating(row, notes)
        row.notes = "; ".join(dict.fromkeys(n for n in notes if n))
        metadata_rel = self._metadata_rel(row)
        row.metadata_file = metadata_rel.as_posix()

        metadata_payload = {
            "id": row.id,
            "timestamp": row.fetched_at,
            "original_url": row.original_url,
            "final_url": row.final_url,
            "http_status": row.http_status,
            "content_type": row.content_type,
            "detected_type": row.detected_type,
            "fetch_method": row.fetch_method,
            "fetch_status": row.fetch_status,
            "source_kind": row.source_kind,
            "error_message": row.error_message,
            "output_files": {
                "raw_file": row.raw_file,
                "rendered_file": row.rendered_file,
                "rendered_pdf_file": row.rendered_pdf_file,
                "markdown_file": row.markdown_file,
                "llm_cleanup_file": row.llm_cleanup_file,
                "catalog_file": row.catalog_file,
                "summary_file": row.summary_file,
                "rating_file": row.rating_file,
                "metadata_file": row.metadata_file,
            },
            "notes": row.notes,
            "sha256": row.sha256,
            "title": row.title,
            "title_status": row.title_status,
            "author_names": row.author_names,
            "publication_date": row.publication_date,
            "publication_year": row.publication_year,
            "document_type": row.document_type,
            "organization_name": row.organization_name,
            "organization_type": row.organization_type,
            "canonical_url": row.canonical_url,
            "extraction_method": row.extraction_method,
            "markdown_char_count": row.markdown_char_count,
            "llm_cleanup_needed": row.llm_cleanup_needed,
            "llm_cleanup_status": row.llm_cleanup_status,
            "catalog_status": row.catalog_status,
            "summary_status": row.summary_status,
            "rating_status": row.rating_status,
            "tags_text": row.tags_text,
            "phase_metadata": {
                key: value.model_dump(mode="json")
                for key, value in row.phase_metadata.items()
            },
        }
        self._write_text(
            metadata_rel,
            json.dumps(metadata_payload, ensure_ascii=False, indent=2),
        )

        event.update(metadata_payload)
        event["completed_at"] = _utc_now_iso()
        self._append_log(event)
        return row

    def _generate_markdown_cleanup(
        self,
        row: SourceManifestRow,
        notes: list[str],
    ) -> None:
        if not self.run_llm_cleanup:
            if row.llm_cleanup_file and _has_output_file(self.output_dir, row.llm_cleanup_file):
                row.llm_cleanup_status = row.llm_cleanup_status or "existing"
            elif not row.llm_cleanup_status:
                row.llm_cleanup_status = "not_requested"
            return

        if not row.markdown_file:
            row.llm_cleanup_status = "missing_markdown"
            row.llm_cleanup_needed = False
            return
        if not self.use_llm:
            row.llm_cleanup_status = "skipped_llm_disabled"
            row.llm_cleanup_needed = False
            return
        if not llm_backend_ready_for_chat(self.llm_backend):
            row.llm_cleanup_status = "skipped_llm_not_configured"
            row.llm_cleanup_needed = False
            notes.append(NOTE_LLM_CLEANUP_SKIPPED_LLM_NOT_CONFIGURED)
            return

        if (
            row.llm_cleanup_file
            and not self.force_llm_cleanup
            and _has_output_file(self.output_dir, row.llm_cleanup_file)
        ):
            row.llm_cleanup_status = row.llm_cleanup_status or "existing"
            return

        markdown_text = self._read_text(Path(row.markdown_file))
        if not markdown_text.strip():
            row.llm_cleanup_status = "missing_markdown"
            row.llm_cleanup_needed = False
            return

        source_text = markdown_text.strip()
        max_chars = self._effective_max_source_chars()
        source_text = self._truncate_for_llm(source_text, max_chars)

        research_purpose = self.research_purpose or (
            "No explicit research purpose was provided. Preserve factual content and clarity."
        )
        user_prompt = SOURCE_MARKDOWN_CLEANUP_USER.format(
            research_purpose=research_purpose,
            source_markdown=source_text,
        )

        try:
            cleanup_response = self._llm_client.sync_chat_completion(
                system_prompt=SOURCE_MARKDOWN_CLEANUP_SYSTEM,
                user_prompt=user_prompt,
                response_format=None,
            ).strip()
            needs_cleanup, cleaned_markdown = parse_cleanup_response(cleanup_response)
            row.llm_cleanup_needed = needs_cleanup

            if needs_cleanup:
                normalized = normalize_cleaned_markdown(cleaned_markdown)
                if not normalized:
                    row.llm_cleanup_status = "failed"
                    notes.append(NOTE_LLM_CLEANUP_FAILED)
                    return
                cleanup_rel = self._llm_cleanup_rel(row)
                self._write_text(cleanup_rel, normalized)
                row.llm_cleanup_file = cleanup_rel.as_posix()
                row.llm_cleanup_status = "cleaned"
            else:
                row.llm_cleanup_status = "not_needed"
        except Exception as exc:
            row.llm_cleanup_status = "failed"
            row.llm_cleanup_needed = False
            notes.append(NOTE_LLM_CLEANUP_FAILED)
            logger.warning("Markdown cleanup failed for %s: %s", row.id, exc)

    def _base_catalog_payload(
        self,
        row: SourceManifestRow,
        existing_payload: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        payload = dict(existing_payload or {})
        payload.update(
            {
                "title": row.title,
                "title_status": row.title_status,
                "author_names": row.author_names,
                "publication_date": row.publication_date,
                "publication_year": row.publication_year,
                "document_type": row.document_type,
                "organization_name": row.organization_name,
                "organization_type": row.organization_type,
                "source_kind": row.source_kind,
            }
        )
        return payload

    def _write_catalog_payload(
        self,
        row: SourceManifestRow,
        payload: dict[str, Any],
    ) -> None:
        catalog_rel = self._catalog_rel(row)
        self._write_text(
            catalog_rel,
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        )
        row.catalog_file = catalog_rel.as_posix()

    def _generate_source_catalog(
        self,
        row: SourceManifestRow,
        notes: list[str],
    ) -> None:
        if not self.run_catalog:
            if not row.catalog_status and not str(row.catalog_file or "").strip():
                row.catalog_status = "not_requested"
            return

        source_digest = self._effective_markdown_digest(row)
        existing_catalog_payload = _load_manifest_json(self.output_dir, row.catalog_file)
        self._begin_row_phase(
            row,
            PHASE_CATALOG,
            prompt_version=PROMPT_VERSION_CATALOG,
        )
        catalog_source_path = self._effective_markdown_rel_path(row)
        if not catalog_source_path or not source_digest:
            row.catalog_status = "missing_markdown"
            self._complete_row_phase(
                row,
                PHASE_CATALOG,
                status="failed",
                error="missing_markdown: no markdown available for cataloging",
                error_code="missing_markdown",
                prompt_version=PROMPT_VERSION_CATALOG,
            )
            return

        if (
            row.catalog_file
            and not self.force_catalog
            and _has_output_file(self.output_dir, row.catalog_file)
            and (catalog_metadata := _get_phase_metadata(row, PHASE_CATALOG)) is not None
            and catalog_metadata.content_digest == source_digest
        ):
            if isinstance(existing_catalog_payload, dict):
                _merge_catalog_payload_into_row(row, existing_catalog_payload, overwrite_existing=False)
            row.catalog_status = "existing"
            self._complete_row_phase(
                row,
                PHASE_CATALOG,
                status="completed",
                content_digest=source_digest,
                prompt_version=PROMPT_VERSION_CATALOG,
            )
            return

        markdown_text = self._read_text(Path(catalog_source_path))
        if not markdown_text.strip():
            row.catalog_status = "missing_markdown"
            self._complete_row_phase(
                row,
                PHASE_CATALOG,
                status="failed",
                error="missing_markdown: empty markdown input",
                error_code="missing_markdown",
                prompt_version=PROMPT_VERSION_CATALOG,
            )
            return

        catalog_html_text = _read_catalog_html_input(row, read_text=self._read_text)
        deterministic = _build_deterministic_catalog_metadata(
            row=row,
            markdown_text=markdown_text,
            html_text=catalog_html_text,
        )
        catalog_updates = {
            key: deterministic.get(key)
            for key in (
                "author_names",
                "publication_date",
                "publication_year",
                "document_type",
                "organization_name",
                "organization_type",
            )
        }
        _merge_catalog_payload_into_row(
            row,
            catalog_updates,
            overwrite_existing=bool(self.force_catalog),
        )
        current_citation = _coerce_citation_metadata(deterministic.get("citation"))
        doi_citation = CitationMetadata()
        if current_citation.doi:
            doi_citation = _resolve_doi_citation_metadata(current_citation.doi)
            current_citation = _merge_citation_metadata(current_citation, doi_citation)
        current_citation = _merge_citation_metadata(
            current_citation,
            {
                "title": row.title,
                "authors": [
                    author.model_dump(mode="json")
                    for author in normalize_citation_authors(row.author_names)
                ],
                "issued": row.publication_date,
                "publisher": row.organization_name,
                "item_type": row.document_type,
                "url": row.original_url or row.final_url,
            },
            overwrite_existing=bool(self.force_catalog),
        )
        current_citation.verification_status = "candidate"
        current_citation.verification_content_digest = source_digest
        current_citation = _finalize_citation_metadata(current_citation)

        candidate_payload = self._build_citation_candidate_payload(
            row=row,
            deterministic_metadata=deterministic,
            html_metadata=extract_html_citation_metadata(
                catalog_html_text,
                base_url=row.original_url or row.final_url,
            ),
            deterministic_citation=current_citation,
            doi_registry_citation=doi_citation,
        )
        catalog_payload = self._base_catalog_payload(row, existing_catalog_payload)
        catalog_payload["evidence_snippets"] = deterministic.get("evidence_snippets", [])
        catalog_payload["citation_candidates"] = candidate_payload
        existing_citation = existing_catalog_payload.get("citation")
        if existing_citation:
            catalog_payload["citation"] = existing_citation
        self._write_catalog_payload(row, catalog_payload)
        row.catalog_status = "generated"
        self._complete_row_phase(
            row,
            PHASE_CATALOG,
            status="completed",
            content_digest=source_digest,
            prompt_version=PROMPT_VERSION_CATALOG,
        )

    def _generate_source_citation_verification(
        self,
        row: SourceManifestRow,
        notes: list[str],
    ) -> None:
        if not self.run_citation_verify:
            return

        source_digest = self._effective_markdown_digest(row)
        existing_catalog_payload = _load_manifest_json(self.output_dir, row.catalog_file)
        existing_citation = _coerce_citation_metadata(existing_catalog_payload.get("citation"))
        self._begin_row_phase(
            row,
            PHASE_CITATION_VERIFY,
            model=self.llm_backend.model,
            prompt_version=PROMPT_VERSION_CITATION_VERIFY,
        )
        citation_source_path = self._effective_markdown_rel_path(row)
        if not citation_source_path or not source_digest:
            self._complete_row_phase(
                row,
                PHASE_CITATION_VERIFY,
                status="failed",
                error="missing_markdown: no markdown available for citation verification",
                error_code="missing_markdown",
                prompt_version=PROMPT_VERSION_CITATION_VERIFY,
            )
            return

        if (
            row.catalog_file
            and not self.force_citation_verify
            and _has_output_file(self.output_dir, row.catalog_file)
            and _citation_verification_is_current(existing_citation, source_digest)
        ):
            self._complete_row_phase(
                row,
                PHASE_CITATION_VERIFY,
                status="completed",
                content_digest=source_digest,
                model=self.llm_backend.model if existing_citation.verification_model else "",
                prompt_version=PROMPT_VERSION_CITATION_VERIFY,
            )
            return

        markdown_text = self._read_text(Path(citation_source_path))
        if not markdown_text.strip():
            self._complete_row_phase(
                row,
                PHASE_CITATION_VERIFY,
                status="failed",
                error="missing_markdown: empty markdown input",
                error_code="missing_markdown",
                prompt_version=PROMPT_VERSION_CITATION_VERIFY,
            )
            return

        catalog_html_text = _read_catalog_html_input(row, read_text=self._read_text)
        deterministic = _build_deterministic_catalog_metadata(
            row=row,
            markdown_text=markdown_text,
            html_text=catalog_html_text,
        )
        current_citation = _coerce_citation_metadata(deterministic.get("citation"))
        doi_citation = CitationMetadata()
        if current_citation.doi:
            doi_citation = _resolve_doi_citation_metadata(current_citation.doi)
            current_citation = _merge_citation_metadata(current_citation, doi_citation)
        current_citation = _merge_citation_metadata(
            current_citation,
            {
                "title": row.title,
                "authors": [
                    author.model_dump(mode="json")
                    for author in normalize_citation_authors(row.author_names)
                ],
                "issued": row.publication_date,
                "publisher": row.organization_name,
                "item_type": row.document_type,
                "url": row.original_url or row.final_url,
            },
            overwrite_existing=False,
        )
        current_citation.verification_status = "candidate"
        current_citation.verification_content_digest = source_digest
        current_citation = _finalize_citation_metadata(current_citation)

        candidate_payload = self._build_citation_candidate_payload(
            row=row,
            deterministic_metadata=deterministic,
            html_metadata=extract_html_citation_metadata(
                catalog_html_text,
                base_url=row.original_url or row.final_url,
            ),
            deterministic_citation=current_citation,
            doi_registry_citation=doi_citation,
        )
        verified_citation, llm_used, llm_error = self._verify_citation_with_llm(
            row=row,
            markdown_text=markdown_text,
            source_digest=source_digest,
            candidate_payload=candidate_payload,
            base_citation=current_citation,
        )
        verified_citation = _apply_citation_manual_overrides(verified_citation, existing_citation)
        if verified_citation.ready_for_ris and not verified_citation.verified_at:
            verified_citation.verified_at = _utc_now_iso()
        if not verified_citation.verification_content_digest:
            verified_citation.verification_content_digest = source_digest
        verified_citation = _finalize_citation_metadata(verified_citation)

        catalog_payload = self._base_catalog_payload(row, existing_catalog_payload)
        if "evidence_snippets" not in catalog_payload:
            catalog_payload["evidence_snippets"] = deterministic.get("evidence_snippets", [])
        if "citation_candidates" not in catalog_payload:
            catalog_payload["citation_candidates"] = candidate_payload
        catalog_payload["citation"] = verified_citation.model_dump(mode="json")
        self._write_catalog_payload(row, catalog_payload)

        if llm_error:
            notes.append("citation_verification_failed")
        citation_phase_status = "completed"
        verification_status = str(verified_citation.verification_status or "").strip().lower()
        if verification_status == "failed" or llm_error:
            citation_phase_status = "failed"
        elif verification_status.startswith("skipped"):
            citation_phase_status = "skipped"
        self._complete_row_phase(
            row,
            PHASE_CITATION_VERIFY,
            status=citation_phase_status,
            content_digest=source_digest,
            error=llm_error,
            error_code=_phase_error_code(llm_error),
            model=self.llm_backend.model if llm_used else "",
            prompt_version=PROMPT_VERSION_CITATION_VERIFY,
        )

    def _build_citation_candidate_payload(
        self,
        *,
        row: SourceManifestRow,
        deterministic_metadata: dict[str, Any],
        html_metadata: dict[str, Any],
        deterministic_citation: CitationMetadata,
        doi_registry_citation: CitationMetadata,
    ) -> dict[str, Any]:
        return {
            "display_metadata": {
                "title": row.title,
                "title_status": row.title_status,
                "author_names": row.author_names,
                "publication_date": row.publication_date,
                "publication_year": row.publication_year,
                "document_type": row.document_type,
                "organization_name": row.organization_name,
                "organization_type": row.organization_type,
            },
            "deterministic_metadata": deterministic_metadata,
            "html_metadata": html_metadata,
            "deterministic_citation": deterministic_citation.model_dump(mode="json"),
            "doi_registry_citation": doi_registry_citation.model_dump(mode="json"),
        }

    def _verify_citation_with_llm(
        self,
        *,
        row: SourceManifestRow,
        markdown_text: str,
        source_digest: str,
        candidate_payload: dict[str, Any],
        base_citation: CitationMetadata,
    ) -> tuple[CitationMetadata, bool, str]:
        if not self.use_llm:
            return (
                _merge_citation_metadata(
                    base_citation,
                    {
                        "verification_status": "skipped_llm_disabled",
                        "verification_content_digest": source_digest,
                        "blocked_reasons": ["Citation verification requires an enabled LLM backend."],
                    },
                    overwrite_existing=True,
                ),
                False,
                "",
            )
        if not llm_backend_ready_for_chat(self.llm_backend):
            return (
                _merge_citation_metadata(
                    base_citation,
                    {
                        "verification_status": "skipped_llm_not_configured",
                        "verification_content_digest": source_digest,
                        "blocked_reasons": ["Citation verification requires a chat-capable LLM backend."],
                    },
                    overwrite_existing=True,
                ),
                False,
                "",
            )

        source_text = self._truncate_for_llm(markdown_text.strip(), self._effective_max_source_chars())
        research_purpose = self.research_purpose or (
            "No explicit research purpose was provided. Verify citation metadata conservatively."
        )
        user_prompt = SOURCE_CITATION_VERIFY_USER.format(
            research_purpose=research_purpose,
            source_kind=row.source_kind,
            original_url=row.original_url or row.final_url or "",
            candidate_metadata_json=json.dumps(candidate_payload, ensure_ascii=False, indent=2),
            source_markdown=source_text,
        )
        try:
            raw_response = self._llm_client.sync_chat_completion(
                system_prompt=SOURCE_CITATION_VERIFY_SYSTEM,
                user_prompt=user_prompt,
                response_format="json",
            ).strip()
            payload = json.loads(raw_response)
            if not isinstance(payload, dict):
                raise ValueError("invalid verification payload")
            citation_payload = payload.get("citation")
            field_evidence_payload = payload.get("field_evidence")
            field_evidence: dict[str, dict[str, Any]] = {}
            if isinstance(field_evidence_payload, dict):
                for field_name, field_data in field_evidence_payload.items():
                    if not isinstance(field_data, dict):
                        continue
                    field_evidence[str(field_name)] = {
                        "source_type": _stringify_manifest_value(field_data.get("source_type")),
                        "source_label": _stringify_manifest_value(field_data.get("source_label")),
                        "evidence": _stringify_manifest_value(field_data.get("evidence")),
                        "confidence": _round_citation_confidence(float(field_data.get("confidence") or 0.0)),
                    }
            verified_citation = _merge_citation_metadata(
                base_citation,
                {
                    **(citation_payload if isinstance(citation_payload, dict) else {}),
                    "field_evidence": field_evidence,
                    "blocked_reasons": payload.get("blocked_reasons"),
                    "notes": payload.get("notes"),
                    "verification_status": "verified",
                    "verification_confidence": payload.get("verification_confidence"),
                    "verification_model": self.llm_backend.model,
                    "verification_content_digest": source_digest,
                    "verified_at": _utc_now_iso(),
                },
                overwrite_existing=True,
            )
            return verified_citation, True, ""
        except Exception as exc:
            logger.warning("Citation verification failed for %s: %s", row.id, exc)
            failed_citation = _merge_citation_metadata(
                base_citation,
                {
                    "verification_status": "failed",
                    "verification_content_digest": source_digest,
                    "blocked_reasons": ["Citation verification failed."],
                },
                overwrite_existing=True,
            )
            return failed_citation, False, f"citation_verification_failed: {type(exc).__name__}: {exc}"

    def _generate_source_title(
        self,
        row: SourceManifestRow,
        notes: list[str],
    ) -> None:
        def sync_catalog_title() -> None:
            existing_catalog_payload = _load_manifest_json(self.output_dir, row.catalog_file)
            catalog_payload = self._base_catalog_payload(row, existing_catalog_payload)
            self._write_catalog_payload(row, catalog_payload)

        should_resolve_title = bool(self.run_llm_title or self.force_title)
        if not should_resolve_title:
            if row.title:
                row.title_status = row.title_status or "existing"
            elif not row.title_status:
                row.title_status = "not_requested"
            return

        if row.title and not self.force_title:
            row.title_status = row.title_status or "existing"
            sync_catalog_title()
            return

        title_source_path = ""
        if row.llm_cleanup_file and _has_output_file(self.output_dir, row.llm_cleanup_file):
            title_source_path = row.llm_cleanup_file
        elif row.markdown_file:
            title_source_path = row.markdown_file

        if not title_source_path:
            row.title_status = "missing_markdown"
            return

        markdown_text = self._read_text(Path(title_source_path))
        if not markdown_text.strip():
            row.title_status = "missing_markdown"
            return

        candidate_title = extract_markdown_title_candidate(markdown_text)
        if candidate_title:
            row.title = candidate_title
            row.title_status = "extracted"
            sync_catalog_title()
            return

        if not self.use_llm:
            row.title_status = "skipped_llm_disabled"
            return
        if not llm_backend_ready_for_chat(self.llm_backend):
            row.title_status = "skipped_llm_not_configured"
            notes.append(NOTE_TITLE_SKIPPED_LLM_NOT_CONFIGURED)
            return

        source_text = markdown_text.strip()
        max_chars = self._effective_max_source_chars()
        source_text = self._truncate_for_llm(source_text, max_chars)
        research_purpose = self.research_purpose or (
            "No explicit research purpose was provided. Prefer the document title when present."
        )
        user_prompt = SOURCE_TITLE_USER.format(
            research_purpose=research_purpose,
            existing_title=row.title or "",
            candidate_title=candidate_title or "",
            source_markdown=source_text,
        )

        try:
            raw_response = self._llm_client.sync_chat_completion(
                system_prompt=SOURCE_TITLE_SYSTEM,
                user_prompt=user_prompt,
                response_format="json",
            ).strip()
            payload = json.loads(raw_response)
            if not isinstance(payload, dict):
                notes.append(NOTE_TITLE_GENERATION_FAILED)
                row.title_status = "failed"
                return

            generated_title = normalize_generated_title(
                _stringify_manifest_value(payload.get("title"))
            )
            if not generated_title:
                notes.append(NOTE_TITLE_GENERATION_FAILED)
                row.title_status = "failed"
                return

            basis = str(payload.get("basis") or "").strip().lower()
            if basis != "document_title":
                generated_title = limit_title_words(generated_title, 10)
            row.title = generated_title
            row.title_status = "generated"
            sync_catalog_title()
        except Exception as exc:
            notes.append(NOTE_TITLE_GENERATION_FAILED)
            row.title_status = "failed"
            logger.warning("Title generation failed for %s: %s", row.id, exc)

    def _generate_source_summary(
        self,
        row: SourceManifestRow,
        notes: list[str],
    ) -> None:
        if not self.run_llm_summary:
            if row.summary_file and _has_output_file(self.output_dir, row.summary_file):
                row.summary_status = row.summary_status or "existing"
            elif not row.summary_status:
                row.summary_status = "not_requested"
            return

        source_digest = self._effective_markdown_digest(row)
        self._begin_row_phase(
            row,
            PHASE_SUMMARY,
            model=self.llm_backend.model,
            prompt_version=PROMPT_VERSION_SUMMARY,
        )
        summary_source_path = self._effective_markdown_rel_path(row)
        if not summary_source_path or not source_digest:
            row.summary_status = "missing_markdown"
            self._complete_row_phase(
                row,
                PHASE_SUMMARY,
                status="failed",
                error="missing_markdown: no markdown available for summarization",
                error_code="missing_markdown",
                prompt_version=PROMPT_VERSION_SUMMARY,
            )
            return
        if not self.use_llm:
            row.summary_status = "failed"
            self._complete_row_phase(
                row,
                PHASE_SUMMARY,
                status="failed",
                content_digest=source_digest,
                error="llm_disabled: summarization requires an enabled LLM backend",
                error_code="llm_disabled",
                prompt_version=PROMPT_VERSION_SUMMARY,
            )
            return
        if not llm_backend_ready_for_chat(self.llm_backend):
            row.summary_status = "failed"
            notes.append(NOTE_SUMMARY_SKIPPED_LLM_NOT_CONFIGURED)
            self._complete_row_phase(
                row,
                PHASE_SUMMARY,
                status="failed",
                content_digest=source_digest,
                error="llm_not_configured: summarization requires a chat-capable model",
                error_code="llm_not_configured",
                prompt_version=PROMPT_VERSION_SUMMARY,
            )
            return
        if (
            row.summary_file
            and not self.force_summary
            and _has_output_file(self.output_dir, row.summary_file)
            and (summary_metadata := _get_phase_metadata(row, PHASE_SUMMARY)) is not None
            and summary_metadata.content_digest == source_digest
        ):
            row.summary_status = "existing"
            self._complete_row_phase(
                row,
                PHASE_SUMMARY,
                status="completed",
                content_digest=source_digest,
                model=self.llm_backend.model,
                prompt_version=PROMPT_VERSION_SUMMARY,
            )
            return

        markdown_text = self._read_text(Path(summary_source_path))
        if not markdown_text.strip():
            row.summary_status = "missing_markdown"
            self._complete_row_phase(
                row,
                PHASE_SUMMARY,
                status="failed",
                error="missing_markdown: empty markdown input",
                error_code="missing_markdown",
                prompt_version=PROMPT_VERSION_SUMMARY,
            )
            return

        source_text = markdown_text.strip()
        max_chars = self._effective_max_source_chars()
        source_text = self._truncate_for_llm(source_text, max_chars)

        research_purpose = self.research_purpose or (
            "No explicit research purpose was provided. "
            "Focus on high-impact findings, methods, limitations, and relevance."
        )
        user_prompt = SOURCE_SUMMARY_USER.format(
            research_purpose=research_purpose,
            source_markdown=source_text,
        )

        try:
            summary = self._llm_client.sync_chat_completion(
                system_prompt=SOURCE_SUMMARY_SYSTEM,
                user_prompt=user_prompt,
                response_format=None,
            ).strip()
            summary = normalize_summary_paragraph(summary)
            if not summary:
                notes.append(NOTE_SUMMARY_GENERATION_FAILED)
                row.summary_status = "failed"
                return

            summary_rel = self._summary_rel(row)
            self._write_text(summary_rel, summary + "\n")
            row.summary_file = summary_rel.as_posix()
            row.summary_status = "generated"
            self._complete_row_phase(
                row,
                PHASE_SUMMARY,
                status="completed",
                content_digest=source_digest,
                model=self.llm_backend.model,
                prompt_version=PROMPT_VERSION_SUMMARY,
            )
        except Exception as exc:
            notes.append(NOTE_SUMMARY_GENERATION_FAILED)
            row.summary_status = "failed"
            self._complete_row_phase(
                row,
                PHASE_SUMMARY,
                status="failed",
                content_digest=source_digest,
                error=f"summary_generation_failed: {type(exc).__name__}: {exc}",
                error_code="summary_generation_failed",
                model=self.llm_backend.model,
                prompt_version=PROMPT_VERSION_SUMMARY,
            )
            logger.warning("Summary generation failed for %s: %s", row.id, exc)


    def _generate_source_rating(
        self,
        row: SourceManifestRow,
        notes: list[str],
    ) -> None:
        if not self.run_llm_rating:
            if row.rating_file and _has_output_file(self.output_dir, row.rating_file):
                row.rating_status = row.rating_status or "existing"
            elif not row.rating_status:
                row.rating_status = "not_requested"
            return

        source_digest = self._effective_markdown_digest(row)
        self._begin_row_phase(
            row,
            PHASE_RATING,
            model=self.llm_backend.model,
            profile_name=self.project_profile_name,
            prompt_version=PROMPT_VERSION_RATING,
        )
        if not self.project_profile_yaml:
            row.rating_status = "failed"
            self._complete_row_phase(
                row,
                PHASE_RATING,
                status="failed",
                content_digest=source_digest,
                error="missing_project_profile: rating requires a project profile",
                error_code="missing_project_profile",
                model=self.llm_backend.model,
                profile_name=self.project_profile_name,
                prompt_version=PROMPT_VERSION_RATING,
            )
            return

        rating_source_path = self._effective_markdown_rel_path(row)
        if not rating_source_path or not source_digest:
            row.rating_status = "missing_markdown"
            self._complete_row_phase(
                row,
                PHASE_RATING,
                status="failed",
                error="missing_markdown: no markdown available for relevance tagging",
                error_code="missing_markdown",
                model=self.llm_backend.model,
                profile_name=self.project_profile_name,
                prompt_version=PROMPT_VERSION_RATING,
            )
            return
        if not self.use_llm:
            row.rating_status = "failed"
            self._complete_row_phase(
                row,
                PHASE_RATING,
                status="failed",
                content_digest=source_digest,
                error="llm_disabled: relevance tagging requires an enabled LLM backend",
                error_code="llm_disabled",
                model=self.llm_backend.model,
                profile_name=self.project_profile_name,
                prompt_version=PROMPT_VERSION_RATING,
            )
            return
        if not llm_backend_ready_for_chat(self.llm_backend):
            row.rating_status = "failed"
            notes.append(NOTE_RATING_SKIPPED_LLM_NOT_CONFIGURED)
            self._complete_row_phase(
                row,
                PHASE_RATING,
                status="failed",
                content_digest=source_digest,
                error="llm_not_configured: relevance tagging requires a chat-capable model",
                error_code="llm_not_configured",
                model=self.llm_backend.model,
                profile_name=self.project_profile_name,
                prompt_version=PROMPT_VERSION_RATING,
            )
            return
        if (
            row.rating_file
            and not self.force_rating
            and _has_output_file(self.output_dir, row.rating_file)
            and (rating_metadata := _get_phase_metadata(row, PHASE_RATING)) is not None
            and rating_metadata.content_digest == source_digest
        ):
            row.rating_status = "existing"
            self._complete_row_phase(
                row,
                PHASE_RATING,
                status="completed",
                content_digest=source_digest,
                model=self.llm_backend.model,
                profile_name=self.project_profile_name,
                prompt_version=PROMPT_VERSION_RATING,
            )
            return

        markdown_text = self._read_text(Path(rating_source_path))
        if not markdown_text.strip():
            row.rating_status = "missing_markdown"
            self._complete_row_phase(
                row,
                PHASE_RATING,
                status="failed",
                error="missing_markdown: empty markdown input",
                error_code="missing_markdown",
                model=self.llm_backend.model,
                profile_name=self.project_profile_name,
                prompt_version=PROMPT_VERSION_RATING,
            )
            return

        source_text = markdown_text.strip()
        max_chars = self._effective_max_source_chars()
        source_text = self._truncate_for_llm(source_text, max_chars)

        research_purpose = self.research_purpose or (
            "No explicit research purpose was provided. "
            "Evaluate the source based on the project profile dimensions."
        )
        system_prompt = SOURCE_RATING_SYSTEM.format(
            project_profile_yaml=self.project_profile_yaml,
        )
        user_prompt = SOURCE_RATING_USER.format(
            research_purpose=research_purpose,
            source_markdown=source_text,
        )

        try:
            raw_response = self._llm_client.sync_chat_completion(
                system_prompt=system_prompt,
                user_prompt=user_prompt,
                response_format="json",
            ).strip()

            rating_data = json.loads(raw_response)
            if not isinstance(rating_data, dict):
                notes.append(NOTE_RATING_GENERATION_FAILED)
                row.rating_status = "failed"
                return

            raw_tags = rating_data.get("tags")
            if isinstance(raw_tags, list):
                row.tags_text = "; ".join(
                    _dedupe_strings([str(item).strip() for item in raw_tags if str(item).strip()])
                )
            elif isinstance(raw_tags, str) and raw_tags.strip():
                row.tags_text = "; ".join(
                    _dedupe_strings([item.strip() for item in raw_tags.split(",") if item.strip()])
                )

            rating_rel = self._rating_rel(row)
            self._write_text(
                rating_rel,
                json.dumps(rating_data, ensure_ascii=False, indent=2) + "\n",
            )
            row.rating_file = rating_rel.as_posix()
            row.rating_status = "generated"
            self._complete_row_phase(
                row,
                PHASE_RATING,
                status="completed",
                content_digest=source_digest,
                model=self.llm_backend.model,
                profile_name=self.project_profile_name,
                prompt_version=PROMPT_VERSION_RATING,
            )
        except Exception as exc:
            notes.append(NOTE_RATING_GENERATION_FAILED)
            row.rating_status = "failed"
            self._complete_row_phase(
                row,
                PHASE_RATING,
                status="failed",
                content_digest=source_digest,
                error=f"rating_generation_failed: {type(exc).__name__}: {exc}",
                error_code="rating_generation_failed",
                model=self.llm_backend.model,
                profile_name=self.project_profile_name,
                prompt_version=PROMPT_VERSION_RATING,
            )
            logger.warning("Rating generation failed for %s: %s", row.id, exc)



    def _write_binary(self, rel_path: Path, content: bytes) -> None:
        dest = self.output_dir / rel_path
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(content)

    def _write_text(self, rel_path: Path, content: str) -> None:
        dest = self.output_dir / rel_path
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_text(content, encoding="utf-8")

    def _read_binary(self, rel_path: Path) -> bytes:
        dest = self.output_dir / rel_path
        if not dest.exists():
            return b""
        return dest.read_bytes()

    def _read_text(self, rel_path: Path) -> str:
        dest = self.output_dir / rel_path
        if not dest.exists():
            return ""
        return dest.read_text(encoding="utf-8", errors="replace")

    def _convert_document_to_markdown(
        self,
        extension: str,
        binary_content: bytes,
    ) -> tuple[str, str, list[str]]:
        ext = extension.lower()
        if ext in {".txt", ".md"}:
            text = decode_bytes_to_text(binary_content)
            return (text.strip(), "plain_text", []) if text.strip() else ("", "", [])

        if ext in {".docx", ".dotx"}:
            markdown_text = convert_docx_bytes_to_markdown(binary_content)
            return (
                (markdown_text, "docx_local", [])
                if markdown_text
                else ("", "", [NOTE_EXTRACTION_FAILURE])
            )

        if ext in {".doc", ".dot"}:
            if not self.runtime_capabilities.textutil_available:
                return "", "", [NOTE_RUNTIME_MISSING_TEXTUTIL, NOTE_DOC_CONVERSION_FAILED]
            markdown_text = convert_doc_bytes_with_textutil(binary_content)
            if markdown_text:
                return markdown_text, "doc_textutil", []
            return "", "", [NOTE_DOC_CONVERSION_FAILED]

        return "", "", []

    def _convert_pdf_to_markdown(
        self,
        pdf_bytes: bytes,
    ) -> tuple[str, str, list[str]]:
        notes: list[str] = []
        pages = extract_pdf_pages(pdf_bytes)
        if not pages:
            return "", "", [NOTE_EXTRACTION_FAILURE]

        native_texts = [p["text"] for p in pages]
        native_metrics = [compute_text_metrics(text) for text in native_texts]
        doc_quality = evaluate_pdf_document_quality(native_metrics)
        if doc_quality["native_good"]:
            markdown_text = format_pages_as_markdown(native_texts)
            if markdown_score(markdown_text) >= MIN_FALLBACK_MARKDOWN_SCORE:
                return markdown_text, "pdf_text", notes

        page_texts = native_texts[:]
        method = "pdf_text"
        low_quality_indexes = doc_quality["low_quality_pages"][:]

        # Tier 2: local OCR
        if low_quality_indexes and self.runtime_capabilities.tesseract_available:
            ocr_applied = False
            for idx in low_quality_indexes:
                ocr_text = run_tesseract_ocr_on_pixmap(pages[idx]["pixmap"])
                if ocr_text and compute_text_metrics(ocr_text)["chars"] >= PDF_OCR_MIN_CHARS:
                    page_texts[idx] = ocr_text
                    ocr_applied = True
            if ocr_applied:
                notes.append(NOTE_OCR_LOCAL_USED)
                method = "pdf_ocr_local"

        # Tier 3: optional LLM vision fallback on remaining low-quality pages
        post_ocr_metrics = [compute_text_metrics(text) for text in page_texts]
        post_ocr_quality = evaluate_pdf_document_quality(post_ocr_metrics)
        remaining_low_quality = post_ocr_quality["low_quality_pages"]
        llm_used = False
        if (
            remaining_low_quality
            and self.use_llm
            and self.runtime_capabilities.llm_vision_enabled
        ):
            for idx in remaining_low_quality:
                llm_text = self._run_llm_vision_ocr_on_page(pages[idx]["pixmap"])
                if llm_text and compute_text_metrics(llm_text)["chars"] >= PDF_OCR_MIN_CHARS:
                    page_texts[idx] = llm_text
                    llm_used = True
            if llm_used:
                notes.append(NOTE_OCR_LLM_FALLBACK_USED)
                method = "pdf_ocr_llm_vision"

        markdown_text = format_pages_as_markdown(page_texts)
        if markdown_score(markdown_text) >= MIN_FALLBACK_MARKDOWN_SCORE:
            return markdown_text, method, notes

        if not self.runtime_capabilities.tesseract_available:
            notes.append(NOTE_RUNTIME_MISSING_TESSERACT)
        if self.use_llm and not self.runtime_capabilities.llm_vision_enabled:
            notes.append(NOTE_RUNTIME_MISSING_LLM_VISION)

        notes.append(NOTE_EXTRACTION_FAILURE)
        return "", "", notes

    def _run_llm_vision_ocr_on_page(self, pixmap: fitz.Pixmap) -> str:
        if not self.runtime_capabilities.llm_vision_enabled:
            return ""
        try:
            image_bytes = pixmap.tobytes("png")
        except Exception:
            return ""

        prompt = (
            "Extract all readable text from this page image. "
            "Return plain text only. Do not summarize, do not add commentary."
        )
        try:
            return run_async_in_sync(
                self._run_llm_vision,
                prompt,
                image_bytes,
            ).strip()
        except Exception:
            return ""

    async def _run_llm_vision(self, prompt: str, image_bytes: bytes) -> str:
        client = UnifiedLLMClient(self.llm_backend)
        try:
            return await client.vision_ocr(prompt=prompt, image_bytes=image_bytes)
        finally:
            await client.close()


def build_manifest_csv(
    rows: list[SourceManifestRow],
    base_dir: Path | None = None,
    *,
    column_configs: Sequence[RepositoryColumnConfig] | None = None,
) -> str:
    fieldnames, records = _build_manifest_records(
        rows,
        base_dir=base_dir,
        column_configs=column_configs,
    )
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for record in records:
        writer.writerow(record)
    return output.getvalue()


def build_manifest_xlsx(
    rows: list[SourceManifestRow],
    base_dir: Path | None = None,
    *,
    column_configs: Sequence[RepositoryColumnConfig] | None = None,
) -> bytes:
    fieldnames, records = _build_manifest_records(
        rows,
        base_dir=base_dir,
        column_configs=column_configs,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "manifest"
    ws.append(fieldnames)
    ws.freeze_panes = "A2"

    link_columns = {
        "raw_file",
        "rendered_file",
        "rendered_pdf_file",
        "markdown_file",
        "llm_cleanup_file",
        "summary_file",
        "rating_file",
        "metadata_file",
    }
    col_idx = {name: idx + 1 for idx, name in enumerate(fieldnames)}

    for record in records:
        ws.append([record.get(column, "") for column in fieldnames])
        row_idx = ws.max_row
        for link_col in link_columns:
            value = record.get(link_col, "")
            if not value:
                continue
            cell = ws.cell(row=row_idx, column=col_idx[link_col])
            cell.value = value
            cell.hyperlink = value
            cell.style = "Hyperlink"

    for idx, column in enumerate(fieldnames, start=1):
        if column.endswith("_url"):
            width = 40
        elif column in {"summary_text", "rating_rationale", "relevant_sections", "rating_raw_json"}:
            width = 56
        else:
            width = 24
        ws.column_dimensions[get_column_letter(idx)].width = width

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def build_manifest_record(
    row: SourceManifestRow,
    base_dir: Path | None = None,
    *,
    column_configs: Sequence[RepositoryColumnConfig] | None = None,
) -> dict[str, str | int | float | bool]:
    data = row.model_dump()
    serialized: dict[str, str | int | float | bool] = {}
    for column in SOURCE_MANIFEST_COLUMNS:
        value = data.get(column)
        if value is None:
            serialized[column] = ""
        else:
            serialized[column] = value

    derived = _derive_manifest_fields(row, base_dir=base_dir)
    serialized.update(derived["base"])
    serialized.update(derived["dynamic"])
    for column in column_configs or []:
        if column.kind != "custom":
            continue
        serialized[column.id] = str((row.custom_fields or {}).get(column.id) or "")
    return serialized


def _build_manifest_records(
    rows: list[SourceManifestRow],
    base_dir: Path | None = None,
    *,
    column_configs: Sequence[RepositoryColumnConfig] | None = None,
) -> tuple[list[str], list[dict[str, str | int | float | bool]]]:
    records: list[dict[str, str | int | float | bool]] = []
    dynamic_columns: list[str] = []
    dynamic_seen: set[str] = set()
    custom_column_ids = [column.id for column in column_configs or [] if column.kind == "custom"]

    for row in rows:
        record = build_manifest_record(
            row,
            base_dir=base_dir,
            column_configs=column_configs,
        )
        records.append(record)
        for column in record.keys():
            if column in SOURCE_MANIFEST_COLUMNS or column in MANIFEST_DERIVED_COLUMNS:
                continue
            if column in dynamic_seen:
                continue
            dynamic_seen.add(column)
            dynamic_columns.append(column)

    custom_fields = [column_id for column_id in custom_column_ids if column_id not in dynamic_seen]
    fieldnames = (
        SOURCE_MANIFEST_COLUMNS
        + MANIFEST_DERIVED_COLUMNS
        + custom_fields
        + sorted(dynamic_columns)
    )
    return fieldnames, records


def _derive_manifest_fields(
    row: SourceManifestRow,
    base_dir: Path | None = None,
) -> dict[str, dict[str, str | int | float | bool]]:
    catalog_payload = _load_manifest_json(base_dir, row.catalog_file)
    citation = _coerce_citation_metadata(catalog_payload.get("citation"))
    summary_text = _load_manifest_text(base_dir, row.summary_file)
    rating_payload = _load_manifest_json(base_dir, row.rating_file)

    if summary_text == "" and isinstance(rating_payload, dict):
        summary_text = _stringify_manifest_value(rating_payload.get("summary"))

    rating_overall = _extract_rating_overall(rating_payload)
    rating_confidence = _extract_rating_confidence(rating_payload)
    rating_rationale = _stringify_manifest_value(
        rating_payload.get("rationale") if isinstance(rating_payload, dict) else ""
    )
    relevant_sections = _format_relevant_sections(
        rating_payload.get("relevant_sections")
        if isinstance(rating_payload, dict)
        else ""
    )
    if not relevant_sections and isinstance(rating_payload, dict):
        relevant_sections = _format_relevant_sections(
            rating_payload.get("sections") or rating_payload.get("relevant_section")
        )

    rating_dimensions = _extract_rating_dimensions(rating_payload)
    flag_scores = _extract_flag_scores(rating_payload)
    tags_text = row.tags_text or _format_tags_text(
        rating_payload.get("tags") if isinstance(rating_payload, dict) else ""
    )

    base: dict[str, str | int | float | bool] = {
        "title": row.title or _stringify_manifest_value(catalog_payload.get("title")),
        "author_names": row.author_names or _stringify_manifest_value(catalog_payload.get("author_names")),
        "publication_date": row.publication_date or _stringify_manifest_value(catalog_payload.get("publication_date")),
        "publication_year": row.publication_year or _stringify_manifest_value(catalog_payload.get("publication_year")),
        "document_type": row.document_type or _stringify_manifest_value(catalog_payload.get("document_type")),
        "organization_name": row.organization_name or _stringify_manifest_value(catalog_payload.get("organization_name")),
        "organization_type": row.organization_type or _stringify_manifest_value(catalog_payload.get("organization_type")),
        "tags_text": tags_text,
        "summary_text": summary_text,
        "rating_overall": rating_overall,
        "rating_confidence": rating_confidence,
        "rating_rationale": rating_rationale,
        "relevant_sections": relevant_sections,
        "rating_dimensions_json": _json_or_blank(rating_dimensions),
        "flag_scores_json": _json_or_blank(flag_scores),
        "rating_raw_json": _json_or_blank(rating_payload),
        "citation_title": citation.title,
        "citation_authors": _citation_author_names(citation),
        "citation_issued": citation.issued,
        "citation_url": citation.url,
        "citation_publisher": citation.publisher,
        "citation_container_title": citation.container_title,
        "citation_volume": citation.volume,
        "citation_issue": citation.issue,
        "citation_pages": citation.pages,
        "citation_language": citation.language,
        "citation_accessed": citation.accessed,
        "citation_type": citation.item_type,
        "citation_doi": citation.doi,
        "citation_report_number": citation.report_number,
        "citation_standard_number": citation.standard_number,
        "citation_verification_status": citation.verification_status,
        "citation_blocked_reasons": _citation_blocked_reasons_text(citation),
        "citation_manual_override_fields": _citation_manual_override_fields_text(citation),
        "citation_field_evidence_json": _citation_field_evidence_json(citation),
        "citation_verified_at": citation.verified_at,
        "citation_ready": citation.ready_for_ris,
        "citation_missing_fields": _citation_missing_fields_text(citation),
        "citation_confidence": citation.verification_confidence or citation.confidence,
    }

    dynamic: dict[str, str | int | float | bool] = {}
    for key, value in rating_dimensions.items():
        dynamic[f"rating_{_manifest_slug(key)}"] = _normalize_manifest_scalar(value)
    for key, value in flag_scores.items():
        dynamic[f"flag_{_manifest_slug(key)}"] = _normalize_manifest_scalar(value)
    return {"base": base, "dynamic": dynamic}


def _load_manifest_text(base_dir: Path | None, rel_path: str | None) -> str:
    full_path = _resolve_manifest_path(base_dir, rel_path)
    if full_path is None or not full_path.is_file():
        return ""
    try:
        return full_path.read_text(encoding="utf-8", errors="replace").strip()
    except OSError:
        return ""


def _load_manifest_json(base_dir: Path | None, rel_path: str | None) -> dict[str, Any]:
    full_path = _resolve_manifest_path(base_dir, rel_path)
    if full_path is None or not full_path.is_file():
        return {}
    try:
        payload = json.loads(full_path.read_text(encoding="utf-8", errors="replace"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _resolve_manifest_path(base_dir: Path | None, rel_path: str | None) -> Path | None:
    value = str(rel_path or "").strip()
    if not value:
        return None
    candidate = Path(value)
    if candidate.is_absolute():
        return candidate
    if base_dir is None:
        return None
    return base_dir / candidate


def _row_citation_verification_status(
    row: SourceManifestRow | None,
    base_dir: Path | None = None,
) -> str:
    if row is None:
        return ""
    catalog_payload = _load_manifest_json(base_dir, row.catalog_file)
    citation = _coerce_citation_metadata(catalog_payload.get("citation"))
    return str(citation.verification_status or "").strip()


def _extract_rating_overall(payload: dict[str, Any]) -> str | int | float | bool:
    for key in ("overall_score", "overall_rating", "overall"):
        if key in payload:
            return _normalize_manifest_scalar(payload.get(key))
    return ""


def _extract_rating_confidence(payload: dict[str, Any]) -> str | int | float | bool:
    if not isinstance(payload, dict):
        return ""
    return _normalize_manifest_scalar(payload.get("confidence"))


def _extract_rating_dimensions(payload: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    extracted: dict[str, Any] = {}
    for container_key in RATING_DIMENSION_CONTAINER_KEYS:
        container = payload.get(container_key)
        if not isinstance(container, dict):
            continue
        for key, value in container.items():
            extracted[str(key)] = value
    if extracted:
        return extracted
    for key, value in payload.items():
        if key in RATING_RESERVED_KEYS or isinstance(value, (dict, list)):
            continue
        extracted[str(key)] = value
    return extracted


def _extract_flag_scores(payload: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    flags = payload.get("flags")
    if not isinstance(flags, dict):
        return {}
    return {str(key): value for key, value in flags.items()}


def _format_relevant_sections(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        parts = [_format_relevant_section_item(item) for item in value]
        return "\n\n".join(part for part in parts if part)
    if isinstance(value, dict):
        return _format_relevant_section_item(value)
    return _stringify_manifest_value(value)


def _format_relevant_section_item(item: Any) -> str:
    if item is None:
        return ""
    if isinstance(item, str):
        return item.strip()
    if isinstance(item, dict):
        ordered_keys = ["section", "title", "heading", "quote", "excerpt", "text", "content"]
        parts: list[str] = []
        for key in ordered_keys:
            piece = _stringify_manifest_value(item.get(key))
            if not piece:
                continue
            if key in {"section", "title", "heading"}:
                parts.append(piece)
            else:
                parts.append(piece)
        if parts:
            return "\n".join(parts)
        return json.dumps(item, ensure_ascii=False, sort_keys=True)
    return _stringify_manifest_value(item)


def _stringify_manifest_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value)
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    except Exception:
        return str(value)


def _json_or_blank(value: Any) -> str:
    if value in ({}, [], "", None):
        return ""
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    except Exception:
        return str(value)


def _citation_field_evidence_json(citation: CitationMetadata) -> str:
    if not citation.field_evidence:
        return ""
    payload = {
        key: value.model_dump(mode="json")
        for key, value in citation.field_evidence.items()
    }
    return _json_or_blank(payload)


def _normalize_manifest_scalar(value: Any) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return _stringify_manifest_value(value)


def _manifest_slug(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower())
    normalized = normalized.strip("_")
    return normalized or "value"


def _coerce_citation_metadata(value: Any) -> CitationMetadata:
    if isinstance(value, CitationMetadata):
        return value
    if not isinstance(value, dict):
        return CitationMetadata()
    try:
        return CitationMetadata.model_validate(value)
    except Exception:
        return CitationMetadata()


def _citation_author_display(author: CitationAuthor) -> str:
    if author.literal:
        return author.literal.strip()
    pieces = [author.given.strip(), author.family.strip()]
    return " ".join(piece for piece in pieces if piece).strip()


def _citation_author_names(citation: CitationMetadata) -> str:
    names = [_citation_author_display(author) for author in citation.authors]
    return "; ".join(name for name in names if name)


def _citation_publication_year(citation: CitationMetadata) -> str:
    return _extract_publication_year(citation.issued)


def _citation_missing_fields_text(citation: CitationMetadata) -> str:
    return "; ".join(_dedupe_strings(citation.missing_fields))


def _citation_blocked_reasons_text(citation: CitationMetadata) -> str:
    return "; ".join(_dedupe_strings(citation.blocked_reasons))


def _citation_manual_override_fields_text(citation: CitationMetadata) -> str:
    return "; ".join(_dedupe_strings(citation.manual_override_fields))


def _round_citation_confidence(value: float) -> float:
    normalized = max(0.0, min(float(value), 1.0))
    step = round(normalized / CITATION_CONFIDENCE_INCREMENT)
    return round(step * CITATION_CONFIDENCE_INCREMENT, 2)


def _citation_field_value_text(citation: CitationMetadata, field_name: str) -> str:
    if field_name == "authors":
        return _citation_author_names(citation)
    return _stringify_manifest_value(getattr(citation, field_name, ""))


def _normalize_citation_field_evidence(value: Any) -> CitationFieldEvidence:
    if isinstance(value, CitationFieldEvidence):
        raw = value.model_dump(mode="json")
    elif isinstance(value, dict):
        raw = dict(value)
    else:
        raw = {}
    return CitationFieldEvidence(
        value=_stringify_manifest_value(raw.get("value")),
        source_type=_stringify_manifest_value(raw.get("source_type")),
        source_label=_stringify_manifest_value(raw.get("source_label")),
        evidence=_stringify_manifest_value(raw.get("evidence")),
        confidence=_round_citation_confidence(float(raw.get("confidence") or 0.0)),
        manual_override=bool(raw.get("manual_override")),
    )


def _empty_citation_field_evidence(
    *,
    value: str = "",
    source_type: str = "",
    source_label: str = "",
    evidence: str = "",
    confidence: float = 0.0,
    manual_override: bool = False,
) -> CitationFieldEvidence:
    return CitationFieldEvidence(
        value=_stringify_manifest_value(value),
        source_type=_stringify_manifest_value(source_type),
        source_label=_stringify_manifest_value(source_label),
        evidence=_stringify_manifest_value(evidence),
        confidence=_round_citation_confidence(confidence),
        manual_override=manual_override,
    )


def _apply_organization_author_fallback(citation: CitationMetadata) -> CitationMetadata:
    if citation.authors:
        return citation
    publisher = _collapse_whitespace(citation.publisher).strip(" ,.;:")
    if not publisher or not _looks_like_corporate_author_name(publisher):
        return citation
    fallback_authors = normalize_citation_authors([publisher])
    if not fallback_authors:
        return citation
    updated = citation.model_copy(deep=True)
    updated.authors = fallback_authors
    updated.notes = _dedupe_strings([*updated.notes, "organization_author_fallback"])
    existing = _normalize_citation_field_evidence(updated.field_evidence.get("authors"))
    if not existing.value:
        existing.value = publisher
    if not existing.source_type:
        existing.source_type = "publisher_fallback"
    if not existing.source_label:
        existing.source_label = "Publisher fallback"
    if not existing.evidence:
        existing.evidence = publisher
    existing.confidence = max(existing.confidence, 0.55)
    updated.field_evidence["authors"] = existing
    return updated


def _collapse_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _normalize_citation_type(
    value: str,
    *,
    fallback_document_type: str = "",
    source_kind: str = "",
    url: str = "",
) -> str:
    for candidate in (value, fallback_document_type):
        normalized = re.sub(r"[^a-z0-9]+", " ", str(candidate or "").strip().lower()).strip()
        if not normalized:
            continue
        if normalized in CANONICAL_CITATION_TYPE_MAP:
            return CANONICAL_CITATION_TYPE_MAP[normalized]
        if "journal" in normalized and "article" in normalized:
            return "journal article"
        if "conference" in normalized:
            return "conference paper"
        if "standard" in normalized:
            return "standard"
        if "thesis" in normalized or "dissertation" in normalized:
            return "thesis"
        if "report" in normalized or "working paper" in normalized:
            return "report"
        if "book" in normalized and "chapter" in normalized:
            return "book chapter"
        if "book" in normalized:
            return "book"
        if "web" in normalized or "site" in normalized or "blog" in normalized:
            return "web page"
    lowered_url = str(url or "").lower()
    if lowered_url.startswith("http"):
        return "web page" if source_kind == "url" else ""
    return ""


def _looks_like_corporate_author_name(text: str) -> bool:
    candidate = _collapse_whitespace(text).strip(" ,.;:")
    if not candidate:
        return True
    lowered = candidate.lower()
    if any(keyword in lowered for keyword in CORPORATE_AUTHOR_KEYWORDS):
        return True
    words = re.findall(r"[A-Za-z][A-Za-z'.-]*", candidate)
    if len(words) > 4:
        return True
    if any(char.isdigit() for char in candidate):
        return True
    if ":" in candidate or "&" in candidate:
        return True
    return False


def _normalize_citation_author(author: Any) -> CitationAuthor | None:
    if isinstance(author, CitationAuthor):
        normalized = author
    elif isinstance(author, dict):
        literal = _collapse_whitespace(str(author.get("literal") or author.get("name") or "")).strip()
        family = _collapse_whitespace(str(author.get("family") or author.get("familyName") or "")).strip()
        given = _collapse_whitespace(str(author.get("given") or author.get("givenName") or "")).strip()
        if literal and not family and not given:
            normalized = CitationAuthor(literal=literal)
        else:
            normalized = CitationAuthor(family=family, given=given, literal=literal if _looks_like_corporate_author_name(literal) else "")
    else:
        text = _collapse_whitespace(str(author or "")).strip(" ,.;:")
        if not text:
            return None
        normalized = _parse_citation_author_text(text)
    if normalized.literal:
        return CitationAuthor(literal=normalized.literal.strip())
    if normalized.family or normalized.given:
        return CitationAuthor(family=normalized.family.strip(), given=normalized.given.strip())
    return None


def _parse_citation_author_text(text: str) -> CitationAuthor:
    candidate = _collapse_whitespace(text).strip(" ,.;:")
    if not candidate:
        return CitationAuthor()
    if _looks_like_corporate_author_name(candidate):
        return CitationAuthor(literal=candidate)
    if "," in candidate:
        family, given = [piece.strip() for piece in candidate.split(",", 1)]
        if family and given and not _looks_like_corporate_author_name(family):
            return CitationAuthor(family=family, given=given)
        return CitationAuthor(literal=candidate)
    words = candidate.split()
    if len(words) == 1:
        return CitationAuthor(literal=candidate)
    family = words[-1]
    given = " ".join(words[:-1]).strip()
    if not family or not given:
        return CitationAuthor(literal=candidate)
    return CitationAuthor(family=family, given=given)


def normalize_citation_authors(value: Any) -> list[CitationAuthor]:
    if value is None:
        return []
    raw_items: list[Any]
    if isinstance(value, str):
        delimiter = ";" if ";" in value else "|"
        raw_items = [item.strip() for item in value.split(delimiter)] if delimiter in value else [value]
    elif isinstance(value, list):
        raw_items = list(value)
    else:
        raw_items = [value]
    authors: list[CitationAuthor] = []
    seen: set[tuple[str, str, str]] = set()
    for item in raw_items:
        normalized = _normalize_citation_author(item)
        if normalized is None:
            continue
        key = (
            normalized.family.lower(),
            normalized.given.lower(),
            normalized.literal.lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        authors.append(normalized)
    return authors


def _clean_doi_candidate(value: str) -> str:
    match = DOI_PATTERN.search(str(value or ""))
    if not match:
        return ""
    return match.group(1).rstrip(".,;:)")


def _extract_publication_year(value: str) -> str:
    match = re.search(r"\b(19|20)\d{2}\b", str(value or ""))
    return match.group(0) if match else ""


def _normalize_date_string(value: str) -> str:
    candidate = _collapse_whitespace(str(value or "")).strip(" ,.;:")
    if not candidate:
        return ""
    iso_match = re.match(r"^(\d{4})[-/](\d{2})[-/](\d{2})", candidate)
    if iso_match:
        return f"{iso_match.group(1)}-{iso_match.group(2)}-{iso_match.group(3)}"
    ym_match = re.match(r"^(\d{4})[-/](\d{2})", candidate)
    if ym_match:
        return f"{ym_match.group(1)}-{ym_match.group(2)}"
    year = _extract_publication_year(candidate)
    if year:
        return year
    return candidate


def _normalize_pages_value(value: Any) -> str:
    text = _collapse_whitespace(str(value or "")).strip(" ,.;:")
    if not text:
        return ""
    match = re.search(r"(\d+)\s*[-\u2013]\s*(\d+)", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return text


def _split_page_range(value: str) -> tuple[str, str]:
    match = re.search(r"^\s*(\S+)\s*[-\u2013]\s*(\S+)\s*$", str(value or ""))
    if not match:
        text = str(value or "").strip()
        return (text, "")
    return match.group(1), match.group(2)


def _citation_has_material_fields(citation: CitationMetadata) -> bool:
    return any(
        [
            citation.item_type,
            citation.title,
            citation.authors,
            citation.issued,
            citation.publisher,
            citation.container_title,
            citation.volume,
            citation.issue,
            citation.pages,
            citation.doi,
            citation.url,
            citation.report_number,
            citation.standard_number,
            citation.language,
            citation.accessed,
            citation.evidence,
        ]
    )


def _normalize_citation_payload(payload: dict[str, Any] | CitationMetadata | None) -> CitationMetadata:
    if isinstance(payload, CitationMetadata):
        raw = payload.model_dump(mode="json")
    elif isinstance(payload, dict):
        raw = dict(payload)
    else:
        raw = {}

    evidence_values = raw.get("evidence") or []
    if isinstance(evidence_values, str):
        evidence_items = [evidence_values]
    elif isinstance(evidence_values, list):
        evidence_items = list(evidence_values)
    else:
        evidence_items = [evidence_values] if evidence_values else []
    missing_field_values = raw.get("missing_fields") or []
    if isinstance(missing_field_values, str):
        missing_items = [item.strip() for item in missing_field_values.split(";") if item.strip()]
    elif isinstance(missing_field_values, list):
        missing_items = list(missing_field_values)
    else:
        missing_items = [missing_field_values] if missing_field_values else []
    blocked_reason_values = raw.get("blocked_reasons") or []
    if isinstance(blocked_reason_values, str):
        blocked_reason_items = [item.strip() for item in blocked_reason_values.split(";") if item.strip()]
    elif isinstance(blocked_reason_values, list):
        blocked_reason_items = list(blocked_reason_values)
    else:
        blocked_reason_items = [blocked_reason_values] if blocked_reason_values else []
    note_values = raw.get("notes") or []
    if isinstance(note_values, str):
        note_items = [item.strip() for item in note_values.split(";") if item.strip()]
    elif isinstance(note_values, list):
        note_items = list(note_values)
    else:
        note_items = [note_values] if note_values else []
    manual_override_values = raw.get("manual_override_fields") or []
    if isinstance(manual_override_values, str):
        manual_override_items = [item.strip() for item in manual_override_values.split(";") if item.strip()]
    elif isinstance(manual_override_values, list):
        manual_override_items = list(manual_override_values)
    else:
        manual_override_items = [manual_override_values] if manual_override_values else []
    raw_field_evidence = raw.get("field_evidence") or {}
    field_evidence: dict[str, CitationFieldEvidence] = {}
    if isinstance(raw_field_evidence, dict):
        for field_name, value in raw_field_evidence.items():
            normalized_field = _normalize_citation_field_evidence(value)
            if (
                normalized_field.value
                or normalized_field.source_type
                or normalized_field.source_label
                or normalized_field.evidence
                or normalized_field.confidence > 0
                or normalized_field.manual_override
            ):
                field_evidence[str(field_name)] = normalized_field

    citation = CitationMetadata(
        item_type=_normalize_citation_type(str(raw.get("item_type") or raw.get("type") or "")),
        title=normalize_generated_title(_stringify_manifest_value(raw.get("title"))),
        authors=normalize_citation_authors(raw.get("authors")),
        issued=_normalize_date_string(_stringify_manifest_value(raw.get("issued") or raw.get("publication_date") or raw.get("date"))),
        publisher=_stringify_manifest_value(raw.get("publisher")),
        container_title=_stringify_manifest_value(raw.get("container_title") or raw.get("journal") or raw.get("publication_title")),
        volume=_stringify_manifest_value(raw.get("volume")),
        issue=_stringify_manifest_value(raw.get("issue")),
        pages=_normalize_pages_value(raw.get("pages")),
        doi=_clean_doi_candidate(_stringify_manifest_value(raw.get("doi"))),
        url=clean_url_candidate(_stringify_manifest_value(raw.get("url"))),
        report_number=_stringify_manifest_value(raw.get("report_number")),
        standard_number=_stringify_manifest_value(raw.get("standard_number")),
        language=_stringify_manifest_value(raw.get("language")),
        accessed=_normalize_date_string(_stringify_manifest_value(raw.get("accessed"))),
        evidence=_dedupe_strings([_stringify_manifest_value(item) for item in evidence_items if _stringify_manifest_value(item)]),
        confidence=_round_citation_confidence(float(raw.get("confidence") or 0.0)),
        missing_fields=[
            _stringify_manifest_value(item)
            for item in missing_items
            if _stringify_manifest_value(item)
        ],
        ready_for_ris=bool(raw.get("ready_for_ris")),
        verification_status=_stringify_manifest_value(raw.get("verification_status")),
        verification_confidence=_round_citation_confidence(float(raw.get("verification_confidence") or 0.0)),
        verification_model=_stringify_manifest_value(raw.get("verification_model")),
        verification_content_digest=_stringify_manifest_value(raw.get("verification_content_digest")),
        verified_at=_stringify_manifest_value(raw.get("verified_at")),
        blocked_reasons=[
            _stringify_manifest_value(item)
            for item in blocked_reason_items
            if _stringify_manifest_value(item)
        ],
        notes=[
            _stringify_manifest_value(item)
            for item in note_items
            if _stringify_manifest_value(item)
        ],
        manual_override_fields=[
            _stringify_manifest_value(item)
            for item in manual_override_items
            if _stringify_manifest_value(item)
        ],
        field_evidence=field_evidence,
    )
    return _finalize_citation_metadata(citation)


def _merge_citation_metadata(
    base: CitationMetadata,
    candidate: dict[str, Any] | CitationMetadata | None,
    *,
    overwrite_existing: bool = False,
) -> CitationMetadata:
    incoming = _normalize_citation_payload(candidate)
    if not _citation_has_material_fields(incoming) and not any(
        [
            incoming.verification_status,
            incoming.blocked_reasons,
            incoming.notes,
            incoming.manual_override_fields,
            incoming.field_evidence,
        ]
    ):
        return _finalize_citation_metadata(base)

    current = base.model_copy(deep=True)
    scalar_fields = (
        "item_type",
        "title",
        "issued",
        "publisher",
        "container_title",
        "volume",
        "issue",
        "pages",
        "doi",
        "url",
        "report_number",
        "standard_number",
        "language",
        "accessed",
    )
    for field_name in scalar_fields:
        current_value = getattr(current, field_name)
        incoming_value = getattr(incoming, field_name)
        if incoming_value and (overwrite_existing or not current_value):
            setattr(current, field_name, incoming_value)
    if incoming.authors and (overwrite_existing or not current.authors):
        current.authors = incoming.authors
    current.evidence = _dedupe_strings([*current.evidence, *incoming.evidence])
    current.confidence = _round_citation_confidence(max(current.confidence, incoming.confidence))
    current.verification_confidence = _round_citation_confidence(
        max(current.verification_confidence, incoming.verification_confidence)
    )
    current.missing_fields = _dedupe_strings([*current.missing_fields, *incoming.missing_fields])
    current.blocked_reasons = _dedupe_strings([*current.blocked_reasons, *incoming.blocked_reasons])
    current.notes = _dedupe_strings([*current.notes, *incoming.notes])
    current.manual_override_fields = _dedupe_strings(
        [*current.manual_override_fields, *incoming.manual_override_fields]
    )
    for field_name, field_data in incoming.field_evidence.items():
        existing = current.field_evidence.get(field_name)
        if (
            overwrite_existing
            or existing is None
            or (not existing.manual_override and field_data.manual_override)
            or (
                not existing.source_type
                and not existing.evidence
                and field_data.confidence >= existing.confidence
            )
        ):
            current.field_evidence[field_name] = field_data
    if incoming.verification_status and (
        overwrite_existing or current.verification_status in {"", "candidate", "blocked", "legacy_unverified"}
    ):
        current.verification_status = incoming.verification_status
    if incoming.verification_model and (overwrite_existing or not current.verification_model):
        current.verification_model = incoming.verification_model
    if incoming.verification_content_digest and (
        overwrite_existing or not current.verification_content_digest
    ):
        current.verification_content_digest = incoming.verification_content_digest
    if incoming.verified_at and (overwrite_existing or not current.verified_at):
        current.verified_at = incoming.verified_at
    return _finalize_citation_metadata(current)


def _finalize_citation_metadata(citation: CitationMetadata) -> CitationMetadata:
    normalized = citation.model_copy(deep=True)
    normalized.item_type = _normalize_citation_type(normalized.item_type)
    normalized.title = normalize_generated_title(normalized.title)
    normalized.issued = _normalize_date_string(normalized.issued)
    normalized.doi = _clean_doi_candidate(normalized.doi)
    normalized.url = clean_url_candidate(normalized.url)
    normalized.pages = _normalize_pages_value(normalized.pages)
    normalized.confidence = _round_citation_confidence(normalized.confidence)
    normalized.verification_confidence = _round_citation_confidence(normalized.verification_confidence)
    normalized.authors = normalize_citation_authors(normalized.authors)
    normalized = _apply_organization_author_fallback(normalized)
    normalized.evidence = _dedupe_strings(normalized.evidence)
    normalized.blocked_reasons = _dedupe_strings(normalized.blocked_reasons)
    normalized.notes = _dedupe_strings(normalized.notes)
    normalized.manual_override_fields = _dedupe_strings(normalized.manual_override_fields)

    normalized_field_evidence: dict[str, CitationFieldEvidence] = {}
    for field_name in CITATION_VERIFIABLE_FIELDS:
        field_data = _normalize_citation_field_evidence(normalized.field_evidence.get(field_name))
        field_data.value = field_data.value or _citation_field_value_text(normalized, field_name)
        if field_name in normalized.manual_override_fields:
            field_data.manual_override = True
            field_data.source_type = field_data.source_type or "manual_override"
            field_data.source_label = field_data.source_label or "Manual override"
            field_data.confidence = max(field_data.confidence, 1.0)
        if field_data.manual_override and field_name not in normalized.manual_override_fields:
            normalized.manual_override_fields.append(field_name)
        normalized_field_evidence[field_name] = _normalize_citation_field_evidence(field_data)
    normalized.manual_override_fields = _dedupe_strings(normalized.manual_override_fields)
    normalized.field_evidence = normalized_field_evidence

    missing_fields: list[str] = []
    for field_name in CITATION_REQUIRED_FIELDS:
        value = getattr(normalized, field_name)
        if field_name == "authors":
            if not normalized.authors:
                missing_fields.append(CITATION_REQUIRED_FIELD_LABELS.get(field_name, field_name))
            continue
        if not str(value or "").strip():
            missing_fields.append(CITATION_REQUIRED_FIELD_LABELS.get(field_name, field_name))
    normalized.missing_fields = _dedupe_strings([*normalized.missing_fields, *missing_fields])
    status = str(normalized.verification_status or "").strip().lower()
    if status == "legacy_complete":
        status = "legacy_unverified"
    if normalized.manual_override_fields and not missing_fields:
        status = "verified"
    elif status == "verified" and missing_fields:
        status = "blocked"
    elif status not in {"verified", "skipped_llm_disabled", "skipped_llm_not_configured", "failed"}:
        if missing_fields or normalized.blocked_reasons:
            status = "blocked" if _citation_has_material_fields(normalized) else status
        elif not status and _citation_has_material_fields(normalized):
            status = "candidate"
        elif status == "legacy_unverified":
            status = "candidate"
    if missing_fields:
        normalized.blocked_reasons = _dedupe_strings(
            [
                *normalized.blocked_reasons,
                f"Missing required citation fields: {', '.join(missing_fields)}",
            ]
        )
    elif status == "verified":
        normalized.blocked_reasons = []
    normalized.verification_status = status
    normalized.ready_for_ris = status == "verified" and len(missing_fields) == 0
    score = 0.0
    if normalized.item_type:
        score += 0.2
    if normalized.title:
        score += 0.2
    if normalized.authors:
        score += 0.2
    if normalized.issued:
        score += 0.15
    if normalized.url:
        score += 0.15
    if normalized.doi:
        score += 0.05
    if normalized.publisher or normalized.container_title:
        score += 0.05
    if normalized.report_number or normalized.standard_number:
        score += 0.05
    normalized.confidence = _round_citation_confidence(
        max(normalized.confidence, normalized.verification_confidence, score)
    )
    if normalized.verification_status == "verified":
        normalized.verification_confidence = _round_citation_confidence(
            max(normalized.verification_confidence, normalized.confidence)
        )
    return normalized


def _effective_manual_override_fields(citation: CitationMetadata) -> list[str]:
    fields = list(citation.manual_override_fields)
    fields.extend(
        field_name
        for field_name, field_data in citation.field_evidence.items()
        if field_data.manual_override
    )
    return _dedupe_strings(fields)


def _apply_citation_manual_overrides(
    citation: CitationMetadata,
    manual_source: CitationMetadata | dict[str, Any] | None,
) -> CitationMetadata:
    manual_citation = _normalize_citation_payload(manual_source)
    override_fields = _effective_manual_override_fields(manual_citation)
    if not override_fields:
        return _finalize_citation_metadata(citation)

    updated = citation.model_copy(deep=True)
    for field_name in override_fields:
        if field_name == "authors":
            updated.authors = manual_citation.authors
        elif hasattr(updated, field_name):
            setattr(updated, field_name, getattr(manual_citation, field_name))
        field_data = _normalize_citation_field_evidence(manual_citation.field_evidence.get(field_name))
        field_data.value = _citation_field_value_text(manual_citation, field_name)
        field_data.manual_override = True
        field_data.source_type = field_data.source_type or "manual_override"
        field_data.source_label = field_data.source_label or "Manual override"
        field_data.confidence = max(field_data.confidence, 1.0)
        updated.field_evidence[field_name] = field_data
    updated.manual_override_fields = _dedupe_strings(
        [*updated.manual_override_fields, *override_fields]
    )
    return _finalize_citation_metadata(updated)


def _citation_verification_is_current(citation: CitationMetadata, source_digest: str) -> bool:
    status = str(citation.verification_status or "").strip().lower()
    if not source_digest:
        return False
    if str(citation.verification_content_digest or "").strip() != source_digest:
        return False
    return status in {
        "verified",
        "blocked",
        "skipped_llm_disabled",
        "skipped_llm_not_configured",
        "failed",
    }


def _citation_payload_to_catalog_fields(citation: CitationMetadata) -> dict[str, Any]:
    author_names = _citation_author_names(citation)
    return {
        "title": citation.title,
        "author_names": author_names,
        "publication_date": citation.issued,
        "publication_year": _citation_publication_year(citation),
        "document_type": citation.item_type,
        "organization_name": citation.publisher,
    }


def _extract_first_html_title(html_text: str) -> str:
    match = HTML_TITLE_PATTERN.search(html_text or "")
    if not match:
        return ""
    return _collapse_whitespace(html.unescape(match.group(1))).strip(" ,.;:")


def _parse_html_meta_tags(html_text: str) -> dict[str, list[str]]:
    values: dict[str, list[str]] = {}
    for key, content in HTML_META_TAG_PATTERN.findall(html_text or ""):
        normalized_key = _collapse_whitespace(str(key or "")).strip().lower()
        normalized_content = _collapse_whitespace(html.unescape(str(content or ""))).strip()
        if not normalized_key or not normalized_content:
            continue
        values.setdefault(normalized_key, []).append(normalized_content)
    return values


def _parse_html_json_ld(html_text: str) -> list[dict[str, Any]]:
    payloads: list[dict[str, Any]] = []
    for raw in HTML_JSON_LD_PATTERN.findall(html_text or ""):
        candidate = str(raw or "").strip()
        if not candidate:
            continue
        try:
            parsed = json.loads(candidate)
        except Exception:
            continue
        payloads.extend(_flatten_json_ld_nodes(parsed))
    return payloads


def _flatten_json_ld_nodes(value: Any) -> list[dict[str, Any]]:
    nodes: list[dict[str, Any]] = []
    if isinstance(value, dict):
        if "@graph" in value and isinstance(value["@graph"], list):
            for item in value["@graph"]:
                nodes.extend(_flatten_json_ld_nodes(item))
        else:
            nodes.append(value)
    elif isinstance(value, list):
        for item in value:
            nodes.extend(_flatten_json_ld_nodes(item))
    return nodes


def _select_json_ld_citation_node(nodes: list[dict[str, Any]]) -> dict[str, Any]:
    scored: list[tuple[int, dict[str, Any]]] = []
    for node in nodes:
        raw_type = node.get("@type") or node.get("type") or ""
        types = [str(item).lower() for item in raw_type] if isinstance(raw_type, list) else [str(raw_type).lower()]
        score = 0
        if any(item in {"scholarlyarticle", "article", "report", "webpage", "book", "thesis", "dataset"} for item in types):
            score += 4
        if node.get("author"):
            score += 1
        if node.get("headline") or node.get("name"):
            score += 1
        scored.append((score, node))
    if not scored:
        return {}
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1]


def extract_html_citation_metadata(html_text: str, *, base_url: str = "") -> dict[str, Any]:
    if not html_text.strip():
        return {}

    meta = _parse_html_meta_tags(html_text)
    title = ""
    authors: list[str] = []
    issued = ""
    doi = ""
    url = ""
    publisher = ""
    container_title = ""
    volume = ""
    issue = ""
    first_page = ""
    last_page = ""
    item_type = ""
    language = ""
    report_number = ""
    standard_number = ""
    evidence: list[str] = []

    def first_value(*keys: str) -> str:
        for key in keys:
            values = meta.get(key.lower()) or []
            for value in values:
                if value:
                    return value
        return ""

    def all_values(*keys: str) -> list[str]:
        items: list[str] = []
        for key in keys:
            items.extend(meta.get(key.lower()) or [])
        return _dedupe_strings(items)

    title = first_value("citation_title", "dc.title", "og:title", "twitter:title") or _extract_first_html_title(html_text)
    authors = all_values("citation_author", "author", "dc.creator", "article:author")
    issued = first_value(
        "citation_publication_date",
        "citation_online_date",
        "dc.date",
        "article:published_time",
        "og:updated_time",
    )
    doi = _clean_doi_candidate(first_value("citation_doi", "dc.identifier", "doi"))
    url = clean_url_candidate(first_value("citation_public_url", "og:url", "twitter:url")) or clean_url_candidate(base_url)
    publisher = first_value("citation_publisher", "publisher", "dc.publisher", "og:site_name")
    container_title = first_value(
        "citation_journal_title",
        "citation_conference_title",
        "citation_inbook_title",
        "citation_collection_title",
        "citation_dissertation_institution",
    )
    volume = first_value("citation_volume")
    issue = first_value("citation_issue")
    first_page = first_value("citation_firstpage")
    last_page = first_value("citation_lastpage")
    item_type = first_value("citation_type", "dc.type", "og:type")
    language = first_value("citation_language", "dc.language", "og:locale")
    report_number = first_value("citation_technical_report_number", "citation_report_number")
    standard_number = first_value("citation_standard_number")

    if not standard_number:
        standard_number = _first_pattern_match(STANDARD_NUMBER_PATTERN, html_text)
    if not report_number:
        report_number = _first_pattern_match(REPORT_NUMBER_PATTERN, html_text)

    json_ld_node = _select_json_ld_citation_node(_parse_html_json_ld(html_text))
    if json_ld_node:
        title = title or _stringify_manifest_value(json_ld_node.get("headline") or json_ld_node.get("name"))
        json_ld_authors = _extract_json_ld_authors(json_ld_node.get("author"))
        if json_ld_authors and not authors:
            authors = json_ld_authors
        issued = issued or _stringify_manifest_value(
            json_ld_node.get("datePublished") or json_ld_node.get("dateCreated")
        )
        publisher = publisher or _extract_json_ld_publisher(json_ld_node.get("publisher"))
        container_title = container_title or _extract_json_ld_container_title(json_ld_node.get("isPartOf"))
        doi = doi or _clean_doi_candidate(
            _extract_json_ld_identifier(json_ld_node.get("identifier"))
        )
        url = url or clean_url_candidate(_stringify_manifest_value(json_ld_node.get("url")))
        volume = volume or _stringify_manifest_value(json_ld_node.get("volumeNumber"))
        issue = issue or _stringify_manifest_value(json_ld_node.get("issueNumber"))
        if not first_page and json_ld_node.get("pageStart"):
            first_page = _stringify_manifest_value(json_ld_node.get("pageStart"))
        if not last_page and json_ld_node.get("pageEnd"):
            last_page = _stringify_manifest_value(json_ld_node.get("pageEnd"))
        item_type = item_type or _normalize_json_ld_type(json_ld_node.get("@type"))
        language = language or _stringify_manifest_value(json_ld_node.get("inLanguage"))
        report_number = report_number or _stringify_manifest_value(
            json_ld_node.get("reportNumber") or json_ld_node.get("identifier")
        )
        evidence.extend(
            _dedupe_strings(
                [
                    _stringify_manifest_value(json_ld_node.get("headline") or json_ld_node.get("name")),
                    _stringify_manifest_value(json_ld_node.get("datePublished")),
                    _extract_json_ld_publisher(json_ld_node.get("publisher")),
                ]
            )
        )

    if trafilatura is not None:
        try:
            meta_doc = trafilatura.extract_metadata(html_text, default_url=base_url or None)
        except Exception:
            meta_doc = None
        if meta_doc is not None:
            title = title or _stringify_manifest_value(getattr(meta_doc, "title", ""))
            issued = issued or _stringify_manifest_value(getattr(meta_doc, "date", ""))
            publisher = publisher or _stringify_manifest_value(getattr(meta_doc, "sitename", ""))
            if not authors:
                authors = _dedupe_strings(
                    [item.strip() for item in re.split(r"[;|]", str(getattr(meta_doc, "author", "") or "")) if item.strip()]
                )
            evidence.extend(
                _dedupe_strings(
                    [
                        _stringify_manifest_value(getattr(meta_doc, "title", "")),
                        _stringify_manifest_value(getattr(meta_doc, "author", "")),
                        _stringify_manifest_value(getattr(meta_doc, "date", "")),
                    ]
                )
            )

    pages = f"{first_page}-{last_page}" if first_page and last_page else first_page or last_page
    return {
        "title": title,
        "authors": authors,
        "issued": issued,
        "publisher": publisher,
        "container_title": container_title,
        "volume": volume,
        "issue": issue,
        "pages": pages,
        "doi": doi,
        "url": url or clean_url_candidate(base_url),
        "report_number": report_number,
        "standard_number": standard_number,
        "language": language,
        "item_type": _normalize_citation_type(item_type),
        "evidence": _dedupe_strings([title, publisher, container_title, issued, *authors, *evidence])[:8],
    }


def _read_catalog_html_input(
    row: SourceManifestRow,
    *,
    read_text: Callable[[Path], str],
) -> str:
    for rel_path in (row.raw_file, row.rendered_file):
        candidate = Path(str(rel_path or ""))
        if not candidate.suffix.lower() in {".html", ".htm"}:
            continue
        html_text = read_text(candidate)
        if html_text.strip():
            return html_text
    return ""


def _extract_json_ld_authors(value: Any) -> list[str]:
    authors: list[str] = []
    items = value if isinstance(value, list) else [value]
    for item in items:
        if isinstance(item, dict):
            name = _stringify_manifest_value(item.get("name"))
            if not name:
                family = _stringify_manifest_value(item.get("familyName"))
                given = _stringify_manifest_value(item.get("givenName"))
                name = ", ".join(part for part in [family, given] if part).strip(", ")
            if name:
                authors.append(name)
        else:
            name = _stringify_manifest_value(item)
            if name:
                authors.append(name)
    return _dedupe_strings(authors)


def _extract_json_ld_publisher(value: Any) -> str:
    if isinstance(value, dict):
        return _stringify_manifest_value(value.get("name"))
    return _stringify_manifest_value(value)


def _extract_json_ld_container_title(value: Any) -> str:
    if isinstance(value, dict):
        return _stringify_manifest_value(value.get("name"))
    return ""


def _extract_json_ld_identifier(value: Any) -> str:
    if isinstance(value, dict):
        return _stringify_manifest_value(value.get("value") or value.get("@id") or value.get("identifier"))
    if isinstance(value, list):
        for item in value:
            identifier = _extract_json_ld_identifier(item)
            if identifier:
                return identifier
        return ""
    return _stringify_manifest_value(value)


def _normalize_json_ld_type(value: Any) -> str:
    if isinstance(value, list):
        for item in value:
            normalized = _normalize_json_ld_type(item)
            if normalized:
                return normalized
        return ""
    raw = _stringify_manifest_value(value).lower()
    if not raw:
        return ""
    if "scholarlyarticle" in raw or "article" == raw:
        return "journal article"
    if "webpage" in raw or "website" in raw:
        return "web page"
    if "report" in raw:
        return "report"
    if "book" in raw:
        return "book"
    if "thesis" in raw or "dissertation" in raw:
        return "thesis"
    if "dataset" in raw:
        return "dataset"
    return ""


def _first_pattern_match(pattern: re.Pattern[str], text: str) -> str:
    match = pattern.search(text or "")
    if not match:
        return ""
    groups = [item for item in match.groups() if item] if match.groups() else [match.group(0)]
    return _collapse_whitespace(str(groups[0] if groups else match.group(0))).strip(" ,.;:")


def _extract_doi_from_text_sources(*values: str) -> str:
    for value in values:
        doi = _clean_doi_candidate(value)
        if doi:
            return doi
    return ""


def _build_citation_metadata(
    *,
    row: SourceManifestRow,
    title: str,
    author_names: str,
    publication_date: str,
    document_type: str,
    organization_name: str,
    html_metadata: dict[str, Any] | None = None,
) -> CitationMetadata:
    html_payload = _normalize_citation_payload(html_metadata or {})
    base_url = clean_url_candidate(row.original_url or row.final_url or "")
    citation = CitationMetadata(
        item_type=_normalize_citation_type(
            html_payload.item_type or document_type,
            fallback_document_type=document_type,
            source_kind=row.source_kind,
            url=base_url,
        ),
        title=html_payload.title or title,
        authors=html_payload.authors or normalize_citation_authors(author_names),
        issued=html_payload.issued or publication_date,
        publisher=html_payload.publisher or organization_name,
        container_title=html_payload.container_title,
        volume=html_payload.volume,
        issue=html_payload.issue,
        pages=html_payload.pages,
        doi=html_payload.doi or _extract_doi_from_text_sources(base_url, row.notes, row.title),
        url=html_payload.url or base_url,
        report_number=html_payload.report_number or _first_pattern_match(REPORT_NUMBER_PATTERN, title),
        standard_number=html_payload.standard_number or _first_pattern_match(STANDARD_NUMBER_PATTERN, title),
        language=html_payload.language,
        evidence=_dedupe_strings(
            [
                title,
                author_names,
                publication_date,
                organization_name,
                *(html_payload.evidence or []),
            ]
        )[:10],
        confidence=html_payload.confidence,
    )
    return _finalize_citation_metadata(citation)


def _citation_needs_llm_review(citation: CitationMetadata) -> bool:
    return (
        str(citation.verification_status or "").strip().lower() != "verified"
        or citation.verification_confidence < 0.75
    )


def _enrich_citation_from_csl_json(citation: CitationMetadata, payload: dict[str, Any]) -> CitationMetadata:
    title_values = payload.get("title")
    title = ""
    if isinstance(title_values, list):
        title = _stringify_manifest_value(title_values[0] if title_values else "")
    else:
        title = _stringify_manifest_value(title_values)

    container_values = payload.get("container-title")
    container_title = ""
    if isinstance(container_values, list):
        container_title = _stringify_manifest_value(container_values[0] if container_values else "")
    else:
        container_title = _stringify_manifest_value(container_values)

    issued = _csl_date_parts_to_string(payload.get("issued"))
    url = clean_url_candidate(_stringify_manifest_value(payload.get("URL") or payload.get("url")))
    doi = _clean_doi_candidate(_stringify_manifest_value(payload.get("DOI") or payload.get("doi")))
    authors = normalize_citation_authors(payload.get("author"))
    citation_payload = {
        "item_type": _normalize_citation_type(_stringify_manifest_value(payload.get("type"))),
        "title": title,
        "authors": [author.model_dump(mode="json") for author in authors],
        "issued": issued,
        "publisher": _stringify_manifest_value(payload.get("publisher")),
        "container_title": container_title,
        "volume": _stringify_manifest_value(payload.get("volume")),
        "issue": _stringify_manifest_value(payload.get("issue")),
        "pages": _normalize_pages_value(payload.get("page")),
        "doi": doi,
        "url": url,
        "language": _stringify_manifest_value(payload.get("language")),
        "report_number": _extract_report_number_from_payload(payload),
        "standard_number": _extract_standard_number_from_payload(payload),
        "evidence": [title, container_title, doi, url],
        "confidence": 0.95,
    }
    return _merge_citation_metadata(citation, citation_payload)


def _csl_date_parts_to_string(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    date_parts = value.get("date-parts")
    if not isinstance(date_parts, list) or not date_parts:
        return ""
    first = date_parts[0]
    if not isinstance(first, list) or not first:
        return ""
    pieces = [str(item) for item in first[:3]]
    if len(pieces) >= 3:
        return f"{pieces[0]}-{str(pieces[1]).zfill(2)}-{str(pieces[2]).zfill(2)}"
    if len(pieces) == 2:
        return f"{pieces[0]}-{str(pieces[1]).zfill(2)}"
    return pieces[0]


def _extract_report_number_from_payload(payload: dict[str, Any]) -> str:
    for key in ("number", "report-number", "reportNumber"):
        value = _stringify_manifest_value(payload.get(key))
        if value:
            return value
    identifiers = payload.get("identifier")
    if isinstance(identifiers, list):
        for item in identifiers:
            value = _stringify_manifest_value(item)
            if value and REPORT_NUMBER_PATTERN.search(value):
                return _first_pattern_match(REPORT_NUMBER_PATTERN, value)
    return ""


def _extract_standard_number_from_payload(payload: dict[str, Any]) -> str:
    for key in ("standard_number", "standardNumber"):
        value = _stringify_manifest_value(payload.get(key))
        if value:
            return value
    for key in ("number", "identifier"):
        value = payload.get(key)
        if isinstance(value, list):
            for item in value:
                text = _stringify_manifest_value(item)
                if STANDARD_NUMBER_PATTERN.search(text):
                    return _first_pattern_match(STANDARD_NUMBER_PATTERN, text)
        text = _stringify_manifest_value(value)
        if STANDARD_NUMBER_PATTERN.search(text):
            return _first_pattern_match(STANDARD_NUMBER_PATTERN, text)
    return ""


def _resolve_doi_citation_metadata(doi: str) -> CitationMetadata:
    cleaned = _clean_doi_candidate(doi)
    if not cleaned:
        return CitationMetadata()
    user_agent = "ResearchAssistant/1.0 (+https://local.app)"
    headers = {"Accept": DOI_CSL_ACCEPT_HEADER, "User-Agent": user_agent}
    try:
        response = httpx.get(
            f"https://doi.org/{quote(cleaned, safe='')}",
            headers=headers,
            timeout=HTTP_TIMEOUT_SECONDS,
            follow_redirects=True,
        )
        if response.status_code < 400:
            payload = response.json()
            if isinstance(payload, dict):
                return _enrich_citation_from_csl_json(CitationMetadata(doi=cleaned), payload)
    except Exception:
        pass

    for url in (
        f"https://api.crossref.org/works/{quote(cleaned, safe='')}",
        f"https://api.datacite.org/dois/{quote(cleaned, safe='')}",
    ):
        try:
            response = httpx.get(
                url,
                headers={"User-Agent": user_agent},
                timeout=HTTP_TIMEOUT_SECONDS,
                follow_redirects=True,
            )
            if response.status_code >= 400:
                continue
            payload = response.json()
        except Exception:
            continue
        candidate = _citation_from_agency_payload(cleaned, payload)
        if candidate.ready_for_ris or candidate.confidence >= 0.85:
            return candidate
    return CitationMetadata(doi=cleaned)


def _citation_from_agency_payload(doi: str, payload: Any) -> CitationMetadata:
    if not isinstance(payload, dict):
        return CitationMetadata(doi=doi)
    if isinstance(payload.get("message"), dict):
        return _enrich_citation_from_csl_json(CitationMetadata(doi=doi), payload["message"])
    data = payload.get("data")
    if not isinstance(data, dict):
        return CitationMetadata(doi=doi)
    attributes = data.get("attributes")
    if not isinstance(attributes, dict):
        return CitationMetadata(doi=doi)
    title = ""
    titles = attributes.get("titles")
    if isinstance(titles, list) and titles:
        first_title = titles[0]
        if isinstance(first_title, dict):
            title = _stringify_manifest_value(first_title.get("title"))
    creators = attributes.get("creators")
    authors: list[dict[str, str]] = []
    if isinstance(creators, list):
        for creator in creators:
            if not isinstance(creator, dict):
                continue
            authors.append(
                {
                    "family": _stringify_manifest_value(creator.get("familyName")),
                    "given": _stringify_manifest_value(creator.get("givenName")),
                    "literal": _stringify_manifest_value(creator.get("name")),
                }
            )
    citation_payload = {
        "item_type": _normalize_citation_type(
            _stringify_manifest_value(
                (attributes.get("types") or {}).get("resourceTypeGeneral")
                if isinstance(attributes.get("types"), dict)
                else ""
            )
        ),
        "title": title,
        "authors": authors,
        "issued": _stringify_manifest_value(attributes.get("publicationYear")),
        "publisher": _stringify_manifest_value(attributes.get("publisher")),
        "container_title": "",
        "doi": doi,
        "url": _stringify_manifest_value(attributes.get("url")),
        "language": _stringify_manifest_value(attributes.get("language")),
        "report_number": _extract_report_number_from_payload(attributes),
        "standard_number": _extract_standard_number_from_payload(attributes),
        "confidence": 0.9,
        "evidence": [title, _stringify_manifest_value(attributes.get("publisher")), doi],
    }
    return _normalize_citation_payload(citation_payload)


def build_ris_records(rows: list[SourceManifestRow], *, base_dir: Path | None = None) -> tuple[str, int, int]:
    records: list[str] = []
    skipped = 0
    for row in rows:
        catalog_payload = _load_manifest_json(base_dir, row.catalog_file)
        citation = _coerce_citation_metadata(catalog_payload.get("citation"))
        citation = _finalize_citation_metadata(citation)
        if not citation.ready_for_ris:
            skipped += 1
            continue
        records.append(build_ris_record(citation))
    return ("\r\n".join(record.rstrip("\r\n") for record in records) + ("\r\n" if records else "")), len(records), skipped


def build_ris_record(citation: CitationMetadata) -> str:
    lines: list[str] = []
    ris_type = RIS_TYPE_MAP.get(citation.item_type, "GEN")
    lines.append(f"TY  - {ris_type}")
    for author in citation.authors:
        if author.literal:
            lines.append(f"AU  - {author.literal}")
        else:
            formatted = ", ".join(part for part in [author.family, author.given] if part)
            if formatted:
                lines.append(f"AU  - {formatted}")
    lines.append(f"TI  - {citation.title}")
    year = _citation_publication_year(citation)
    if year:
        lines.append(f"PY  - {year}")
    if citation.issued:
        lines.append(f"DA  - {citation.issued.replace('-', '/')}")
    if citation.item_type == "journal article":
        if citation.container_title:
            lines.append(f"T2  - {citation.container_title}")
            lines.append(f"JO  - {citation.container_title}")
            lines.append(f"JF  - {citation.container_title}")
    elif citation.item_type in {"web page", "report", "conference paper", "book chapter"} and citation.container_title:
        lines.append(f"T2  - {citation.container_title}")
    if citation.publisher:
        lines.append(f"PB  - {citation.publisher}")
    if citation.volume:
        lines.append(f"VL  - {citation.volume}")
    if citation.issue:
        lines.append(f"IS  - {citation.issue}")
    start_page, end_page = _split_page_range(citation.pages)
    if start_page:
        lines.append(f"SP  - {start_page}")
    if end_page:
        lines.append(f"EP  - {end_page}")
    if citation.doi:
        lines.append(f"DO  - {citation.doi}")
    if citation.url:
        lines.append(f"UR  - {citation.url}")
    if citation.report_number:
        lines.append(f"SN  - {citation.report_number}")
    if citation.standard_number:
        lines.append("M3  - Standard")
        lines.append(f"VO  - {citation.standard_number}")
    lines.append("ER  -")
    return "\r\n".join(lines)


def _count_fetch_outcomes(rows: list[SourceManifestRow]) -> dict[str, int]:
    counts = {"success": 0, "failed": 0, "partial": 0}
    for row in rows:
        outcome = _row_task_outcome(row)
        if outcome == "success":
            counts["success"] += 1
        elif outcome == "partial":
            counts["partial"] += 1
        elif outcome == "failed":
            counts["failed"] += 1
    return counts


def summarize_output_rows(rows: list[SourceManifestRow]) -> SourceOutputSummary:
    catalog_missing = 0
    catalog_failed = 0
    markdown_ready = 0
    summary_missing = 0
    summary_failed = 0
    rating_missing = 0
    rating_failed = 0
    llm_cleanup_failed = 0

    for row in rows:
        if row.markdown_file or row.llm_cleanup_file:
            markdown_ready += 1
        if (row.catalog_status or "").strip().lower() == "failed":
            catalog_failed += 1
        if (row.summary_status or "").strip().lower() == "failed":
            summary_failed += 1
        if (row.rating_status or "").strip().lower() == "failed":
            rating_failed += 1
        if (row.llm_cleanup_status or "").strip().lower() == "failed":
            llm_cleanup_failed += 1

        has_catalog = bool(row.catalog_file)
        if (row.markdown_file or row.llm_cleanup_file) and not has_catalog:
            catalog_missing += 1

        has_summary = bool(row.summary_file)
        if (row.markdown_file or row.llm_cleanup_file) and not has_summary:
            summary_missing += 1

        has_rating = bool(row.rating_file)
        if (row.markdown_file or row.llm_cleanup_file) and not has_rating:
            rating_missing += 1

    return SourceOutputSummary(
        total_rows=len(rows),
        raw_file_count=sum(1 for row in rows if row.raw_file),
        rendered_html_count=sum(1 for row in rows if row.rendered_file),
        rendered_pdf_count=sum(1 for row in rows if row.rendered_pdf_file),
        markdown_count=sum(1 for row in rows if row.markdown_file),
        llm_cleanup_file_count=sum(1 for row in rows if row.llm_cleanup_file),
        llm_cleanup_needed_count=sum(1 for row in rows if row.llm_cleanup_needed),
        llm_cleanup_failed_count=llm_cleanup_failed,
        catalog_file_count=sum(1 for row in rows if row.catalog_file),
        catalog_missing_count=catalog_missing,
        catalog_failed_count=catalog_failed,
        summary_file_count=sum(1 for row in rows if row.summary_file),
        summary_missing_count=summary_missing,
        summary_failed_count=summary_failed,
        rating_file_count=sum(1 for row in rows if row.rating_file),
        rating_missing_count=rating_missing,
        rating_failed_count=rating_failed,
    )


def parse_notes(value: str) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in value.split(";") if part.strip()]


def _phase_error_code(value: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    code = normalized.split(":", 1)[0].strip().lower()
    code = re.sub(r"[^a-z0-9_]+", "_", code)
    return code.strip("_")


def _phase_completion_status(
    value: str,
    *,
    success: set[str],
    failed: set[str],
    skipped: set[str] | None = None,
) -> str:
    normalized = str(value or "").strip().lower()
    skipped_values = skipped or set()
    if normalized in success:
        return "completed"
    if normalized in failed:
        return "failed"
    if normalized in skipped_values or normalized.startswith("skipped"):
        return "skipped"
    return "pending"


def _dedupe_strings(values: list[str]) -> list[str]:
    deduped: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        lowered = text.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        deduped.append(text)
    return deduped


def _phase_status_outcome(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if normalized in {"success", "generated", "existing", "completed", "extracted"}:
        return "success"
    if normalized in {"partial", "stale"}:
        return "partial"
    if normalized in {"not_applicable", "skipped", "not_requested"} or normalized.startswith("skipped"):
        return "skipped"
    if normalized in {
        "failed",
        "missing_markdown",
        "missing_project_profile",
        "invalid_url",
        "timeout",
        "network_failure",
        "unsupported_content",
    }:
        return "failed"
    return "pending"


def _row_task_outcome(
    row: SourceManifestRow,
    requested_phases: list[str] | None = None,
) -> str:
    phases = requested_phases or [PHASE_FETCH]
    outcomes: list[str] = []
    for phase in phases:
        if phase == PHASE_FETCH:
            outcomes.append(_phase_status_outcome(row.fetch_status))
        elif phase == PHASE_CONVERT:
            metadata = _get_phase_metadata(row, PHASE_CONVERT)
            if metadata is not None:
                outcomes.append(_phase_status_outcome(metadata.status))
            elif row.llm_cleanup_file or row.markdown_file:
                outcomes.append("success")
            else:
                outcomes.append("pending")
        elif phase == PHASE_CLEANUP:
            metadata = _get_phase_metadata(row, PHASE_CLEANUP)
            if metadata is not None:
                outcomes.append(_phase_status_outcome(metadata.status))
            else:
                outcomes.append(_phase_status_outcome(row.llm_cleanup_status))
        elif phase == PHASE_TITLE:
            metadata = _get_phase_metadata(row, PHASE_TITLE)
            if metadata is not None:
                outcomes.append(_phase_status_outcome(metadata.status))
            else:
                outcomes.append(_phase_status_outcome(row.title_status))
        elif phase == PHASE_CATALOG:
            metadata = _get_phase_metadata(row, PHASE_CATALOG)
            if metadata is not None:
                outcomes.append(_phase_status_outcome(metadata.status))
            else:
                outcomes.append(_phase_status_outcome(row.catalog_status))
        elif phase == PHASE_CITATION_VERIFY:
            metadata = _get_phase_metadata(row, PHASE_CITATION_VERIFY)
            if metadata is not None:
                outcomes.append(_phase_status_outcome(metadata.status))
            else:
                outcomes.append(_phase_status_outcome(_row_citation_verification_status(row)))
        elif phase == PHASE_SUMMARY:
            metadata = _get_phase_metadata(row, PHASE_SUMMARY)
            if metadata is not None:
                outcomes.append(_phase_status_outcome(metadata.status))
            else:
                outcomes.append(_phase_status_outcome(row.summary_status))
        elif phase == PHASE_RATING:
            metadata = _get_phase_metadata(row, PHASE_RATING)
            if metadata is not None:
                outcomes.append(_phase_status_outcome(metadata.status))
            else:
                outcomes.append(_phase_status_outcome(row.rating_status))

    if any(item == "failed" for item in outcomes):
        return "failed"
    if any(item == "partial" for item in outcomes) or any(
        item == "pending" for item in outcomes
    ):
        return "partial"
    if outcomes and all(item == "skipped" for item in outcomes):
        return "skipped"
    return "success"


def _catalog_missing_core_fields(row: SourceManifestRow) -> bool:
    return any(
        not str(value or "").strip()
        for value in (
            row.title,
            row.author_names,
            row.document_type,
            row.organization_name,
            row.organization_type,
        )
    )


def _merge_catalog_payload_into_row(
    row: SourceManifestRow,
    payload: dict[str, Any],
    *,
    overwrite_existing: bool = False,
) -> None:
    def should_set(current: str) -> bool:
        return overwrite_existing or not str(current or "").strip()

    title = normalize_generated_title(_stringify_manifest_value(payload.get("title")))
    title_basis = str(payload.get("title_basis") or payload.get("basis") or "").strip().lower()
    if title and should_set(row.title):
        row.title = limit_title_words(title, 10) if title_basis == "generated" else title
        row.title_status = "generated" if title_basis == "generated" else "extracted"

    author_names = payload.get("author_names")
    normalized_authors = ""
    if isinstance(author_names, list):
        normalized_authors = "; ".join(
            _dedupe_strings([str(item).strip() for item in author_names if str(item).strip()])
        )
    elif isinstance(author_names, str):
        normalized_authors = "; ".join(
            _dedupe_strings([item.strip() for item in re.split(r"[;|,]", author_names) if item.strip()])
        )
    if normalized_authors and should_set(row.author_names):
        row.author_names = normalized_authors

    publication_date = _stringify_manifest_value(payload.get("publication_date"))
    if publication_date and should_set(row.publication_date):
        row.publication_date = publication_date

    publication_year = _stringify_manifest_value(payload.get("publication_year"))
    if not publication_year and publication_date:
        match = re.search(r"\b(19|20)\d{2}\b", publication_date)
        if match:
            publication_year = match.group(0)
    if publication_year and should_set(row.publication_year):
        row.publication_year = publication_year

    document_type = _stringify_manifest_value(payload.get("document_type"))
    if document_type and should_set(row.document_type):
        row.document_type = document_type

    organization_name = _stringify_manifest_value(payload.get("organization_name"))
    if organization_name and should_set(row.organization_name):
        row.organization_name = organization_name

    organization_type = _stringify_manifest_value(payload.get("organization_type"))
    if organization_type and should_set(row.organization_type):
        row.organization_type = organization_type

def _build_deterministic_catalog_metadata(
    *,
    row: SourceManifestRow,
    markdown_text: str,
    html_text: str = "",
) -> dict[str, Any]:
    front_matter = _parse_markdown_front_matter(markdown_text)
    lines = [line.strip() for line in markdown_text.splitlines() if line.strip()]
    title_candidate = extract_markdown_title_candidate(markdown_text) or _stringify_manifest_value(
        front_matter.get("title")
    )
    author_names = _front_matter_list(front_matter, ("authors", "author", "byline"))
    if not author_names:
        author_names = _extract_byline_authors(lines[:10])

    publication_date = _front_matter_scalar(
        front_matter,
        ("date", "published", "publication_date", "updated"),
    )
    if not publication_date:
        publication_date = _extract_publication_date(lines[:30])

    publication_year = ""
    if publication_date:
        match = re.search(r"\b(19|20)\d{2}\b", publication_date)
        if match:
            publication_year = match.group(0)
    if not publication_year:
        publication_year = _extract_publication_year(markdown_text)

    organization_name = _front_matter_scalar(
        front_matter,
        ("organization", "publisher", "site_name", "institution", "company"),
    )
    if not organization_name:
        organization_name = _organization_name_from_url(row.original_url or row.final_url)

    document_type = _infer_document_type(row=row, markdown_text=markdown_text)
    html_metadata = extract_html_citation_metadata(html_text, base_url=row.original_url or row.final_url)
    if html_metadata.get("authors") and not author_names:
        author_names = _citation_author_names(_normalize_citation_payload(html_metadata))
    if html_metadata.get("issued") and not publication_date:
        publication_date = _normalize_date_string(_stringify_manifest_value(html_metadata.get("issued")))
    if html_metadata.get("publisher") and not organization_name:
        organization_name = _stringify_manifest_value(html_metadata.get("publisher"))
    if html_metadata.get("item_type"):
        document_type = _normalize_citation_type(
            _stringify_manifest_value(html_metadata.get("item_type")),
            fallback_document_type=document_type,
            source_kind=row.source_kind,
            url=row.original_url or row.final_url,
        ) or document_type
    if publication_date:
        publication_year = _extract_publication_year(publication_date) or publication_year
    organization_type = _infer_organization_type(
        organization_name=organization_name,
        url=row.original_url or row.final_url,
        source_kind=row.source_kind,
    )

    citation = _build_citation_metadata(
        row=row,
        title=title_candidate,
        author_names=author_names,
        publication_date=publication_date,
        document_type=document_type,
        organization_name=organization_name,
        html_metadata=html_metadata,
    )

    evidence_snippets = _dedupe_strings(
        [
            title_candidate,
            _stringify_manifest_value(front_matter.get("title")),
            _stringify_manifest_value(front_matter.get("authors")),
            _stringify_manifest_value(front_matter.get("date")),
            *(html_metadata.get("evidence") or []),
            *(lines[:3]),
        ]
    )[:5]

    return {
        "title": title_candidate,
        "author_names": author_names,
        "publication_date": publication_date,
        "publication_year": publication_year,
        "document_type": document_type,
        "organization_name": organization_name,
        "organization_type": organization_type,
        "evidence_snippets": evidence_snippets,
        "citation": citation.model_dump(mode="json"),
    }


def _parse_markdown_front_matter(markdown_text: str) -> dict[str, Any]:
    lines = markdown_text.splitlines()
    if not lines or lines[0].strip() != "---":
        return {}
    payload: dict[str, Any] = {}
    current_key = ""
    for line in lines[1:]:
        stripped = line.strip()
        if stripped == "---":
            break
        match = re.match(r"^([A-Za-z0-9_-]+)\s*:\s*(.*)$", line)
        if match:
            current_key = match.group(1).strip().lower()
            raw_value = match.group(2).strip()
            if not raw_value:
                payload[current_key] = []
            elif raw_value.startswith("[") and raw_value.endswith("]"):
                payload[current_key] = [
                    item.strip().strip("'\"")
                    for item in raw_value[1:-1].split(",")
                    if item.strip()
                ]
            else:
                payload[current_key] = raw_value.strip("'\"")
            continue
        if current_key and re.match(r"^\s*-\s+.+$", line):
            payload.setdefault(current_key, [])
            if isinstance(payload[current_key], list):
                payload[current_key].append(line.split("-", 1)[1].strip().strip("'\""))
    return payload


def _front_matter_scalar(payload: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = payload.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _front_matter_list(payload: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = payload.get(key)
        if isinstance(value, list):
            return "; ".join(
                _dedupe_strings([str(item).strip() for item in value if str(item).strip()])
            )
        if isinstance(value, str) and value.strip():
            return "; ".join(
                _dedupe_strings([item.strip() for item in re.split(r"[;|,]", value) if item.strip()])
            )
    return ""


def _extract_byline_authors(lines: list[str]) -> str:
    for line in lines:
        if not line.lower().startswith("by "):
            continue
        raw = line[3:].strip()
        return "; ".join(
            _dedupe_strings([item.strip() for item in re.split(r"\band\b|[;,|]", raw) if item.strip()])
        )
    return ""


def _extract_publication_date(lines: list[str]) -> str:
    patterns = [
        r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},\s+(19|20)\d{2}\b",
        r"\b(19|20)\d{2}-\d{2}-\d{2}\b",
        r"\b(19|20)\d{2}/\d{2}/\d{2}\b",
    ]
    for line in lines:
        for pattern in patterns:
            match = re.search(pattern, line, flags=re.IGNORECASE)
            if match:
                return match.group(0)
    return ""


def _extract_publication_year(markdown_text: str) -> str:
    match = re.search(r"\b(19|20)\d{2}\b", markdown_text[:4000])
    return match.group(0) if match else ""


def _organization_name_from_url(url: str) -> str:
    normalized = str(url or "").strip()
    if not normalized:
        return ""
    try:
        parsed = urlsplit(normalized if "://" in normalized else f"https://{normalized}")
    except Exception:
        return ""
    host = parsed.netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    if not host:
        return ""
    label = host.split(":")[0].split(".")[0].replace("-", " ").replace("_", " ").strip()
    return label.title()


def _infer_document_type(*, row: SourceManifestRow, markdown_text: str) -> str:
    normalized_text = markdown_text[:1600].lower()
    normalized_title = str(row.title or "").strip().lower()
    if row.source_kind == "uploaded_document":
        suffix = Path(row.raw_file or row.source_document_name or "").suffix.lower()
        if suffix == ".pdf":
            return "report"
        if suffix in {".html", ".htm"}:
            return "web page"
        if suffix in {".doc", ".docx", ".rtf", ".txt"}:
            return "document"
        if suffix == ".md":
            return "markdown document"
    url = (row.original_url or row.final_url or "").lower()
    if "journal article" in normalized_title or "journal article" in normalized_text:
        return "journal article"
    if "working paper" in normalized_title or "working paper" in normalized_text:
        return "working paper"
    if "report" in normalized_title or re.search(r"\breport\b", normalized_text):
        return "report"
    if ".gov" in url and ("report" in normalized_text or row.detected_type == "pdf"):
        return "report"
    if row.detected_type == "html":
        return "web page"
    if row.detected_type == "pdf":
        return "report"
    if row.detected_type == "document":
        return "document"
    return "document"


def _infer_organization_type(
    *,
    organization_name: str,
    url: str,
    source_kind: str,
) -> str:
    normalized_url = str(url or "").lower()
    normalized_name = str(organization_name or "").lower()
    if ".edu" in normalized_url or "university" in normalized_name or "college" in normalized_name:
        return "university"
    if ".gov" in normalized_url:
        if ".state." in normalized_url or ".ca.gov" in normalized_url:
            return "state agency"
        return "federal agency"
    if any(token in normalized_name for token in ("ministry", "department", "agency")):
        return "government agency"
    if any(token in normalized_name for token in ("policy", "council", "commission", "board")):
        return "policy body"
    if source_kind == "uploaded_document" and not normalized_url:
        return "uploaded document"
    if ".org" in normalized_url:
        return "organization"
    if any(token in normalized_name for token in ("blog", "substack")):
        return "blog"
    if normalized_url:
        return "company"
    return ""


def _format_tags_text(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(
            _dedupe_strings([str(item).strip() for item in value if str(item).strip()])
        )
    if isinstance(value, str):
        return "; ".join(
            _dedupe_strings([item.strip() for item in re.split(r"[;,|]", value) if item.strip()])
        )
    return ""


def _has_output_file(output_dir: Path, rel_path: str) -> bool:
    if not rel_path:
        return False
    return (output_dir / rel_path).exists()


def png_images_to_pdf_bytes(images: list[bytes]) -> bytes:
    if not images:
        return b""

    doc = fitz.open()
    try:
        for image_bytes in images:
            if not image_bytes:
                continue
            image_doc = fitz.open(stream=image_bytes, filetype="png")
            try:
                rect = image_doc[0].rect
            finally:
                image_doc.close()

            page = doc.new_page(width=rect.width, height=rect.height)
            page.insert_image(page.rect, stream=image_bytes)

        if doc.page_count == 0:
            return b""
        return doc.tobytes(deflate=True, garbage=3)
    finally:
        doc.close()


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def detect_runtime_capabilities(
    use_llm: bool,
    llm_backend: LLMBackendConfig,
) -> RuntimeCapabilities:
    runtime_notes: list[str] = []
    runtime_guidance: list[dict[str, str]] = []

    trafilatura_available = trafilatura is not None
    if not trafilatura_available:
        runtime_notes.append(NOTE_RUNTIME_MISSING_TRAFILATURA)
        runtime_guidance.append(
            {
                "code": NOTE_RUNTIME_MISSING_TRAFILATURA,
                "title": "Install missing parser dependency",
                "detail": (
                    "Python package `trafilatura` is not available, so HTML-to-markdown "
                    "extraction uses a fallback path. Run from the project root."
                ),
                "command": INSTALL_BOOTSTRAP_COMMAND,
            }
        )

    playwright_python_available, playwright_browser_available, playwright_error = (
        check_playwright_runtime()
    )
    if not playwright_browser_available:
        runtime_notes.append(NOTE_RUNTIME_MISSING_PLAYWRIGHT)
        if not playwright_python_available:
            runtime_guidance.append(
                {
                    "code": NOTE_RUNTIME_MISSING_PLAYWRIGHT,
                    "title": "Install Playwright Python package",
                    "detail": (
                        "Playwright is not importable from the Python environment, so "
                        "rendered HTML and visual captures are unavailable. "
                        "Run from the project root."
                    ),
                    "command": INSTALL_BOOTSTRAP_COMMAND,
                }
            )
        else:
            runtime_guidance.append(
                {
                    "code": NOTE_RUNTIME_MISSING_PLAYWRIGHT,
                    "title": "Install Playwright Chromium browser",
                    "detail": (
                        "Playwright is installed but Chromium is missing, so rendered HTML "
                        "and screenshot-based PDF capture are unavailable."
                    ),
                    "command": INSTALL_PLAYWRIGHT_BROWSER_COMMAND,
                }
            )
            if playwright_error:
                runtime_guidance[-1]["detail"] += f" ({playwright_error})"

    textutil_available = check_textutil_available()
    if not textutil_available:
        runtime_notes.append(NOTE_RUNTIME_MISSING_TEXTUTIL)

    tesseract_available = check_tesseract_available()
    if not tesseract_available:
        runtime_notes.append(NOTE_RUNTIME_MISSING_TESSERACT)

    llm_vision_enabled = False
    if use_llm:
        llm_vision_enabled = llm_backend_ready_for_vision(llm_backend)
        if not llm_vision_enabled:
            runtime_notes.append(NOTE_RUNTIME_MISSING_LLM_VISION)

    return RuntimeCapabilities(
        trafilatura_available=trafilatura_available,
        playwright_python_available=playwright_python_available,
        playwright_browser_available=playwright_browser_available,
        textutil_available=textutil_available,
        tesseract_available=tesseract_available,
        llm_vision_enabled=llm_vision_enabled,
        runtime_notes=runtime_notes,
        runtime_guidance=runtime_guidance,
    )


def check_playwright_browser_available() -> bool:
    _, browser_available, _ = check_playwright_runtime()
    return browser_available


def check_playwright_runtime() -> tuple[bool, bool, str]:
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        return False, False, f"import_failure: {exc}"

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            browser.close()
            return True, True, ""
    except Exception as exc:
        normalized = _normalize_playwright_error(exc)
        return True, False, normalized


def check_textutil_available() -> bool:
    return shutil.which("textutil") is not None


def check_tesseract_available() -> bool:
    return shutil.which("tesseract") is not None


def llm_backend_ready_for_vision(config: LLMBackendConfig) -> bool:
    return llm_backend_ready_for_chat(config)


def llm_backend_ready_for_chat(config: LLMBackendConfig) -> bool:
    if not config.model.strip():
        return False
    if not config.base_url.strip():
        return False
    return config.kind in {"openai", "ollama"}


def parse_cleanup_response(response_text: str) -> tuple[bool, str]:
    text = (response_text or "").strip()
    if not text:
        return False, ""

    needs_match = re.search(r"NEEDS_CLEANUP:\s*(yes|no)", text, flags=re.IGNORECASE)
    needs_cleanup = False
    if needs_match:
        needs_cleanup = needs_match.group(1).strip().lower() == "yes"

    cleaned_markdown = ""
    marker_match = re.search(r"CLEANED_MARKDOWN:\s*", text, flags=re.IGNORECASE)
    if marker_match:
        cleaned_markdown = text[marker_match.end() :].strip()
    elif needs_cleanup:
        # Fallback: if the model ignored the format, treat full text as cleaned output.
        cleaned_markdown = text

    return needs_cleanup, cleaned_markdown


def normalize_cleaned_markdown(text: str) -> str:
    cleaned = (text or "").strip()
    if not cleaned:
        return ""
    return cleaned + "\n"


def normalize_summary_paragraph(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", (text or "")).strip()
    if not cleaned:
        return ""

    sentences = split_sentences(cleaned)
    if not sentences:
        return ""
    if len(sentences) > 4:
        sentences = sentences[:4]
    return " ".join(sentences).strip()


def normalize_generated_title(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", (text or "")).strip()
    cleaned = cleaned.strip("\"'`")
    cleaned = re.sub(r"^[Tt]itle:\s*", "", cleaned)
    return cleaned.strip()


def limit_title_words(text: str, max_words: int) -> str:
    words = [word for word in re.split(r"\s+", (text or "").strip()) if word]
    if len(words) <= max_words:
        return " ".join(words)
    return " ".join(words[:max_words]).strip()


def split_sentences(text: str) -> list[str]:
    parts = re.split(r"(?<=[.!?])\s+", text)
    return [p.strip() for p in parts if p.strip()]


def clean_url_candidate(url: str) -> str:
    candidate = html.unescape((url or "").strip())
    if not candidate:
        return ""

    # Trim common wrappers added by markdown/table formatting.
    if candidate.startswith("<") and candidate.endswith(">"):
        candidate = candidate[1:-1].strip()

    candidate = candidate.strip("\"'`")
    candidate = _trim_unbalanced_trailing_closers(candidate)
    return candidate


def _trim_unbalanced_trailing_closers(candidate: str) -> str:
    value = (candidate or "").strip()
    if not value:
        return ""

    matching_openers = {")": "(", "]": "[", "}": "{", ">": "<"}
    while value:
        last = value[-1]
        if last in "\"'`":
            value = value[:-1].rstrip()
            continue
        opener = matching_openers.get(last)
        if not opener:
            break
        if value.count(last) <= value.count(opener):
            break
        value = value[:-1].rstrip()
    return value


def normalize_url(url: str) -> tuple[str, str]:
    candidate = clean_url_candidate(url)
    if not candidate:
        return "", "empty_url"
    if "://" not in candidate:
        candidate = f"https://{candidate}"

    try:
        parsed = urlsplit(candidate)
    except Exception as exc:
        return "", str(exc)

    if parsed.scheme not in {"http", "https"}:
        return "", f"unsupported_scheme:{parsed.scheme}"
    if not parsed.netloc:
        return "", "missing_hostname"
    normalized = urlunsplit(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path or "/",
            parsed.query,
            parsed.fragment,
        )
    )
    return normalized, ""


def dedupe_url_key(url: str) -> str:
    normalized_url, _ = normalize_url(url)
    candidate = clean_url_candidate(normalized_url or url or "")
    if not candidate:
        return ""
    try:
        parsed = urlsplit(candidate)
    except Exception:
        return candidate.lower()

    if not parsed.netloc:
        return candidate.lower()

    filtered_params: list[tuple[str, str]] = []
    for key, value in parse_qsl(parsed.query, keep_blank_values=True):
        key_lower = key.lower()
        if key_lower in TRACKING_PARAM_EXACT:
            continue
        if any(key_lower.startswith(prefix) for prefix in TRACKING_PARAM_PREFIXES):
            continue
        filtered_params.append((key, value))

    filtered_params.sort(key=lambda item: (item[0].lower(), item[1]))
    query = urlencode(filtered_params, doseq=True)
    canonical_path = quote(unquote(parsed.path or "/"), safe="/:@!$&'()*+,;=-._~")

    return urlunsplit(
        (
            parsed.scheme.lower() or "https",
            parsed.netloc.lower(),
            canonical_path,
            query,
            "",
        )
    )


def classify_http_status(status_code: int) -> str:
    if status_code in {401, 403, 407, 429}:
        return "blocked_request"
    if 400 <= status_code < 500:
        return "network_failure"
    if status_code >= 500:
        return "network_failure"
    return "ok"


def detect_source_type(content_type: str, final_url: str, body: bytes) -> str:
    ct = (content_type or "").lower()
    ct_base = ct.split(";", 1)[0].strip()
    final = (final_url or "").lower()
    body_head = (body or b"")[:1024].lstrip()

    if "application/pdf" in ct:
        return "pdf"
    if final.endswith(".pdf"):
        return "pdf"
    if body_head.startswith(b"%PDF-"):
        return "pdf"

    if "text/html" in ct or "application/xhtml+xml" in ct:
        return "html"
    if body_head.startswith(b"<!doctype html") or body_head.startswith(b"<html"):
        return "html"

    if is_document_source(ct_base, final):
        return "document"
    return "unsupported"


def is_document_source(content_type: str, final_url: str) -> bool:
    suffix = Path(urlsplit(final_url).path).suffix.lower()
    if suffix in DOCUMENT_EXTENSIONS:
        return True
    if content_type in DOCUMENT_CONTENT_TYPE_EXT:
        return True
    # Treat remaining text-like responses as document text if not HTML.
    return content_type.startswith("text/")


def infer_document_extension(
    final_url: str,
    content_type: str,
    content_disposition: str,
) -> str:
    url_ext = Path(urlsplit(final_url).path).suffix.lower()
    if url_ext in DOCUMENT_EXTENSIONS:
        return url_ext

    disposition_ext = _extract_extension_from_content_disposition(content_disposition)
    if disposition_ext:
        return disposition_ext

    ct_base = (content_type or "").lower().split(";", 1)[0].strip()
    if ct_base in DOCUMENT_CONTENT_TYPE_EXT:
        return DOCUMENT_CONTENT_TYPE_EXT[ct_base]
    if ct_base.startswith("text/"):
        return ".txt"
    return ".bin"


def _extract_extension_from_content_disposition(content_disposition: str) -> str:
    if not content_disposition:
        return ""
    # Matches filename=report.docx or filename*=UTF-8''report.docx
    match = re.search(
        r"filename\*?=(?:UTF-8''|\"|')?([^\"';]+)",
        content_disposition,
        re.IGNORECASE,
    )
    if not match:
        return ""
    filename = match.group(1).strip()
    suffix = Path(filename).suffix.lower()
    if suffix in DOCUMENT_EXTENSIONS:
        return suffix
    return ""


def decode_document_text(response: httpx.Response) -> str:
    if not response.content:
        return ""
    try:
        return response.text
    except Exception:
        for encoding in ("utf-8", "latin-1"):
            try:
                return response.content.decode(encoding, errors="replace")
            except Exception:
                continue
    return ""


def first_nonempty_line(text: str) -> str:
    for line in (text or "").splitlines():
        stripped = line.strip()
        if stripped:
            return stripped
    return ""


def _normalize_playwright_error(exc: Exception) -> str:
    details = f"{type(exc).__name__}: {exc}"
    lowered = details.lower()
    if "no module named 'playwright'" in lowered or "modulenotfounderror" in lowered:
        return f"playwright_not_installed: run `{INSTALL_BOOTSTRAP_COMMAND}`"
    if "executable doesn't exist" in lowered or "playwright install" in lowered:
        return (
            f"playwright_not_installed: run `{INSTALL_PLAYWRIGHT_BROWSER_COMMAND}`"
        )
    return f"rendering_failure: {details}"


def normalize_render_error_note(render_error: str) -> str:
    lowered = (render_error or "").lower()
    if "playwright_not_installed" in lowered or "playwright install chromium" in lowered:
        return NOTE_RUNTIME_MISSING_PLAYWRIGHT
    return "rendering_failure"


def blocked_error_message(status_code: int) -> str:
    if status_code >= 400:
        return f"{NOTE_BLOCKED_REQUEST}: http_status_{status_code}"
    return f"{NOTE_BLOCKED_REQUEST}: challenge_page_detected"


def detect_blocked_page(html_text: str, title: str, final_url: str) -> bool:
    sample = (html_text or "")[:20000]
    title_lower = (title or "").strip().lower()
    url_lower = (final_url or "").strip().lower()

    if "cdn-cgi/challenge-platform" in url_lower:
        return True
    if "just a moment" in title_lower:
        return True

    signals = 0
    for pattern in BLOCKED_PAGE_PATTERNS:
        if pattern.search(sample):
            signals += 1

    if "cf-ray" in sample.lower() or "__cf_bm" in sample.lower():
        signals += 1
    if "hcaptcha" in sample.lower() or "g-recaptcha" in sample.lower():
        signals += 1

    return signals >= 2


def decode_bytes_to_text(data: bytes) -> str:
    if not data:
        return ""
    for encoding in ("utf-8", "utf-16", "latin-1"):
        try:
            return data.decode(encoding)
        except Exception:
            continue
    return data.decode("utf-8", errors="replace")


def convert_docx_bytes_to_markdown(data: bytes) -> str:
    try:
        from docx import Document

        doc = Document(io.BytesIO(data))
    except Exception:
        return ""

    lines: list[str] = []
    for para in doc.paragraphs:
        text = re.sub(r"\s+", " ", (para.text or "").strip())
        if not text:
            continue
        style_name = (para.style.name or "").lower()
        if "heading" in style_name:
            level = 2
            for ch in style_name:
                if ch.isdigit():
                    level = max(1, min(int(ch), 6))
                    break
            lines.append(f"{'#' * level} {text}")
        else:
            lines.append(text)
        lines.append("")

    markdown_text = "\n".join(lines).strip()
    return markdown_text


def convert_doc_bytes_with_textutil(data: bytes) -> str:
    with tempfile.TemporaryDirectory() as tmpdir:
        doc_path = Path(tmpdir) / "input.doc"
        doc_path.write_bytes(data)
        cmd = ["textutil", "-convert", "txt", "-stdout", str(doc_path)]
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=False,
                check=False,
                timeout=30,
            )
        except Exception:
            return ""
        if result.returncode != 0:
            return ""
        return decode_bytes_to_text(result.stdout).strip()


def extract_pdf_pages(pdf_bytes: bytes) -> list[dict]:
    try:
        with suppress_mupdf_messages():
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")

            pages: list[dict] = []
            try:
                for page_index in range(doc.page_count):
                    page = doc[page_index]
                    text = page.get_text("text")
                    pixmap = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
                    pages.append(
                        {
                            "index": page_index + 1,
                            "text": text.strip(),
                            "pixmap": pixmap,
                        }
                    )
            finally:
                doc.close()
        return pages
    except Exception:
        return []


def compute_text_metrics(text: str) -> dict[str, float]:
    compact = re.sub(r"\s+", " ", text or "").strip()
    chars = len(compact)
    alpha_chars = sum(1 for ch in compact if ch.isalpha())
    alpha_ratio = (alpha_chars / chars) if chars else 0.0
    return {
        "chars": float(chars),
        "alpha_ratio": alpha_ratio,
    }


def evaluate_pdf_document_quality(page_metrics: list[dict[str, float]]) -> dict:
    if not page_metrics:
        return {"native_good": False, "low_quality_pages": []}

    low_quality_pages: list[int] = []
    total_chars = 0.0
    for idx, metrics in enumerate(page_metrics):
        chars = metrics.get("chars", 0.0)
        alpha_ratio = metrics.get("alpha_ratio", 0.0)
        total_chars += chars
        if chars < PDF_NATIVE_PAGE_MIN_CHARS or alpha_ratio < PDF_TEXT_ALPHA_MIN_RATIO:
            low_quality_pages.append(idx)

    avg_chars = total_chars / len(page_metrics)
    native_good = (
        avg_chars >= PDF_NATIVE_DOC_MIN_AVG_CHARS
        and len(low_quality_pages) <= max(1, int(len(page_metrics) * 0.25))
    )
    return {
        "native_good": native_good,
        "low_quality_pages": low_quality_pages,
        "avg_chars": avg_chars,
    }


def run_tesseract_ocr_on_pixmap(pixmap: fitz.Pixmap) -> str:
    with tempfile.TemporaryDirectory() as tmpdir:
        image_path = Path(tmpdir) / "page.png"
        image_path.write_bytes(pixmap.tobytes("png"))
        cmd = [
            "tesseract",
            str(image_path),
            "stdout",
            "--psm",
            "3",
            "-l",
            "eng",
        ]
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                check=False,
                timeout=45,
            )
        except Exception:
            return ""
        if result.returncode != 0:
            return ""
        return (result.stdout or "").strip()


def format_pages_as_markdown(page_texts: list[str]) -> str:
    lines: list[str] = []
    for idx, page_text in enumerate(page_texts, start=1):
        cleaned = re.sub(r"\n{3,}", "\n\n", (page_text or "").strip())
        if not cleaned:
            continue
        lines.append(f"## Page {idx}")
        lines.append("")
        lines.append(cleaned)
        lines.append("")
    return "\n".join(lines).strip()


def decode_html(response: httpx.Response) -> str:
    if not response.content:
        return ""
    try:
        return response.text
    except Exception:
        for encoding in ("utf-8", "latin-1"):
            try:
                return response.content.decode(encoding, errors="replace")
            except Exception:
                continue
    return ""


def extract_markdown(html: str) -> tuple[str, str]:
    if not html:
        return "", "extraction_failure"
    if trafilatura is None:
        return "", "extraction_failure"
    try:
        md = trafilatura.extract(
            html,
            output_format="markdown",
            include_links=True,
            include_images=False,
            include_tables=True,
            favor_recall=True,
            deduplicate=True,
        )
    except Exception:
        return "", "extraction_failure"

    if not md:
        return "", "extraction_failure"
    return md.strip(), ""


def extract_markdown_with_fallback(
    html_text: str,
    runtime_capabilities: RuntimeCapabilities,
) -> tuple[str, bool, list[str]]:
    notes: list[str] = []
    if not html_text:
        notes.append(NOTE_EXTRACTION_FAILURE)
        return "", False, notes

    if runtime_capabilities.trafilatura_available and trafilatura is not None:
        markdown_text, _ = extract_markdown(html_text)
        if markdown_text:
            return markdown_text, False, notes
    else:
        notes.append(NOTE_RUNTIME_MISSING_TRAFILATURA)

    fallback_markdown = extract_markdown_fallback(html_text)
    if fallback_markdown:
        return fallback_markdown, True, notes

    notes.append(NOTE_EXTRACTION_FAILURE)
    return "", False, notes


def extract_markdown_fallback(html_text: str) -> str:
    if not html_text:
        return ""

    candidate = select_main_content_fragment(html_text)

    # Drop high-noise page chrome and scripts before text conversion.
    candidate = re.sub(
        r"(?is)<(script|style|noscript|svg|canvas|iframe|form|button|header|footer|nav|aside)[^>]*>.*?</\1>",
        " ",
        candidate,
    )
    candidate = re.sub(r"(?is)<br\s*/?>", "\n", candidate)
    candidate = re.sub(
        r"(?is)<h([1-6])[^>]*>(.*?)</h\1>",
        _heading_replacer,
        candidate,
    )
    candidate = re.sub(r"(?is)<li[^>]*>(.*?)</li>", _list_item_replacer, candidate)
    candidate = re.sub(r"(?is)</(p|div|section|article|main|blockquote|tr|table)>", "\n\n", candidate)
    candidate = re.sub(r"(?is)<(p|div|section|article|main|blockquote|tr|table)[^>]*>", "\n", candidate)

    text = re.sub(r"(?is)<[^>]+>", " ", candidate)
    text = html.unescape(text)
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    cleaned_lines: list[str] = []
    seen: set[str] = set()
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line:
            if cleaned_lines and cleaned_lines[-1] != "":
                cleaned_lines.append("")
            continue

        if is_boilerplate_line(line):
            continue

        dedupe_key = line.lower()
        if dedupe_key in seen and len(line) < 120:
            continue
        seen.add(dedupe_key)
        cleaned_lines.append(line)

    markdown_text = "\n".join(cleaned_lines).strip()
    markdown_text = re.sub(r"\n{3,}", "\n\n", markdown_text)
    if markdown_score(markdown_text) < MIN_FALLBACK_MARKDOWN_SCORE:
        return ""
    return markdown_text


def select_main_content_fragment(html_text: str) -> str:
    for pattern in [
        r"(?is)<main[^>]*>(.*?)</main>",
        r"(?is)<article[^>]*>(.*?)</article>",
        r'(?is)<div[^>]+(?:id|class)=["\'][^"\']*(?:content|article|post|main)[^"\']*["\'][^>]*>(.*?)</div>',
        r'(?is)<section[^>]+(?:id|class)=["\'][^"\']*(?:content|article|post|main)[^"\']*["\'][^>]*>(.*?)</section>',
    ]:
        match = re.search(pattern, html_text)
        if match:
            return match.group(1)
    return html_text


def _heading_replacer(match: re.Match) -> str:
    level = int(match.group(1))
    content = strip_tags(match.group(2))
    if not content:
        return "\n"
    return f"\n{'#' * level} {content}\n"


def _list_item_replacer(match: re.Match) -> str:
    content = strip_tags(match.group(1))
    if not content:
        return "\n"
    return f"\n- {content}"


def strip_tags(text: str) -> str:
    cleaned = re.sub(r"(?is)<[^>]+>", " ", text or "")
    cleaned = html.unescape(cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def is_boilerplate_line(line: str) -> bool:
    normalized = line.strip().lower()
    if not normalized:
        return True

    if normalized in {"menu", "navigation", "skip to content", "back to top"}:
        return True

    boilerplate_tokens = [
        "all rights reserved",
        "privacy policy",
        "cookie policy",
        "terms of use",
        "accept all cookies",
        "manage cookies",
        "sign in",
        "subscribe",
    ]
    return any(token in normalized for token in boilerplate_tokens)


def markdown_score(markdown_text: str) -> int:
    if not markdown_text:
        return 0
    compact = re.sub(r"\s+", " ", markdown_text).strip()
    return len(compact)


def extract_title(html: str) -> str:
    if not html:
        return ""
    match = re.search(r"<title[^>]*>(.*?)</title>", html, re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    title = re.sub(r"\s+", " ", match.group(1)).strip()
    return title[:500]


def extract_canonical_url(html: str) -> str:
    if not html:
        return ""
    match = re.search(
        r'<link[^>]+rel=["\'][^"\']*canonical[^"\']*["\'][^>]*href=["\']([^"\']+)["\']',
        html,
        re.IGNORECASE,
    )
    if match:
        return match.group(1).strip()
    return ""
