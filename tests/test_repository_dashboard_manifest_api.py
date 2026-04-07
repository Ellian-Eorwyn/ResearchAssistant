from __future__ import annotations

import json
import io
import tempfile
import unittest
from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.routers import repository
from backend.storage.attached_repository import AttachedRepositoryService
from backend.storage.file_store import FileStore


class RepositoryDashboardManifestApiTests(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory(prefix="repo-dashboard-manifest-api-tests-")
        self.tmp_path = Path(self._tmp.name)
        self.store = FileStore(base_dir=self.tmp_path / "app_data")
        self.service = AttachedRepositoryService(store=self.store)

        self.repo_dir = self.tmp_path / "repo"
        self.repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(self.repo_dir))
        self.first_import = self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
                "https://example.com/c,Gamma Source\n"
            ).encode("utf-8"),
        )

        state_path = self.repo_dir / ".ra_repo" / "repository_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        rows = state.get("sources", [])
        self.assertGreaterEqual(len(rows), 3)

        rows[0]["title"] = "Alpha Source"
        rows[0]["source_kind"] = "url"
        rows[0]["fetch_status"] = "success"
        rows[0]["detected_type"] = "pdf"
        rows[0]["author_names"] = "Jane Doe; John Roe"
        rows[0]["publication_date"] = "2024-03-15"
        rows[0]["publication_year"] = "2024"
        rows[0]["document_type"] = "report"
        rows[0]["organization_name"] = "Alpha Agency"
        rows[0]["organization_type"] = "state agency"
        rows[0]["catalog_status"] = "generated"
        rows[0]["tags_text"] = "housing; retrofit"
        rows[0]["summary_file"] = "summaries/000001_summary.md"
        rows[0]["summary_status"] = "generated"
        rows[0]["rating_file"] = "ratings/000001_rating.json"
        rows[0]["rating_status"] = "generated"

        rows[1]["title"] = "Beta Source"
        rows[1]["source_kind"] = "url"
        rows[1]["fetch_status"] = "failed"
        rows[1]["detected_type"] = "html"
        rows[1]["author_names"] = "Beta Team"
        rows[1]["publication_date"] = "2023"
        rows[1]["publication_year"] = "2023"
        rows[1]["document_type"] = "web page"
        rows[1]["organization_name"] = "Beta Blog"
        rows[1]["organization_type"] = "blog"
        rows[1]["catalog_status"] = "generated"
        rows[1]["tags_text"] = "blog; commentary"
        rows[1]["summary_file"] = ""
        rows[1]["summary_status"] = "failed"
        rows[1]["rating_file"] = ""
        rows[1]["rating_status"] = ""

        rows[2]["title"] = "Gamma Source"
        rows[2]["source_kind"] = "uploaded_document"
        rows[2]["fetch_status"] = "queued"
        rows[2]["detected_type"] = "document"
        rows[2]["author_names"] = "Gamma Lab"
        rows[2]["publication_date"] = "2025-01-07"
        rows[2]["publication_year"] = "2025"
        rows[2]["document_type"] = "journal article"
        rows[2]["organization_name"] = "Gamma University"
        rows[2]["organization_type"] = "university"
        rows[2]["catalog_status"] = "generated"
        rows[2]["tags_text"] = "heat pump; workforce"
        rows[2]["summary_file"] = ""
        rows[2]["summary_status"] = ""
        rows[2]["rating_file"] = "ratings/000003_rating.json"
        rows[2]["rating_status"] = "generated"

        state_path.write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        (self.repo_dir / "summaries").mkdir(parents=True, exist_ok=True)
        (self.repo_dir / "ratings").mkdir(parents=True, exist_ok=True)
        (self.repo_dir / "summaries" / "000001_summary.md").write_text(
            "Alpha summary text",
            encoding="utf-8",
        )
        (self.repo_dir / "ratings" / "000001_rating.json").write_text(
            json.dumps(
                {
                    "overall_score": 0.9,
                    "confidence": 0.85,
                    "rationale": "Strong direct evidence.",
                    "relevant_sections": ["Efficiency findings"],
                    "ratings": {
                        "relevance": 0.95,
                        "overall_relevance": 0.93,
                        "depth_score": 0.91,
                        "relevant_detail_score": 0.89,
                    },
                    "flags": {"needs_manual_review": 0.1},
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        (self.repo_dir / "ratings" / "000003_rating.json").write_text(
            json.dumps(
                {
                    "overall_score": 0.3,
                    "confidence": 0.5,
                    "rationale": "Only tangentially related.",
                    "ratings": {
                        "overall_relevance": 0.25,
                        "depth_score": 0.2,
                        "relevant_detail_score": 0.1,
                    },
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        app = FastAPI()
        app.state.file_store = self.store
        app.state.repository_service = self.service
        app.include_router(repository.router, prefix="/api")
        self.client = TestClient(app)

    def tearDown(self):
        self._tmp.cleanup()

    def test_repository_dashboard_contains_expected_sections(self):
        resp = self.client.get("/api/repository/dashboard")
        self.assertEqual(resp.status_code, 200)

        payload = resp.json()
        self.assertIn("status", payload)
        self.assertIn("metrics", payload)
        self.assertIn("output_formats", payload)
        self.assertIn("warning_aggregates", payload)
        self.assertIn("recent_imports", payload)
        self.assertIn("recent_jobs", payload)

        self.assertEqual(payload["status"]["attached"], True)
        self.assertGreaterEqual(payload["metrics"]["total_sources"], 3)
        self.assertGreaterEqual(len(payload["recent_imports"]), 1)

    def test_repository_manifest_filters_by_fetch_status(self):
        resp = self.client.get("/api/repository/manifest?fetch_status=failed")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        rows = payload["rows"]
        self.assertGreaterEqual(len(rows), 1)
        self.assertTrue(all((row.get("fetch_status") or "") == "failed" for row in rows))

    def test_repository_manifest_filters_by_summary_and_rating(self):
        with_summary = self.client.get("/api/repository/manifest?has_summary=true")
        self.assertEqual(with_summary.status_code, 200)
        summary_rows = with_summary.json()["rows"]
        self.assertGreaterEqual(len(summary_rows), 1)
        self.assertTrue(all((row.get("summary_file") or "") for row in summary_rows))

        without_rating = self.client.get("/api/repository/manifest?has_rating=false")
        self.assertEqual(without_rating.status_code, 200)
        rating_rows = without_rating.json()["rows"]
        self.assertGreaterEqual(len(rating_rows), 1)
        self.assertTrue(all(not (row.get("rating_file") or "") for row in rating_rows))

    def test_repository_manifest_supports_sort_and_pagination(self):
        resp = self.client.get(
            "/api/repository/manifest?sort_by=title&sort_dir=desc&limit=1&offset=0"
        )
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        self.assertEqual(payload["limit"], 1)
        self.assertEqual(payload["offset"], 0)
        self.assertEqual(payload["sort_by"], "title")
        self.assertEqual(payload["sort_dir"], "desc")
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["title"], "Gamma Source")

    def test_repository_manifest_q_filter(self):
        resp = self.client.get("/api/repository/manifest?q=alpha")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        self.assertGreaterEqual(len(payload["rows"]), 1)
        self.assertTrue(any((row.get("title") or "").lower() == "alpha source" for row in payload["rows"]))

    def test_repository_manifest_q_filter_matches_catalog_fields(self):
        resp = self.client.get("/api/repository/manifest?q=jane%20doe")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["id"], "000001")

    def test_repository_manifest_filters_by_rating_thresholds(self):
        resp = self.client.get(
            "/api/repository/manifest?rating_overall_relevance_min=0.9&rating_depth_score_min=0.9"
        )
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["id"], "000001")

    def test_repository_manifest_supports_sort_by_dynamic_rating_field(self):
        resp = self.client.get(
            "/api/repository/manifest?sort_by=rating_depth_score&sort_dir=desc"
        )
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        self.assertEqual(payload["sort_by"], "rating_depth_score")
        self.assertGreaterEqual(len(payload["rows"]), 2)
        self.assertEqual(payload["rows"][0]["id"], "000001")

    def test_repository_manifest_sorts_dates_type_aware(self):
        resp = self.client.get(
            "/api/repository/manifest?sort_by=publication_date&sort_dir=desc"
        )
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        self.assertEqual(payload["sort_by"], "publication_date")
        self.assertEqual(payload["rows"][0]["id"], "000003")
        self.assertEqual(payload["rows"][1]["id"], "000001")
        self.assertEqual(payload["rows"][2]["id"], "000002")

    def test_repository_manifest_returns_column_metadata(self):
        resp = self.client.get("/api/repository/manifest")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        column_map = {item["key"]: item for item in payload["columns"]}
        keys = set(column_map)
        self.assertIn("id", keys)
        self.assertIn("source_kind", keys)
        self.assertIn("author_names", keys)
        self.assertIn("document_type", keys)
        self.assertIn("organization_type", keys)
        self.assertIn("tags_text", keys)
        self.assertIn("rating_depth_score", keys)
        self.assertIn("rating_overall_relevance", keys)
        self.assertTrue(column_map["author_names"]["instruction_prompt"])
        self.assertTrue(column_map["publication_year"]["instruction_prompt"])
        self.assertEqual(column_map["publication_year"]["sort_type"], "number")
        self.assertEqual(column_map["title"]["include_source_text"], True)
        self.assertEqual(column_map["title"]["include_row_context"], False)

    def test_repository_manifest_export_endpoint_supports_csv_and_xlsx(self):
        csv_resp = self.client.post(
            "/api/repository/manifest/export",
            json={
                "scope": "selected",
                "format": "csv",
                "source_ids": ["000002", "000001"],
                "filters": {},
            },
        )
        self.assertEqual(csv_resp.status_code, 200)
        self.assertEqual(csv_resp.headers["content-type"], "text/csv; charset=utf-8")
        csv_lines = csv_resp.content.decode("utf-8-sig").splitlines()
        self.assertTrue(csv_lines[0].startswith("id,repository_source_id"))
        self.assertTrue(csv_lines[1].startswith("000002,"))
        self.assertTrue(csv_lines[2].startswith("000001,"))

        xlsx_resp = self.client.post(
            "/api/repository/manifest/export",
            json={
                "scope": "selected",
                "format": "xlsx",
                "source_ids": ["000002", "000001"],
                "filters": {},
            },
        )
        self.assertEqual(xlsx_resp.status_code, 200)
        self.assertEqual(
            xlsx_resp.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(io.BytesIO(xlsx_resp.content))
        worksheet = workbook.active
        self.assertEqual(worksheet.cell(row=2, column=1).value, "000002")
        self.assertEqual(worksheet.cell(row=3, column=1).value, "000001")

    def test_repository_manifest_export_endpoint_can_limit_visible_columns(self):
        resp = self.client.post(
            "/api/repository/manifest/export",
            json={
                "scope": "selected",
                "format": "csv",
                "column_scope": "visible",
                "column_keys": ["title", "author_names"],
                "source_ids": ["000001"],
                "filters": {},
            },
        )
        self.assertEqual(resp.status_code, 200)
        lines = resp.content.decode("utf-8-sig").splitlines()
        self.assertEqual(lines[0], "title,author_names")
        self.assertTrue(lines[1].startswith("Alpha Source,"))

    def test_repository_manifest_filters_by_catalog_fields(self):
        resp = self.client.get(
            "/api/repository/manifest?source_kind=uploaded_document&organization_type=university&document_type=journal"
        )
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        self.assertEqual(len(payload["rows"]), 1)
        row = payload["rows"][0]
        self.assertEqual(row["id"], "000003")
        self.assertEqual(row["source_kind"], "uploaded_document")
        self.assertEqual(row["organization_type"], "university")

    def test_repository_manifest_returns_enriched_summary_and_rating_fields(self):
        resp = self.client.get("/api/repository/manifest?q=alpha")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()
        row = payload["rows"][0]

        self.assertEqual(row["summary_text"], "Alpha summary text")
        self.assertEqual(row["rating_overall"], 0.9)
        self.assertEqual(row["rating_confidence"], 0.85)
        self.assertEqual(row["rating_relevance"], 0.95)
        self.assertEqual(row["rating_overall_relevance"], 0.93)
        self.assertEqual(row["rating_depth_score"], 0.91)
        self.assertEqual(row["rating_relevant_detail_score"], 0.89)
        self.assertEqual(row["flag_needs_manual_review"], 0.1)
        self.assertIn("Efficiency findings", row["relevant_sections"])

    def test_repository_manifest_rejects_invalid_sort_field(self):
        resp = self.client.get("/api/repository/manifest?sort_by=unknown_field")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("Invalid sort_by", resp.json()["detail"])

    def test_repository_citation_data_endpoint_returns_data_shapes(self):
        resp = self.client.get("/api/repository/citation-data")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()

        self.assertIn("bibliography", payload)
        self.assertIn("citations", payload)
        self.assertIn("entries", payload["bibliography"])
        self.assertIn("citations", payload["citations"])
        self.assertIn("sentences", payload["citations"])
        self.assertIn("matches", payload["citations"])


if __name__ == "__main__":
    unittest.main()
