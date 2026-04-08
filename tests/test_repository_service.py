from __future__ import annotations

import csv
import io
import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from backend.models.export import EXPORT_COLUMNS, ExportRow
from backend.models.ingestion_profiles import DocumentNormalizationResult, IngestionProfile
from backend.models.repository import (
    RepositoryBundleExportRequest,
    RepositoryCitationRisExportRequest,
    RepositoryColumnRunRequest,
    RepositoryManifestExportRequest,
    RepositoryManifestFilterRequest,
)
from backend.models.settings import RepoSettings
from backend.models.sources import SourceManifestRow
from backend.pipeline.standardized_markdown import NormalizedDocumentOutput
from backend.storage.attached_repository import AttachedRepositoryService, repository_dedupe_key
from backend.storage.file_store import FileStore


class _ImmediateThread:
    def __init__(self, target=None, args=(), kwargs=None, daemon=None):
        self._target = target
        self._args = args
        self._kwargs = kwargs or {}

    def start(self) -> None:
        if self._target is not None:
            self._target(*self._args, **self._kwargs)

    def is_alive(self) -> bool:
        return False


class _NoOpThread:
    def __init__(self, *args, **kwargs):
        return None

    def start(self) -> None:
        return None

    def is_alive(self) -> bool:
        return False


class RepositoryServiceTests(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory(prefix="repo-tests-")
        self.tmp_path = Path(self._tmp.name)
        self.store = FileStore(base_dir=self.tmp_path / "app_data")
        self.service = AttachedRepositoryService(store=self.store)

    def tearDown(self):
        self._tmp.cleanup()

    def _attach_repo(self, name: str = "repo") -> Path:
        repo_dir = self.tmp_path / name
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))
        return repo_dir

    def _repo_job_store(self) -> FileStore:
        return self.service.repo_job_store()

    def _create_repo_job(self) -> tuple[str, FileStore]:
        job_store = self._repo_job_store()
        return job_store.create_job(prefix="repo"), job_store

    def _write_repo_document(
        self,
        repo_dir: Path,
        import_id: str,
        filename: str,
        content: bytes,
    ) -> dict[str, str]:
        doc_dir = repo_dir / "documents" / import_id
        doc_dir.mkdir(parents=True, exist_ok=True)
        target = doc_dir / filename
        target.write_bytes(content)
        import hashlib

        return {
            "filename": filename,
            "source_document_name": filename,
            "repository_path": f"documents/{import_id}/{filename}",
            "sha256": hashlib.sha256(content).hexdigest(),
            "document_import_id": import_id,
        }

    def _save_repo_state(
        self,
        *,
        sources: list[SourceManifestRow],
        citations: list[ExportRow],
        imports: list[dict],
    ) -> None:
        with self.service._writer_lock():
            self.service._save_state_locked(
                sources=sources,
                citations=citations,
                imports=imports,
            )
            meta = self.service._load_meta_locked()
            self.service._save_meta_locked(
                {
                    **meta,
                    "next_source_id": max(
                        int(meta.get("next_source_id") or 1),
                        len(sources) + 1,
                    ),
                }
            )
            self.service._rebuild_outputs_locked(sources, citations)

    def _normalization_output(
        self,
        *,
        filename: str,
        source_document_path: str,
        status: str,
    ) -> NormalizedDocumentOutput:
        standardized_path = ""
        metadata_path = f"{Path(source_document_path).with_suffix('').as_posix()}.standardized.json"
        markdown_text = ""
        if status != "failed":
            standardized_path = f"{Path(source_document_path).with_suffix('').as_posix()}.standardized.md"
            markdown_text = "# Standardized\n"
        return NormalizedDocumentOutput(
            markdown_text=markdown_text,
            result=DocumentNormalizationResult(
                filename=filename,
                source_document_path=source_document_path,
                standardized_markdown_path=standardized_path,
                metadata_path=metadata_path,
                selected_profile_id="auto_detect",
                selected_profile_label="Auto-detect",
                status=status,
            ),
            suggestion=None,
        )

    def test_repository_dedupe_key_strips_tracking_params(self):
        a = "https://example.com/path?a=1&utm_source=test#section"
        b = "https://example.com/path?a=1"
        c = "https://example.com/path?a=1&fbclid=abc"

        self.assertEqual(repository_dedupe_key(a), repository_dedupe_key(b))
        self.assertEqual(repository_dedupe_key(b), repository_dedupe_key(c))

    def test_merge_source_rows_continues_numeric_ids(self):
        rows = [
            SourceManifestRow(id="000010", original_url="https://example.com/a", fetch_status="success"),
            SourceManifestRow(id="", original_url="https://example.com/b", fetch_status="queued"),
        ]

        merged = self.service._merge_source_rows(rows)
        ids = [row.id for row in merged.rows]

        self.assertEqual(ids, ["000010", "000011"])
        self.assertEqual(merged.next_source_id, 12)

    def test_attach_scans_legacy_manifest_and_dedupes(self):
        repo_dir = self.tmp_path / "repo"
        repo_dir.mkdir(parents=True, exist_ok=True)
        manifest = repo_dir / "manifest.csv"

        with manifest.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "id",
                    "original_url",
                    "fetch_status",
                    "fetched_at",
                ],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "id": "000001",
                    "original_url": "https://example.com/a?utm_source=one",
                    "fetch_status": "success",
                    "fetched_at": "2026-01-01T00:00:00+00:00",
                }
            )
            writer.writerow(
                {
                    "id": "000099",
                    "original_url": "https://example.com/a",
                    "fetch_status": "failed",
                    "fetched_at": "2026-01-01T00:00:00+00:00",
                }
            )

        status = self.service.attach(str(repo_dir))
        self.assertTrue(status.attached)
        self.assertEqual(status.total_sources, 1)
        self.assertEqual(status.next_source_id, 2)

    def test_create_export_job_scope_all_creates_job_bibliography(self):
        repo_dir = self.tmp_path / "repo_all"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL\n"
                "https://example.com/a\n"
                "https://example.com/b\n"
            ).encode("utf-8"),
        )

        result = self.service.create_export_job(scope="all")
        self.assertEqual(result.scope, "all")
        self.assertEqual(result.total_urls, 2)
        self.assertTrue(result.job_id)

        bib = self.service.job_store_for(result.job_id).load_artifact(result.job_id, "03_bibliography")
        self.assertIsNotNone(bib)
        urls = [entry.get("url") for entry in bib.get("entries", [])]
        self.assertEqual(urls, ["https://example.com/a", "https://example.com/b"])

    def test_import_source_list_preserves_titles_from_upload(self):
        repo_dir = self.tmp_path / "repo_import_titles"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))

        result = self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
            ).encode("utf-8"),
        )

        self.assertEqual(result.accepted_new, 2)
        state = json.loads((repo_dir / ".ra_repo" / "repository_state.json").read_text(encoding="utf-8"))
        titles = [row.get("title") for row in state.get("sources", [])]
        self.assertEqual(titles, ["Alpha Source", "Beta Source"])

    def test_import_seed_files_from_markdown_harvests_links_without_citations(self):
        repo_dir = self.tmp_path / "repo_seed_markdown"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))

        result = self.service.import_seed_files(
            [
                (
                    "report.md",
                    (
                        "# Deep Research Report\n\n"
                        "- [Alpha Study](https://example.com/a)\n"
                        "- Beta source https://example.com/b\n"
                    ).encode("utf-8"),
                )
            ]
        )

        self.assertEqual(result.import_type, "source_seed")
        self.assertEqual(result.accepted_new, 2)
        state = json.loads(
            (repo_dir / ".ra_repo" / "repository_state.json").read_text(encoding="utf-8")
        )
        self.assertEqual(len(state["sources"]), 2)
        self.assertEqual(state["citations"], [])
        self.assertEqual(state["sources"][0]["source_kind"], "url")
        self.assertEqual(state["sources"][0]["title"], "Alpha Study")
        self.assertEqual(state["sources"][1]["title"], "Beta source")

    def test_import_manual_documents_creates_uploaded_document_rows_and_dedupes_by_sha(self):
        repo_dir = self.tmp_path / "repo_manual_documents"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))

        content = b"# Local Memo\n\nDocument body text.\n"
        result = self.service.import_manual_documents(
            [
                ("memo.md", content),
                ("duplicate.md", content),
            ]
        )

        self.assertEqual(result.import_type, "document_source")
        self.assertEqual(result.accepted_new, 1)
        self.assertEqual(result.duplicates_skipped, 1)
        state = json.loads(
            (repo_dir / ".ra_repo" / "repository_state.json").read_text(encoding="utf-8")
        )
        self.assertEqual(len(state["sources"]), 1)
        self.assertEqual(state["citations"], [])
        row = state["sources"][0]
        self.assertEqual(row["source_kind"], "uploaded_document")
        self.assertEqual(row["import_type"], "document_source")
        self.assertEqual(row["fetch_status"], "not_applicable")
        self.assertEqual(row["detected_type"], "document")
        self.assertEqual(row["title"], "Local Memo")
        self.assertTrue(row["raw_file"].startswith("sources/000001/"))
        self.assertTrue(row["sha256"])
        self.assertTrue((repo_dir / row["raw_file"]).is_file())

    def test_create_export_job_scope_import_selects_only_that_import(self):
        repo_dir = self.tmp_path / "repo_import"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))

        first = self.service.import_source_list(
            filename="sources.csv",
            content=("URL\nhttps://example.com/a\n").encode("utf-8"),
        )
        second = self.service.import_source_list(
            filename="sources2.csv",
            content=("URL\nhttps://example.com/b\n").encode("utf-8"),
        )
        self.assertNotEqual(first.import_id, second.import_id)

        result = self.service.create_export_job(scope="import", import_id=second.import_id)
        self.assertEqual(result.scope, "import")
        self.assertEqual(result.import_id, second.import_id)
        self.assertEqual(result.total_urls, 1)

        bib = self.service.job_store_for(result.job_id).load_artifact(result.job_id, "03_bibliography")
        self.assertIsNotNone(bib)
        urls = [entry.get("url") for entry in bib.get("entries", [])]
        self.assertEqual(urls, ["https://example.com/b"])

    def test_create_export_job_scope_import_rejects_unknown_import_id(self):
        repo_dir = self.tmp_path / "repo_unknown_import"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))
        self.service.import_source_list(
            filename="sources.csv",
            content=("URL\nhttps://example.com/a\n").encode("utf-8"),
        )

        with self.assertRaises(ValueError):
            self.service.create_export_job(scope="import", import_id="does-not-exist")

    def test_create_export_job_scope_import_duplicate_batch_uses_existing_rows(self):
        repo_dir = self.tmp_path / "repo_empty_import"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))
        self.service.import_source_list(
            filename="sources.csv",
            content=("URL\nhttps://example.com/a\n").encode("utf-8"),
        )
        duplicate_import = self.service.import_source_list(
            filename="sources_dup.csv",
            content=("URL\nhttps://example.com/a\n").encode("utf-8"),
        )
        self.assertEqual(duplicate_import.accepted_new, 0)

        result = self.service.create_export_job(scope="import", import_id=duplicate_import.import_id)
        self.assertEqual(result.scope, "import")
        self.assertEqual(result.import_id, duplicate_import.import_id)
        self.assertEqual(result.total_urls, 1)

        bib = self.service.job_store_for(result.job_id).load_artifact(result.job_id, "03_bibliography")
        self.assertIsNotNone(bib)
        urls = [entry.get("url") for entry in bib.get("entries", [])]
        self.assertEqual(urls, ["https://example.com/a"])

    def test_find_duplicate_source_candidates_groups_exact_and_similar_matches(self):
        self._attach_repo("repo_duplicate_candidates")
        self._save_repo_state(
            sources=[
                SourceManifestRow(
                    id="000001",
                    original_url="https://example.com/report?utm_source=newsletter",
                    fetch_status="success",
                    title="Housing Retrofit Program Evaluation",
                    author_names="Jane Doe",
                    publication_date="2025-02-01",
                ),
                SourceManifestRow(
                    id="000002",
                    original_url="https://example.com/report",
                    fetch_status="failed",
                    title="Housing Retrofit Program Evaluation",
                    author_names="Jane Doe",
                    publication_date="2025-02-01",
                ),
                SourceManifestRow(
                    id="000003",
                    original_url="https://data.example.org/grid-modernization-final",
                    fetch_status="success",
                    title="Grid Modernization Assessment Final Report",
                    author_names="Alex Smith; Taylor Jones",
                    publication_year="2024",
                ),
                SourceManifestRow(
                    id="000004",
                    original_url="https://mirror.example.net/grid-modernization-report",
                    fetch_status="failed",
                    title="Grid Modernization Assessment Report",
                    author_names="Taylor Jones; Alex Smith",
                    publication_date="2024-05-20",
                ),
            ],
            citations=[],
            imports=[],
        )

        response = self.service.find_duplicate_source_candidates()

        self.assertEqual(response.scanned_sources, 4)
        self.assertEqual(response.total_groups, 2)
        self.assertFalse(response.truncated)
        self.assertEqual(response.groups[0].match_reason, "Matching normalized URL")
        self.assertEqual(response.groups[0].confidence, "high")
        self.assertEqual(
            {row.id for row in response.groups[0].rows},
            {"000001", "000002"},
        )
        self.assertEqual(response.groups[0].suggested_keep_id, "000001")
        self.assertEqual(response.groups[1].match_reason, "Similar title, year, and authors")
        self.assertEqual(response.groups[1].confidence, "medium")
        self.assertEqual(
            {row.id for row in response.groups[1].rows},
            {"000003", "000004"},
        )
        self.assertEqual(response.groups[1].suggested_keep_id, "000003")

    def test_attach_scans_citations_xlsx(self):
        repo_dir = self.tmp_path / "repo_citations_xlsx"
        repo_dir.mkdir(parents=True, exist_ok=True)

        manifest = repo_dir / "manifest.csv"
        with manifest.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["id", "original_url", "fetch_status", "fetched_at"],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "id": "000001",
                    "original_url": "https://example.com/a",
                    "fetch_status": "success",
                    "fetched_at": "2026-01-01T00:00:00+00:00",
                }
            )

        wb = Workbook()
        ws = wb.active
        ws.append(EXPORT_COLUMNS)
        ws.append(
            [
                "000001",  # repository_source_id
                "legacy_scan",  # import_type
                "2026-01-01T00:00:00+00:00",  # imported_at
                "legacy",  # provenance_ref
                "doc.pdf",  # source_document
                "1",  # page_in_source
                "Sentence",  # citing_sentence
                "Paragraph",  # citing_paragraph
                "",  # context_before
                "",  # context_after
                "[1]",  # citation_raw
                "1",  # citation_ref_numbers
                "Author",  # cited_authors
                "Title",  # cited_title
                "2025",  # cited_year
                "Journal",  # cited_source
                "",  # cited_volume
                "",  # cited_issue
                "",  # cited_pages
                "",  # cited_doi
                "https://example.com/a",  # cited_url
                "Raw entry",  # cited_raw_entry
                0.9,  # match_confidence
                "legacy",  # match_method
                "",  # warnings
                "",  # cited_abstract
                "",  # cited_summary
                "",  # research_purpose
            ]
        )
        wb.save(repo_dir / "citations.xlsx")
        wb.close()

        status = self.service.attach(str(repo_dir))
        self.assertEqual(status.total_sources, 1)
        self.assertEqual(status.total_citations, 1)

    def test_seed_job_output_run_from_repository(self):
        repo_dir = self.tmp_path / "repo_seed_output"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))

        imported = self.service.import_source_list(
            filename="sources.csv",
            content=("URL\nhttps://example.com/a\n").encode("utf-8"),
        )
        self.assertEqual(imported.accepted_new, 1)

        markdown_rel = "markdown/000001.md"
        markdown_abs = repo_dir / markdown_rel
        markdown_abs.parent.mkdir(parents=True, exist_ok=True)
        markdown_abs.write_text("# Existing markdown\n", encoding="utf-8")

        state_path = repo_dir / ".ra_repo" / "repository_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["sources"][0]["fetch_status"] = "success"
        state["sources"][0]["markdown_file"] = markdown_rel
        state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

        export_job = self.service.create_export_job(scope="all")
        seeded = self.service.seed_job_output_run(export_job.job_id)
        self.assertEqual(seeded["seeded_rows"], 1)

        artifact = self.service.job_store_for(export_job.job_id).load_artifact(
            export_job.job_id,
            "06_sources_manifest",
        )
        self.assertIsNotNone(artifact)
        rows = artifact.get("rows", [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].get("fetch_status"), "success")
        self.assertEqual(rows[0].get("markdown_file"), markdown_rel)

    def test_prepare_standardized_pipeline_documents_uploads_standardized_markdown_only(self):
        repo_dir = self._attach_repo("repo_preprocess_pipeline")
        stored_document = self._write_repo_document(
            repo_dir,
            "import123",
            "report.md",
            (
                "# Housing Retrofit Findings\n\n"
                "Major benefits were reported [1].\n\n"
                "## Works Cited\n"
                "1. Agency. Housing Retrofit Findings. https://example.com/findings\n"
            ).encode("utf-8"),
        )
        job_id, job_store = self._create_repo_job()

        prepared_documents, normalization_outputs = (
            self.service._prepare_standardized_pipeline_documents(
                job_id=job_id,
                documents=[stored_document],
                settings=RepoSettings(),
            )
        )

        self.assertEqual(len(prepared_documents), 1)
        self.assertEqual(len(normalization_outputs), 1)
        self.assertTrue(prepared_documents[0]["filename"].endswith(".standardized.md"))
        upload_names = sorted(
            path.name for path in job_store.get_upload_dir(job_id).iterdir() if path.is_file()
        )
        self.assertEqual(upload_names, [prepared_documents[0]["filename"]])
        standardized_path = repo_dir / prepared_documents[0]["standardized_markdown_path"]
        self.assertTrue(standardized_path.is_file())
        self.assertIn(
            "Works Cited",
            standardized_path.read_text(encoding="utf-8"),
        )

        preprocess_artifact = job_store.load_artifact(job_id, "00_repository_preprocess")
        self.assertIsNotNone(preprocess_artifact)
        self.assertEqual(len(preprocess_artifact.get("prepared_documents", [])), 1)

        status = job_store.get_job_status(job_id)
        self.assertEqual(status["repository_preprocess_state"], "completed")
        self.assertIn("Prepared 1/1 standardized markdown file(s)", status["repository_preprocess_message"])

    def test_prepare_standardized_pipeline_documents_reuses_existing_standardized_markdown(self):
        repo_dir = self._attach_repo("repo_preprocess_reuse")
        stored_document = self._write_repo_document(
            repo_dir,
            "import123",
            "report.md",
            b"# Original\n\nBody text.\n",
        )
        standardized_path = repo_dir / "documents" / "import123" / "report.standardized.md"
        metadata_path = repo_dir / "documents" / "import123" / "report.standardized.json"
        standardized_path.write_text("# Reused Standardized\n\nClaim. [1]\n", encoding="utf-8")
        metadata_path.write_text(
            json.dumps(
                {
                    "filename": "report.md",
                    "source_document_path": stored_document["repository_path"],
                    "standardized_markdown_path": "documents/import123/report.standardized.md",
                    "metadata_path": "documents/import123/report.standardized.json",
                    "selected_profile_id": "llm_deep_research_markdown",
                    "selected_profile_label": "LLM Deep Research Markdown",
                    "status": "normalized",
                    "confidence_score": 0.91,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        job_id, job_store = self._create_repo_job()

        prepared_documents, normalization_outputs = (
            self.service._prepare_standardized_pipeline_documents(
                job_id=job_id,
                documents=[stored_document],
                settings=RepoSettings(),
                profile_override="generic_numeric_academic",
            )
        )

        self.assertEqual(len(prepared_documents), 1)
        self.assertEqual(normalization_outputs[0].markdown_text, "# Reused Standardized\n\nClaim. [1]\n")
        self.assertIn(
            "Reused existing standardized markdown file.",
            normalization_outputs[0].result.warnings,
        )
        uploaded = job_store.get_upload_dir(job_id) / prepared_documents[0]["filename"]
        self.assertEqual(uploaded.read_text(encoding="utf-8"), "# Reused Standardized\n\nClaim. [1]\n")
        status = job_store.get_job_status(job_id)
        self.assertEqual(status["repository_preprocess_state"], "completed")
        self.assertIn("1 reused", status["repository_preprocess_message"])

    def test_rebuild_manifest_includes_summary_and_rating_details(self):
        repo_dir = self.tmp_path / "repo_manifest_details"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))
        self.service.import_source_list(
            filename="sources.csv",
            content=("URL\nhttps://example.com/a\n").encode("utf-8"),
        )

        summary_rel = "summaries/000001_summary.md"
        rating_rel = "ratings/000001_rating.json"
        (repo_dir / "summaries").mkdir(parents=True, exist_ok=True)
        (repo_dir / "ratings").mkdir(parents=True, exist_ok=True)
        (repo_dir / summary_rel).write_text(
            "This is the full source summary.",
            encoding="utf-8",
        )
        (repo_dir / rating_rel).write_text(
            json.dumps(
                {
                    "overall_score": 0.85,
                    "confidence": 0.9,
                    "rationale": "High relevance to the project profile.",
                    "relevant_sections": [
                        {
                            "section": "Results",
                            "text": "Heat pump retrofits cut operating costs.",
                        }
                    ],
                    "ratings": {
                        "relevance": 0.95,
                        "evidence_quality": 0.8,
                    },
                    "flags": {
                        "needs_manual_review": 0.25,
                    },
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        state_path = repo_dir / ".ra_repo" / "repository_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["sources"][0]["summary_file"] = summary_rel
        state["sources"][0]["summary_status"] = "generated"
        state["sources"][0]["rating_file"] = rating_rel
        state["sources"][0]["rating_status"] = "generated"
        state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

        self.service.rebuild()

        with (repo_dir / "manifest.csv").open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            fieldnames = reader.fieldnames or []
            rows = list(reader)

        self.assertIn("summary_text", fieldnames)
        self.assertIn("rating_overall", fieldnames)
        self.assertIn("rating_confidence", fieldnames)
        self.assertIn("rating_relevance", fieldnames)
        self.assertIn("rating_evidence_quality", fieldnames)
        self.assertIn("flag_needs_manual_review", fieldnames)
        self.assertEqual(rows[0]["summary_text"], "This is the full source summary.")
        self.assertEqual(rows[0]["rating_overall"], "0.85")
        self.assertEqual(rows[0]["rating_confidence"], "0.9")
        self.assertIn("Heat pump retrofits cut operating costs.", rows[0]["relevant_sections"])
        self.assertEqual(rows[0]["rating_relevance"], "0.95")
        self.assertEqual(rows[0]["flag_needs_manual_review"], "0.25")

    def test_cleanup_repository_layout_moves_legacy_files_into_sources_folder(self):
        repo_dir = self.tmp_path / "repo_cleanup_layout"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))
        self.service.import_source_list(
            filename="sources.csv",
            content=("URL\nhttps://example.com/a\n").encode("utf-8"),
        )

        markdown_rel = "markdown/000001_clean.md"
        summary_rel = "summaries/000001_summary.md"
        rating_rel = "ratings/000001_rating.json"
        for rel_path, content in [
            (markdown_rel, "# Clean markdown\n"),
            (summary_rel, "Legacy summary text\n"),
            (rating_rel, json.dumps({"confidence": 0.8, "ratings": {"relevance": 0.9}}, ensure_ascii=False)),
        ]:
            abs_path = repo_dir / rel_path
            abs_path.parent.mkdir(parents=True, exist_ok=True)
            abs_path.write_text(content, encoding="utf-8")

        state_path = repo_dir / ".ra_repo" / "repository_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        source = state["sources"][0]
        source["markdown_file"] = markdown_rel
        source["summary_file"] = summary_rel
        source["summary_status"] = "generated"
        source["rating_file"] = rating_rel
        source["rating_status"] = "generated"
        state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

        response = self.service.cleanup_repository_layout()
        self.assertEqual(response.status, "completed")

        updated_state = json.loads(state_path.read_text(encoding="utf-8"))
        updated_source = updated_state["sources"][0]
        self.assertTrue(updated_source["markdown_file"].startswith("sources/000001/"))
        self.assertTrue(updated_source["summary_file"].startswith("sources/000001/"))
        self.assertTrue(updated_source["rating_file"].startswith("sources/000001/"))
        self.assertFalse((repo_dir / markdown_rel).exists())
        self.assertFalse((repo_dir / summary_rel).exists())
        self.assertFalse((repo_dir / rating_rel).exists())
        self.assertTrue((repo_dir / updated_source["markdown_file"]).is_file())
        self.assertTrue((repo_dir / updated_source["summary_file"]).is_file())
        self.assertTrue((repo_dir / updated_source["rating_file"]).is_file())

    def test_create_scaffolds_repository_first_layout(self):
        repo_dir = self.tmp_path / "repo_scaffold"
        status = self.service.create(str(repo_dir))

        self.assertTrue(status.attached)
        self.assertTrue((repo_dir / ".ra_repo" / "repository.json").exists())
        self.assertTrue((repo_dir / ".ra_repo" / "repository_state.json").exists())
        self.assertTrue((repo_dir / ".ra_repo" / "settings.json").exists())
        self.assertTrue((repo_dir / ".ra_repo" / "bundled_ingestion_profiles.json").exists())
        self.assertTrue((repo_dir / ".ra_repo" / "backups").is_dir())
        self.assertTrue((repo_dir / "sources").is_dir())
        self.assertTrue((repo_dir / "documents").is_dir())
        self.assertTrue((repo_dir / "project_profiles").is_dir())
        self.assertTrue((repo_dir / "project_profiles" / "default_project_profile.yaml").is_file())

        manifest_text = (repo_dir / "manifest.csv").read_text(encoding="utf-8-sig")
        citations_text = (repo_dir / "citations.csv").read_text(encoding="utf-8-sig")
        self.assertTrue(manifest_text.startswith("id,repository_source_id,"))
        self.assertTrue(citations_text.startswith(",".join(EXPORT_COLUMNS)))
        self.assertTrue((repo_dir / "manifest.xlsx").exists())

    def test_repo_settings_round_trip_persists_default_project_profile_name(self):
        repo_dir = self.tmp_path / "repo_settings_round_trip"
        self.service.create(str(repo_dir))

        settings = RepoSettings(
            use_llm=True,
            research_purpose="Track housing retrofit workforce evidence.",
            default_project_profile_name="housing_retrofit.yaml",
            fetch_delay=3.5,
        )
        self.service.save_repo_settings(settings)

        loaded = self.service.load_repo_settings()
        raw = json.loads((repo_dir / ".ra_repo" / "settings.json").read_text(encoding="utf-8"))

        self.assertEqual(loaded.default_project_profile_name, "housing_retrofit.yaml")
        self.assertEqual(loaded.fetch_delay, 3.5)
        self.assertEqual(raw["default_project_profile_name"], "housing_retrofit.yaml")
        self.assertEqual(raw["research_purpose"], "Track housing retrofit workforce evidence.")

    def test_column_configs_persist_and_blank_custom_columns_render_in_manifest(self):
        repo_dir = self._attach_repo("repo_column_configs")
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
            ).encode("utf-8"),
        )

        custom = self.service.create_column("POC")
        builtin = self.service.update_column(
            "title",
            patch={"instruction_prompt": "Normalize the title string and keep it concise."},
        )

        manifest = self.service.list_manifest(limit=10, offset=0, sort_by="id", sort_dir="asc")
        columns = {column["key"]: column for column in manifest["columns"]}

        self.assertEqual(custom.kind, "custom")
        self.assertIn(custom.id, columns)
        self.assertEqual(columns[custom.id]["label"], "POC")
        self.assertEqual(columns[custom.id]["kind"], "custom")
        self.assertEqual(columns[custom.id]["processable"], True)
        self.assertIn(custom.id, manifest["rows"][0])
        self.assertEqual(manifest["rows"][0][custom.id], "")
        self.assertEqual(builtin.instruction_prompt, "Normalize the title string and keep it concise.")
        self.assertEqual(columns["title"]["instruction_prompt"], builtin.instruction_prompt)
        self.assertTrue(columns["author_names"]["instruction_prompt"])
        self.assertTrue(columns["publication_year"]["instruction_prompt"])

        self.service.attach(str(repo_dir))
        reloaded = self.service.list_manifest(limit=10, offset=0, sort_by="id", sort_dir="asc")
        reloaded_columns = {column["key"]: column for column in reloaded["columns"]}
        self.assertEqual(reloaded_columns[custom.id]["label"], "POC")
        self.assertEqual(reloaded_columns["title"]["instruction_prompt"], builtin.instruction_prompt)
        self.assertTrue(reloaded_columns["publication_year"]["instruction_prompt"])

    def test_column_context_flags_persist_in_manifest_metadata(self):
        self._attach_repo("repo_column_context_flags")
        custom = self.service.create_column("POC")

        updated = self.service.update_column(
            custom.id,
            patch={
                "instruction_prompt": "Answer yes or no only.",
                "include_source_text": False,
                "include_row_context": True,
            },
        )

        manifest = self.service.list_manifest(limit=10, offset=0, sort_by="id", sort_dir="asc")
        columns = {column["key"]: column for column in manifest["columns"]}

        self.assertFalse(updated.include_source_text)
        self.assertTrue(updated.include_row_context)
        self.assertEqual(columns[custom.id]["include_source_text"], False)
        self.assertEqual(columns[custom.id]["include_row_context"], True)
        self.assertEqual(columns["title"]["include_source_text"], True)
        self.assertEqual(columns["title"]["include_row_context"], False)

    def test_export_manifest_supports_selected_csv_and_xlsx(self):
        self._attach_repo("repo_manifest_export")
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
            ).encode("utf-8"),
        )

        csv_bytes, csv_headers, csv_media_type = self.service.export_manifest(
            RepositoryManifestExportRequest(
                scope="selected",
                format="csv",
                source_ids=["000002", "000001"],
            )
        )
        csv_rows = list(csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig"))))
        self.assertEqual(csv_media_type, "text/csv; charset=utf-8")
        self.assertEqual(csv_headers["Content-Disposition"], 'attachment; filename="selected-manifest.csv"')
        self.assertEqual([row["id"] for row in csv_rows], ["000002", "000001"])

        xlsx_bytes, xlsx_headers, xlsx_media_type = self.service.export_manifest(
            RepositoryManifestExportRequest(
                scope="selected",
                format="xlsx",
                source_ids=["000002", "000001"],
            )
        )
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        worksheet = workbook.active
        exported_ids = [worksheet.cell(row=row_index, column=1).value for row_index in range(2, 4)]
        self.assertEqual(
            xlsx_media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(xlsx_headers["Content-Disposition"], 'attachment; filename="selected-manifest.xlsx"')
        self.assertEqual(exported_ids, ["000002", "000001"])

    def test_export_manifest_can_limit_to_visible_columns(self):
        self._attach_repo("repo_manifest_export_visible_columns")
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
            ).encode("utf-8"),
        )

        csv_bytes, _headers, _media_type = self.service.export_manifest(
            RepositoryManifestExportRequest(
                scope="selected",
                format="csv",
                column_scope="visible",
                column_keys=["title", "author_names", "file_md"],
                source_ids=["000001"],
            )
        )
        csv_rows = list(csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig"))))

        self.assertEqual(list(csv_rows[0].keys()), ["title", "author_names", "file_md"])
        self.assertEqual(csv_rows[0]["title"], "Alpha Source")

    def test_export_repository_bundle_packages_files_csv_ris_and_custom_columns(self):
        repo_dir = self._attach_repo("repo_bundle_export")
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
            ).encode("utf-8"),
        )

        custom = self.service.create_column("Priority")
        self.service.update_source(
            "000001",
            patch={
                "author_names": "Jane Doe; John Roe",
                "publication_date": "2025-03-15",
                "document_type": "report",
                "organization_name": "Alpha Agency",
                "organization_type": "government",
                "citation_title": "Alpha Policy Memo",
                "citation_authors": "Jane Doe; John Roe",
                "citation_issued": "2025-03-15",
                "citation_url": "https://example.com/a/citation",
                "citation_type": "report",
                "citation_report_number": "RPT-001",
                "overall_relevance": 0.9,
                "custom_fields": {custom.id: "Core source"},
            },
        )
        self.service.update_source(
            "000002",
            patch={
                "author_names": "Beta Org",
                "publication_date": "2024",
                "document_type": "brief",
                "organization_name": "Beta Org",
                "organization_type": "nonprofit",
                "citation_title": "Beta Research Brief",
                "citation_authors": "Beta Org",
                "citation_issued": "2024",
                "citation_url": "https://example.com/b/citation",
                "citation_type": "webpage",
                "overall_relevance": 0.7,
            },
        )

        state_path = repo_dir / ".ra_repo" / "repository_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        rows = state["sources"]
        rows[0]["raw_file"] = "sources/000001/alpha.pdf"
        rows[0]["llm_cleanup_file"] = "sources/000001/alpha_clean.md"
        rows[0]["summary_file"] = "sources/000001/alpha_summary.txt"
        rows[0]["rating_file"] = "sources/000001/alpha_rating.json"
        rows[0]["fetch_status"] = "success"
        rows[0]["markdown_char_count"] = 1200
        rows[1]["raw_file"] = "sources/000002/beta.html"
        rows[1]["rendered_pdf_file"] = "sources/000002/beta_rendered.pdf"
        rows[1]["markdown_file"] = "sources/000002/beta.md"
        rows[1]["fetch_status"] = "success"
        rows[1]["markdown_char_count"] = 800
        state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

        (repo_dir / "sources" / "000001").mkdir(parents=True, exist_ok=True)
        (repo_dir / "sources" / "000002").mkdir(parents=True, exist_ok=True)
        (repo_dir / "sources" / "000001" / "alpha.pdf").write_bytes(b"%PDF-1.4 alpha\n")
        (repo_dir / "sources" / "000001" / "alpha_clean.md").write_text(
            "# Alpha clean markdown\n",
            encoding="utf-8",
        )
        (repo_dir / "sources" / "000001" / "alpha_summary.txt").write_text(
            "Alpha summary for exported browser.",
            encoding="utf-8",
        )
        (repo_dir / "sources" / "000001" / "alpha_rating.json").write_text(
            json.dumps(
                {
                    "summary": "Alpha summary from rating fallback.",
                    "rationale": "Alpha rationale from rating payload.",
                    "relevant_sections": ["Executive Summary", "Policy Recommendations"],
                    "overall_relevance": 0.9,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        (repo_dir / "sources" / "000002" / "beta.html").write_text(
            "<html><body>Beta</body></html>",
            encoding="utf-8",
        )
        (repo_dir / "sources" / "000002" / "beta_rendered.pdf").write_bytes(
            b"%PDF-1.4 beta rendered\n"
        )
        (repo_dir / "sources" / "000002" / "beta.md").write_text(
            "# Beta markdown\n",
            encoding="utf-8",
        )
        self.service.rebuild()

        bundle_bytes, headers = self.service.export_repository_bundle(
            RepositoryBundleExportRequest(
                scope="selected",
                source_ids=["000001", "000002"],
                file_kinds=["pdf", "rendered", "html", "md"],
            )
        )

        self.assertEqual(headers["Content-Disposition"], 'attachment; filename="selected-repository-export.zip"')
        archive = zipfile.ZipFile(io.BytesIO(bundle_bytes))
        names = set(archive.namelist())
        self.assertIn("index.html", names)
        self.assertIn("research-export.csv", names)
        self.assertIn("citations.ris", names)
        self.assertIn("PDF/Jane Doe et al - 2025-03-15 - Alpha Policy Memo.pdf", names)
        self.assertIn("RENDERED/Beta Org - 2024 - Beta Research Brief.pdf", names)
        self.assertIn("MD/Jane Doe et al - 2025-03-15 - Alpha Policy Memo.md", names)
        self.assertIn("HTML/Beta Org - 2024 - Beta Research Brief.html", names)
        self.assertIn("MD/Beta Org - 2024 - Beta Research Brief.md", names)

        viewer_html = archive.read("index.html").decode("utf-8")
        self.assertIn("Offline Repository Browser", viewer_html)
        self.assertIn("Export Selected Bundle", viewer_html)
        self.assertIn("Export All Bundle", viewer_html)
        self.assertIn("Export Selected RIS", viewer_html)
        self.assertIn("Alpha Policy Memo", viewer_html)
        self.assertIn("Alpha summary for exported browser.", viewer_html)
        self.assertIn("Alpha rationale from rating payload.", viewer_html)
        self.assertIn("Policy Recommendations", viewer_html)
        self.assertIn("# Alpha clean markdown", viewer_html)
        self.assertIn("PDF/Jane Doe et al - 2025-03-15 - Alpha Policy Memo.pdf", viewer_html)
        self.assertIn("RENDERED/Beta Org - 2024 - Beta Research Brief.pdf", viewer_html)
        self.assertIn('"Priority"', viewer_html)
        self.assertIn("Core source", viewer_html)
        self.assertIn("showDirectoryPicker", viewer_html)
        self.assertNotIn(str(repo_dir), viewer_html)

        csv_rows = list(
            csv.DictReader(io.StringIO(archive.read("research-export.csv").decode("utf-8-sig")))
        )
        self.assertEqual(
            list(csv_rows[0].keys()),
            [
                "Title",
                "Authors",
                "Publication Date",
                "Organization",
                "Organization Type",
                "URL",
                "Markdown Char Count",
                "Report Number",
                "Document Type",
                "Citation Type",
                "Overall Rating",
                "Priority",
            ],
        )
        self.assertEqual(csv_rows[0]["Title"], "Alpha Source")
        self.assertEqual(csv_rows[0]["URL"], "https://example.com/a/citation")
        self.assertEqual(csv_rows[0]["Priority"], "Core source")
        self.assertEqual(csv_rows[0]["Overall Rating"], "0.9")

        ris_text = archive.read("citations.ris").decode("utf-8")
        self.assertIn("TI  - Alpha Policy Memo", ris_text)
        self.assertIn("TI  - Beta Research Brief", ris_text)

    def test_export_repository_bundle_cloud_mode_writes_manifest_and_storage_filenames(self):
        repo_dir = self._attach_repo("repo_bundle_cloud_export")
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
            ).encode("utf-8"),
        )

        self.service.update_source(
            "000001",
            patch={
                "author_names": "Jane Doe",
                "publication_date": "2025-03-15",
                "document_type": "report",
                "organization_name": "Alpha Agency",
                "organization_type": "government",
                "citation_title": "Alpha Policy Memo",
                "citation_authors": "Jane Doe",
                "citation_issued": "2025-03-15",
                "citation_url": "https://example.com/a/citation",
                "citation_type": "report",
                "overall_relevance": 0.9,
            },
        )
        self.service.update_source(
            "000002",
            patch={
                "author_names": "Beta Org",
                "publication_date": "2024",
                "document_type": "brief",
                "organization_name": "Beta Org",
                "organization_type": "nonprofit",
                "citation_title": "Budget Brief",
                "citation_authors": "Beta Org",
                "citation_issued": "2024",
                "citation_url": "https://example.com/b/citation",
                "citation_type": "webpage",
                "overall_relevance": 0.7,
            },
        )

        state_path = repo_dir / ".ra_repo" / "repository_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        rows = state["sources"]
        rows[0]["raw_file"] = "sources/000001/Résumé, Final (A).pdf"
        rows[0]["fetch_status"] = "success"
        rows[1]["rendered_pdf_file"] = "sources/000002/Budget (FY2024) – Final.pdf"
        rows[1]["markdown_file"] = "sources/000002/Budget Notes.md"
        rows[1]["fetch_status"] = "success"
        state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

        (repo_dir / "sources" / "000001").mkdir(parents=True, exist_ok=True)
        (repo_dir / "sources" / "000002").mkdir(parents=True, exist_ok=True)
        (repo_dir / "sources" / "000001" / "Résumé, Final (A).pdf").write_bytes(b"%PDF-1.4 alpha\n")
        (repo_dir / "sources" / "000002" / "Budget (FY2024) – Final.pdf").write_bytes(
            b"%PDF-1.4 beta rendered\n"
        )
        (repo_dir / "sources" / "000002" / "Budget Notes.md").write_text(
            "# Budget markdown\n",
            encoding="utf-8",
        )
        self.service.rebuild()

        bundle_bytes, headers = self.service.export_repository_bundle(
            RepositoryBundleExportRequest(
                mode="cloud",
                scope="selected",
                source_ids=["000001", "000002"],
                file_kinds=["pdf", "rendered", "md"],
                base_url="https://cdn.example.com/client-a",
            )
        )

        self.assertEqual(
            headers["Content-Disposition"],
            'attachment; filename="selected-repository-cloud-export.zip"',
        )
        archive = zipfile.ZipFile(io.BytesIO(bundle_bytes))
        names = set(archive.namelist())
        self.assertIn("index.html", names)
        self.assertIn("manifest.json", names)
        self.assertIn("research-export.csv", names)
        self.assertIn("citations.ris", names)
        self.assertIn("files/resume-final-a-000001-pdf.pdf", names)
        self.assertIn("files/budget-fy2024-final-000002-rendered.pdf", names)
        self.assertIn("files/budget-notes-000002-md.md", names)

        viewer_html = archive.read("index.html").decode("utf-8")
        self.assertIn('const BASE_URL = "https://cdn.example.com/client-a/";', viewer_html)
        self.assertIn("Cloud Repository Browser", viewer_html)
        self.assertIn("cloud export", viewer_html)
        self.assertIn("encodeURIComponent", viewer_html)
        self.assertIn("resume-final-a-000001-pdf.pdf", viewer_html)
        self.assertIn("const label = file.label;", viewer_html)
        self.assertIn('let lastAnchorId = "";', viewer_html)
        self.assertIn("Boolean(event.shiftKey)", viewer_html)
        self.assertIn("Direct file links can still open without cross-origin byte access.", viewer_html)
        self.assertIn('window.location.protocol === "file:"', viewer_html)
        self.assertNotIn("Offline Repository Browser", viewer_html)
        self.assertNotIn("offline export", viewer_html)
        self.assertNotIn("Choose Package Folder", viewer_html)
        self.assertNotIn("showDirectoryPicker", viewer_html)
        self.assertNotIn("Edit BASE_URL near the top of this file after upload.", viewer_html)
        self.assertNotIn('folderStatus.textContent = "BASE_URL "', viewer_html)
        self.assertNotIn("Custom Columns", viewer_html)
        self.assertLess(viewer_html.index("<h3>Exported Files</h3>"), viewer_html.index('renderRichTextSection("Summary"'))
        self.assertLess(viewer_html.index("<h3>Full Markdown</h3>"), viewer_html.index('renderRichTextSection("Relevance Rationale"'))
        self.assertNotIn(str(repo_dir), viewer_html)

        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
        self.assertEqual(manifest["exportMode"], "cloud")
        self.assertEqual(manifest["rows"][0]["files"]["pdf"]["displayName"], "Résumé, Final (A).pdf")
        self.assertEqual(
            manifest["rows"][0]["files"]["pdf"]["storageName"],
            "resume-final-a-000001-pdf.pdf",
        )
        self.assertEqual(
            manifest["rows"][1]["files"]["rendered"]["storageName"],
            "budget-fy2024-final-000002-rendered.pdf",
        )
        self.assertEqual(
            manifest["rows"][1]["files"]["md"]["storageName"],
            "budget-notes-000002-md.md",
        )
        self.assertNotIn("relativePath", manifest["rows"][0]["files"]["pdf"])

    def test_export_repository_bundle_cloud_mode_requires_base_url(self):
        self._attach_repo("repo_bundle_cloud_requires_url")
        self.service.import_source_list(
            filename="sources.csv",
            content=("URL,Title\nhttps://example.com/a,Alpha Source\n").encode("utf-8"),
        )

        with self.assertRaisesRegex(ValueError, "Base URL"):
            self.service.export_repository_bundle(
                RepositoryBundleExportRequest(
                    mode="cloud",
                    scope="all",
                    source_ids=[],
                    file_kinds=["pdf"],
                    base_url="",
                )
            )

    def test_column_run_uses_all_filtered_rows_across_pages(self):
        self._attach_repo("repo_column_scope")
        csv_rows = ["URL,Title"]
        csv_rows.extend(
            f"https://example.com/source-{index:03d},Scope Source {index:03d}"
            for index in range(260)
        )
        self.service.import_source_list(
            filename="sources.csv",
            content=("\n".join(csv_rows) + "\n").encode("utf-8"),
        )

        settings = self.service.load_repo_settings()
        settings.use_llm = True
        settings.llm_backend.model = "test-model"
        self.service.save_repo_settings(settings)

        column = self.service.create_column("Classification")
        self.service.update_column(
            column.id,
            patch={"instruction_prompt": "Answer yes or no only based on the document."},
        )

        page = self.service.list_manifest(q="Scope Source", limit=250, offset=0)
        self.assertEqual(page["total"], 260)
        self.assertEqual(len(page["rows"]), 250)

        with patch("backend.storage.attached_repository.llm_backend_ready_for_chat", return_value=True), patch(
            "backend.storage.attached_repository.threading.Thread",
            _NoOpThread,
        ):
            response = self.service.start_column_run(
                column.id,
                payload=RepositoryColumnRunRequest(
                    filters=RepositoryManifestFilterRequest(q="Scope Source"),
                    confirm_overwrite=True,
                ),
            )

        self.assertEqual(response.status, "started")
        self.assertEqual(response.total_rows, 260)

    def test_column_run_returns_confirmation_when_custom_cells_are_populated(self):
        self._attach_repo("repo_column_confirmation")
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
            ).encode("utf-8"),
        )

        settings = self.service.load_repo_settings()
        settings.use_llm = True
        settings.llm_backend.model = "test-model"
        self.service.save_repo_settings(settings)

        column = self.service.create_column("POC")
        self.service.update_column(
            column.id,
            patch={"instruction_prompt": "Answer yes or no only based on the document."},
        )
        self.service.update_source("000001", patch={"custom_fields": {column.id: "yes"}})

        with patch("backend.storage.attached_repository.llm_backend_ready_for_chat", return_value=True):
            response = self.service.start_column_run(
                column.id,
                payload=RepositoryColumnRunRequest(
                    filters=RepositoryManifestFilterRequest(),
                    confirm_overwrite=False,
                ),
            )

        self.assertEqual(response.status, "confirmation_required")
        self.assertEqual(response.total_rows, 2)
        self.assertEqual(response.populated_rows, 1)

    def test_column_run_can_target_only_selected_rows(self):
        self._attach_repo("repo_column_selected_scope")
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
                "https://example.com/c,Gamma Source\n"
            ).encode("utf-8"),
        )

        settings = self.service.load_repo_settings()
        settings.use_llm = True
        settings.llm_backend.model = "test-model"
        self.service.save_repo_settings(settings)

        column = self.service.create_column("POC")
        self.service.update_column(
            column.id,
            patch={"instruction_prompt": "Answer yes or no only based on the document."},
        )

        with patch("backend.storage.attached_repository.llm_backend_ready_for_chat", return_value=True), patch(
            "backend.storage.attached_repository.threading.Thread",
            _NoOpThread,
        ):
            response = self.service.start_column_run(
                column.id,
                payload=RepositoryColumnRunRequest(
                    scope="selected",
                    source_ids=["000002", "000003"],
                    confirm_overwrite=True,
                ),
            )

        self.assertEqual(response.status, "started")
        self.assertEqual(response.total_rows, 2)

    def test_builtin_column_run_isolates_row_failures_and_updates_catalog_artifacts(self):
        repo_dir = self._attach_repo("repo_builtin_column_run")
        self.service.import_source_list(
            filename="sources.csv",
            content=(
                "URL,Title\n"
                "https://example.com/a,Alpha Source\n"
                "https://example.com/b,Beta Source\n"
            ).encode("utf-8"),
        )

        settings = self.service.load_repo_settings()
        settings.use_llm = True
        settings.llm_backend.model = "test-model"
        self.service.save_repo_settings(settings)
        manifest = self.service.list_manifest(limit=10, offset=0, sort_by="id", sort_dir="asc")
        columns = {column["key"]: column for column in manifest["columns"]}
        self.assertTrue(columns["title"]["instruction_prompt"])

        class DummyClient:
            call_count = 0

            def __init__(self, *_args, **_kwargs):
                return None

            def sync_chat_completion(self, **_kwargs):
                type(self).call_count += 1
                if type(self).call_count == 1:
                    return json.dumps({"value": "ALPHA SOURCE REWRITTEN", "status": "ok"})
                raise RuntimeError("Simulated LLM failure")

            def sync_close(self):
                return None

        with patch("backend.storage.attached_repository.llm_backend_ready_for_chat", return_value=True), patch(
            "backend.storage.attached_repository.UnifiedLLMClient",
            DummyClient,
        ), patch(
            "backend.storage.attached_repository.threading.Thread",
            _NoOpThread,
        ):
            response = self.service.start_column_run(
                "title",
                payload=RepositoryColumnRunRequest(
                    filters=RepositoryManifestFilterRequest(),
                    confirm_overwrite=True,
                ),
            )
            self.service._repository_column_run_worker(
                response.job_id,
                "title",
                ["000001", "000002"],
                settings,
            )

        status = self.service.get_column_run_status(response.job_id)
        self.assertEqual(status.state, "completed")
        self.assertEqual(status.succeeded_rows, 1)
        self.assertEqual(status.failed_rows, 1)
        self.assertEqual(len(status.row_errors), 1)

        manifest = self.service.list_manifest(limit=10, offset=0, sort_by="id", sort_dir="asc")
        self.assertEqual(manifest["rows"][0]["title"], "ALPHA SOURCE REWRITTEN")
        self.assertEqual(manifest["rows"][1]["title"], "Beta Source")

        catalog_path = repo_dir / "sources" / "000001" / "000001_catalog.json"
        self.assertTrue(catalog_path.exists())
        catalog_payload = json.loads(catalog_path.read_text(encoding="utf-8"))
        self.assertEqual(catalog_payload["title"], "ALPHA SOURCE REWRITTEN")

    def test_author_column_run_falls_back_to_organization_and_updates_ris_export(self):
        repo_dir = self._attach_repo("repo_author_column_fallback")
        self.service.import_source_list(
            filename="sources.csv",
            content=("URL,Title\nhttps://example.com/a,Alpha Source\n").encode("utf-8"),
        )

        self.service.update_source(
            "000001",
            patch={
                "title": "Alpha Source",
                "publication_date": "2025",
                "document_type": "report",
                "organization_name": "California Energy Commission",
            },
        )

        settings = self.service.load_repo_settings()
        settings.use_llm = True
        settings.llm_backend.model = "test-model"
        self.service.save_repo_settings(settings)

        class DummyClient:
            def __init__(self, *_args, **_kwargs):
                return None

            def sync_chat_completion(self, **_kwargs):
                return json.dumps({"value": "", "status": "insufficient_evidence"})

            def sync_close(self):
                return None

        with patch("backend.storage.attached_repository.llm_backend_ready_for_chat", return_value=True), patch(
            "backend.storage.attached_repository.UnifiedLLMClient",
            DummyClient,
        ), patch(
            "backend.storage.attached_repository.threading.Thread",
            _NoOpThread,
        ):
            response = self.service.start_column_run(
                "author_names",
                payload=RepositoryColumnRunRequest(
                    filters=RepositoryManifestFilterRequest(),
                    confirm_overwrite=True,
                ),
            )
            self.service._repository_column_run_worker(
                response.job_id,
                "author_names",
                ["000001"],
                settings,
            )

        manifest = self.service.list_manifest(limit=10, offset=0, sort_by="id", sort_dir="asc")
        row = manifest["rows"][0]
        self.assertEqual(row["author_names"], "California Energy Commission")
        self.assertEqual(row["citation_authors"], "California Energy Commission")
        self.assertEqual(row["citation_verification_status"], "verified")
        self.assertTrue(row["citation_ready"])

        catalog = json.loads(
            (repo_dir / "sources" / "000001" / "000001_catalog.json").read_text(encoding="utf-8")
        )
        self.assertEqual(catalog["citation"]["authors"][0]["literal"], "California Energy Commission")
        self.assertEqual(catalog["citation"]["verification_status"], "verified")
        self.assertTrue(catalog["citation"]["ready_for_ris"])

        ris_bytes, headers = self.service.export_citations_ris(
            RepositoryCitationRisExportRequest(scope="all")
        )
        decoded = ris_bytes.decode("utf-8")
        self.assertIn("AU  - California Energy Commission", decoded)
        self.assertEqual(headers["X-ResearchAssistant-Exported-Count"], "1")

    def test_citation_ready_column_run_recomputes_using_publication_year_without_llm(self):
        repo_dir = self._attach_repo("repo_citation_ready_column_run")
        self.service.import_source_list(
            filename="sources.csv",
            content=("URL,Title\nhttps://example.com/a,Alpha Source\n").encode("utf-8"),
        )

        state_path = repo_dir / ".ra_repo" / "repository_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["sources"][0]["title"] = "Alpha Source"
        state["sources"][0]["author_names"] = "Jane Doe"
        state["sources"][0]["publication_date"] = ""
        state["sources"][0]["publication_year"] = "2024"
        state["sources"][0]["document_type"] = "report"
        state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
        self.service.rebuild()

        manifest = self.service.list_manifest(limit=10, offset=0, sort_by="id", sort_dir="asc")
        columns = {column["key"]: column for column in manifest["columns"]}
        self.assertTrue(columns["citation_ready"]["processable"])
        self.assertFalse(columns["citation_ready"]["requires_llm"])

        settings = self.service.load_repo_settings()
        with patch("backend.storage.attached_repository.threading.Thread", _NoOpThread):
            response = self.service.start_column_run(
                "citation_ready",
                payload=RepositoryColumnRunRequest(
                    scope="empty_only",
                    confirm_overwrite=True,
                ),
            )
            self.service._repository_column_run_worker(
                response.job_id,
                "citation_ready",
                ["000001"],
                settings,
            )

        manifest = self.service.list_manifest(limit=10, offset=0, sort_by="id", sort_dir="asc")
        row = manifest["rows"][0]
        self.assertTrue(row["citation_ready"])
        self.assertEqual(row["citation_issued"], "2024")
        self.assertEqual(row["citation_missing_fields"], "")

        catalog = json.loads(
            (repo_dir / "sources" / "000001" / "000001_catalog.json").read_text(encoding="utf-8")
        )
        self.assertEqual(catalog["citation"]["issued"], "2024")
        self.assertTrue(catalog["citation"]["ready_for_ris"])

    def test_list_ingestion_profiles_reads_repo_bundled_snapshot(self):
        repo_dir = self._attach_repo("repo_bundled_profiles")
        bundled_path = repo_dir / ".ra_repo" / "bundled_ingestion_profiles.json"
        bundled_path.write_text(
            json.dumps(
                [
                    IngestionProfile(
                        profile_id="repo_local_bundled_profile",
                        label="Repo Local Bundled Profile",
                        built_in=True,
                        file_type_hints=["md"],
                    ).model_dump(mode="json")
                ],
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        profiles = self.service.list_ingestion_profiles().profiles
        self.assertEqual([profile.profile_id for profile in profiles], ["repo_local_bundled_profile"])
        self.assertTrue(profiles[0].built_in)

    def test_merge_processing_job_results_persists_repository_sources_and_citations(self):
        repo_dir = self.tmp_path / "repo_documents"
        repo_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(repo_dir))

        job_id, job_store = self._create_repo_job()
        job_store.save_artifact(
            job_id,
            "01_ingestion",
            {
                "documents": [
                    {
                        "filename": "alpha.md",
                        "file_type": "md",
                        "total_pages": None,
                        "blocks": [
                            {
                                "text": "Policy Memo",
                                "page_number": None,
                                "block_index": 0,
                                "is_heading": True,
                                "heading_level": 1,
                                "char_offset_start": 0,
                                "char_offset_end": 11,
                            },
                            {
                                "text": "Alpha claim [1] is important.",
                                "page_number": None,
                                "block_index": 1,
                                "is_heading": False,
                                "heading_level": None,
                                "char_offset_start": 12,
                                "char_offset_end": 41,
                            },
                            {
                                "text": "References",
                                "page_number": None,
                                "block_index": 2,
                                "is_heading": True,
                                "heading_level": 2,
                                "char_offset_start": 42,
                                "char_offset_end": 52,
                            },
                            {
                                "text": "[1] Example Org. Alpha. 2024. https://example.com/a",
                                "page_number": None,
                                "block_index": 3,
                                "is_heading": False,
                                "heading_level": None,
                                "char_offset_start": 53,
                                "char_offset_end": 106,
                            },
                        ],
                        "full_text": "Policy Memo\nAlpha claim [1] is important.\nReferences\n[1] Example Org. Alpha. 2024. https://example.com/a",
                        "warnings": [],
                        "inline_citation_urls": {},
                    }
                ]
            },
        )
        job_store.save_artifact(
            job_id,
            "03_bibliography",
            {
                "sections": [
                    {
                        "document_filename": "alpha.md",
                        "start_block_index": 2,
                        "end_block_index": 3,
                        "heading_text": "References",
                        "raw_text": "[1] Example Org. Alpha. 2024. https://example.com/a",
                        "detection_method": "heading_match",
                        "confidence": 1.0,
                    }
                ],
                "entries": [
                    {
                        "ref_number": 1,
                        "raw_text": "[1] Example Org. Alpha. 2024. https://example.com/a",
                        "source_document_name": "alpha.md",
                        "authors": ["Example Org"],
                        "title": "Alpha",
                        "year": "2024",
                        "url": "https://example.com/a",
                        "parse_confidence": 1.0,
                        "parse_warnings": [],
                    }
                ],
                "total_raw_entries": 1,
                "parse_failures": 0,
            },
        )
        job_store.save_artifact(
            job_id,
            "05_export",
            {
                "rows": [
                    {
                        "repository_source_id": "",
                        "import_type": "",
                        "imported_at": "",
                        "provenance_ref": "",
                        "source_document": "alpha.md",
                        "page_in_source": "3",
                        "citing_sentence": "Alpha sentence",
                        "citing_paragraph": "Alpha paragraph",
                        "context_before": "",
                        "context_after": "",
                        "citation_raw": "[1]",
                        "citation_ref_numbers": "1",
                        "cited_authors": "Author",
                        "cited_title": "Alpha",
                        "cited_year": "2024",
                        "cited_source": "Journal",
                        "cited_volume": "",
                        "cited_issue": "",
                        "cited_pages": "",
                        "cited_doi": "",
                        "cited_url": "https://example.com/a",
                        "cited_raw_entry": "Alpha raw",
                        "match_confidence": 1.0,
                        "match_method": "exact",
                        "warnings": "",
                        "cited_abstract": "",
                        "cited_summary": "",
                        "research_purpose": "",
                    }
                ],
                "total_citations_found": 1,
                "total_bib_entries": 1,
                "matched_count": 1,
                "unmatched_count": 0,
            },
        )

        result = self.service.merge_processing_job_results(
            job_id=job_id,
            import_id="import123",
            documents=[
                {
                    "filename": "alpha.md",
                    "repository_path": "documents/import123/alpha.md",
                    "sha256": "abcdef1234567890",
                }
            ],
        )

        self.assertEqual(result["accepted_new"], 1)
        state = json.loads((repo_dir / ".ra_repo" / "repository_state.json").read_text(encoding="utf-8"))
        self.assertEqual(len(state["sources"]), 1)
        self.assertEqual(state["sources"][0]["fetch_status"], "queued")
        self.assertEqual(state["sources"][0]["original_url"], "https://example.com/a")
        self.assertEqual(state["sources"][0]["title"], "Alpha")
        self.assertEqual(len(state["citations"]), 1)
        self.assertEqual(state["citations"][0]["repository_source_id"], "000001")
        self.assertIn("document:abcdef123456:alpha.md", state["citations"][0]["provenance_ref"])

        citations_csv = (repo_dir / "citations.csv").read_text(encoding="utf-8-sig")
        self.assertIn("Alpha sentence", citations_csv)

        standardized_markdown = repo_dir / "documents" / "import123" / "alpha.standardized.md"
        self.assertTrue(standardized_markdown.exists())
        standardized_text = standardized_markdown.read_text(encoding="utf-8")
        self.assertIn("# Policy Memo", standardized_text)
        self.assertIn("Alpha claim is important. [1]", standardized_text)
        self.assertIn("## Works Cited", standardized_text)
        self.assertIn("[Source](https://example.com/a)", standardized_text)

    def test_merge_repositories_preserves_existing_ids_and_appends_new_rows(self):
        dest_dir = self.tmp_path / "repo_dest"
        dest_dir.mkdir(parents=True, exist_ok=True)
        self.service.attach(str(dest_dir))
        self.service.import_source_list(
            filename="dest.csv",
            content=("URL\nhttps://example.com/a\nhttps://example.com/b\n").encode("utf-8"),
        )

        dest_state_path = dest_dir / ".ra_repo" / "repository_state.json"
        dest_state = json.loads(dest_state_path.read_text(encoding="utf-8"))
        dest_state["sources"][0]["fetch_status"] = "queued"
        dest_state["sources"][1]["fetch_status"] = "queued"
        dest_state_path.write_text(json.dumps(dest_state, ensure_ascii=False, indent=2), encoding="utf-8")

        other_store = FileStore(base_dir=self.tmp_path / "other_app_data")
        other_service = AttachedRepositoryService(store=other_store)
        src_dir = self.tmp_path / "repo_src"
        src_dir.mkdir(parents=True, exist_ok=True)
        other_service.attach(str(src_dir))
        other_service.import_source_list(
            filename="src.csv",
            content=("URL\nhttps://example.com/a\nhttps://example.com/c\n").encode("utf-8"),
        )

        src_state_path = src_dir / ".ra_repo" / "repository_state.json"
        src_state = json.loads(src_state_path.read_text(encoding="utf-8"))
        src_state["sources"][0]["fetch_status"] = "success"
        src_state["sources"][0]["markdown_file"] = "sources/000001/000001_clean.md"
        src_state["sources"][0]["markdown_char_count"] = 42
        src_state["sources"][1]["fetch_status"] = "success"
        src_state["sources"][1]["markdown_file"] = "sources/000002/000002_clean.md"
        src_state["sources"][1]["markdown_char_count"] = 24
        src_state["citations"] = [
            {
                "repository_source_id": "000001",
                "import_type": "document_process",
                "imported_at": "2026-01-01T00:00:00+00:00",
                "provenance_ref": "document:abc:alpha.md",
                "source_document": "alpha.md",
                "page_in_source": "1",
                "citing_sentence": "Sentence",
                "citing_paragraph": "Paragraph",
                "context_before": "",
                "context_after": "",
                "citation_raw": "[1]",
                "citation_ref_numbers": "1",
                "cited_authors": "Author",
                "cited_title": "Title",
                "cited_year": "2024",
                "cited_source": "Journal",
                "cited_volume": "",
                "cited_issue": "",
                "cited_pages": "",
                "cited_doi": "",
                "cited_url": "https://example.com/a",
                "cited_raw_entry": "Raw",
                "match_confidence": 1.0,
                "match_method": "exact",
                "warnings": "",
                "cited_abstract": "",
                "cited_summary": "",
                "research_purpose": "",
            }
        ]
        src_state_path.write_text(json.dumps(src_state, ensure_ascii=False, indent=2), encoding="utf-8")

        (src_dir / "sources" / "000001").mkdir(parents=True, exist_ok=True)
        (src_dir / "sources" / "000002").mkdir(parents=True, exist_ok=True)
        (src_dir / "sources" / "000001" / "000001_clean.md").write_text("# A\n", encoding="utf-8")
        (src_dir / "sources" / "000002" / "000002_clean.md").write_text("# C\n", encoding="utf-8")

        result = self.service._merge_repositories([str(src_dir)])

        self.assertEqual(result.sources_merged, 1)
        merged_state = json.loads(dest_state_path.read_text(encoding="utf-8"))
        merged_sources = {row["original_url"]: row for row in merged_state["sources"]}
        self.assertEqual(merged_sources["https://example.com/a"]["id"], "000001")
        self.assertEqual(merged_sources["https://example.com/c"]["id"], "000003")
        self.assertEqual(merged_sources["https://example.com/a"]["fetch_status"], "success")
        self.assertTrue((dest_dir / merged_sources["https://example.com/a"]["markdown_file"]).exists())
        self.assertEqual(merged_state["citations"][0]["repository_source_id"], "000001")

    def test_reprocess_replaces_target_document_rows_and_keeps_public_citations_csv(self):
        repo_dir = self._attach_repo("repo_reprocess_replace")
        import_id = "import123"
        stored_document = self._write_repo_document(
            repo_dir,
            import_id,
            "report.md",
            b"# Report\n\nAlpha claim.\n",
        )

        existing_sources = [
            SourceManifestRow(
                id="000001",
                repository_source_id="000001",
                original_url="https://example.com/a",
                title="",
                source_document_name="",
                fetch_status="success",
            ),
            SourceManifestRow(
                id="000002",
                repository_source_id="000002",
                original_url="https://example.com/b",
                title="Existing Title",
                fetch_status="success",
            ),
        ]
        existing_citations = [
            ExportRow(
                repository_source_id="000001",
                import_type="document_process",
                imported_at="2026-01-01T00:00:00+00:00",
                provenance_ref=f"document:{stored_document['sha256'][:12]}:report.md",
                source_document="report.md",
                citing_sentence="Old extracted sentence",
                citation_ref_numbers="1",
                cited_title="Old Alpha",
                cited_url="https://example.com/a",
            ),
            ExportRow(
                repository_source_id="000002",
                import_type="document_process",
                imported_at="2026-01-01T00:00:00+00:00",
                provenance_ref="import123:report.md",
                source_document="report.md",
                citing_sentence="Old placeholder sentence",
                citation_ref_numbers="2",
                cited_title="Old Beta",
                cited_url="https://example.com/b",
            ),
            ExportRow(
                repository_source_id="000002",
                import_type="document_process",
                imported_at="2026-01-01T00:00:00+00:00",
                provenance_ref="import999:other.md",
                source_document="other.md",
                citing_sentence="Keep me",
                citation_ref_numbers="9",
                cited_title="Other",
                cited_url="https://example.com/b",
            ),
        ]
        imports = [
            {
                "import_id": import_id,
                "import_type": "document_process",
                "imported_at": "2026-01-01T00:00:00+00:00",
                "documents": [stored_document],
            }
        ]
        self._save_repo_state(
            sources=existing_sources,
            citations=existing_citations,
            imports=imports,
        )

        job_id, job_store = self._create_repo_job()
        documents = [
            {
                **stored_document,
                "filename": "import123__report.md",
            }
        ]
        job_store.save_artifact(
            job_id,
            "03_bibliography",
            {
                "entries": [
                    {
                        "ref_number": 1,
                        "authors": ["Alpha Author"],
                        "title": "Filled Title",
                        "year": "2024",
                        "journal_or_source": "Journal",
                        "volume": "",
                        "issue": "",
                        "pages": "",
                        "doi": "",
                        "url": "https://example.com/a",
                        "raw_text": "Alpha entry",
                        "parse_confidence": 0.9,
                        "parse_warnings": [],
                        "source_document_name": "import123__report.md",
                    },
                    {
                        "ref_number": 2,
                        "authors": ["Beta Author"],
                        "title": "Should Not Replace Existing",
                        "year": "2024",
                        "journal_or_source": "Journal",
                        "volume": "",
                        "issue": "",
                        "pages": "",
                        "doi": "",
                        "url": "https://example.com/b",
                        "raw_text": "Beta entry",
                        "parse_confidence": 0.9,
                        "parse_warnings": [],
                        "source_document_name": "import123__report.md",
                    },
                    {
                        "ref_number": 3,
                        "authors": ["Gamma Author"],
                        "title": "New Source",
                        "year": "2025",
                        "journal_or_source": "Journal",
                        "volume": "",
                        "issue": "",
                        "pages": "",
                        "doi": "",
                        "url": "https://example.com/c",
                        "raw_text": "Gamma entry",
                        "parse_confidence": 0.9,
                        "parse_warnings": [],
                        "source_document_name": "import123__report.md",
                    },
                ]
            },
        )
        job_store.save_artifact(
            job_id,
            "05_export",
            {
                "rows": [
                    {
                        "source_document": "import123__report.md",
                        "citing_sentence": "New alpha sentence",
                        "citation_ref_numbers": "1",
                        "cited_title": "Filled Title",
                        "cited_url": "https://example.com/a",
                    },
                    {
                        "source_document": "import123__report.md",
                        "citing_sentence": "New gamma sentence",
                        "citation_ref_numbers": "3",
                        "cited_title": "New Source",
                        "cited_url": "https://example.com/c",
                    },
                ]
            },
        )
        job_store.save_artifact(job_id, "01_ingestion", {"documents": []})

        with patch.object(
            self.service,
            "_write_standardized_markdown_documents",
            return_value=[
                self._normalization_output(
                    filename="report.md",
                    source_document_path=stored_document["repository_path"],
                    status="normalized",
                )
            ],
        ):
            result = self.service.merge_reprocessed_documents_results(
                job_id=job_id,
                reprocess_id="reprocess123",
                target_import_ids=[import_id],
                documents=documents,
            )

        self.assertEqual(result["accepted_new"], 1)
        state = json.loads(
            (repo_dir / ".ra_repo" / "repository_state.json").read_text(encoding="utf-8")
        )
        sources_by_url = {row["original_url"]: row for row in state["sources"]}
        self.assertEqual(sources_by_url["https://example.com/a"]["id"], "000001")
        self.assertEqual(sources_by_url["https://example.com/a"]["title"], "Filled Title")
        self.assertEqual(sources_by_url["https://example.com/b"]["title"], "Existing Title")
        self.assertEqual(sources_by_url["https://example.com/c"]["id"], "000003")
        self.assertEqual(sources_by_url["https://example.com/c"]["fetch_status"], "queued")

        report_rows = [
            row
            for row in state["citations"]
            if row.get("document_repository_path") == stored_document["repository_path"]
        ]
        self.assertEqual(len(report_rows), 2)
        self.assertEqual(
            sorted(row["repository_source_id"] for row in report_rows),
            ["000001", "000003"],
        )
        self.assertTrue(all(row["document_import_id"] == import_id for row in report_rows))
        self.assertNotIn("Old extracted sentence", json.dumps(report_rows))
        self.assertNotIn("Old placeholder sentence", json.dumps(report_rows))
        self.assertIn("Keep me", json.dumps(state["citations"]))

        citations_csv = (repo_dir / "citations.csv").read_text(encoding="utf-8-sig")
        self.assertNotIn("document_repository_path", citations_csv.splitlines()[0])
        self.assertIn("New alpha sentence", citations_csv)

    def test_reprocess_preserves_failed_document_rows_and_replaces_successful_ones(self):
        repo_dir = self._attach_repo("repo_reprocess_partial")
        first_document = self._write_repo_document(
            repo_dir,
            "import123",
            "first.md",
            b"# First\n\nAlpha.\n",
        )
        second_document = self._write_repo_document(
            repo_dir,
            "import456",
            "second.md",
            b"# Second\n\nBeta.\n",
        )
        sources = [
            SourceManifestRow(
                id="000001",
                repository_source_id="000001",
                original_url="https://example.com/a",
                title="Alpha",
                fetch_status="success",
            ),
            SourceManifestRow(
                id="000002",
                repository_source_id="000002",
                original_url="https://example.com/b",
                title="Beta",
                fetch_status="success",
            ),
        ]
        citations = [
            ExportRow(
                repository_source_id="000001",
                provenance_ref=f"document:{first_document['sha256'][:12]}:first.md",
                source_document="first.md",
                citing_sentence="Old first sentence",
                cited_url="https://example.com/a",
            ),
            ExportRow(
                repository_source_id="000002",
                provenance_ref=f"document:{second_document['sha256'][:12]}:second.md",
                source_document="second.md",
                citing_sentence="Old second sentence",
                cited_url="https://example.com/b",
            ),
        ]
        self._save_repo_state(
            sources=sources,
            citations=citations,
            imports=[
                {
                    "import_id": "import123",
                    "import_type": "document_process",
                    "imported_at": "2026-01-01T00:00:00+00:00",
                    "documents": [first_document],
                },
                {
                    "import_id": "import456",
                    "import_type": "document_process",
                    "imported_at": "2026-01-01T00:00:00+00:00",
                    "documents": [second_document],
                },
            ],
        )

        job_id, job_store = self._create_repo_job()
        documents = [
            {**first_document, "filename": "import123__first.md"},
            {**second_document, "filename": "import456__second.md"},
        ]
        job_store.save_artifact(
            job_id,
            "03_bibliography",
            {
                "entries": [
                    {
                        "ref_number": 1,
                        "authors": ["Alpha Author"],
                        "title": "Alpha",
                        "year": "2024",
                        "journal_or_source": "Journal",
                        "volume": "",
                        "issue": "",
                        "pages": "",
                        "doi": "",
                        "url": "https://example.com/a",
                        "raw_text": "Alpha",
                        "parse_confidence": 0.9,
                        "parse_warnings": [],
                        "source_document_name": "import123__first.md",
                    }
                ]
            },
        )
        job_store.save_artifact(
            job_id,
            "05_export",
            {
                "rows": [
                    {
                        "source_document": "import123__first.md",
                        "citing_sentence": "New first sentence",
                        "cited_url": "https://example.com/a",
                    }
                ]
            },
        )
        job_store.save_artifact(job_id, "01_ingestion", {"documents": []})

        with patch.object(
            self.service,
            "_write_standardized_markdown_documents",
            return_value=[
                self._normalization_output(
                    filename="first.md",
                    source_document_path=first_document["repository_path"],
                    status="partial",
                ),
                self._normalization_output(
                    filename="second.md",
                    source_document_path=second_document["repository_path"],
                    status="failed",
                ),
            ],
        ):
            result = self.service.merge_reprocessed_documents_results(
                job_id=job_id,
                reprocess_id="reprocess456",
                target_import_ids=["import123", "import456"],
                documents=documents,
            )

        self.assertEqual(result["replaced_documents"], 1)
        self.assertEqual(result["preserved_failed_documents"], 1)
        state = json.loads(
            (repo_dir / ".ra_repo" / "repository_state.json").read_text(encoding="utf-8")
        )
        first_rows = [
            row
            for row in state["citations"]
            if row.get("document_repository_path") == first_document["repository_path"]
        ]
        self.assertEqual(len(first_rows), 1)
        self.assertEqual(first_rows[0]["citing_sentence"], "New first sentence")
        self.assertIn(
            "Old second sentence",
            json.dumps(state["citations"]),
        )

    def test_reprocess_can_run_twice_without_duplicate_rows(self):
        repo_dir = self._attach_repo("repo_reprocess_repeat")
        stored_document = self._write_repo_document(
            repo_dir,
            "import123",
            "report.md",
            b"# Report\n\nAlpha.\n",
        )
        self._save_repo_state(
            sources=[
                SourceManifestRow(
                    id="000001",
                    repository_source_id="000001",
                    original_url="https://example.com/a",
                    title="Alpha",
                    fetch_status="success",
                )
            ],
            citations=[
                ExportRow(
                    repository_source_id="000001",
                    provenance_ref=f"document:{stored_document['sha256'][:12]}:report.md",
                    source_document="report.md",
                    citing_sentence="Legacy row",
                    cited_url="https://example.com/a",
                )
            ],
            imports=[
                {
                    "import_id": "import123",
                    "import_type": "document_process",
                    "imported_at": "2026-01-01T00:00:00+00:00",
                    "documents": [stored_document],
                }
            ],
        )
        documents = [{**stored_document, "filename": "import123__report.md"}]

        def run_merge(job_suffix: str, sentence: str) -> None:
            job_id, job_store = self._create_repo_job()
            job_store.save_artifact(
                job_id,
                "03_bibliography",
                {
                    "entries": [
                        {
                            "ref_number": 1,
                            "authors": ["Alpha Author"],
                            "title": "Alpha",
                            "year": "2024",
                            "journal_or_source": "Journal",
                            "volume": "",
                            "issue": "",
                            "pages": "",
                            "doi": "",
                            "url": "https://example.com/a",
                            "raw_text": "Alpha",
                            "parse_confidence": 0.9,
                            "parse_warnings": [],
                            "source_document_name": "import123__report.md",
                        }
                    ]
                },
            )
            job_store.save_artifact(
                job_id,
                "05_export",
                {
                    "rows": [
                        {
                            "source_document": "import123__report.md",
                            "citing_sentence": sentence,
                            "cited_url": "https://example.com/a",
                        }
                    ]
                },
            )
            job_store.save_artifact(job_id, "01_ingestion", {"documents": []})
            with patch.object(
                self.service,
                "_write_standardized_markdown_documents",
                return_value=[
                    self._normalization_output(
                        filename="report.md",
                        source_document_path=stored_document["repository_path"],
                        status="normalized",
                    )
                ],
            ):
                self.service.merge_reprocessed_documents_results(
                    job_id=job_id,
                    reprocess_id=f"reprocess-{job_suffix}",
                    target_import_ids=["import123"],
                    documents=documents,
                )

        run_merge("one", "First rerun sentence")
        run_merge("two", "Second rerun sentence")

        state = json.loads(
            (repo_dir / ".ra_repo" / "repository_state.json").read_text(encoding="utf-8")
        )
        report_rows = [
            row
            for row in state["citations"]
            if row.get("document_repository_path") == stored_document["repository_path"]
        ]
        self.assertEqual(len(report_rows), 1)
        self.assertEqual(report_rows[0]["citing_sentence"], "Second rerun sentence")

    def test_clear_citations_removes_repository_rows_and_resets_csv(self):
        repo_dir = self._attach_repo("repo_clear_citations")
        self._save_repo_state(
            sources=[
                SourceManifestRow(
                    id="000001",
                    repository_source_id="000001",
                    original_url="https://example.com/a",
                    fetch_status="queued",
                )
            ],
            citations=[
                ExportRow(
                    repository_source_id="000001",
                    citation_raw="[1]",
                    citation_ref_numbers="1",
                    citing_paragraph="Alpha paragraph",
                    cited_url="https://example.com/a",
                    cited_raw_entry="Alpha raw entry",
                )
            ],
            imports=[],
        )

        response = self.service.clear_citations()

        self.assertEqual(response.status, "completed")
        self.assertEqual(response.total_citations, 0)
        state = json.loads(
            (repo_dir / ".ra_repo" / "repository_state.json").read_text(encoding="utf-8")
        )
        self.assertEqual(state["citations"], [])
        citations_text = (repo_dir / "citations.csv").read_text(encoding="utf-8-sig")
        self.assertEqual(citations_text.strip(), ",".join(EXPORT_COLUMNS))


if __name__ == "__main__":
    unittest.main()
